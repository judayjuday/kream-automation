"""
판매 입찰 디버그 v2 - 체크박스 문제 해결
"""
import asyncio
import os
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

PARTNER_URL = "https://partner.kream.co.kr"
STATE_FILE = "auth_state.json"
EXCEL_PATH = "kream_data_template.xlsx"
SCREENSHOT_DIR = "debug_screenshots"

os.makedirs(SCREENSHOT_DIR, exist_ok=True)
step_num = 0

async def ss(page, name):
    global step_num
    step_num += 1
    fname = f"{SCREENSHOT_DIR}/{step_num:02d}_{name}.png"
    await page.screenshot(path=fname, full_page=False)
    print(f"  📸 [{step_num}] {name}")
    return fname


async def main():
    global step_num
    step_num = 0

    # 엑셀 데이터
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["입찰데이터"]
    headers = [cell.value for cell in ws[1]]
    bid = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            bid = dict(zip(headers, row))
            break
    wb.close()

    product_id = str(bid["product_id"])
    price = str(int(bid.get("입찰가격", 0)))
    qty = int(bid.get("수량", 1))
    print(f"📋 입찰: #{product_id}, {price}원, {qty}개")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            channel="chrome", headless=False,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        context = await browser.new_context(
            storage_state=STATE_FILE if Path(STATE_FILE).exists() else None,
            viewport={"width": 1440, "height": 900},
            locale="ko-KR",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await context.new_page()
        await Stealth().apply_stealth_async(page)

        # 로그인 확인
        await page.goto(f"{PARTNER_URL}/c2c")
        await page.wait_for_timeout(3000)
        if "/sign-in" in page.url:
            print("❌ 로그인 필요!")
            await browser.close()
            return
        print("✅ 로그인 OK")

        # 팝업 닫기
        for _ in range(5):
            await page.wait_for_timeout(500)
            try:
                cb = page.locator('text="다시 보지 않기"')
                if await cb.is_visible(timeout=300): await cb.click()
            except: pass
            try:
                btn = page.locator('button:has-text("확인")').first
                if await btn.is_visible(timeout=300): await btn.click(); continue
            except: pass
            try:
                x = page.locator('button:has-text("✕"), button:has-text("×")').first
                if await x.is_visible(timeout=300): await x.click(); continue
            except: pass
            break

        # ── 1. 상품 검색 ──
        print("\n[1] 상품 검색...")
        await page.goto(
            f"{PARTNER_URL}/business/products?page=1&perPage=10"
            f"&startDate=&endDate=&keyword={product_id}"
            f"&categoryId=&brandId=&productId=&sort="
        )
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)

        try:
            all_tab = page.locator('button:has-text("전체")').first
            if await all_tab.is_visible(timeout=1000):
                await all_tab.click()
                await page.wait_for_timeout(1500)
        except: pass

        # ── 2. "판매 입찰하기" 버튼 클릭 ──
        print("[2] '판매 입찰하기' 클릭...")
        bid_buttons = page.locator('button:has-text("판매 입찰하기")')
        if await bid_buttons.count() == 0:
            print("  ❌ 버튼 없음")
            await browser.close()
            return

        await bid_buttons.first.click()
        await page.wait_for_timeout(2000)

        # 고시정보 팝업 체크
        try:
            if await page.locator('text="상품 정보 입력 필요"').is_visible(timeout=1000):
                print("  ⚠ 고시정보 필요")
                await browser.close()
                return
        except: pass

        # ── 3. 옵션/수량 모달 ──
        print("[3] 옵션/수량 모달...")
        try:
            await page.locator('text="옵션/수량 선택"').wait_for(timeout=5000)
        except:
            print("  ❌ 모달 안 열림")
            await browser.close()
            return

        plus_btn = page.locator('[class*="Counter_plus"]')
        if await plus_btn.count() == 0:
            plus_btn = page.locator('button:has-text("+")').first
        for i in range(qty):
            await plus_btn.first.click()
            await page.wait_for_timeout(300)

        # ── 4. "판매 입찰 계속" ──
        print("[4] '판매 입찰 계속' 클릭...")
        await page.locator('button:has-text("판매 입찰 계속")').click(timeout=3000)
        await page.wait_for_timeout(2000)
        await ss(page, "price_input_page")

        # ── 5. 가격 입력 ──
        print("[5] 가격 입력...")
        price_inputs = page.locator('input[placeholder*="판매 희망가"]')
        if await price_inputs.count() == 0:
            price_inputs = page.locator('input[placeholder*="희망가"]')

        for i in range(await price_inputs.count()):
            inp = price_inputs.nth(i)
            await inp.click()
            await page.keyboard.press("Meta+a")
            await page.keyboard.press("Backspace")
            await inp.type(price, delay=50)
            await page.wait_for_timeout(300)
        print(f"  ✅ {price}원 입력")

        # 가격 입력 후 blur 이벤트 발생 (React에서 값 반영)
        await page.keyboard.press("Tab")
        await page.wait_for_timeout(1000)
        await ss(page, "after_price_tab")

        # ── 6. 체크박스 선택 (핵심 수정) ──
        print("\n[6] 체크박스 선택 (다양한 방법 시도)...")

        # 먼저 DOM 구조 파악
        cb_info = await page.evaluate("""() => {
            const results = [];
            // 모든 input[type=checkbox] 찾기
            const checkboxes = document.querySelectorAll('input[type="checkbox"]');
            checkboxes.forEach((cb, i) => {
                const rect = cb.getBoundingClientRect();
                const parent = cb.parentElement;
                const grandparent = parent ? parent.parentElement : null;
                results.push({
                    index: i,
                    checked: cb.checked,
                    visible: rect.width > 0 && rect.height > 0,
                    rect: { x: rect.x, y: rect.y, w: rect.width, h: rect.height },
                    parentTag: parent ? parent.tagName : null,
                    parentClass: parent ? parent.className.substring(0, 80) : null,
                    grandparentTag: grandparent ? grandparent.tagName : null,
                    grandparentClass: grandparent ? grandparent.className.substring(0, 80) : null,
                    id: cb.id,
                    name: cb.name,
                });
            });
            // BaseCheckbox도 확인
            const baseCheckboxes = document.querySelectorAll('[data-sentry-element="BaseCheckbox"]');
            baseCheckboxes.forEach((bc, i) => {
                const rect = bc.getBoundingClientRect();
                const innerInput = bc.querySelector('input[type="checkbox"]');
                results.push({
                    index: 'base_' + i,
                    isBaseCheckbox: true,
                    rect: { x: rect.x, y: rect.y, w: rect.width, h: rect.height },
                    hasInnerInput: !!innerInput,
                    innerChecked: innerInput ? innerInput.checked : null,
                    tag: bc.tagName,
                    className: bc.className.substring(0, 80),
                });
            });
            return results;
        }""")

        print("  DOM 분석:")
        for info in cb_info:
            print(f"    {info}")

        # 방법 A: label/wrapper 클릭 (체크박스의 부모 요소 클릭)
        print("\n  → 방법 A: 체크박스 부모(label) 클릭...")
        checkboxes = page.locator('input[type="checkbox"]')
        cb_count = await checkboxes.count()

        # 헤더 체크박스(전체 선택) 시도 - 보통 첫 번째
        for i in range(cb_count):
            cb = checkboxes.nth(i)
            is_visible = await cb.is_visible()
            if not is_visible:
                continue

            # 부모 label 클릭 시도
            try:
                parent = page.locator(f'input[type="checkbox"] >> nth={i} >> xpath=..')
                await parent.click(timeout=2000)
                await page.wait_for_timeout(500)
                checked = await cb.is_checked()
                print(f"    checkbox[{i}] 부모 클릭 → checked={checked}")
                if checked:
                    print(f"    ✅ 방법 A 성공!")
                    break
            except Exception as e:
                print(f"    checkbox[{i}] 부모 클릭 실패: {e}")

        await ss(page, "after_method_A")

        # 상태 확인
        any_checked = False
        for i in range(cb_count):
            try:
                checked = await checkboxes.nth(i).is_checked()
                if checked:
                    any_checked = True
            except: pass

        if not any_checked:
            print("\n  → 방법 B: force click 직접 시도...")
            for i in range(cb_count):
                cb = checkboxes.nth(i)
                try:
                    await cb.click(force=True)
                    await page.wait_for_timeout(500)
                    checked = await cb.is_checked()
                    print(f"    checkbox[{i}] force click → checked={checked}")
                except Exception as e:
                    print(f"    checkbox[{i}] force click 실패: {e}")
            await ss(page, "after_method_B")

        # 다시 확인
        any_checked = False
        for i in range(cb_count):
            try:
                checked = await checkboxes.nth(i).is_checked()
                if checked:
                    any_checked = True
            except: pass

        if not any_checked:
            print("\n  → 방법 C: dispatchEvent + React fiber...")
            await page.evaluate("""() => {
                const checkboxes = document.querySelectorAll('input[type="checkbox"]');
                checkboxes.forEach(cb => {
                    // React 16+에서는 fiber를 통해 onChange를 직접 호출해야 함
                    const key = Object.keys(cb).find(k => k.startsWith('__reactFiber$') || k.startsWith('__reactInternalInstance$'));
                    if (key) {
                        const fiber = cb[key];
                        // React fiber tree를 탐색하며 onChange 핸들러 찾기
                        let current = fiber;
                        for (let i = 0; i < 10 && current; i++) {
                            if (current.memoizedProps && current.memoizedProps.onChange) {
                                const event = { target: { checked: true }, preventDefault: () => {}, stopPropagation: () => {} };
                                current.memoizedProps.onChange(event);
                                break;
                            }
                            current = current.return;
                        }
                    }
                });
            }""")
            await page.wait_for_timeout(1000)
            await ss(page, "after_method_C")

            for i in range(cb_count):
                try:
                    checked = await checkboxes.nth(i).is_checked()
                    print(f"    checkbox[{i}] 방법C 후 checked={checked}")
                    if checked: any_checked = True
                except: pass

        if not any_checked:
            print("\n  → 방법 D: 좌표 기반 클릭 (체크박스 영역)...")
            # 스크린샷에서 체크박스는 ONE SIZE 행의 맨 왼쪽
            # 헤더 체크박스를 먼저 시도 (전체 선택)
            for i in range(cb_count):
                cb = checkboxes.nth(i)
                try:
                    box = await cb.bounding_box()
                    if box and box['width'] > 0:
                        # 체크박스 중앙 클릭
                        x = box['x'] + box['width'] / 2
                        y = box['y'] + box['height'] / 2
                        print(f"    checkbox[{i}] 좌표 ({x:.0f}, {y:.0f}) 클릭...")
                        await page.mouse.click(x, y)
                        await page.wait_for_timeout(500)
                        checked = await cb.is_checked()
                        print(f"    → checked={checked}")
                        if checked:
                            any_checked = True
                except Exception as e:
                    print(f"    checkbox[{i}] 좌표 클릭 실패: {e}")

            await ss(page, "after_method_D")

        if not any_checked:
            print("\n  → 방법 E: BaseCheckbox 요소 직접 클릭...")
            base_cbs = page.locator('[data-sentry-element="BaseCheckbox"]')
            bcnt = await base_cbs.count()
            for i in range(bcnt):
                try:
                    await base_cbs.nth(i).click(force=True)
                    await page.wait_for_timeout(500)
                    print(f"    BaseCheckbox[{i}] 클릭")
                except Exception as e:
                    print(f"    BaseCheckbox[{i}] 실패: {e}")
            await ss(page, "after_method_E")

            # 상태 확인
            for i in range(cb_count):
                try:
                    checked = await checkboxes.nth(i).is_checked()
                    print(f"    checkbox[{i}] checked={checked}")
                    if checked: any_checked = True
                except: pass

        # 총 N건 확인
        await page.wait_for_timeout(500)
        try:
            total_text = await page.locator('text=/총.*건/').first.inner_text()
            print(f"\n  📊 {total_text}")
        except:
            print("\n  총 N건 텍스트 못 찾음")

        await ss(page, "before_final_button")

        # ── 7. "판매 입찰하기" 최종 버튼 ──
        print("\n[7] '판매 입찰하기' 최종 클릭...")
        final_btn = page.locator('button:has-text("판매 입찰하기")').last
        is_enabled = await final_btn.is_enabled()
        print(f"  enabled={is_enabled}")

        await final_btn.click(timeout=3000)
        await page.wait_for_timeout(2000)
        await ss(page, "after_final_click")

        # ── 8. 확인 팝업 ──
        print("\n[8] 확인 팝업 대기...")
        await page.wait_for_timeout(1000)

        # 페이지 텍스트에서 "총" 관련 확인
        try:
            body_text = await page.locator('body').inner_text()
            for line in body_text.split('\n'):
                line = line.strip()
                if '총' in line and '건' in line:
                    print(f"  텍스트: '{line}'")
                if '옵션' in line and '선택' in line:
                    print(f"  에러: '{line}'")
        except: pass

        await ss(page, "popup_state")

        # 확인 버튼 찾기
        confirm_btns = page.locator('button:has-text("확인")')
        ccnt = await confirm_btns.count()
        print(f"  '확인' 버튼: {ccnt}개")

        if ccnt > 0:
            for i in range(ccnt):
                vis = await confirm_btns.nth(i).is_visible()
                if vis:
                    await confirm_btns.nth(i).click()
                    print(f"  ✅ 확인 클릭")
                    break
            await page.wait_for_timeout(3000)
            await ss(page, "after_confirm")

        # ── 9. 결과 확인 ──
        print("\n[9] 결과...")
        try:
            await page.locator('text="입찰 신청 결과"').wait_for(timeout=5000)
            result_text = await page.locator('text=/성공.*건/').first.inner_text()
            print(f"  🎉 {result_text}")
        except:
            print("  ⚠ 결과 팝업 없음")

        await ss(page, "result")

        # ── 10. asks 페이지 확인 ──
        print("\n[10] 입찰 내역 확인...")
        await page.goto(f"{PARTNER_URL}/business/asks")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        await ss(page, "asks_page")

        pcnt = await page.locator('text="78,000"').count()
        print(f"  78,000원: {pcnt}건")
        if pcnt > 0:
            print("  🎉🎉🎉 성공!")

        await context.storage_state(path=STATE_FILE)
        await page.wait_for_timeout(5000)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())
