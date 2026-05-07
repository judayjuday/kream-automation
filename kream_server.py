"""
KREAM 판매자센터 자동화 - Flask 백엔드 서버
- 대시보드 UI 서빙
- 가격 수집 / 입찰 등록 / 고시정보 등록 API
- Playwright 기반 자동화를 비동기 호출

실행: python3 kream_server.py
접속: http://localhost:5001
"""

import asyncio
import json
import logging
import os
import math
import random
import re
import sqlite3
import smtplib
import subprocess
import tempfile
import threading
import time
import traceback
import urllib.request
import urllib.error
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

from flask import Flask, request, jsonify, send_file, send_from_directory, Response
from werkzeug.utils import secure_filename
import openpyxl

# ── Step 18-D: APScheduler (일일 자동화 + 운영 가시성) ──
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    scheduler = BackgroundScheduler(timezone='Asia/Seoul')
    _APSCHEDULER_AVAILABLE = True
except Exception as _ape:
    print(f"[SCHEDULER] APScheduler import 실패: {_ape}")
    scheduler = None
    _APSCHEDULER_AVAILABLE = False

# ── 기존 모듈 import ──
from kream_collector import collect_prices
from kream_adjuster import full_adjust_flow, modify_bid_price
from kream_bot import (
    create_browser, create_context, apply_stealth,
    ensure_logged_in, fill_product_info, place_bid, place_bids_batch, dismiss_popups,
    save_state_with_localstorage, collect_shipments,
    STATE_FILE, PARTNER_URL, KREAM_URL,
)
from playwright.async_api import async_playwright
from health_alert import HealthAlert
from kream_hubnet_bot import (
    ensure_hubnet_logged_in,
    hubnet_login,
    save_hubnet_session,
    fetch_hubnet_orders,
    upsert_hubnet_orders,
    match_all_unmatched,
    download_invoice_pdf,
    download_pending_invoices,
)

app = Flask(__name__)
logger = logging.getLogger(__name__)

# ── 경보 시스템 ──
health_alerter = HealthAlert()

BASE_DIR = Path(__file__).parent
HISTORY_FILE = BASE_DIR / "execution_history.json"
BATCH_HISTORY_FILE = BASE_DIR / "batch_history.json"
SETTINGS_FILE = BASE_DIR / "settings.json"
DISCOVERY_FILE = BASE_DIR / "kream_discovery_data.xlsx"
MY_BIDS_FILE = BASE_DIR / "my_bids_local.json"
QUEUE_FILE = BASE_DIR / "queue_data.json"

# ── 실행 상태 관리 ──
tasks = {}  # task_id → { status, logs, result }
task_counter = 0
task_lock = threading.Lock()

# ── 자동 입찰 제어 ──
auto_bid_control = {"state": "idle"}  # idle, running, paused, stopping
auto_bid_lock = threading.Lock()
auto_bid_event = threading.Event()
auto_bid_event.set()  # 초기 상태: 실행 가능

# ── 상품 큐 ──
product_queue = []  # [{id, model, cny, category, ...}, ...]
queue_counter = 0
queue_lock = threading.Lock()


def save_queue():
    """product_queue를 queue_data.json에 저장"""
    try:
        QUEUE_FILE.write_text(json.dumps(
            {"counter": queue_counter, "queue": product_queue},
            ensure_ascii=False, indent=2
        ))
    except Exception:
        pass


def load_queue():
    """서버 시작 시 queue_data.json 복원"""
    global product_queue, queue_counter
    if not QUEUE_FILE.exists():
        return
    try:
        data = json.loads(QUEUE_FILE.read_text())
        product_queue = data.get("queue", [])
        queue_counter = data.get("counter", len(product_queue))
        print(f"[큐 복원] {len(product_queue)}건 로드됨")
    except Exception as e:
        print(f"[큐 복원 실패] {e}")


load_queue()

# ── 입찰 순위 모니터링 ──
PRICE_DB = BASE_DIR / "price_history.db"
MONITOR_HOURS = [8, 10, 12, 14, 16, 18, 20, 22]
EMAIL_SENDER = "judaykream@gmail.com"
EMAIL_RECEIVER = "judaykream@gmail.com"

monitor_state = {
    "running": False,
    "last_run": None,
    "next_run": None,
    "total_checks": 0,
    "total_adjustments": 0,
}
_monitor_timer = None
_monitor_lock = threading.Lock()

# ── 자동 백업 스케줄러 (운영안정화 §1.1) ──
backup_state = {
    "running": False,
    "last_run": None,
    "last_status": None,
}
_backup_timer = None
BACKUP_INTERVAL_SEC = 24 * 3600  # 24시간 주기


def _backup_run_once():
    """backup_db.sh 1회 실행. 다른 스케줄러에 영향 안 가도록 try/except 격리."""
    try:
        script = BASE_DIR / "backup_db.sh"
        if not script.exists():
            backup_state["last_status"] = f"script not found: {script}"
            print(f"[backup] {backup_state['last_status']}")
            return
        result = subprocess.run(
            ["/bin/bash", str(script)],
            capture_output=True, text=True, timeout=180,
            cwd=str(BASE_DIR),
        )
        backup_state["last_run"] = datetime.now().isoformat()
        if result.returncode == 0:
            backup_state["last_status"] = "success"
            tail = (result.stdout or "").strip().splitlines()
            print(f"[backup] OK {tail[-1] if tail else ''}")
        else:
            backup_state["last_status"] = f"failed rc={result.returncode}"
            print(
                f"[backup] FAIL rc={result.returncode} "
                f"stderr={(result.stderr or '')[:200]}"
            )
            try:
                health_alerter.alert(
                    "backup_failed",
                    f"백업 실패 rc={result.returncode}: {(result.stderr or '')[:200]}",
                )
            except Exception:
                pass
    except subprocess.TimeoutExpired:
        backup_state["last_status"] = "timeout"
        print("[backup] TIMEOUT (180s 초과)")
    except Exception as e:
        backup_state["last_status"] = f"exception: {e}"
        print(f"[backup] exception: {e}")


def _backup_tick():
    """24h 주기 트리거. 실행 후 다음 트리거 재등록."""
    global _backup_timer
    if not backup_state.get("running"):
        return
    try:
        _backup_run_once()
    except Exception as e:
        print(f"[backup] tick exception: {e}")
    if backup_state.get("running"):
        _backup_timer = threading.Timer(BACKUP_INTERVAL_SEC, _backup_tick)
        _backup_timer.daemon = True
        _backup_timer.start()


# ── 세션 사전 갱신 스케줄러 (Step 17-D Phase 2-B) ──
_session_refresh_lock = threading.Lock()
_session_refresh_thread = None
_session_refresh_stop = threading.Event()
_session_refresh_status = {
    "enabled": True,
    "last_run": None,
    "last_result": None,
    "next_run": None,
    "interval_hours": 12,
    "trigger_threshold_hours": 18,
}


def _refresh_session_if_stale(target):
    """target='partner'|'kream'|'hubnet'. 18h 초과 + 토큰 valid 시 사전 재로그인."""
    state_paths = {
        "partner": BASE_DIR / "auth_state.json",
        "kream": BASE_DIR / "auth_state_kream.json",
        "hubnet": BASE_DIR / "auth_state_hubnet.json",
    }
    path = state_paths.get(target)
    if not path or not path.exists():
        return {"target": target, "action": "skip", "success": False,
                "message": "state file not found"}

    age_hours = (time.time() - path.stat().st_mtime) / 3600
    threshold = _session_refresh_status["trigger_threshold_hours"]
    if age_hours < threshold:
        return {"target": target, "action": "skip", "success": True,
                "message": f"still fresh (age={age_hours:.1f}h)"}

    try:
        if target == "hubnet":
            sess = ensure_hubnet_logged_in()
            cookie_count = len(sess.cookies) if sess is not None else 0
            return {"target": target, "action": "refreshed", "success": True,
                    "message": f"cookies={cookie_count}"}
        elif target == "partner":
            async def _do():
                from kream_bot import login_auto_partner
                async with async_playwright() as p:
                    return await login_auto_partner(p)
            asyncio.run(_do())
            return {"target": target, "action": "refreshed", "success": True,
                    "message": "partner re-logged in"}
        elif target == "kream":
            return {"target": target, "action": "skip", "success": True,
                    "message": "kream auto-relogin not implemented"}
    except Exception as e:
        msg = str(e)[:200]
        try:
            safe_send_alert(
                subject=f"[KREAM] 세션 사전 갱신 실패: {target}",
                body=f"target={target}\nmessage={msg}",
                alert_type="auth_failure",
            )
        except Exception as _ae:
            print(f"[session_refresh] safe_send_alert 실패: {_ae}")
        return {"target": target, "action": "failed", "success": False,
                "message": msg}


def _session_refresh_worker():
    """백그라운드 스레드. 12h 주기로 partner/hubnet 사전 갱신 점검."""
    while not _session_refresh_stop.is_set():
        try:
            with _session_refresh_lock:
                if _session_refresh_status["enabled"]:
                    results = []
                    for target in ("partner", "hubnet"):
                        results.append(_refresh_session_if_stale(target))
                    _session_refresh_status["last_run"] = datetime.now().isoformat()
                    _session_refresh_status["last_result"] = results

                    failures = [r for r in results if r.get("action") == "failed"]
                    if failures:
                        try:
                            health_alerter.alert(
                                "session_refresh_failed",
                                f"세션 사전 갱신 실패: {failures}",
                            )
                        except Exception as ae:
                            print(f"[session_refresh] alert 실패: {ae}")

            interval_sec = _session_refresh_status["interval_hours"] * 3600
            _session_refresh_status["next_run"] = (
                datetime.now() + timedelta(seconds=interval_sec)
            ).isoformat()

            for _ in range(int(interval_sec)):
                if _session_refresh_stop.is_set():
                    break
                time.sleep(1)
        except Exception as e:
            print(f"[session_refresh] worker exception: {e}")
            time.sleep(60)


def start_session_refresh_scheduler():
    """서버 시작 시 호출. 이미 실행 중이면 no-op."""
    global _session_refresh_thread
    if _session_refresh_thread is not None and _session_refresh_thread.is_alive():
        return
    _session_refresh_stop.clear()
    _session_refresh_thread = threading.Thread(
        target=_session_refresh_worker,
        daemon=True,
        name="session_refresh_scheduler",
    )
    _session_refresh_thread.start()
    print("[session_refresh] 사전 갱신 스케줄러 시작 (12h 주기, 18h 임계)")


# ── SQLite WAL 모드 활성화 ──
def _enable_wal_mode():
    """price_history.db를 WAL 모드로 변경 (동시 읽기/쓰기 성능 향상)"""
    conn = sqlite3.connect(str(PRICE_DB))
    before = conn.execute("PRAGMA journal_mode;").fetchone()[0]
    conn.execute("PRAGMA journal_mode=WAL;")
    after = conn.execute("PRAGMA journal_mode;").fetchone()[0]
    conn.close()
    print(f"[DB] journal_mode: {before} → {after}")

_enable_wal_mode()


def _init_adjustments_table():
    """price_adjustments 테이블 생성"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS price_adjustments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        product_id TEXT,
        model TEXT,
        name_kr TEXT,
        size TEXT,
        old_price INTEGER,
        competitor_price INTEGER,
        new_price INTEGER,
        expected_profit INTEGER,
        status TEXT DEFAULT 'pending',
        created_at TEXT NOT NULL,
        executed_at TEXT
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_pa_status ON price_adjustments(status)")
    conn.commit()
    conn.close()


_init_adjustments_table()


def _init_bid_cost_table():
    """bid_cost 테이블 생성 — 입찰 시점의 원가 정보 보관"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS bid_cost (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT UNIQUE,
        model TEXT,
        size TEXT,
        cny_price REAL,
        exchange_rate REAL,
        overseas_shipping INTEGER DEFAULT 8000,
        other_costs INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bc_model ON bid_cost(model)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bc_model_size ON bid_cost(model, size)")

    # ── Step 16-A: cny_source 컬럼 추가 (idempotent, 트랜잭션 보장) ──
    cols = [r[1] for r in c.execute("PRAGMA table_info(bid_cost)").fetchall()]
    if "cny_source" not in cols:
        c.execute("BEGIN")
        try:
            c.execute("ALTER TABLE bid_cost ADD COLUMN cny_source TEXT")
            c.execute("UPDATE bid_cost SET cny_source='unknown' WHERE cny_source IS NULL")
            conn.commit()
        except Exception:
            conn.rollback()
            raise
    else:
        conn.commit()
    conn.close()


_init_bid_cost_table()


# ── bid_cleanup_log 테이블 ──
def _init_bid_cleanup_table():
    """입찰 정리 이력 테이블"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS bid_cleanup_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT, model TEXT, size TEXT, price INTEGER,
        cleanup_type TEXT,
        reason TEXT,
        status TEXT,
        detected_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        scheduled_delete_at DATETIME,
        executed_at DATETIME,
        cancel_reason TEXT,
        snapshot TEXT
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cleanup_status ON bid_cleanup_log(status, scheduled_delete_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cleanup_order ON bid_cleanup_log(order_id, status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cleanup_detected ON bid_cleanup_log(detected_at)")
    conn.commit()
    conn.close()


_init_bid_cleanup_table()


# ═══════════════════════════════════════════
# Step 17-A: 카테고리 판정 + 사이즈 유효성 검증
# ═══════════════════════════════════════════

_ONE_SIZE_TOKENS = ("ONE SIZE", "ONESIZE", "ONE_SIZE", "FREE", "OS")


# Step 17-E: model_category DB의 영문 카테고리 → 큐 시스템 한글 카테고리 매핑
DB_CATEGORY_TO_KR = {
    "shoes": "신발",
    "bags": "가방",
    "clothing": "의류",
    "unknown": "",  # 미결정 명시
}


def map_db_category_to_kr(db_category):
    """DB의 영문 카테고리를 큐 시스템의 한글로 변환. unknown/None은 빈 문자열."""
    if not db_category:
        return ""
    return DB_CATEGORY_TO_KR.get(str(db_category).lower(), "")


def get_model_category(model):
    """모델의 카테고리/사이즈 필수 여부 반환.

    우선순위: model_category 캐시 → shihuo_prices(active=1) 직접 조회 → 보수적 디폴트.
    반환: {"category": str, "needs_size": bool, "source": str}
    """
    if not model:
        return {"category": "unknown", "needs_size": True, "source": "default"}

    conn = sqlite3.connect(str(PRICE_DB))
    try:
        row = conn.execute(
            "SELECT category, needs_size, source FROM model_category WHERE model=?",
            (model,)
        ).fetchone()
        if row:
            return {"category": row[0], "needs_size": bool(row[1]), "source": row[2]}

        row = conn.execute(
            "SELECT category FROM shihuo_prices WHERE active=1 AND model=? LIMIT 1",
            (model,)
        ).fetchone()
        if row:
            cat = row[0] or "unknown"
            needs = 0 if cat == "bags" else 1
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO model_category (model, category, source, needs_size, notes) VALUES (?,?,?,?,?)",
                    (model, cat, "shihuo", needs, "shihuo_prices 활성 batch 추론")
                )
                conn.commit()
            except Exception:
                pass
            return {"category": cat, "needs_size": bool(needs), "source": "shihuo"}
    finally:
        conn.close()

    # 보수적 디폴트: 사이즈 필수
    return {"category": "unknown", "needs_size": True, "source": "default"}


def validate_size_for_bid(model, size, raise_on_error=False):
    """입찰 전 사이즈 유효성 검증.

    needs_size=1 카테고리에서 size가 빈값/ONE SIZE/FREE → 차단.
    needs_size=0(가방) 카테고리는 통과.

    반환: (is_valid, error_msg, category_info)
    """
    cat_info = get_model_category(model)
    size_clean = (size or "").strip().upper()

    if cat_info["needs_size"]:
        if not size_clean or size_clean in _ONE_SIZE_TOKENS:
            msg = (f"카테고리 '{cat_info['category']}'은(는) 사이즈 필수입니다 "
                   f"(model={model}, size='{size}')")
            if raise_on_error:
                raise ValueError(msg)
            return (False, msg, cat_info)

    return (True, None, cat_info)


def validate_category_for_bid(model, category=None):
    """Step 17-E: 입찰 전 카테고리 유효성 검증.

    차단 조건:
    - category가 None / 빈문자열 / "미분류" → 차단
    - DB 조회 결과와 다르면 logger.warning만 (DB가 진실, 큐 execute에서 0순위로 적용됨)

    반환: (is_valid: bool, error_msg: str)
    """
    if not category or str(category).strip() in ("", "미분류"):
        return (False, f"카테고리 미결정 입찰 차단 (model={model}, category='{category}')")

    if model:
        try:
            db_info = get_model_category(model)
            db_kr = map_db_category_to_kr(db_info.get("category", ""))
            if db_kr and db_kr != category:
                logger.warning(
                    "category mismatch model=%s db=%s current=%s",
                    model, db_kr, category
                )
        except Exception as e:
            logger.warning("category DB lookup failed model=%s err=%s", model, e)

    return (True, "")


def validate_gosi_for_bid(gosi):
    """Step 17-E 보완: 입찰 전 gosi(고시정보) 결정 검증.

    차단 조건:
    - gosi가 None / 빈 dict
    - gosi.type이 None / 빈 문자열 / 공백만

    반환: (is_valid: bool, error_msg: str)
    """
    if not gosi:
        return (False, "고시정보 미결정 (auto_fill_gosi가 None 또는 빈 dict 반환)")
    if not isinstance(gosi, dict):
        return (False, f"고시정보 형식 오류 (type={type(gosi).__name__})")
    gosi_type = gosi.get("type", "")
    if not gosi_type or not str(gosi_type).strip():
        return (False, "고시정보 type 미결정")
    return (True, "")


def _save_bid_cost(order_id, model, size, cny_price, exchange_rate,
                   overseas_shipping=8000, other_costs=0,
                   cny_source=None):
    """입찰 시점의 원가 저장 (UPSERT by order_id).

    cny_price가 None/0이면 shihuo_prices(active=1) 매칭으로 자동 채택 시도.
    매칭 키: model 정확 일치 + CAST(size AS INTEGER) = kream_mm.
    매칭 실패 + manual 없음 → 저장 스킵 (가짜 값 채우기 금지).
    """
    if not order_id:
        return None
    # Step 17-A: size 빈값/None 차단 (ONE SIZE 디폴트 제거 후 명시 강제)
    if size is None or not str(size).strip():
        raise ValueError(
            f"size required for bid_cost (order_id={order_id}, model={model})"
        )

    resolved_cny = float(cny_price) if cny_price and float(cny_price) > 0 else None
    resolved_source = cny_source

    # manual 명시 입력이 우선 — 식货 절대 덮어쓰지 않음 (절대 규칙 #3)
    if resolved_cny is not None:
        if not resolved_source:
            resolved_source = "manual"

    rate_f = float(exchange_rate) if exchange_rate else 0.0

    conn = sqlite3.connect(str(PRICE_DB))
    try:
        cur = conn.cursor()

        # manual 미지정 → shihuo 자동 채택 시도
        if resolved_cny is None:
            try:
                size_int = int(str(size).strip())
            except (ValueError, TypeError):
                size_int = None

            if model and size_int is not None:
                row = cur.execute(
                    """SELECT cny_price FROM shihuo_prices
                       WHERE active=1 AND model=? AND kream_mm=?
                       ORDER BY imported_at DESC LIMIT 1""",
                    (model, size_int)
                ).fetchone()
                if row and row[0]:
                    resolved_cny = float(row[0])
                    resolved_source = "shihuo"

        if resolved_cny is None:
            # 매칭 실패 + manual 없음 → 저장 스킵 (절대 규칙 #4)
            print(f"[bid_cost] 스킵: order_id={order_id} model={model} size={size} — cny_price 없음 + 식货 매칭 실패")
            return None

        if not resolved_source:
            resolved_source = "unknown"

        cur.execute(
            """INSERT INTO bid_cost (order_id, model, size, cny_price, exchange_rate,
                  overseas_shipping, other_costs, cny_source)
               VALUES (?,?,?,?,?,?,?,?)
               ON CONFLICT(order_id) DO UPDATE SET
                 cny_price=excluded.cny_price,
                 exchange_rate=excluded.exchange_rate,
                 overseas_shipping=excluded.overseas_shipping,
                 other_costs=excluded.other_costs,
                 cny_source=excluded.cny_source""",
            (order_id, model or "", size or "", resolved_cny, rate_f,
             int(overseas_shipping), int(other_costs), resolved_source)
        )
        conn.commit()
        return {"cny_price": resolved_cny, "cny_source": resolved_source}
    finally:
        conn.close()


def _init_auto_adjust_log_table():
    """auto_adjust_log 테이블 — 자동 실행 이력"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS auto_adjust_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        model TEXT,
        size TEXT,
        old_price INTEGER,
        new_price INTEGER,
        expected_profit INTEGER,
        action TEXT,
        skip_reason TEXT,
        modify_result TEXT,
        executed_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_auto_adjust_executed ON auto_adjust_log(executed_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_auto_adjust_order ON auto_adjust_log(order_id, executed_at)")
    conn.commit()
    conn.close()


_init_auto_adjust_log_table()


def _init_auto_rebid_log_table():
    """auto_rebid_log 테이블 — 자동 재입찰 이력
    action: auto_rebid_success | skipped_no_cost | skipped_loop_guard |
            skipped_margin_low | skipped_price_shift | skipped_blacklist |
            skipped_daily_limit | skipped_disabled | rebid_failed
    """
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS auto_rebid_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        original_order_id TEXT,
        model TEXT,
        size TEXT,
        sold_price INTEGER,
        new_bid_price INTEGER,
        expected_profit INTEGER,
        action TEXT,
        skip_reason TEXT,
        new_order_id TEXT,
        executed_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_rebid_executed ON auto_rebid_log(executed_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_rebid_model_size ON auto_rebid_log(model, size, executed_at)")
    conn.commit()
    conn.close()


_init_auto_rebid_log_table()


def _init_model_price_book_table():
    """model_price_book — 모델별 마스터 단가표 (Step 37).

    size=NULL 레코드는 해당 모델 모든 사이즈에 동일 단가 적용 (가방류 등).
    UNIQUE(model, size) 제약, INSERT OR IGNORE로 시드 안전.
    """
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS model_price_book (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        model TEXT NOT NULL,
        size TEXT,
        cny_price REAL NOT NULL,
        category TEXT,
        brand TEXT,
        is_bulk_item INTEGER DEFAULT 0,
        notes TEXT,
        source TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(model, size)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_pb_model ON model_price_book(model)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_pb_bulk ON model_price_book(is_bulk_item)")

    # 시드 데이터 (사장님 제공) — NULL-safe 존재 체크 후 INSERT.
    # SQLite UNIQUE 제약은 NULL을 서로 다른 값으로 취급하므로
    # INSERT OR IGNORE만으로는 size=NULL 시드의 중복을 막지 못함.
    seeds = [
        ("IX7693", "ONE SIZE", 205, "가방", None, 1, "대량 구매 모델", "사장님 직접 입력"),
        ("IX7694", "ONE SIZE", 205, "가방", None, 1, "대량 구매 모델", "사장님 직접 입력"),
        ("JQ4110", None, 370, "가방", None, 1,
         "대량 구매 모델 / W215~W255 전 사이즈 동일", "사장님 직접 입력"),
        # Step 40: 1203A243 시리즈 (bid_cost 21건 실데이터 검증값)
        # 사장님 회상값 385 대신 bid_cost 일치값 380/375 채택
        ("1203A243-021", "225", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-021", "230", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-021", "235", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-021", "240", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-021", "245", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-021", "250", 375, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 250사이즈는 375", "bid_cost 검증"),
        ("1203A243-100", "225", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-100", "230", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 누락 사이즈 신규", "bid_cost 검증"),
        ("1203A243-100", "235", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-100", "240", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-100", "245", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
        ("1203A243-100", "250", 380, "신발", "Onitsuka Tiger", 1,
         "대량 구매 / bid_cost 21건 평균", "bid_cost 검증"),
    ]
    for seed in seeds:
        model, size = seed[0], seed[1]
        if size is None:
            cnt = c.execute(
                "SELECT COUNT(*) FROM model_price_book WHERE model = ? AND size IS NULL",
                (model,),
            ).fetchone()[0]
        else:
            cnt = c.execute(
                "SELECT COUNT(*) FROM model_price_book WHERE model = ? AND size = ?",
                (model, size),
            ).fetchone()[0]
        if cnt == 0:
            c.execute(
                """INSERT INTO model_price_book
                   (model, size, cny_price, category, brand, is_bulk_item, notes, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                seed,
            )
    conn.commit()
    conn.close()


_init_model_price_book_table()


# ── 조건부 입찰 (conditional_bids) DB ──
def _init_conditional_bids_table():
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS conditional_bids (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id TEXT NOT NULL,
        model TEXT,
        size TEXT DEFAULT 'ONE SIZE',
        condition_type TEXT NOT NULL,
        condition_value INTEGER NOT NULL,
        bid_price INTEGER NOT NULL,
        status TEXT DEFAULT 'active',
        created_at TEXT NOT NULL,
        triggered_at TEXT
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cb_status ON conditional_bids(status)")
    conn.commit()
    conn.close()


_init_conditional_bids_table()


# ── 수정 이력 (edit_log) DB ──
def _init_edit_log_table():
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS edit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_type TEXT NOT NULL,
        item_id TEXT NOT NULL,
        field_name TEXT NOT NULL,
        old_value TEXT,
        new_value TEXT,
        edited_at TEXT NOT NULL
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_el_edited ON edit_log(edited_at)")
    conn.commit()
    conn.close()


_init_edit_log_table()


def save_edit_log(item_type, item_id, field_name, old_value, new_value):
    """수정 이력 저장"""
    if str(old_value) == str(new_value):
        return
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        "INSERT INTO edit_log (item_type, item_id, field_name, old_value, new_value, edited_at) VALUES (?,?,?,?,?,?)",
        (item_type, str(item_id), field_name, str(old_value), str(new_value),
         datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()


# ── 물류 관리 (logistics) DB ──
def _init_logistics_tables():
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    # 협력사
    c.execute("""CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        contact TEXT,
        phone TEXT,
        wechat TEXT,
        notes TEXT,
        created_at TEXT NOT NULL
    )""")
    # 발송 요청
    c.execute("""CREATE TABLE IF NOT EXISTS shipment_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        product_id TEXT,
        model TEXT,
        size TEXT,
        supplier_id INTEGER,
        hubnet_hbl TEXT,
        request_date TEXT,
        tracking_number TEXT,
        status TEXT DEFAULT '발송대기',
        proof_image TEXT,
        notes TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sr_status ON shipment_requests(status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sr_order ON shipment_requests(order_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sr_hbl ON shipment_requests(hubnet_hbl)")
    # 물류 비용
    c.execute("""CREATE TABLE IF NOT EXISTS shipment_costs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shipment_id INTEGER NOT NULL,
        cost_type TEXT NOT NULL,
        amount REAL NOT NULL,
        currency TEXT DEFAULT 'KRW',
        notes TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY (shipment_id) REFERENCES shipment_requests(id)
    )""")
    conn.commit()
    conn.close()


_init_logistics_tables()


# ── 판매 이력 (sales_history) DB ──
def _init_sales_history_table():
    """sales_history 테이블 생성"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS sales_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT UNIQUE,
        product_id TEXT,
        model TEXT,
        product_info TEXT,
        size TEXT,
        sale_price INTEGER,
        trade_date TEXT,
        ship_date TEXT,
        ship_status TEXT,
        collected_at TEXT NOT NULL
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sh_order ON sales_history(order_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sh_model ON sales_history(model)")
    conn.commit()
    conn.close()


_init_sales_history_table()


# ── 알림 센터 DB ──
def _init_notifications_table():
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL,
        title TEXT NOT NULL,
        message TEXT,
        action_url TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        dismissed INTEGER DEFAULT 0
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_notif_read ON notifications(is_read)")
    # 기존 DB에 dismissed 컬럼 없으면 추가 (idempotent)
    try:
        cols = [r[1] for r in c.execute("PRAGMA table_info(notifications)").fetchall()]
        if "dismissed" not in cols:
            c.execute("ALTER TABLE notifications ADD COLUMN dismissed INTEGER DEFAULT 0")
    except Exception as _e:
        print(f"[notifications] dismissed migrate fail: {_e}")
    conn.commit()
    conn.close()


_init_notifications_table()


def add_notification(ntype, title, message="", action_url=""):
    """알림 추가 (서버 내부에서 호출)"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute(
        "INSERT INTO notifications (type, title, message, action_url, created_at) VALUES (?,?,?,?,?)",
        (ntype, title, message, action_url, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()


_alert_dedup_lock = {}


def _should_send_alert_dedupe(subject, body, window_sec=60):
    """동일 알림 디바운싱: window_sec 내 동일 (subject, body) 조합은 차단."""
    import time
    key = (str(subject)[:200], str(body)[:500])
    now = time.time()
    last = _alert_dedup_lock.get(key, 0)
    if now - last < window_sec:
        return False
    _alert_dedup_lock[key] = now
    if len(_alert_dedup_lock) > 500:
        cutoff = now - 3600
        for k in list(_alert_dedup_lock.keys()):
            if _alert_dedup_lock[k] < cutoff:
                del _alert_dedup_lock[k]
    return True


def safe_send_alert(subject, body, alert_type='info'):
    """알림 발송 안전 wrapper. 실패해도 서버 동작에 영향 X.
    1) DB notifications 테이블에 누적, 2) 이메일 시도, 두 단계 모두 try/except.
    기존 notifications 스키마(title/message)를 그대로 사용 (subject→title, body→message).
    """
    if not _should_send_alert_dedupe(subject, body):
        return
    import sys
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute(
            "INSERT INTO notifications (type, title, message, created_at) VALUES (?,?,?,?)",
            (alert_type, str(subject)[:500], str(body)[:4000],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[ALERT-DB-FAIL] {e}", file=sys.stderr)

    try:
        # 기존 health_alerter(HealthAlert) 우선 사용, 없으면 send_alert 시도
        try:
            health_alerter.alert(alert_type, f"{subject}\n{body}")
        except Exception:
            from health_alert import send_alert  # type: ignore
            send_alert(subject=subject, body=body)
    except ImportError:
        print(f"[ALERT-EMAIL] (health_alert 없음) {subject}: {str(body)[:200]}", file=sys.stderr)
    except Exception as e:
        print(f"[ALERT-EMAIL-FAIL] {e}", file=sys.stderr)

    # Discord 4채널 발송 (이메일 실패와 무관하게 시도, 실패는 stderr만)
    try:
        from services.discord_notifier import send_for_alert_type
        send_for_alert_type(alert_type, str(subject)[:256], str(body)[:4000])
    except Exception as e:
        print(f"[ALERT-DISCORD-FAIL] {e}", file=sys.stderr)


@app.route("/api/discord/test", methods=["POST"])
def api_discord_test():
    """Discord 4채널 연결 테스트. 사장님 폰 알림 수신 확인용."""
    try:
        from services.discord_notifier import send_discord, get_loaded_channels, COLORS
    except Exception as e:
        return jsonify({"error": f"discord_notifier import 실패: {e}"}), 500

    loaded = get_loaded_channels()
    plan = {
        "bids":   ("✅ KREAM 입찰 채널 테스트",  "이 채널은 입찰 갱신/가격 자동조정/경쟁자 침입 감지 알림용", COLORS["info"]),
        "sales":  ("💰 KREAM 판매 채널 테스트",  "이 채널은 체결 감지/자동 재입찰 트리거 알림용", COLORS["success"]),
        "errors": ("⚠️ KREAM 에러 채널 테스트",  "이 채널은 헬스체크 critical/sync 멈춤/자동 재로그인/자동 토글 차단 알림용", COLORS["error"]),
        "daily":  ("📊 KREAM 일일 채널 테스트",  "이 채널은 매일 23:55 일일 리포트(매출/마진/체결률) 알림용", COLORS["warn"]),
    }
    results = {}
    for ch, (title, body, color) in plan.items():
        if ch not in loaded:
            results[ch] = "missing_webhook"
            continue
        ok = send_discord(ch, title, body, color=color, dedupe=False)
        results[ch] = "ok" if ok else "fail"
    return jsonify(results)


# ── Step 33-A: 자동 재로그인 인프라 ──
def _check_session_and_relogin():
    """sync 1h+ 멈추면 자동 재로그인. 6h 쿨다운."""
    from datetime import datetime, timedelta
    from pathlib import Path

    state_file = Path(__file__).parent / '.relogin_state.json'
    state = {}
    if state_file.exists():
        try:
            state = json.loads(state_file.read_text())
        except Exception:
            pass

    # 6h 쿨다운
    last_attempt = state.get('last_attempt')
    if last_attempt:
        try:
            last_dt = datetime.fromisoformat(last_attempt)
            if datetime.now() - last_dt < timedelta(hours=6):
                return
        except Exception:
            pass

    # sync 시각 확인
    local_path = Path(__file__).parent / 'my_bids_local.json'
    if not local_path.exists():
        return

    try:
        local = json.loads(local_path.read_text(encoding='utf-8'))
        last_sync = local.get('lastSync') or local.get('last_sync')
        if last_sync:
            try:
                last_sync_dt = datetime.strptime(last_sync, '%Y/%m/%d %H:%M')
            except Exception:
                last_sync_dt = datetime.fromisoformat(last_sync)
            if datetime.now() - last_sync_dt < timedelta(hours=1):
                return
    except Exception:
        return

    print("[AUTO-RELOGIN] 세션 만료 추정 → 자동 재로그인", flush=True)
    state['last_attempt'] = datetime.now().isoformat()
    state_file.write_text(json.dumps(state, indent=2))

    try:
        import subprocess
        result = subprocess.run(
            ['python3', 'kream_bot.py', '--mode', 'auto-login-partner'],
            capture_output=True, text=True, timeout=180,
            cwd=str(Path(__file__).parent)
        )
        if result.returncode == 0:
            print("[AUTO-RELOGIN] ✅ 성공", flush=True)
            state['last_success'] = datetime.now().isoformat()
            state_file.write_text(json.dumps(state, indent=2))
            try:
                safe_send_alert('[KREAM] 자동 재로그인 성공', '세션 만료 → 자동 재로그인 완료', 'auto_relogin_success')
            except Exception:
                pass
        else:
            print(f"[AUTO-RELOGIN] ❌ 실패: {result.stderr[:300]}", flush=True)
            state['last_failure'] = datetime.now().isoformat()
            state['last_failure_reason'] = result.stderr[:500]
            state_file.write_text(json.dumps(state, indent=2))
            try:
                safe_send_alert('[KREAM] 자동 재로그인 실패', f'수동 점검 필요\n\n{result.stderr[:500]}', 'auto_relogin_failure')
            except Exception:
                pass
    except Exception as e:
        print(f"[AUTO-RELOGIN] ❌ 예외: {e}", flush=True)


# 판매 수집 스케줄러 상태
sales_scheduler_state = {
    "running": False,
    "last_run": None,
    "next_run": None,
    "total_syncs": 0,
    "last_new_count": 0,
}
_sales_timer = None
_sales_lock = threading.Lock()


# ── 得物 가격 & 사이즈 변환 DB ──
def _init_dewu_tables():
    """dewu_prices, size_conversion 테이블 생성 + 초기 데이터"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS dewu_prices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        model TEXT NOT NULL,
        brand TEXT,
        eu_size TEXT,
        kr_size TEXT,
        cny_price REAL,
        updated_at TEXT NOT NULL
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_dewu_model ON dewu_prices(model)")

    c.execute("""CREATE TABLE IF NOT EXISTS size_conversion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        brand TEXT NOT NULL,
        eu_size TEXT NOT NULL,
        kr_size TEXT NOT NULL
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_sc_brand ON size_conversion(brand)")

    # 초기 데이터 삽입 (테이블이 비어있을 때만)
    c.execute("SELECT COUNT(*) FROM size_conversion")
    if c.fetchone()[0] == 0:
        _SIZE_MAP = {
            "onitsuka": {
                "EU36": "225", "EU37": "230", "EU37.5": "235", "EU38": "240",
                "EU39": "245", "EU39.5": "250", "EU40": "252.5", "EU40.5": "255",
                "EU41.5": "260", "EU42": "265", "EU42.5": "270", "EU43.5": "275",
                "EU44": "280", "EU44.5": "282.5", "EU45": "285", "EU46": "290",
            },
            "newbalance": {
                "EU35.5": "215", "EU36": "220", "EU37": "225", "EU37.5": "230",
                "EU38": "235", "EU38.5": "240", "EU39.5": "245", "EU40": "250",
                "EU40.5": "255", "EU41.5": "260", "EU42": "265", "EU42.5": "270",
                "EU43": "275", "EU44": "280", "EU44.5": "285", "EU45": "290",
            },
            "mizuno": {
                "EU36": "225", "EU36.5": "230", "EU37": "235", "EU38": "240",
                "EU38.5": "245", "EU39": "250", "EU40": "255", "EU40.5": "260",
                "EU41": "265", "EU42": "270", "EU42.5": "275", "EU43": "280",
                "EU44": "285", "EU44.5": "290", "EU45": "295",
            },
        }
        rows = []
        for brand, sizes in _SIZE_MAP.items():
            for eu, kr in sizes.items():
                rows.append((brand, eu, kr))
        c.executemany("INSERT INTO size_conversion (brand, eu_size, kr_size) VALUES (?, ?, ?)", rows)
        print(f"[DB] size_conversion 초기 데이터 {len(rows)}건 삽입")

    c.execute("SELECT COUNT(*) FROM dewu_prices")
    if c.fetchone()[0] == 0:
        _PRODUCTS = [
            {"model": "1183B480-250", "brand": "onitsuka", "dewu_prices": {
                "EU36": 532, "EU37": 544, "EU37.5": 565, "EU38": 565,
                "EU39": 499, "EU39.5": 502, "EU40": 484, "EU40.5": 480,
                "EU41.5": 479, "EU42": 498, "EU42.5": 491, "EU43.5": 516,
                "EU44": 505, "EU44.5": 514, "EU45": 561, "EU46": 574,
            }},
            {"model": "M1906AD", "brand": "newbalance", "dewu_prices": {
                "EU36": 768, "EU37": 768, "EU37.5": 838, "EU38": 843,
                "EU38.5": 886, "EU39.5": 820, "EU40": 847, "EU40.5": 860,
                "EU41.5": 834, "EU42": 829, "EU42.5": 847, "EU43": 894,
                "EU44": 805, "EU44.5": 918, "EU45": 997,
            }},
            {"model": "M1906AG", "brand": "newbalance", "dewu_prices": {
                "EU36": 1018, "EU37": 959, "EU37.5": 1014, "EU38": 1022,
                "EU38.5": 1048, "EU39.5": 931, "EU40": 919, "EU40.5": 949,
                "EU41.5": 857, "EU42": 838, "EU42.5": 879, "EU43": 853,
                "EU44": 857, "EU44.5": 1141, "EU45": 1029,
            }},
            {"model": "1183B799-101", "brand": "onitsuka", "dewu_prices": {
                "EU36": 485, "EU37": 422, "EU37.5": 434, "EU38": 437,
                "EU39": 423, "EU39.5": 438, "EU40": 441, "EU40.5": 438,
                "EU41.5": 429, "EU42": 497, "EU42.5": 599, "EU43.5": 494,
                "EU44": 482, "EU44.5": 476, "EU45": 548, "EU46": 534,
            }},
            {"model": "1203A714-020", "brand": "onitsuka", "dewu_prices": {
                "EU37": 1067, "EU39": 1058, "EU39.5": 940, "EU40": 538,
                "EU40.5": 529, "EU41.5": 530, "EU42": 530, "EU42.5": 1422,
                "EU43.5": 538, "EU44": 699,
            }},
            {"model": "D1GH241906", "brand": "mizuno", "dewu_prices": {
                "EU36": 798, "EU36.5": 760, "EU37": 649, "EU38": 649,
                "EU38.5": 680, "EU39": 666, "EU40": 488, "EU40.5": 488,
                "EU41": 488, "EU42": 488, "EU42.5": 488, "EU43": 488,
                "EU44": 488, "EU44.5": 488,
            }},
        ]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = []
        # 사이즈 변환표 로드
        sc_map = {}
        c.execute("SELECT brand, eu_size, kr_size FROM size_conversion")
        for b, eu, kr in c.fetchall():
            sc_map[(b, eu)] = kr
        for p in _PRODUCTS:
            for eu, cny in p["dewu_prices"].items():
                kr = sc_map.get((p["brand"], eu), "")
                rows.append((p["model"], p["brand"], eu, kr, cny, now))
        c.executemany(
            "INSERT INTO dewu_prices (model, brand, eu_size, kr_size, cny_price, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            rows
        )
        print(f"[DB] dewu_prices 초기 데이터 {len(rows)}건 삽입")

    conn.commit()
    conn.close()


_init_dewu_tables()


def _init_trade_volume_table():
    """trade_volume 테이블 생성 (주간 거래량 추적용)"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS trade_volume (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id TEXT NOT NULL,
        size TEXT,
        weekly_trades INTEGER DEFAULT 0,
        collected_at TEXT NOT NULL
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_tv_pid ON trade_volume(product_id)")
    conn.commit()
    conn.close()


_init_trade_volume_table()


# ── 탈환률 추적 (bid_competition_log) DB ──
def _init_bid_competition_log():
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS bid_competition_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id TEXT,
        model TEXT,
        size TEXT,
        my_price INTEGER,
        market_lowest INTEGER,
        am_i_lowest BOOLEAN,
        my_margin INTEGER,
        competitor_count INTEGER,
        checked_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bid_comp_model ON bid_competition_log(model, checked_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bid_comp_checked ON bid_competition_log(checked_at)")
    conn.commit()
    conn.close()

_init_bid_competition_log()


def _init_size_charts_tables():
    """Step 12: size_charts + size_conversion_log."""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS size_charts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chart_name TEXT NOT NULL,
        brand TEXT NOT NULL,
        gender TEXT NOT NULL,
        category TEXT NOT NULL DEFAULT 'shoes',
        purchase_country TEXT DEFAULT 'ALL',
        eu_size TEXT NOT NULL,
        us_size TEXT,
        uk_size TEXT,
        kream_mm INTEGER NOT NULL,
        notes TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(brand, gender, category, eu_size, purchase_country)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_size_charts_brand ON size_charts(brand, gender, category)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_size_charts_eu ON size_charts(eu_size)")

    c.execute("""CREATE TABLE IF NOT EXISTS size_conversion_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        brand TEXT NOT NULL,
        model TEXT,
        raw_size TEXT NOT NULL,
        normalized_size TEXT,
        kream_mm INTEGER,
        rule_applied TEXT,
        decision_notes TEXT,
        logged_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_size_log_brand_model ON size_conversion_log(brand, model)")
    conn.commit()
    conn.close()

_init_size_charts_tables()


def _init_shihuo_prices_table():
    """Step 15: 識货 시장가 임포트 결과 저장 테이블."""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS shihuo_prices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id TEXT NOT NULL,
        active INTEGER DEFAULT 1,
        brand_raw TEXT,
        brand_normalized TEXT,
        category TEXT,
        model TEXT NOT NULL,
        color TEXT,
        size_eu TEXT,
        size_normalized TEXT,
        kream_mm INTEGER,
        cny_price REAL NOT NULL,
        supplier TEXT,
        platform TEXT,
        source_created_at DATETIME,
        imported_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        mapping_status TEXT,
        mapping_note TEXT,
        UNIQUE(batch_id, model, size_eu)
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_shihuo_active_model ON shihuo_prices(active, model)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_shihuo_active_kream ON shihuo_prices(active, model, kream_mm)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_shihuo_batch ON shihuo_prices(batch_id)")
    conn.commit()
    conn.close()


_init_shihuo_prices_table()


def calc_customer_total(bid_price, category="신발"):
    """해외배송 고객 총 결제금액 계산.
    고객 결제 = 입찰가 + 배송비(3,000원) + 관부가세(USD $150 초과 시)
    """
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    usd_rate = float(settings.get("usdRate", 1495.76))
    vat_rate = float(settings.get("vatRate", 0.10))

    tariff_rate = 0.08 if category == "가방" else 0.13
    shipping_customer = 3000  # 해외배송 고객 배송비

    usd_value = bid_price / usd_rate
    customs = 0
    import_vat = 0
    if usd_value > 150:
        customs = round(bid_price * tariff_rate)
        import_vat = round((bid_price + customs) * vat_rate)

    total = bid_price + shipping_customer + customs + import_vat
    return {
        "bid_price": bid_price,
        "shipping": shipping_customer,
        "customs": customs,
        "import_vat": import_vat,
        "customer_total": total,
        "usd_value": round(usd_value, 1),
        "over_limit": usd_value > 150,
    }


def analyze_competitiveness(bid_price, category, sdp_entry):
    """사이즈별 입찰 경쟁력 분석.
    sdp_entry: sizeDeliveryPrices의 한 항목 {buyFast, buyNormal, buyOverseas, ...}
    반환: {domestic_min, overseas_min, customer_total, competitiveness, ...}
    """
    buy_fast = sdp_entry.get("buyFast") or 0
    buy_normal = sdp_entry.get("buyNormal") or 0
    buy_overseas = sdp_entry.get("buyOverseas") or 0

    # 국내 최저가 (빠른배송 + 일반배송)
    domestic_prices = [p for p in [buy_fast, buy_normal] if p > 0]
    domestic_min = min(domestic_prices) if domestic_prices else 0

    # 해외 최저가 (다른 해외배송 판매자 = buyOverseas)
    overseas_min = buy_overseas if buy_overseas > 0 else 0

    # 해외배송 고객 총 결제금액 (우리 입찰가 기준)
    ct = calc_customer_total(bid_price, category)
    customer_total = ct["customer_total"]

    # 경쟁력 판단
    if domestic_min == 0:
        competitiveness = "독점 가능"
        comp_color = "green"
        diff = 0
        diff_pct = 0
    elif customer_total <= domestic_min:
        competitiveness = "경쟁력 있음"
        comp_color = "green"
        diff = customer_total - domestic_min
        diff_pct = round(diff / domestic_min * 100, 1)
    elif customer_total <= domestic_min * 1.2:
        competitiveness = "보통"
        comp_color = "yellow"
        diff = customer_total - domestic_min
        diff_pct = round(diff / domestic_min * 100, 1)
    else:
        competitiveness = "경쟁 어려움"
        comp_color = "red"
        diff = customer_total - domestic_min
        diff_pct = round(diff / domestic_min * 100, 1)

    return {
        "domestic_min": domestic_min,
        "domestic_fast": buy_fast,
        "domestic_normal": buy_normal,
        "overseas_min": overseas_min,
        "customer_total": customer_total,
        "customer_shipping": ct["shipping"],
        "customer_customs": ct["customs"],
        "customer_vat": ct["import_vat"],
        "customer_usd": ct["usd_value"],
        "customer_over_limit": ct["over_limit"],
        "competitiveness": competitiveness,
        "comp_color": comp_color,
        "diff_vs_domestic": diff,
        "diff_pct": diff_pct,
    }


def get_dewu_prices(model):
    """DB에서 모델별 得物 가격 조회 → {kr_size: cny_price, ...}"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("SELECT kr_size, cny_price, eu_size, brand FROM dewu_prices WHERE model = ?", (model,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return None
    result = {"sizes": {}, "brand": rows[0][3]}
    for kr, cny, eu, brand in rows:
        key = kr if kr else eu
        result["sizes"][key] = {"cny": cny, "eu_size": eu, "kr_size": kr}
    return result


def classify_market(size_margins_with_dewu):
    """
    시장 경쟁 상태 분류
    - size_margins_with_dewu: [{size, totalCost, instantBuyPrice, ...}, ...]
    - 각 사이즈별로 마진율 계산 → 평균 마진율로 분류
    반환: {market_type, market_color, avg_margin_rate, profitable_count, total_count, details}
    """
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    fee_rate = float(settings.get("feeRate", 0.06))
    fixed_fee = int(settings.get("fixedFee", 2500))
    vat_rate = float(settings.get("vatRate", 0.10))

    margin_rates = []
    profitable = 0
    details = []

    for sz in size_margins_with_dewu:
        sell_price = sz.get("instantBuyPrice") or 0
        total_cost = sz.get("totalCost", 0)
        if not sell_price or not total_cost:
            details.append({"size": sz.get("size", "?"), "margin_rate": None, "profitable": False})
            continue

        # 정산액 계산
        commission = round(sell_price * fee_rate)
        comm_vat = round(commission * vat_rate)
        total_fee = commission + comm_vat + fixed_fee
        settlement = sell_price - total_fee
        margin = settlement - total_cost
        margin_rate = round(margin / total_cost * 100, 1) if total_cost > 0 else 0

        margin_rates.append(margin_rate)
        is_profit = margin_rate >= 0
        if margin_rate >= 10:
            profitable += 1
        details.append({
            "size": sz.get("size", "?"),
            "margin_rate": margin_rate,
            "margin": margin,
            "sell_price": sell_price,
            "total_cost": total_cost,
            "profitable": is_profit,
        })

    if not margin_rates:
        return {
            "market_type": "데이터 부족",
            "market_color": "gray",
            "avg_margin_rate": None,
            "profitable_count": 0,
            "total_count": len(size_margins_with_dewu),
            "details": details,
        }

    avg_rate = round(sum(margin_rates) / len(margin_rates), 1)
    ok_count = sum(1 for r in margin_rates if r >= 0)

    if avg_rate >= 10:
        mtype, mcolor = "정상 시장", "green"
    elif avg_rate >= 0:
        mtype, mcolor = "혼합 시장", "yellow"
    else:
        mtype, mcolor = "비정상 시장", "red"

    return {
        "market_type": mtype,
        "market_color": mcolor,
        "avg_margin_rate": avg_rate,
        "profitable_count": ok_count,
        "total_count": len(margin_rates),
        "details": details,
    }


# ── 환율 캐시 ──
def _load_initial_rates():
    """settings.json에서 마지막으로 저장된 환율 읽기"""
    d = {"cny": 218.12, "usd": 1495.76, "updated_at": None}
    if SETTINGS_FILE.exists():
        try:
            s = json.loads(SETTINGS_FILE.read_text())
            if "cnyRate" in s:
                d["cny"] = float(s["cnyRate"])
            if "usdRate" in s:
                d["usd"] = float(s["usdRate"])
        except Exception:
            pass
    return d


_exchange_rate_cache = _load_initial_rates()
_exchange_rate_lock = threading.Lock()


def fetch_exchange_rates():
    """CNY→KRW, USD→KRW 환율을 외부 API에서 가져와 settings.json에 반영"""
    global _exchange_rate_cache

    cny_rate = None
    usd_rate = None

    # CNY: primary → open.er-api.com, fallback → fawazahmed0 currency-api
    try:
        url = "https://open.er-api.com/v6/latest/CNY"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            cny_rate = round(float(data["rates"]["KRW"]), 2)
        print(f"[환율] CNY 조회 성공 (open.er-api): {cny_rate}")
    except Exception as e:
        print(f"[환율] CNY primary 조회 실패: {e}, fallback 시도...")
        try:
            url = "https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@latest/v1/currencies/cny.json"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read())
                cny_rate = round(float(data["cny"]["krw"]), 2)
            print(f"[환율] CNY 조회 성공 (fallback): {cny_rate}")
        except Exception as e2:
            print(f"[환율] CNY fallback도 실패: {e2}")

    # USD
    try:
        url = "https://open.er-api.com/v6/latest/USD"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            usd_rate = round(float(data["rates"]["KRW"]), 2)
    except Exception as e:
        print(f"[환율] USD 조회 실패: {e}")

    if cny_rate is None and usd_rate is None:
        # 실패 시 기본값 사용
        print("[환율] 모든 API 실패 — 기본값 사용 (CNY=218.12)")
        cny_rate = 218.12
        usd_rate = _exchange_rate_cache.get("usd", 1495.76)

    now = datetime.now().strftime("%m/%d %H:%M")

    with _exchange_rate_lock:
        if cny_rate is not None:
            _exchange_rate_cache["cny"] = cny_rate
        if usd_rate is not None:
            _exchange_rate_cache["usd"] = usd_rate
        _exchange_rate_cache["updated_at"] = now

    try:
        settings = {}
        if SETTINGS_FILE.exists():
            settings = json.loads(SETTINGS_FILE.read_text())
        if cny_rate is not None:
            settings["cnyRate"] = cny_rate
        if usd_rate is not None:
            settings["usdRate"] = usd_rate
        SETTINGS_FILE.write_text(json.dumps(settings, ensure_ascii=False, indent=2))
        print(f"[환율] CNY={_exchange_rate_cache['cny']}, USD={_exchange_rate_cache['usd']} ({now})")
    except Exception as e:
        print(f"[환율] settings 저장 실패: {e}")

    return _exchange_rate_cache.copy()


def get_headless():
    """설정에서 headless 모드 읽기"""
    if SETTINGS_FILE.exists():
        try:
            s = json.loads(SETTINGS_FILE.read_text())
            return s.get("headless", True)
        except Exception:
            pass
    return True


def new_task():
    global task_counter
    with task_lock:
        task_counter += 1
        tid = f"task_{task_counter}"
        tasks[tid] = {"status": "running", "logs": [], "result": None}
        return tid


def add_log(tid, level, msg):
    ts = datetime.now().strftime("%H:%M:%S")
    entry = {"time": ts, "level": level, "msg": msg}
    if tid in tasks:
        tasks[tid]["logs"].append(entry)


def finish_task(tid, result=None, error=None):
    if tid in tasks:
        tasks[tid]["status"] = "error" if error else "done"
        tasks[tid]["result"] = result
        if error:
            add_log(tid, "error", str(error))


# ═══════════════════════════════════════════
# Cloudflare 잔여 코드 자동 정리
# ═══════════════════════════════════════════

def _cleanup_cloudflare():
    """서버 시작 시 HTML 파일에서 Cloudflare 이메일 보호 코드 자동 제거"""
    patterns = [
        (r'<script[^>]*email-decode[^>]*></script>', ''),
        (r'<script[^>]*cfasync[^>]*src="[^"]*cloudflare[^"]*"[^>]*></script>', ''),
        (r'<a[^>]*class="__cf_email__"[^>]*>\[email[^<]*\]</a>', ''),
        (r'<a[^>]*href="/cdn-cgi/l/email-protection"[^>]*>[^<]*</a>', ''),
    ]
    files = [BASE_DIR / "kream_dashboard.html"] + list((BASE_DIR / "tabs").glob("*.html"))
    for fpath in files:
        try:
            content = fpath.read_text(encoding="utf-8")
            changed = False
            for pattern, replacement in patterns:
                new_content = re.sub(pattern, replacement, content)
                if new_content != content:
                    content = new_content
                    changed = True
            if changed:
                fpath.write_text(content, encoding="utf-8")
                print(f"[Cloudflare 정리] {fpath.name}: 잔여 코드 제거됨")
        except Exception as e:
            print(f"[Cloudflare 정리] {fpath.name}: 오류 {e}")


_cleanup_cloudflare()


# ═══════════════════════════════════════════
# 페이지 서빙
# ═══════════════════════════════════════════

@app.route("/")
def index():
    return send_file(BASE_DIR / "kream_dashboard.html")


@app.route("/tabs/<path:filename>")
def serve_tab(filename):
    """탭 HTML 파일 서빙"""
    tab_path = BASE_DIR / "tabs" / filename
    if not tab_path.exists():
        return "Not Found", 404
    return send_file(str(tab_path))


# ═══════════════════════════════════════════
# API: 상품 검색 (가격 수집)
# ═══════════════════════════════════════════

@app.route("/api/search", methods=["POST"])
def api_search():
    """모델번호 또는 상품번호로 KREAM 가격 수집"""
    data = request.json or {}
    product_id = str(data.get("productId", "")).strip()
    model = str(data.get("model", "")).strip()

    if not product_id and not model:
        return jsonify({"error": "productId 또는 model 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"검색 시작: productId={product_id}, model={model}")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            if product_id:
                add_log(tid, "info", f"상품 #{product_id} 가격 수집 중...")
                results = loop.run_until_complete(
                    collect_prices([product_id], headless=get_headless(), include_partner=False)
                )
            else:
                # 모델번호로 검색 → 상품번호 찾기
                add_log(tid, "info", f"모델번호 '{model}' 검색 중...")
                results = loop.run_until_complete(
                    search_by_model(model)
                )

            loop.close()

            if results and len(results) > 0:
                kream = results[0].get("kream", {})
                # 세션 만료 감지
                if kream.get("session_expired"):
                    add_log(tid, "error", f"세션 만료! {kream['session_expired']} 재로그인 필요")
                else:
                    add_log(tid, "success", f"수집 완료: {kream.get('product_name', 'N/A')}")
                finish_task(tid, result=results[0])
            else:
                add_log(tid, "error", "검색 결과 없음")
                finish_task(tid, error="검색 결과 없음")
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    thread = threading.Thread(target=run, daemon=True)
    thread.start()

    return jsonify({"taskId": tid})


async def search_by_model(model: str):
    """모델번호로 KREAM 검색 → 상품번호 찾기 → 가격 수집"""
    from playwright.async_api import async_playwright
    from kream_collector import (
        create_browser, create_context, apply_stealth,
        STATE_FILE_KREAM,
    )

    kream_session = STATE_FILE_KREAM if Path(STATE_FILE_KREAM).exists() else None

    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, kream_session)
        page = await context.new_page()
        await apply_stealth(page)

        # KREAM 검색 페이지
        search_url = f"https://kream.co.kr/search?keyword={model}&tab=products"
        await page.goto(search_url, wait_until="domcontentloaded")
        # 스마트 대기: 첫 검색결과 등장 즉시 진행, 없으면 3초 후 실패 처리
        try:
            await page.wait_for_selector('a[href*="/products/"]', timeout=3000)
        except Exception:
            pass  # 결과 없음 → product_id = None → 빈 배열 반환

        # 첫 번째 검색 결과에서 상품번호 + 브랜드 + 카테고리 추출
        search_info = await page.evaluate(r"""() => {
            const link = document.querySelector('a[href*="/products/"]');
            let productId = null, brand = '';
            if (link) {
                const m = link.href.match(/\/products\/(\d+)/);
                if (m) productId = m[1];
                // 검색 결과 카드에서 브랜드 (첫번째 텍스트 줄)
                const lines = link.innerText.trim().split('\n').map(s => s.trim()).filter(s => s);
                if (lines.length > 0) brand = lines[0];
            }
            // 카테고리
            let category = '';
            const cats = document.querySelectorAll('.category, [class*="category"], [class*="tag"]');
            for (const el of cats) {
                const t = el.innerText.trim();
                if (t) { category = t; break; }
            }
            return { productId, category, brand };
        }""")

        product_id = search_info.get("productId") if search_info else None
        kream_category = search_info.get("category", "") if search_info else ""
        kream_brand = search_info.get("brand", "") if search_info else ""

        if kream_session:
            await save_state_with_localstorage(page, context, STATE_FILE_KREAM, "https://kream.co.kr")
        await browser.close()

    if product_id:
        results = await collect_prices([product_id], headless=get_headless(), include_partner=False)
        # 카테고리/브랜드 정보를 kream 데이터에 주입
        if results:
            if kream_brand:
                results[0].setdefault("kream", {})["brand"] = kream_brand
        if results and kream_category:
            results[0].setdefault("kream", {})["category"] = kream_category
        return results
    return []


# ═══════════════════════════════════════════
# API: 입찰 등록
# ═══════════════════════════════════════════

@app.route("/api/bid", methods=["POST"])
def api_bid():
    """판매 입찰 등록"""
    data = request.json or {}
    product_id = str(data.get("productId", "")).strip()
    price = int(data.get("price", 0))
    size = str(data.get("size", "")).strip()
    qty = int(data.get("quantity", 1))
    model = str(data.get("model", "")).strip()

    if not product_id or not price:
        return jsonify({"error": "productId, price 필요"}), 400

    # Step 17-A: 카테고리별 사이즈 필수 검증 (신발은 ONE SIZE 차단)
    is_valid, err_msg, _cat = validate_size_for_bid(model, size)
    if not is_valid:
        return jsonify({"ok": False, "error": err_msg, "code": "SIZE_REQUIRED"}), 400

    # Step 17-E: 카테고리 미결정 차단
    bid_category = str(data.get("category", "")).strip()
    cat_ok, cat_err = validate_category_for_bid(model, bid_category)
    if not cat_ok:
        return jsonify({"ok": False, "error": cat_err, "code": "CATEGORY_UNDECIDED"}), 400

    tid = new_task()
    add_log(tid, "info", f"입찰 시작: #{product_id} {price:,}원 × {qty}개")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(
                run_bid(product_id, price, size, qty, tid, model=model)
            )
            loop.close()
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return jsonify({"taskId": tid})


async def run_bid(product_id, price, size, qty, tid, model=""):
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page, context):
            add_log(tid, "error", "판매자센터 로그인 필요 (python3 kream_bot.py --mode login)")
            await browser.close()
            return {"success": False, "error": "로그인 필요"}

        bid_data = {
            "product_id": product_id,
            "model": model,
            "사이즈": size,
            "입찰가격": price,
            "수량": qty,
        }

        add_log(tid, "info", f"판매 입찰 등록 중... #{product_id} {price:,}원 × {qty}개")
        success = await place_bid(page, bid_data, delay=2.0)

        # 성공 시에만 세션 저장 (실패 시 빈 세션으로 덮어쓰기 방지)
        if success:
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        await browser.close()

        if success:
            add_log(tid, "success", f"입찰 등록 완료! #{product_id} → {price:,}원")
            save_history("입찰", product_id, price, qty, True)
            save_bid_local(product_id, model="", size=size, price=price, source="placed")
        else:
            add_log(tid, "error", f"입찰 등록 실패: #{product_id}")

        return {"success": success}


# ═══════════════════════════════════════════
# API: 고시정보 등록
# ═══════════════════════════════════════════

@app.route("/api/product-info", methods=["POST"])
def api_product_info():
    """상품 고시정보 등록"""
    data = request.json or {}
    product_id = str(data.get("productId", "")).strip()

    if not product_id:
        return jsonify({"error": "productId 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"고시정보 등록 시작: #{product_id}")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(
                run_product_info(product_id, data, tid)
            )
            loop.close()
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return jsonify({"taskId": tid})


async def run_product_info(product_id, data, tid):
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page, context):
            add_log(tid, "error", "판매자센터 로그인 필요")
            await browser.close()
            return {"success": False, "error": "로그인 필요"}

        gosi = data.get("gosi", {})
        category = gosi.get("category", "가방")
        product_data = build_gosi_data(product_id, gosi, category)

        add_log(tid, "info", f"고시정보 입력 중... #{product_id}")
        await fill_product_info(page, product_data, delay=2.0)

        await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        await browser.close()

        add_log(tid, "success", f"고시정보 등록 완료: #{product_id}")
        return {"success": True}


# ═══════════════════════════════════════════
# API: 고시정보 + 입찰 통합 실행
# ═══════════════════════════════════════════

@app.route("/api/register", methods=["POST"])
def api_register():
    """고시정보 등록 (필요시) + 판매 입찰 통합 실행"""
    data = request.json or {}
    product_id = str(data.get("productId", "")).strip()
    price = int(data.get("price", 0))
    size = str(data.get("size", "")).strip()
    qty = int(data.get("quantity", 1))
    gosi_already = data.get("gosiAlready", False)
    gosi = data.get("gosi", {})

    if not product_id or not price:
        return jsonify({"error": "productId, price 필요"}), 400

    # 원가 정보 (있으면 입찰 성공 시 bid_cost에 저장)
    cny_price = data.get("cny_price", 0)
    exchange_rate = data.get("exchange_rate", 0)
    overseas_shipping = data.get("overseas_shipping", 8000)
    model = str(data.get("model", "")).strip()

    # Step 17-A: 카테고리별 사이즈 필수 검증 (신발은 ONE SIZE 차단)
    is_valid, err_msg, _cat = validate_size_for_bid(model, size)
    if not is_valid:
        return jsonify({"ok": False, "error": err_msg, "code": "SIZE_REQUIRED"}), 400

    # Step 17-E: 카테고리 미결정 차단 (data.category > gosi.category > "")
    bid_category = str(data.get("category", "") or gosi.get("category", "")).strip()
    cat_ok, cat_err = validate_category_for_bid(model, bid_category)
    if not cat_ok:
        return jsonify({"ok": False, "error": cat_err, "code": "CATEGORY_UNDECIDED"}), 400

    # Step 17-E 보완: gosi 미결정 차단 (gosi=None / 빈 dict / type 빈값)
    gosi_ok, gosi_err = validate_gosi_for_bid(gosi)
    if not gosi_ok:
        return jsonify({"ok": False, "error": gosi_err, "code": "GOSI_UNDECIDED"}), 400

    # CNY 필수 검증
    require_cny = True
    if SETTINGS_FILE.exists():
        try:
            _s = json.loads(SETTINGS_FILE.read_text())
            require_cny = _s.get("require_cny_on_bid", True)
        except Exception:
            pass
    if require_cny and (not cny_price or float(cny_price) <= 0):
        return jsonify({"error": "원가(CNY)는 필수입니다. 설정에서 해제 가능"}), 400

    tid = new_task()
    add_log(tid, "info", f"자동화 시작: #{product_id} → {price:,}원 × {qty}개")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(
                run_full_register(product_id, price, size, qty,
                                  gosi_already, gosi, tid, model=model)
            )
            loop.close()

            # 입찰 성공 시 bid_cost 저장 (식货 매칭 시 자동 채택)
            if result and result.get("success"):
                try:
                    saved = _save_bid_cost(
                        order_id=result.get("orderId") or f"{product_id}_{size}",
                        model=model, size=size,
                        cny_price=float(cny_price) if cny_price else None,
                        exchange_rate=float(exchange_rate) if exchange_rate else 0,
                        overseas_shipping=int(overseas_shipping),
                        cny_source=("manual" if cny_price and float(cny_price) > 0 else None),
                    )
                    if not saved:
                        print(f"[bid_cost] #{product_id} 원가 미저장 (manual 없음 + 식货 매칭 실패)")
                except Exception as e:
                    print(f"[bid_cost] 저장 실패: {e}")

            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return jsonify({"taskId": tid})


GOSI_DEFAULTS = {
    # Step 17-E: "가방" 폴백 제거 — 카테고리 미결정 시 빈 문자열 (호출자가 처리)
    "type": "",
    "material": "상품별 상이",
    "color": "상품별 상이",
    "size_info": "상품별 상이",
    "maker": "상품별 상이",
    "country": "상품별 상이 (케어라벨 참고)",
    "caution": "제품 라벨 참조",
    "warranty": "관련 법 및 소비자 분쟁 해결 기준에 따름",
    "phone": "010-7544-6127",
    "origin_bag": "China (중국) (CN)",
    "origin_shoe": "China (중국) (CN)",
    "hs_bag": "4202.92",
    "hs_shoe": "6404.11",
    # 신발 전용 필수 필드
    "foot_length": "사이즈별 상이",
    "heel_height": "사이즈별 상이",
    "manufacture_date": "상품별 상이",
    # 고시카테고리명
    "gosi_category_bag": "가방",
    "gosi_category_shoe": "구두/신발",
    "gosi_category_clothing": "의류",
}


def build_gosi_data(product_id, gosi, category="가방"):
    """고시정보 dict 조립 (기본값 적용, 카테고리별 필드 자동 처리)"""
    is_shoe = "신발" in category or "sneaker" in category.lower()
    is_clothing = "의류" in category

    def _val(key, default_key):
        """gosi에서 값 꺼내되, 빈 문자열이면 GOSI_DEFAULTS 사용"""
        v = gosi.get(key, "")
        return v if v and str(v).strip() else GOSI_DEFAULTS[default_key]

    # 고시카테고리 자동 설정 (KREAM 드롭다운에 맞는 정확한 이름 사용)
    # "신발" → "구두/신발", "가방" → "가방", "의류" → "의류"
    if is_shoe:
        gosi_cat = GOSI_DEFAULTS["gosi_category_shoe"]  # "구두/신발"
    elif is_clothing:
        gosi_cat = GOSI_DEFAULTS["gosi_category_clothing"]  # "의류"
    else:
        gosi_cat = GOSI_DEFAULTS["gosi_category_bag"]  # "가방"

    result = {
        "product_id": product_id,
        "고시카테고리": gosi_cat,
        "소재": _val("material", "material"),
        "색상": _val("color", "color"),
        "제조자_수입자": _val("maker", "maker") if gosi.get("maker") else
                        _val("manufacturer", "maker"),
        "제조국": _val("country", "country"),
        "취급시_주의사항": _val("caution", "caution"),
        "품질보증기준": _val("warranty", "warranty"),
        "AS_전화번호": _val("phone", "phone"),
        "제조년월": _val("manufacture_date", "manufacture_date"),
        "원산지": gosi.get("origin", "") or
                 (GOSI_DEFAULTS["origin_shoe"] if is_shoe else GOSI_DEFAULTS["origin_bag"]),
        "HS코드": gosi.get("hsCode", "") or
                 (GOSI_DEFAULTS["hs_shoe"] if is_shoe else GOSI_DEFAULTS["hs_bag"]),
    }

    # 카테고리별 필수 필드 추가
    if is_shoe:
        # 구두/신발: 발길이, 굽높이 필수
        result["발길이"] = _val("foot_length", "foot_length")
        result["굽높이"] = _val("heel_height", "heel_height")
    else:
        # 가방/의류: 종류, 크기
        result["종류"] = _val("type", "type")
        result["크기"] = _val("size", "size_info")

    return result


async def run_full_register(product_id, price, size, qty, gosi_already, gosi, tid, model="", bid_days=30):
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page, context):
            add_log(tid, "error", "판매자센터 로그인 필요 (python3 kream_bot.py --mode login)")
            await browser.close()
            return {"success": False, "error": "로그인 필요"}

        # ── 1) 고시정보 자동 감지 + 등록 ──
        if not gosi_already:
            add_log(tid, "info", f"고시정보 등록 중... #{product_id}")
            category = gosi.get("category", "가방")
            product_data = build_gosi_data(product_id, gosi, category)
            try:
                await fill_product_info(page, product_data, delay=2.0)
                add_log(tid, "success", "고시정보 등록 완료")
            except Exception as e:
                add_log(tid, "error", f"고시정보 등록 실패: {e}")
                await browser.close()
                return {"success": False, "error": f"고시정보 실패: {e}"}
        else:
            add_log(tid, "info", "고시정보 건너뜀 (이미 등록됨)")

        # ── 2) 판매 입찰 등록 ──
        add_log(tid, "info", f"판매 입찰 등록 중... {price:,}원 × {qty}개 ({bid_days}일)")
        bid_data = {
            "product_id": product_id,
            "model": model,
            "사이즈": size,
            "입찰가격": price,
            "수량": qty,
            "bid_days": bid_days,
        }

        try:
            success = await place_bid(page, bid_data, delay=2.0)
        except Exception as e:
            add_log(tid, "error", f"입찰 등록 실패: {e}")
            await browser.close()
            return {"success": False, "error": f"입찰 실패: {e}"}

        # 성공 시에만 세션 저장 (실패 시 빈 세션으로 덮어쓰기 방지)
        if success:
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        await browser.close()

        if success:
            add_log(tid, "success", f"입찰 등록 완료! #{product_id} → {price:,}원 × {qty}개")
            task_type = "고시정보+입찰" if not gosi_already else "입찰"
            save_history(task_type, product_id, price, qty, True)
            save_bid_local(product_id, model=model, size=size, price=price, source="placed")
        else:
            add_log(tid, "error", "입찰 등록 실패")
            save_history("입찰실패", product_id, price, qty, False)

        return {"success": success}


async def _run_gosi_only(product_id, gosi, category, tid):
    """고시정보만 등록 (입찰 없이)"""
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)
        if not await ensure_logged_in(page, context):
            await browser.close()
            return False
        product_data = build_gosi_data(product_id, gosi, category)
        try:
            await fill_product_info(page, product_data, delay=2.0)
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        except Exception as e:
            add_log(tid, "error", f"고시정보 실패: {e}")
            await browser.close()
            return False
        await browser.close()
    return True


async def _run_bid_only(product_id, price, size, qty, bid_days, tid, model=""):
    """입찰만 실행 (고시정보 없이)"""
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)
        if not await ensure_logged_in(page, context):
            await browser.close()
            return {"success": False}
        bid_data = {
            "product_id": product_id, "model": model, "사이즈": size,
            "입찰가격": price, "수량": qty, "bid_days": bid_days,
        }
        success = await place_bid(page, bid_data, delay=2.0)
        if success:
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
            save_history("입찰", product_id, price, qty, True)
        await browser.close()
    return {"success": success}


async def _run_batch_bid(product_id, bids, bid_days, tid):
    """같은 상품의 여러 사이즈 일괄 입찰"""
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)
        if not await ensure_logged_in(page, context):
            await browser.close()
            return {"success": 0, "fail": len(bids), "results": []}
        result = await place_bids_batch(page, product_id, bids,
                                        bid_days=bid_days, delay=2.0)
        if result.get("success", 0) > 0:
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
            for r in result.get("results", []):
                if r.get("ok"):
                    save_history("입찰", product_id, r["price"], 1, True)
        await browser.close()
    return result


# ═══════════════════════════════════════════
# API: 태스크 상태 폴링
# ═══════════════════════════════════════════

@app.route("/api/task/<task_id>")
def api_task_status(task_id):
    """태스크 실행 상태 및 로그 조회"""
    task = tasks.get(task_id)
    if not task:
        return jsonify({"error": "태스크 없음"}), 404
    return jsonify(task)


# ═══════════════════════════════════════════
# API: 가격 자동 조정
# ═══════════════════════════════════════════

@app.route("/api/adjust/scan", methods=["POST"])
def api_adjust_scan():
    """1~3단계: 내 입찰 수집 → 시장 분석 → 추천 생성"""
    tid = new_task()
    add_log(tid, "info", "가격 조정 스캔 시작...")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            add_log(tid, "info", "1단계: 내 입찰 목록 수집 중...")
            result = loop.run_until_complete(full_adjust_flow(headless=get_headless()))
            loop.close()

            n_bids = len(result.get("bids", []))
            n_recs = len(result.get("recommendations", []))
            raises = sum(1 for r in result.get("recommendations", []) if r["action"] == "raise")
            add_log(tid, "success",
                    f"스캔 완료: 입찰 {n_bids}건, 상향 제안 {raises}건, 유지 {n_recs - raises}건")
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/adjust/execute", methods=["POST"])
def api_adjust_execute():
    """5단계: 승인된 항목들의 가격 수정 실행"""
    data = request.json or {}
    items = data.get("items", [])  # [{orderId, newPrice}, ...]

    if not items:
        return jsonify({"error": "수정할 항목 없음"}), 400

    tid = new_task()
    add_log(tid, "info", f"가격 수정 실행: {len(items)}건")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            results = []
            for i, item in enumerate(items, 1):
                oid = item["orderId"]
                price = item["newPrice"]
                add_log(tid, "info", f"[{i}/{len(items)}] {oid} → {price:,}원 수정 중...")

                ok = loop.run_until_complete(modify_bid_price(oid, price, headless=get_headless()))
                results.append({"orderId": oid, "success": ok})

                if ok:
                    add_log(tid, "success", f"{oid} → {price:,}원 수정 완료")
                else:
                    add_log(tid, "error", f"{oid} 수정 실패")

            loop.close()

            success = sum(1 for r in results if r["success"])
            add_log(tid, "success", f"완료: {success}/{len(items)}건 성공")
            finish_task(tid, result={"results": results, "success": success, "total": len(items)})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# API: 대량 입찰 엑셀 생성
# ═══════════════════════════════════════════

@app.route("/api/bulk/generate", methods=["POST"])
def api_bulk_generate():
    """KREAM 대량입찰 양식 엑셀 생성"""
    data = request.json or {}
    items = data.get("items", [])
    if not items:
        return jsonify({"error": "items 필요"}), 400

    output_path = BASE_DIR / "kream_bulk_bid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "대량입찰"

    # 1행: 유의사항
    ws.append(["[유의사항] 상품번호, 모델번호, 영문상품명은 수정하지 마세요. 판매희망가와 수량만 입력하세요."])
    # 2행: 헤더
    ws.append(["상품번호", "모델번호", "영문상품명", "옵션명", "판매희망가", "수량", "입찰기한", "창고보관"])
    # 3행: 필수/선택
    ws.append(["필수", "필수", "필수", "필수", "필수", "필수", "선택", "선택"])

    # 4행~: 데이터 (Step 17-A: 신발인데 사이즈 누락된 항목은 skip)
    skipped = []
    for item in items:
        _model = item.get("model", "")
        _size = item.get("size", "")
        is_valid, err_msg, _cat = validate_size_for_bid(_model, _size)
        if not is_valid:
            skipped.append({"model": _model, "size": _size, "reason": err_msg})
            continue
        ws.append([
            item.get("productId", ""),
            _model,
            item.get("nameEn", ""),
            _size or "ONE SIZE",
            item.get("price", ""),
            item.get("quantity", 1),
            item.get("deadline", ""),
            item.get("warehouse", ""),
        ])

    wb.save(str(output_path))
    return jsonify({
        "ok": True,
        "path": str(output_path),
        "count": len(items) - len(skipped),
        "skipped": skipped,
    })


@app.route("/api/bulk/download")
def api_bulk_download():
    """생성된 대량입찰 엑셀 다운로드"""
    path = BASE_DIR / "kream_bulk_bid.xlsx"
    if not path.exists():
        return jsonify({"error": "파일 없음"}), 404
    return send_file(str(path), as_attachment=True, download_name="kream_bulk_bid.xlsx")


@app.route("/api/bulk/upload", methods=["POST"])
def api_bulk_upload():
    """KREAM 판매자센터에 대량입찰 엑셀 업로드 자동화"""
    tid = new_task()
    add_log(tid, "info", "대량입찰 엑셀 업로드 시작...")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(upload_bulk_excel(tid))
            loop.close()
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


async def upload_bulk_excel(tid):
    """partner.kream.co.kr/asks/bulk 에 엑셀 업로드"""
    excel_path = str(BASE_DIR / "kream_bulk_bid.xlsx")
    if not Path(excel_path).exists():
        add_log(tid, "error", "대량입찰 엑셀 파일 없음 → 먼저 생성하세요")
        return {"success": False}

    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page, context):
            add_log(tid, "error", "판매자센터 로그인 필요")
            await browser.close()
            return {"success": False}

        add_log(tid, "info", "대량 입찰/수정 페이지 이동...")
        await page.goto(f"{PARTNER_URL}/asks/bulk", wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)

        # 파일 업로드 input 찾기
        try:
            file_input = page.locator('input[type="file"]').first
            await file_input.set_input_files(excel_path)
            await page.wait_for_timeout(3000)
            add_log(tid, "success", "엑셀 파일 업로드 완료")

            # 업로드 확인/등록 버튼 클릭
            for btn_text in ["일괄 등록", "등록", "확인", "업로드"]:
                try:
                    btn = page.locator(f'button:has-text("{btn_text}")').first
                    if await btn.is_visible(timeout=2000):
                        await btn.click()
                        await page.wait_for_timeout(3000)
                        add_log(tid, "info", f"'{btn_text}' 클릭")
                        break
                except Exception:
                    continue

            # 최종 확인
            try:
                confirm = page.locator('button:has-text("확인")').first
                if await confirm.is_visible(timeout=2000):
                    await confirm.click()
                    await page.wait_for_timeout(2000)
            except Exception:
                pass

            add_log(tid, "success", "대량입찰 등록 완료!")
            add_log(tid, "warn", "대량입찰은 원가(bid_cost) 자동 저장 불가 — 가격 조정 탭에서 원가를 수동 등록하세요")
        except Exception as e:
            add_log(tid, "error", f"업로드 실패: {e}")
            await browser.close()
            return {"success": False}

        await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        await browser.close()

    return {"success": True}


# ═══════════════════════════════════════════
# API: 입찰 내역 관리 (7단계)
# ═══════════════════════════════════════════

@app.route("/api/my-bids")
def api_my_bids():
    """내 입찰 목록 조회"""
    tid = new_task()
    add_log(tid, "info", "내 입찰 목록 수집 중...")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from kream_adjuster import collect_my_bids_via_menu
            bids = loop.run_until_complete(collect_my_bids_via_menu(headless=get_headless()))
            loop.close()
            # DB에 내 입찰 이력 저장
            try:
                from kream_collector import save_my_bids_to_db
                save_my_bids_to_db(bids)
            except Exception:
                pass
            add_log(tid, "success", f"입찰 {len(bids)}건 수집")
            finish_task(tid, result={"bids": bids})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/my-bids/delete", methods=["POST"])
def api_delete_bids():
    """선택한 입찰 삭제"""
    data = request.json or {}
    order_ids = data.get("orderIds", [])
    verify = data.get("verify", False)
    wait_seconds = data.get("wait_seconds", 300)
    if not order_ids:
        return jsonify({"error": "orderIds 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"입찰 삭제 시작: {len(order_ids)}건")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(delete_bids(order_ids, tid))
            loop.close()
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    response_data = {"taskId": tid}
    if verify:
        response_data["verify_after_seconds"] = wait_seconds
        response_data["hint"] = f"task 완료 후 {wait_seconds}초 대기 → /api/my-bids/sync 호출 → 잔존 확인 권장"
    return jsonify(response_data)


@app.route("/api/my-bids/verify-deleted", methods=["POST"])
def api_verify_deleted():
    """삭제 검증: order_ids 리스트가 my_bids_local.json에서 사라졌는지 확인."""
    data = request.get_json() or {}
    order_ids = data.get("orderIds", [])
    if not order_ids:
        return jsonify({"ok": False, "error": "orderIds required"}), 400
    try:
        local_path = MY_BIDS_FILE
        if not local_path.exists():
            return jsonify({"ok": True, "remaining": [], "remaining_count": 0, "all_deleted": True, "note": "local cache 없음"})
        local = json.loads(local_path.read_text(encoding="utf-8"))
        bids = local.get("bids", []) if isinstance(local, dict) else []
        existing_ids = {str(b.get("orderId")) for b in bids}
        remaining = [oid for oid in order_ids if str(oid) in existing_ids]
        return jsonify({
            "ok": True,
            "requested": len(order_ids),
            "remaining": remaining,
            "remaining_count": len(remaining),
            "all_deleted": len(remaining) == 0,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


async def delete_bids(order_ids, tid):
    """판매자센터에서 입찰 삭제"""
    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        url = f"{PARTNER_URL}/business/asks?page=1&perPage=100&startDate=&endDate="
        await page.goto(url, wait_until="domcontentloaded")
        await page.wait_for_timeout(2500)

        if "/sign-in" in page.url:
            add_log(tid, "error", "로그인 필요")
            await browser.close()
            return {"success": 0, "total": len(order_ids)}

        success = 0
        for i, oid in enumerate(order_ids, 1):
            add_log(tid, "info", f"[{i}/{len(order_ids)}] {oid} 삭제 중...")
            deleted = await page.evaluate("""(orderId) => {
                const allEls = document.querySelectorAll('*');
                for (const el of allEls) {
                    const direct = Array.from(el.childNodes)
                        .filter(n => n.nodeType === 3)
                        .map(n => n.textContent.trim()).join('');
                    if (direct === orderId) {
                        let parent = el;
                        for (let i = 0; i < 15; i++) {
                            parent = parent.parentElement;
                            if (!parent) break;
                            const btns = parent.querySelectorAll('button');
                            for (const btn of btns) {
                                if (btn.innerText.includes('삭제')) {
                                    btn.click();
                                    return true;
                                }
                            }
                        }
                    }
                }
                return false;
            }""", oid)

            if deleted:
                await page.wait_for_timeout(1000)
                # 확인 팝업
                try:
                    confirm = page.locator('button:has-text("확인")').last
                    if await confirm.is_visible(timeout=2000):
                        await confirm.click()
                        await page.wait_for_timeout(1200)
                except Exception:
                    pass
                success += 1
                add_log(tid, "success", f"{oid} 삭제 완료")
            else:
                add_log(tid, "error", f"{oid} 삭제 버튼 못 찾음")

        await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
        await browser.close()

    return {"success": success, "total": len(order_ids)}


@app.route("/api/my-bids/modify", methods=["POST"])
def api_modify_bid():
    """입찰가 수정"""
    data = request.json or {}
    order_id = data.get("orderId", "")
    new_price = int(data.get("newPrice", 0))
    if not order_id or not new_price:
        return jsonify({"error": "orderId, newPrice 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"입찰가 수정: {order_id} → {new_price:,}원")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            ok = loop.run_until_complete(modify_bid_price(order_id, new_price, headless=get_headless()))
            loop.close()
            if ok:
                add_log(tid, "success", f"수정 완료: {new_price:,}원")
            else:
                add_log(tid, "error", "수정 실패")
            finish_task(tid, result={"success": ok})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# API: 중국 가격 수집 (识货/得物 앱)
# ═══════════════════════════════════════════

@app.route("/api/china-price", methods=["POST"])
def api_china_price():
    """识货/得物 앱에서 중국 가격 수집"""
    data = request.json or {}
    model = str(data.get("model", "")).strip()
    app_name = data.get("app", "识货")
    if not model:
        return jsonify({"error": "model 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"중국 가격 검색: {model} ({app_name})")

    def run():
        try:
            from china_price import search_price, load_config
            config = load_config()
            config["app_name"] = app_name
            result = search_price(model, config)
            add_log(tid, "success" if not result.get("error") else "error",
                    f"검색 완료: {model}")
            finish_task(tid, result=result)
        except Exception as e:
            import traceback; traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# API: KREAM 키워드 검색
# ═══════════════════════════════════════════

@app.route("/api/keyword-search", methods=["POST"])
def api_keyword_search():
    """KREAM에서 키워드로 상품 검색"""
    data = request.json or {}
    keyword = str(data.get("keyword", "")).strip()
    max_scroll = int(data.get("maxScroll", 3))
    if not keyword:
        return jsonify({"error": "keyword 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"KREAM 검색: '{keyword}'")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(
                kream_keyword_search(keyword, max_scroll, tid)
            )
            loop.close()
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


async def kream_keyword_search(keyword, max_scroll, tid):
    """kream.co.kr 검색 → 상품 목록 수집"""
    from kream_collector import (
        create_browser as col_browser,
        create_context as col_context,
        apply_stealth as col_stealth,
        STATE_FILE_KREAM,
    )
    import urllib.parse

    kream_session = STATE_FILE_KREAM if Path(STATE_FILE_KREAM).exists() else None

    async with async_playwright() as p:
        browser = await col_browser(p, headless=get_headless())
        context = await col_context(browser, kream_session)
        page = await context.new_page()
        await col_stealth(page)

        encoded = urllib.parse.quote(keyword)
        url = f"https://kream.co.kr/search?keyword={encoded}&tab=products"
        add_log(tid, "info", f"검색 페이지 로딩: {keyword}")
        await page.goto(url, wait_until="domcontentloaded")
        try:
            await page.wait_for_selector('a[href*="/products/"]', timeout=3500)
            await page.wait_for_timeout(300)
        except Exception:
            await page.wait_for_timeout(2000)

        # 스크롤로 더 많은 상품 로딩
        for i in range(max_scroll):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1500)

        add_log(tid, "info", "상품 목록 파싱 중...")

        products = await page.evaluate(r"""() => {
            const results = [];
            const seen = new Set();
            const cards = document.querySelectorAll('a[href*="/products/"]');

            for (const card of cards) {
                const href = card.href || '';
                const pidMatch = href.match(/\/products\/(\d+)/);
                if (!pidMatch) continue;
                const pid = pidMatch[1];
                if (seen.has(pid)) continue;
                seen.add(pid);

                const text = card.innerText.trim();
                const lines = text.split('\n').map(s => s.trim()).filter(s => s);
                const img = card.querySelector('img');
                const imgAlt = (img && img.alt) || '';

                // imgAlt: "한글명(English Name)" 형태
                let nameKr = '', nameEn = '';
                const altMatch = imgAlt.match(/^(.+?)\((.+)\)$/);
                if (altMatch) {
                    nameKr = altMatch[1].trim();
                    nameEn = altMatch[2].trim();
                } else {
                    nameKr = imgAlt;
                }

                // 브랜드: 첫 줄
                let brand = lines[0] || '';

                // 가격: "숫자,숫자원" 패턴
                let price = 0;
                for (const line of lines) {
                    const pm = line.match(/^([0-9,]+)\uC6D0$/);
                    if (pm) { price = parseInt(pm[1].replace(/,/g, '')); break; }
                }

                // 거래수, 관심수
                let trades = 0, interest = 0;
                for (const line of lines) {
                    const tm = line.match(/\uAC70\uB798\s*([0-9,.]+\uB9CC?)/);
                    if (tm) {
                        let v = tm[1].replace(/,/g, '');
                        if (v.includes('\uB9CC')) trades = parseFloat(v) * 10000;
                        else trades = parseInt(v);
                    }
                    const im = line.match(/\uAD00\uC2EC\s*([0-9,.]+\uB9CC?)/);
                    if (im) {
                        let v = im[1].replace(/,/g, '');
                        if (v.includes('\uB9CC')) interest = parseFloat(v) * 10000;
                        else interest = parseInt(v);
                    }
                }

                results.push({
                    productId: pid,
                    brand: brand,
                    nameKr: nameKr,
                    nameEn: nameEn,
                    price: price,
                    trades: trades,
                    interest: interest,
                });
            }
            return results;
        }""")

        # ── 2단계: 각 상품 상세 페이지에서 모델번호 수집 ──
        add_log(tid, "info", f"{len(products)}건 모델번호 수집 중...")
        for i, prod in enumerate(products):
            try:
                detail_url = f"https://kream.co.kr/products/{prod['productId']}"
                await page.goto(detail_url, wait_until="domcontentloaded")
                await page.wait_for_timeout(1000)

                model_info = await page.evaluate(r"""() => {
                    const body = document.body.innerText;
                    const m = body.match(/모델번호\s*([A-Za-z0-9][A-Za-z0-9_\-\/. ]+)/);
                    return m ? m[1].trim() : null;
                }""")

                if model_info:
                    prod["model"] = model_info

                # 10건마다 로그
                if (i + 1) % 10 == 0:
                    add_log(tid, "info", f"모델번호 수집 {i+1}/{len(products)}건...")
            except Exception:
                pass

        if kream_session:
            await save_state_with_localstorage(page, context, STATE_FILE_KREAM, "https://kream.co.kr")
        await browser.close()

    collected = sum(1 for p in products if p.get("model"))
    add_log(tid, "success", f"검색 완료: '{keyword}' → {len(products)}건 (모델번호 {collected}건)")
    return {"keyword": keyword, "products": products, "count": len(products)}


@app.route("/api/keyword-search/download", methods=["POST"])
def api_keyword_download():
    """키워드 검색 결과를 엑셀로 다운로드"""
    data = request.json or {}
    products = data.get("products", [])
    if not products:
        return jsonify({"error": "데이터 없음"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검색결과"
    ws.append(["키워드", "상품ID", "브랜드", "상품명", "영문명", "모델번호",
               "표시가", "거래수", "관심수", "즉시구매가", "즉시판매가"])
    for p in products:
        ws.append([
            p.get("keyword", ""),
            p.get("productId", ""),
            p.get("brand", ""),
            p.get("nameKr", ""),
            p.get("nameEn", ""),
            p.get("model", ""),
            p.get("price", ""),
            p.get("trades", ""),
            p.get("interest", ""),
            p.get("instantBuy", ""),
            p.get("instantSell", ""),
        ])

    path = str(BASE_DIR / "kream_search_result.xlsx")
    wb.save(path)
    return send_file(path, as_attachment=True, download_name="kream_search_result.xlsx")


# ═══════════════════════════════════════════
# API: 상품 발굴 데이터
# ═══════════════════════════════════════════

def parse_discovery_excel(path):
    """KREAM 데이터 엑셀 파싱"""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"overseas_top100": [], "search_surge": [], "brand_top100": []}

    # 해외직구 TOP 100
    if "해외직구 TOP 100" in wb.sheetnames:
        ws = wb["해외직구 TOP 100"]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[3]:
                continue
            result["overseas_top100"].append({
                "category": row[0] or "",
                "brand": row[2] or "",
                "productId": str(row[3]) if row[3] else "",
                "name": row[4] or "",
                "model": row[5] or "",
                "rank": row[6] if row[6] else 999,
            })

    # 크림 내 검색량 급등
    if "크림 내 검색량 급등" in wb.sheetnames:
        ws = wb["크림 내 검색량 급등"]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row[0]:
                continue
            result["search_surge"].append({
                "productId": str(row[0]) if row[0] else "",
                "brand": row[1] or "",
                "model": row[2] or "",
                "name": row[3] or "",
                "category": row[4] or "",
                "surge": row[6] if row[6] else 0,
            })

    # BRAND TOP 100
    if "BRAND TOP 100" in wb.sheetnames:
        ws = wb["BRAND TOP 100"]
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if not row[1]:
                continue
            result["brand_top100"].append({
                "brand": row[0] or "",
                "productId": str(row[1]) if row[1] else "",
                "name": row[2] or "",
            })

    wb.close()
    return result


@app.route("/api/discovery")
def api_discovery():
    """엑셀에서 상품 발굴 데이터 조회"""
    path = DISCOVERY_FILE
    if not path.exists():
        return jsonify({"error": "엑셀 파일 없음", "overseas_top100": [], "search_surge": [], "brand_top100": []}), 200
    try:
        data = parse_discovery_excel(path)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/discovery/upload", methods=["POST"])
def api_discovery_upload():
    """새 엑셀 파일 업로드"""
    if "file" not in request.files:
        return jsonify({"error": "파일 없음"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "xlsx 파일만 가능"}), 400
    f.save(str(DISCOVERY_FILE))
    return jsonify({"ok": True, "filename": f.filename})


# ── 상품 자동 발굴 ──
AUTO_SCAN_KEYWORDS = {
    "sneakers": [
        "오니츠카 타이거", "뉴발란스 1906", "미즈노", "아식스",
        "나이키 덩크", "아디다스 삼바", "뉴발란스 530",
    ],
    "bag": [
        "마르니 트렁크", "메종키츠네", "르메르",
    ],
    "apparel": [
        "스투시", "칼하트", "그라미치",
    ],
}


@app.route("/api/discovery/auto-scan", methods=["POST"])
def api_discovery_auto_scan():
    """자동 상품 발굴 — 인기 키워드 검색 → 점수 계산"""
    data = request.json or {}
    category = data.get("category", "sneakers")

    if category == "all":
        keywords = []
        for kws in AUTO_SCAN_KEYWORDS.values():
            keywords.extend(kws)
    else:
        keywords = AUTO_SCAN_KEYWORDS.get(category, AUTO_SCAN_KEYWORDS["sneakers"])

    tid = new_task()
    add_log(tid, "info", f"자동 스캔 시작: {category} ({len(keywords)}개 키워드)")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            all_products = []

            for i, kw in enumerate(keywords):
                add_log(tid, "info", f"[{i+1}/{len(keywords)}] '{kw}' 검색 중...")
                try:
                    result = loop.run_until_complete(
                        kream_keyword_search(kw, max_scroll=2, tid=tid)
                    )
                    products = result.get("products", [])
                    for p in products:
                        p["search_keyword"] = kw
                    all_products.extend(products)
                except Exception as e:
                    add_log(tid, "warning", f"'{kw}' 검색 실패: {e}")

            loop.close()

            # 중복 제거 (productId 기준)
            seen = set()
            unique = []
            for p in all_products:
                pid = p.get("productId", "")
                if pid and pid not in seen:
                    seen.add(pid)
                    unique.append(p)

            # 점수 계산
            for p in unique:
                score = 0
                trades = p.get("trades", 0)
                price = p.get("price", 0)

                # 거래량 점수 (0~40)
                if trades >= 10000:
                    score += 40
                elif trades >= 5000:
                    score += 30
                elif trades >= 1000:
                    score += 20
                elif trades >= 100:
                    score += 10

                # 가격대 점수 (5만~30만 사이가 마진 잡기 좋은 구간) (0~30)
                if 50000 <= price <= 300000:
                    score += 30
                elif 30000 <= price <= 500000:
                    score += 20
                elif price > 0:
                    score += 10

                # 관심수 점수 (0~20)
                interest = p.get("interest", 0)
                if interest >= 10000:
                    score += 20
                elif interest >= 1000:
                    score += 15
                elif interest >= 100:
                    score += 10

                # 모델번호 있으면 보너스 (분석 가능)
                if p.get("model"):
                    score += 10

                p["score"] = score

            # 점수 순 정렬
            unique.sort(key=lambda x: x.get("score", 0), reverse=True)

            add_log(tid, "success", f"자동 스캔 완료: {len(unique)}건 발굴 (상위 50건 표시)")
            finish_task(tid, result={
                "products": unique[:50],
                "total_scanned": len(unique),
                "keywords_used": keywords,
            })

        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# API: 설정
# ═══════════════════════════════════════════

def _check_session_file(path, token_prefix="_token."):
    """세션 파일에서 유효한 토큰이 있는지 확인"""
    if not Path(path).exists():
        return False
    try:
        data = json.loads(Path(path).read_text())
        for o in data.get("origins", []):
            for item in o.get("localStorage", []):
                name = item.get("name", "")
                val = item.get("value", "")
                if name.startswith(token_prefix) and val not in ("false", "", "null"):
                    return True
    except Exception:
        pass
    return False


@app.route("/api/notifications/unread")
def api_notifications_unread():
    """미확인 알림 목록"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM notifications WHERE is_read=0 ORDER BY created_at DESC LIMIT 50")
    items = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({"ok": True, "notifications": items, "count": len(items)})


@app.route("/api/notifications/recent")
def api_notifications_recent():
    """최근 알림 (읽음 포함)"""
    limit = request.args.get("limit", 30, type=int)
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT ?", (limit,))
    items = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({"ok": True, "notifications": items})


@app.route("/api/notifications/read", methods=["POST"])
def api_notifications_read():
    """알림 읽음 처리"""
    data = request.json or {}
    ids = data.get("ids", [])
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    if ids:
        placeholders = ",".join("?" * len(ids))
        c.execute(f"UPDATE notifications SET is_read=1 WHERE id IN ({placeholders})", ids)
    else:
        c.execute("UPDATE notifications SET is_read=1 WHERE is_read=0")
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/notifications/auth-failures", methods=["GET"])
def api_auth_failures():
    """auth_failure 알림 최근 24시간 조회 (대시보드 배너용)."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            SELECT id, title, message, created_at, dismissed
            FROM notifications
            WHERE type = 'auth_failure'
              AND datetime(created_at) > datetime('now', '-24 hours', 'localtime')
              AND (dismissed IS NULL OR dismissed = 0)
            ORDER BY created_at DESC
            LIMIT 10
        """)
        rows = c.fetchall()
        conn.close()
        return jsonify({
            "ok": True,
            "count": len(rows),
            "items": [
                {"id": r[0], "subject": r[1], "body": r[2],
                 "created_at": r[3], "dismissed": r[4]}
                for r in rows
            ],
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/notifications/dismiss", methods=["POST"])
def api_notifications_dismiss():
    """알림 dismiss (배너 닫기)."""
    data = request.get_json(silent=True) or {}
    nid = data.get("id")
    if not nid:
        return jsonify({"ok": False, "error": "id required"}), 400
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("UPDATE notifications SET dismissed = 1 WHERE id = ?", (nid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/session/status")
def api_session_status():
    """KREAM + 판매자센터 세션 상태 확인"""
    from kream_collector import STATE_FILE_KREAM
    partner_valid = _check_session_file(STATE_FILE, "accessToken") or _check_session_file(STATE_FILE, "_token.")
    kream_valid = _check_session_file(STATE_FILE_KREAM, "_token.")

    warnings = []
    if not partner_valid:
        warnings.append("판매자센터 세션 만료 — 고시정보/입찰 불가")
    if not kream_valid:
        warnings.append("KREAM 세션 만료 — 사이즈별 즉시구매가 수집 불가")

    return jsonify({
        "ok": True,
        "partner": Path(STATE_FILE).exists(),
        "partner_valid": partner_valid,
        "kream": Path(STATE_FILE_KREAM).exists(),
        "kream_valid": kream_valid,
        "warning": " | ".join(warnings) if warnings else None,
    })


@app.route("/api/session/relogin", methods=["POST"])
def api_session_relogin():
    """자동 재로그인 트리거"""
    data = request.json or {}
    target = data.get("target", "both")  # "partner", "kream", "both"
    tid = new_task()
    add_log(tid, "info", f"자동 재로그인 시작: {target}")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            results = {}

            async def _do_relogin():
                from kream_bot import login_auto_partner as _lap, login_auto_kream as _lak
                async with async_playwright() as pw:
                    if target in ("partner", "both"):
                        add_log(tid, "info", "판매자센터 재로그인...")
                        try:
                            await _lap(pw)
                            results["partner"] = True
                            add_log(tid, "success", "판매자센터 로그인 성공")
                        except Exception as e:
                            results["partner"] = False
                            add_log(tid, "error", f"판매자센터 로그인 실패: {e}")
                            try:
                                health_alerter.alert("auth_partner_login_failed", f"판매자센터 자동 로그인 실패: {e}")
                            except Exception:
                                pass
                    if target in ("kream", "both"):
                        add_log(tid, "info", "KREAM 재로그인...")
                        try:
                            await _lak(pw)
                            results["kream"] = True
                            add_log(tid, "success", "KREAM 로그인 성공")
                        except Exception as e:
                            results["kream"] = False
                            add_log(tid, "error", f"KREAM 로그인 실패: {e}")
                            try:
                                health_alerter.alert("auth_kream_login_failed", f"KREAM 자동 로그인 실패: {e}")
                            except Exception:
                                pass

            loop.run_until_complete(_do_relogin())

            loop.close()
            finish_task(tid, result=results)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ── 사전 갱신 스케줄러 API (Step 17-D Phase 2-B) ──

@app.route("/api/session/refresh-status", methods=["GET"])
def api_session_refresh_status():
    """사전 갱신 스케줄러 상태 조회."""
    alive = bool(_session_refresh_thread and _session_refresh_thread.is_alive())
    return jsonify({"ok": True, "thread_alive": alive, **_session_refresh_status})


@app.route("/api/session/refresh-toggle", methods=["POST"])
def api_session_refresh_toggle():
    """사전 갱신 스케줄러 ON/OFF (런타임)."""
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get("enabled", True))
    with _session_refresh_lock:
        _session_refresh_status["enabled"] = enabled
    return jsonify({"ok": True, "enabled": enabled})


@app.route("/api/session/refresh-run-once", methods=["POST"])
def api_session_refresh_run_once():
    """수동 1회 실행 (디버깅용). target='all'|'partner'|'hubnet'."""
    data = request.get_json(silent=True) or {}
    target = data.get("target", "all")
    targets = ["partner", "hubnet"] if target == "all" else [target]
    results = []
    with _session_refresh_lock:
        for t in targets:
            results.append(_refresh_session_if_stale(t))
    return jsonify({"ok": True, "results": results})


@app.route("/api/settings", methods=["GET"])
def api_get_settings():
    if SETTINGS_FILE.exists():
        return jsonify(json.loads(SETTINGS_FILE.read_text()))
    return jsonify({})


@app.route("/api/settings", methods=["POST"])
def api_save_settings():
    data = request.json or {}
    # 기존 설정과 머지 (덮어쓰기 방지)
    existing = {}
    if SETTINGS_FILE.exists():
        try:
            existing = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    existing.update(data)
    SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))
    return jsonify({"ok": True})


# ═══════════════════════════════════════════
# API: 환율
# ═══════════════════════════════════════════

@app.route("/api/exchange-rate", methods=["GET"])
def api_get_exchange_rate():
    with _exchange_rate_lock:
        return jsonify(_exchange_rate_cache.copy())


@app.route("/api/exchange-rate/refresh", methods=["POST"])
def api_refresh_exchange_rate():
    result = fetch_exchange_rates()
    if result:
        return jsonify({"ok": True, **result})
    return jsonify({"ok": False, "error": "환율 조회 실패"}), 500


# ═══════════════════════════════════════════
# API: 가격 이력 조회
# ═══════════════════════════════════════════

@app.route("/api/price-history/<product_id>")
def api_price_history(product_id):
    """상품의 가격 수집 이력 조회"""
    import sqlite3
    from kream_collector import DB_PATH
    if not DB_PATH.exists():
        return jsonify({"records": [], "summary": {}})
    try:
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        # 최근 수집 이력 (최근 7일, 최대 500건)
        c.execute(
            "SELECT size, delivery_type, buy_price, sell_price, "
            "recent_trade_price, collected_at FROM price_history "
            "WHERE product_id=? ORDER BY collected_at DESC LIMIT 500",
            (product_id,)
        )
        rows = [dict(r) for r in c.fetchall()]

        # 최신 수집 시간 기준 요약 (사이즈×배송타입 매트릭스)
        summary = {}
        if rows:
            latest_time = rows[0]["collected_at"][:19]  # 초 단위
            latest = [r for r in rows if r["collected_at"][:19] == latest_time]
            for r in latest:
                sz = r["size"]
                if sz not in summary:
                    summary[sz] = {"size": sz}
                dt = r["delivery_type"]
                summary[sz][dt] = {
                    "buy": r["buy_price"],
                    "sell": r["sell_price"],
                }
            summary = list(summary.values())
        else:
            summary = []

        conn.close()
        return jsonify({
            "records": rows[:100],  # 최근 100건만 전달
            "summary": summary,
            "total": len(rows),
            "latestAt": rows[0]["collected_at"] if rows else None,
        })
    except Exception as e:
        return jsonify({"error": str(e), "records": [], "summary": []})


# ═══════════════════════════════════════════
# API: 상품 큐 + 일괄 실행
# ═══════════════════════════════════════════

def detect_category(english_name):
    """KREAM 영문명으로 카테고리 자동 판별"""
    name = (english_name or "").lower()

    bag_kw = ['bag', 'backpack', 'tote', 'pouch', 'wallet', 'clutch',
              'purse', 'satchel', 'rucksack', 'crossbody', 'shoulder bag',
              'duffle', 'messenger', 'fanny pack', 'waist bag']
    # Step 17-E 보강: 신발 일반 키워드 추가 (training/athletic/performance)
    # ⚠️ court/gym/sport는 부분 매칭 부작용 위험으로 제외
    shoe_kw = ['shoe', 'sneaker', 'boot', 'sandal', 'slipper',
               'runner', 'trainer', 'loafer', 'mule', 'clog',
               'slide', 'flip flop', 'oxford', 'derby',
               'training', 'athletic', 'performance']
    clothing_kw = ['jacket', 'hoodie', 'shirt', 'pants', 'shorts',
                   'dress', 'skirt', 'coat', 'sweater', 'cardigan',
                   'vest', 'tee', 't-shirt', 'jogger', 'track']

    for kw in bag_kw:
        if kw in name:
            return {"category": "가방", "tariff": 0.08, "auto": True}
    for kw in shoe_kw:
        if kw in name:
            return {"category": "신발", "tariff": 0.13, "auto": True}
    for kw in clothing_kw:
        if kw in name:
            return {"category": "의류", "tariff": 0.13, "auto": True}
    return {"category": None, "tariff": None, "auto": False}


def _map_kream_category(kream_cat):
    """KREAM에서 반환하는 카테고리 문자열을 우리 카테고리로 매핑"""
    cat = (kream_cat or "").lower()
    if any(k in cat for k in ['bag', 'wallet', 'acc', '가방', '지갑', '액세서리']):
        return "가방"
    if any(k in cat for k in ['shoe', 'sneaker', 'sandal', 'boot', '신발', '스니커즈']):
        return "신발"
    if any(k in cat for k in ['apparel', 'clothing', 'top', 'bottom', 'outer',
                               '의류', '상의', '하의', '아우터']):
        return "의류"
    return None


def detect_category_kr(korean_name):
    """한글 상품명에서 카테고리 판별"""
    name = korean_name or ""
    bag_kw = ['숄더백', '토트백', '크로스백', '백팩', '파우치', '지갑', '클러치',
              '더플백', '메신저백', '웨이스트백', '버킷백', '호보백', '에코백',
              '가방', '백 ', '월렛']
    # Step 17-E 보강: 신발 일반 키워드 추가 (트레이닝/러닝/운동/스포츠)
    # ⚠️ 코트화는 부분 매칭 부작용 위험으로 제외
    shoe_kw = ['러닝화', '스니커즈', '슬라이드', '샌들', '부츠', '로퍼',
               '슬리퍼', '트레이너', '운동화', '스니커', '구두',
               '트레이닝', '러닝', '운동', '스포츠']
    clothing_kw = ['후드', '자켓', '티셔츠', '팬츠', '쇼츠', '스웨터',
                   '코트', '셔츠', '조거', '베스트', '드레스']

    for kw in bag_kw:
        if kw in name:
            return "가방"
    for kw in shoe_kw:
        if kw in name:
            return "신발"
    for kw in clothing_kw:
        if kw in name:
            return "의류"
    return None


def auto_fill_gosi(kream_data):
    """KREAM 상품명/영문명에서 고시정보 자동 추출.
    반환 dict 키: type, color, manufacturer, material, size, hs_code, tariff
    build_gosi_data()에서 빈 문자열이면 GOSI_DEFAULTS로 대체됨.
    """
    eng_name_raw = (kream_data.get("product_name_en")
                    or kream_data.get("english_name")
                    or kream_data.get("nameEn") or "")
    eng_name = eng_name_raw.lower()
    kor_name = kream_data.get("product_name") or kream_data.get("nameKr") or ""
    brand_raw = kream_data.get("brand") or ""

    # ── 제조자/수입자: 브랜드 추출 ──
    BRAND_MAP = {
        # 한글명 → 영문 브랜드 (한글 상품명에서 매칭)
        "아디다스": "Adidas", "나이키": "Nike", "뉴발란스": "New Balance",
        "아식스": "Asics", "퓨마": "Puma", "컨버스": "Converse",
        "반스": "Vans", "리복": "Reebok", "필라": "Fila",
        "스케쳐스": "Skechers", "디스커버리": "Discovery",
        "노스페이스": "The North Face", "파타고니아": "Patagonia",
        "스투시": "Stussy", "팔라스": "Palace", "슈프림": "Supreme",
        "발렌시아가": "Balenciaga", "구찌": "Gucci", "프라다": "Prada",
        "로에베": "Loewe", "몽클레르": "Moncler", "디올": "Dior",
        "셀린느": "Celine", "보테가": "Bottega Veneta",
        "메종키츠네": "Maison Kitsune", "아크네": "Acne Studios",
        "마르지엘라": "Maison Margiela",
    }
    # 영문명 첫 단어 or 한글명에서 브랜드 감지
    manufacturer = brand_raw
    if not manufacturer:
        # 한글 상품명에서 브랜드 매칭
        for kr_brand, en_brand in BRAND_MAP.items():
            if kr_brand in kor_name:
                manufacturer = en_brand
                break
        # 폴백: 영문명 첫 단어
        if not manufacturer and eng_name_raw:
            first_word = eng_name_raw.split()[0] if eng_name_raw.split() else ""
            known_en = [
                "Adidas", "Nike", "New Balance", "Asics", "Puma",
                "Converse", "Vans", "Reebok", "Fila", "Skechers",
                "Balenciaga", "Gucci", "Prada", "Loewe", "Moncler",
                "Dior", "Celine", "Supreme", "Stussy", "Palace",
                "Discovery", "Patagonia",
            ]
            for b in known_en:
                if b.lower() == first_word.lower():
                    manufacturer = b
                    break
            if not manufacturer:
                manufacturer = first_word

    info = {"manufacturer": manufacturer}

    # ── 종류: 상품명에서 가방 종류 매칭 ──
    # 영문 매칭 (longer phrases first)
    TYPE_MAP_EN = [
        ("shoulder bag", "숄더백"), ("tote bag", "토트백"),
        ("crossbody", "크로스백"), ("cross body", "크로스백"),
        ("messenger bag", "메신저백"), ("messenger", "메신저백"),
        ("duffle bag", "더플백"), ("duffle", "더플백"), ("duffel", "더플백"),
        ("boston bag", "보스턴백"), ("bucket bag", "버킷백"),
        ("waist bag", "웨이스트백"), ("belt bag", "웨이스트백"),
        ("fanny pack", "힙색"), ("hip pack", "힙색"),
        ("backpack", "백팩"), ("rucksack", "백팩"),
        ("pouch", "파우치"), ("clutch", "클러치"),
        ("eco bag", "에코백"), ("shopper", "쇼퍼백"),
        ("hobo", "호보백"), ("tote", "토트백"),
        ("wallet", "지갑"), ("card holder", "카드홀더"),
        # 신발
        ("running shoe", "러닝화"), ("sneaker", "스니커즈"),
        ("slide", "슬라이드"), ("sandal", "샌들"), ("boot", "부츠"),
        ("loafer", "로퍼"), ("trainer", "트레이너"), ("slipper", "슬리퍼"),
        # Step 17-E 보강: 신발 일반 용어 (모델명 매칭은 X, 일반 카테고리 용어만)
        ("training", "운동화"), ("athletic", "운동화"),
        ("performance", "운동화"), ("runner", "러닝화"),
        # 의류
        ("hoodie", "후드"), ("jacket", "자켓"), ("t-shirt", "티셔츠"),
        ("pants", "팬츠"), ("shorts", "쇼츠"), ("sweater", "스웨터"),
    ]
    # 한글 매칭
    TYPE_MAP_KR = [
        "숄더백", "크로스백", "토트백", "백팩", "클러치", "파우치",
        "힙색", "웨이스트백", "더플백", "메신저백", "보스턴백", "에코백",
        "쇼퍼백", "호보백", "버킷백",
        "러닝화", "스니커즈", "슬라이드", "샌들", "부츠", "로퍼",
        # Step 17-E 보강: 신발 일반 한글 용어
        "트레이닝", "러닝", "운동화", "스포츠",
        "후드", "자켓", "티셔츠", "팬츠", "쇼츠",
    ]

    detected_type = ""
    for eng, kor in TYPE_MAP_EN:
        if eng in eng_name:
            detected_type = kor
            break
    if not detected_type:
        for kt in TYPE_MAP_KR:
            if kt in kor_name:
                detected_type = kt
                break
    # Step 17-E: detected_type 못 찾으면 None 반환 (가방 폴백 제거)
    # 호출자(큐 execute)가 None을 받으면 카테고리 미결정으로 처리
    if not detected_type:
        logger.warning(
            "auto_fill_gosi: type detection failed name_en=%r name_kr=%r — returning None",
            eng_name_raw, kor_name
        )
        return None
    info["type"] = detected_type

    # ── 색상: 상품명에서 매칭 ──
    COLOR_MAP_EN = {
        "black": "블랙", "white": "화이트", "red": "레드",
        "blue": "블루", "navy": "네이비", "green": "그린",
        "grey": "그레이", "gray": "그레이", "pink": "핑크",
        "beige": "베이지", "brown": "브라운", "cream": "크림",
        "orange": "오렌지", "yellow": "옐로우", "purple": "퍼플",
        "silver": "실버", "gold": "골드", "olive": "올리브",
        "burgundy": "버건디", "khaki": "카키", "ivory": "아이보리",
        "coral": "코랄", "mint": "민트", "charcoal": "차콜",
        "multi": "멀티",
    }
    COLOR_MAP_KR = [
        "블랙", "화이트", "네이비", "블루", "레드", "그린", "핑크",
        "베이지", "브라운", "그레이", "카키", "옐로우", "퍼플",
        "오렌지", "실버", "골드", "멀티", "아이보리", "크림",
        "올리브", "버건디", "차콜", "민트", "코랄",
    ]

    colors = []
    for eng, kor in COLOR_MAP_EN.items():
        if eng in eng_name and kor not in colors:
            colors.append(kor)
    if not colors:
        for kc in COLOR_MAP_KR:
            if kc in kor_name and kc not in colors:
                colors.append(kc)
    info["color"] = ", ".join(colors) if colors else GOSI_DEFAULTS["color"]

    # ── 소재/크기: 고정 기본값 ──
    info["material"] = GOSI_DEFAULTS["material"]
    info["size"] = GOSI_DEFAULTS["size_info"]

    # ── 신발 필수 필드: 발길이, 굽높이 ──
    ext_category = kream_data.get("category", "")
    is_shoe = "신발" in ext_category
    if is_shoe:
        info["foot_length"] = GOSI_DEFAULTS["foot_length"]
        info["heel_height"] = GOSI_DEFAULTS["heel_height"]

    # ── HS코드/관세: 외부 카테고리 우선, 폴백은 영문명 감지 ──
    if is_shoe:
        resolved_cat = "신발"
    elif "가방" in ext_category:
        resolved_cat = "가방"
    else:
        cat = detect_category(eng_name)
        resolved_cat = cat.get("category", "")

    if resolved_cat == "가방":
        info["hs_code"] = GOSI_DEFAULTS["hs_bag"]
        info["tariff"] = 8
    elif resolved_cat == "신발":
        info["hs_code"] = GOSI_DEFAULTS["hs_shoe"]
        info["tariff"] = 13
    else:
        info["hs_code"] = ""
        info["tariff"] = 13

    return info


def _calc_profit_simple(sell_price, total_cost):
    """판매가, 원가로 간단한 마진 문자열 반환"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    fee_rate = float(settings.get("feeRate", 0.06))
    fixed_fee = int(settings.get("fixedFee", 2500))
    vat_rate = float(settings.get("vatRate", 0.10))
    effective_rate = 1 - fee_rate * (1 + vat_rate)
    settlement = round(sell_price * effective_rate - fixed_fee)
    profit = settlement - total_cost
    if total_cost > 0:
        rate = profit / total_cost * 100
        return f"{'+' if profit >= 0 else ''}{profit:,.0f} ({rate:.1f}%)"
    return f"{'+' if profit >= 0 else ''}{profit:,.0f}"


def calculate_margin_for_queue(cny_price, category, shipping_krw=8000):
    """큐 일괄 실행용 마진 계산"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass

    cny_rate = float(settings.get("cnyRate", 218.12))
    usd_rate = float(settings.get("usdRate", 1495.76))
    fee_rate = float(settings.get("feeRate", 0.06))
    fixed_fee = int(settings.get("fixedFee", 2500))
    cny_margin = float(settings.get("cnyMargin", 1.03))
    vat_rate = float(settings.get("vatRate", 0.10))
    usd_limit = float(settings.get("usdLimit", 150))

    tariff_rate = 0.08 if category == "가방" else 0.13
    krw_price = round(cny_price * cny_rate * cny_margin)
    usd_equiv = round(cny_price * cny_rate / usd_rate, 2)

    # 관부가세는 고객 부담 → 원가에서 제외, 참고용으로만 계산
    customs = 0
    import_vat = 0
    if usd_equiv > usd_limit:
        customs = round(cny_price * cny_rate * tariff_rate)
        import_vat = round((cny_price * cny_rate + customs) * vat_rate)

    total_cost = krw_price + shipping_krw  # 관부가세 제외

    margins = {}
    for pct in [0, 10, 15, 20]:
        target_profit = total_cost * (pct / 100)
        required_net = total_cost + target_profit
        # settlement = sell_price - (sell_price * fee * 1.1) - fixed
        # settlement = sell_price * (1 - fee * 1.1) - fixed
        effective_rate = 1 - fee_rate * (1 + vat_rate)
        raw_price = (required_net + fixed_fee) / effective_rate
        # KREAM은 1,000원 단위만 가능 → 올림
        sell_price = int(math.ceil(raw_price / 1000) * 1000)
        settlement = round(sell_price * effective_rate - fixed_fee)
        margins[f"margin_{pct}"] = {
            "sell_price": sell_price,
            "profit": settlement - total_cost,
        }

    return {
        "krw_price": krw_price,
        "customs": customs,
        "import_vat": import_vat,
        "shipping": shipping_krw,
        "total_cost": total_cost,
        "margins": margins,
    }


@app.route("/api/queue/verify-model", methods=["POST"])
def api_queue_verify_model():
    """모델번호로 KREAM에 상품 존재 여부를 빠르게 확인 (큐 추가 전)"""
    data = request.json or {}
    model = str(data.get("model", "")).strip().upper()
    if not model:
        return jsonify({"error": "model 필요"}), 400

    # 큐에 이미 같은 모델로 완료된 결과가 있으면 재활용
    for q in product_queue:
        if q.get("model", "").upper() == model and q.get("status") == "완료" and q.get("result"):
            r = q["result"]
            return jsonify({
                "ok": True, "exists": True, "cached": True,
                "productId": r.get("productId"),
                "nameKr": r.get("nameKr") or r.get("nameEn", ""),
                "brand": r.get("brand", ""),
            })

    # KREAM 검색 (비동기)
    tid = new_task()
    add_log(tid, "info", f"모델 확인: {model}")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            result = loop.run_until_complete(search_by_model(model))
            loop.close()
            if result:
                kream = result[0].get("kream", {})
                pid = kream.get("product_id")
                name = kream.get("product_name", "") or kream.get("product_name_en", "")
                finish_task(tid, result={
                    "exists": bool(pid), "productId": pid,
                    "nameKr": name, "brand": kream.get("brand", ""),
                })
            else:
                finish_task(tid, result={"exists": False})
        except Exception as e:
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/queue/add", methods=["POST"])
def api_queue_add():
    """큐에 상품 추가.
    가방/의류: {"model":"IX7694", "cny":220}
    신발(사이즈별): {"model":"ID6016", "sizes":[{"size":"38","cny_price":314},...],"sizeSystem":"EU"}
    """
    global queue_counter
    data = request.json or {}
    model = str(data.get("model", "")).strip()

    if not model:
        return jsonify({"error": "model 필요"}), 400

    # sizes 배열이 있으면 신발(사이즈별 가격), 없으면 단일 가격
    sizes = data.get("sizes", [])
    cny = data.get("cny")

    # result가 함께 전달되면 cny/sizes 없어도 허용 (완료 항목 복사)
    if not sizes and cny is None and not data.get("result"):
        return jsonify({"error": "cny 또는 sizes 필요"}), 400

    with queue_lock:
        queue_counter += 1
        # result/status/gosi를 함께 전달하면 KREAM 검색 없이 바로 완료 상태로 추가 (복사 기능)
        preset_result = data.get("result", None)
        preset_status = data.get("status", "대기") if preset_result else "대기"
        item = {
            "id": queue_counter,
            "model": model.upper(),
            "cny": float(cny) if cny is not None else None,
            "category": data.get("category", ""),
            "categoryAuto": False,
            "sizes": sizes,  # [{"size":"38","cny_price":314}, ...]
            "sizeSystem": data.get("sizeSystem", ""),  # "EU" or "JP"
            "size": data.get("size", ""),
            "shipping": int(data.get("shipping", 8000)),
            "quantity": int(data.get("quantity", 1)),
            "bid_days": int(data.get("bid_days", 30)),
            "bid_strategy": data.get("bid_strategy", "undercut"),
            "status": preset_status,
            "result": preset_result,
            "gosi": data.get("gosi", None),
        }
        product_queue.append(item)

    save_queue()
    return jsonify({"ok": True, "item": item})


@app.route("/api/queue/upload-excel", methods=["POST"])
def api_queue_upload_excel():
    """XLSX 파일 업로드 → 파싱 → 큐에 추가"""
    global queue_counter
    if "file" not in request.files:
        return jsonify({"error": "파일 없음"}), 400

    file = request.files["file"]
    tmp_path = BASE_DIR / f"_tmp_upload_{file.filename}"
    try:
        file.save(str(tmp_path))
        wb = openpyxl.load_workbook(str(tmp_path), data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]

        # 컬럼 인덱스 매핑
        def find_col(keywords):
            for i, h in enumerate(headers):
                if any(k in h for k in keywords):
                    return i
            return -1

        i_model = find_col(["모델", "model", "Model"])
        i_category = find_col(["카테고리", "category"])
        i_size = find_col(["사이즈", "size"])
        i_cny = find_col(["CNY", "cny", "중국가"])
        i_qty = find_col(["수량", "qty", "quantity"])
        i_shipping = find_col(["배송비", "shipping"])
        i_bid_days = find_col(["만료", "bid_days"])

        if i_model == -1:
            wb.close()
            return jsonify({"error": "모델번호 컬럼을 찾을 수 없습니다"}), 400

        # 각 행을 독립적으로 큐에 추가 (같은 품번도 별도 항목)
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= i_model:
                continue
            model = str(row[i_model] or "").strip()
            if not model:
                continue

            category = str(row[i_category] or "").strip() if i_category >= 0 else ""
            size = str(row[i_size] or "").strip() if i_size >= 0 else ""
            cny = float(row[i_cny] or 0) if i_cny >= 0 else 0
            qty = int(row[i_qty] or 1) if i_qty >= 0 else 1
            shipping = int(row[i_shipping] or 8000) if i_shipping >= 0 else 8000
            bid_days = int(row[i_bid_days] or 30) if i_bid_days >= 0 else 30

            sizes = []
            row_cny = None
            if size and size != "ONE SIZE":
                sizes.append({"size": size, "cny_price": cny})
            else:
                row_cny = cny if cny else None

            rows.append({
                "model": model.upper(), "category": category, "sizes": sizes,
                "cny": row_cny, "quantity": qty, "shipping": shipping, "bid_days": bid_days,
            })

        wb.close()

        # 큐에 추가
        added = []
        with queue_lock:
            for m in rows:
                queue_counter += 1
                item = {
                    "id": queue_counter,
                    "model": m["model"],
                    "cny": m["cny"],
                    "category": m["category"],
                    "categoryAuto": False,
                    "sizes": m["sizes"],
                    "sizeSystem": "",
                    "size": "전사이즈" if m["sizes"] else "",
                    "shipping": m["shipping"],
                    "quantity": m["quantity"],
                    "bid_days": m["bid_days"],
                    "status": "대기",
                    "result": None,
                    "gosi": None,
                }
                product_queue.append(item)
                added.append(item)

        save_queue()
        return jsonify({"ok": True, "count": len(added), "items": added, "queue": product_queue})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"파일 파싱 실패: {e}"}), 400
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


@app.route("/api/queue/bulk-add", methods=["POST"])
def api_queue_bulk_add():
    """CSV/엑셀에서 파싱한 상품 목록을 큐에 일괄 추가"""
    global queue_counter
    data = request.json or {}
    items = data.get("items", [])
    if not items:
        return jsonify({"error": "items 필요"}), 400

    added = []
    with queue_lock:
        for row in items:
            model = str(row.get("model", "")).strip()
            if not model:
                continue

            queue_counter += 1

            sizes = row.get("sizes", [])
            cny = row.get("cny")

            item = {
                "id": queue_counter,
                "model": model.upper(),
                "cny": float(cny) if cny is not None else None,
                "category": row.get("category", ""),
                "categoryAuto": False,
                "sizes": sizes,
                "sizeSystem": row.get("sizeSystem", ""),
                "size": row.get("size", ""),
                "shipping": int(row.get("shipping", 8000)),
                "quantity": int(row.get("quantity", 1)),
                "bid_days": int(row.get("bid_days", 30)),
                "status": "대기",
                "result": None,
                "gosi": None,
            }
            product_queue.append(item)
            added.append(item)

    save_queue()
    return jsonify({"ok": True, "count": len(added), "items": added})


@app.route("/api/queue/download-excel")
def api_queue_download_excel():
    """현재 큐를 XLSX 파일로 다운로드"""
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품큐"
    headers = ["모델번호", "카테고리", "사이즈", "CNY", "배송비", "수량", "만료일(일)", "상태"]
    ws.append(headers)

    for item in product_queue:
        sizes = item.get("sizes", [])
        if sizes:
            for s in sizes:
                ws.append([
                    item.get("model", ""),
                    item.get("category", ""),
                    s.get("size", ""),
                    s.get("cny_price", 0),
                    item.get("shipping", 8000),
                    item.get("quantity", 1),
                    item.get("bid_days", 30),
                    item.get("status", "대기"),
                ])
        else:
            ws.append([
                item.get("model", ""),
                item.get("category", ""),
                item.get("size", ""),
                item.get("cny", 0) or 0,
                item.get("shipping", 8000),
                item.get("quantity", 1),
                item.get("bid_days", 30),
                item.get("status", "대기"),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"queue_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/api/queue/template")
def api_queue_template():
    """업로드용 빈 엑셀 양식 다운로드 (예시 1행 포함)"""
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "큐 양식"
    headers = ["모델번호", "카테고리", "사이즈", "중국가(CNY)", "배송비", "수량", "만료일(일)"]
    ws.append(headers)
    ws.append(["IX7694", "가방", "ONE SIZE", 220, 8000, 1, 30])

    # 컬럼 너비 조정
    widths = [14, 10, 12, 14, 10, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="kream_queue_template.xlsx")


@app.route("/api/queue/list")
def api_queue_list():
    """큐 목록 조회"""
    return jsonify({"queue": product_queue})


@app.route("/api/queue/<int:item_id>", methods=["DELETE"])
def api_queue_delete(item_id):
    """큐에서 삭제"""
    with queue_lock:
        idx = next((i for i, q in enumerate(product_queue) if q["id"] == item_id), None)
        if idx is not None:
            product_queue.pop(idx)
            save_queue()
            return jsonify({"ok": True})
    return jsonify({"error": "항목 없음"}), 404


@app.route("/api/queue/<int:item_id>", methods=["PUT"])
def api_queue_update(item_id):
    """큐 항목 수정"""
    data = request.json or {}
    with queue_lock:
        item = next((q for q in product_queue if q["id"] == item_id), None)
        if not item:
            return jsonify({"error": "항목 없음"}), 404
        tracked_fields = ["cny", "quantity", "bid_strategy", "selectedMargin", "bid_days", "shipping"]
        for key in ["model", "cny", "category", "size", "shipping", "quantity",
                     "sizes", "sizeSystem", "gosi", "selectedMargin", "bid_strategy", "bid_days",
                     "status", "result", "categoryAuto"]:
            if key in data:
                if key in tracked_fields and item.get(key) != data[key]:
                    save_edit_log("queue", f"{item.get('model','')}", key, item.get(key), data[key])
                item[key] = data[key]
        if "model" in data:
            item["model"] = str(data["model"]).upper()
        save_queue()
        return jsonify({"ok": True, "item": item})


@app.route("/api/queue/clear", methods=["DELETE"])
def api_queue_clear():
    """큐 전체 삭제"""
    with queue_lock:
        product_queue.clear()
    save_queue()
    return jsonify({"ok": True})


@app.route("/api/queue/execute", methods=["POST"])
def api_queue_execute():
    """큐 일괄 실행: KREAM 검색 + 카테고리 판별 + 마진 계산"""
    if not product_queue:
        return jsonify({"error": "큐가 비어있음"}), 400

    # 대기 상태인 항목만 실행
    pending = [q for q in product_queue if q["status"] in ("대기", "실패")]
    if not pending:
        return jsonify({"error": "실행할 항목 없음"}), 400

    tid = new_task()
    add_log(tid, "info", f"큐 일괄 실행 시작: {len(pending)}건")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            search_cache = {}  # 품번별 KREAM 검색 결과 캐시

            # 내 입찰 로드 (즉시구매가에서 내 입찰 제외용)
            my_bids_for_filter = {}  # key: f"{productId}_{size}" → [price, price, ...]
            try:
                mb_file = BASE_DIR / "my_bids_local.json"
                if mb_file.exists():
                    mb_data = json.loads(mb_file.read_text())
                    for b in mb_data.get("bids", []):
                        pid = str(b.get("productId", ""))
                        sz = str(b.get("size", "ONE SIZE")).strip()
                        price = b.get("price", 0)
                        if pid and price:
                            key = f"{pid}_{sz}"
                            if key not in my_bids_for_filter:
                                my_bids_for_filter[key] = []
                            my_bids_for_filter[key].append(price)
                    add_log(tid, "info", f"내 입찰 {len(mb_data.get('bids',[]))}건 로드 (즉시구매가 필터용)")
            except Exception as e:
                add_log(tid, "warn", f"내 입찰 로드 실패: {e} — 필터 없이 진행")

            for i, item in enumerate(pending, 1):
                model = item["model"]
                item["status"] = "검색 중"

                try:
                    # 같은 품번이면 캐시된 검색 결과 재사용
                    if model in search_cache:
                        cached = search_cache[model]
                        if cached is None:
                            item["status"] = "실패"
                            item["result"] = {"error": "KREAM 검색 결과 없음 (캐시)"}
                            add_log(tid, "info", f"[{i}/{len(pending)}] {model} 캐시 사용 → 검색 결과 없음")
                            continue
                        if cached == "session_expired":
                            item["status"] = "실패"
                            item["result"] = {"error": "세션 만료 (캐시)"}
                            add_log(tid, "info", f"[{i}/{len(pending)}] {model} 캐시 사용 → 세션 만료")
                            continue
                        kream = cached
                        add_log(tid, "info", f"[{i}/{len(pending)}] {model} 캐시 사용 (검색 생략)")
                    else:
                        add_log(tid, "info", f"[{i}/{len(pending)}] {model} KREAM 검색 중...")
                        results = loop.run_until_complete(search_by_model(model))

                        if not results or len(results) == 0:
                            search_cache[model] = None
                            item["status"] = "실패"
                            item["result"] = {"error": "KREAM 검색 결과 없음"}
                            add_log(tid, "error", f"{model}: 검색 결과 없음")
                            continue

                        kream = results[0].get("kream", {})
                        if kream.get("session_expired"):
                            search_cache[model] = "session_expired"
                            item["status"] = "실패"
                            item["result"] = {"error": "세션 만료"}
                            add_log(tid, "error", f"{model}: 세션 만료")
                            continue

                        search_cache[model] = kream

                    product_id = str(kream.get("product_id", ""))
                    # collector는 product_name_en 키를 사용
                    name_en = kream.get("product_name_en", "") or kream.get("english_name", "") or ""
                    name_kr = kream.get("product_name", "")
                    # 상품명에서 발매가 정보 제거 (예: "발매가 $65 (약 96,700원)")
                    name_kr = re.sub(r'\s*발매가\s*\$?\d[\d,]*\s*(\(약\s*[\d,]+원\))?\s*', '', name_kr).strip()
                    name_en = re.sub(r'\s*Retail\s*Price\s*\$?\d[\d,]*\s*', '', name_en, flags=re.IGNORECASE).strip()
                    # 즉시구매가 = 현재 판매입찰 최저가 (과거 체결가 아님)
                    instant_buy = kream.get("instant_buy_price")  # sell_bids 최저가
                    recent_trade = kream.get("recent_trade_price") or kream.get("display_price")  # 과거 체결가

                    # 사이즈 자동 설정
                    # Step 17-A: 사이즈 정보가 없을 때 신발 모델은 ONE SIZE 자동 설정 차단
                    if not item.get("sizes") and not item.get("size"):
                        kream_sizes = kream.get("sizes", [])
                        if kream_sizes:
                            item["size"] = "전사이즈"
                        else:
                            _cat_info = get_model_category(model)
                            if _cat_info["needs_size"]:
                                item["status"] = "사이즈 필요"
                                item["result"] = {
                                    "error": (f"카테고리 '{_cat_info['category']}'은 사이즈 필수인데 "
                                              f"KREAM 사이즈 정보가 없습니다 (model={model})")
                                }
                                add_log(tid, "error",
                                        f"{model}: 신발 카테고리 사이즈 누락 — 자동 등록 스킵")
                                continue
                            item["size"] = "ONE SIZE"

                    # 카테고리 자동 판별 — Step 17-E: 0순위 model_category DB 추가
                    item["status"] = "계산 중"
                    decision_source = "preset"  # 사용자가 미리 지정한 경우
                    if not item["category"]:
                        decision_source = "unresolved"
                        # 0순위: model_category DB (식货 mapped 등 신뢰 가능 출처)
                        db_info = get_model_category(model)
                        db_kr = map_db_category_to_kr(db_info.get("category", ""))
                        if db_kr:
                            item["category"] = db_kr
                            item["categoryAuto"] = True
                            decision_source = "db"
                        # 1순위: KREAM 카테고리 정보
                        if not item["category"]:
                            kream_cat = kream.get("category", "")
                            if kream_cat:
                                cat_mapped = _map_kream_category(kream_cat)
                                if cat_mapped:
                                    item["category"] = cat_mapped
                                    item["categoryAuto"] = True
                                    decision_source = "kream"
                        # 2순위: 영문 상품명 파싱
                        if not item["category"]:
                            cat_info = detect_category(name_en)
                            if cat_info["category"]:
                                item["category"] = cat_info["category"]
                                item["categoryAuto"] = True
                                decision_source = "detect_en"
                        # 3순위: 한글 상품명에서도 시도
                        if not item["category"] and name_kr:
                            cat_info = detect_category_kr(name_kr)
                            if cat_info:
                                item["category"] = cat_info
                                item["categoryAuto"] = True
                                decision_source = "detect_kr"
                        # 못 찾으면 미분류 (decision_source는 unresolved 유지)
                        if not item["category"]:
                            item["category"] = "미분류"
                            item["categoryAuto"] = True

                    item["category_decision_source"] = decision_source

                    # 고시정보 자동 채움 (카테고리 전달)
                    # Step 17-E: auto_fill_gosi가 None 반환 = type 결정 실패 → gosi=None
                    gosi = auto_fill_gosi({
                        "english_name": name_en,
                        "product_name": name_kr,
                        "brand": kream.get("brand", ""),
                        "category": item["category"],
                    })
                    item["gosi"] = gosi
                    if gosi is None:
                        add_log(tid, "warn",
                                f"{model}: gosi.type 결정 실패 — 카테고리 미분류 또는 키워드 매칭 실패")

                    # 마진 계산 — 사이즈별 가격이 있으면 각각 계산
                    input_sizes = item.get("sizes", [])
                    # 사이즈별 KREAM 즉시구매가 맵 구축 (sizeDeliveryPrices에서)
                    # KREAM API 사이즈 형식: "W215", "260", "ONE SIZE" 등
                    # 사용자 입력 형식: "215", "260", "ONE SIZE" 등
                    # → 숫자만 추출하여 매칭 (W215 ↔ 215)
                    sdp_list = kream.get("size_delivery_prices", [])
                    sdp_map = {}  # size → buyPrice (원본 키 + 숫자만 키 모두 등록)
                    sdp_full_map = {}  # size → full sdp entry (배송타입별 가격 포함)
                    for sdp in sdp_list:
                        sdp_size = str(sdp.get("size", "")).strip()
                        sdp_buy = sdp.get("buyPrice") or sdp.get("buyNormal") or 0
                        if sdp_size:
                            sdp_full_map[sdp_size] = sdp
                            digits = re.sub(r'[^0-9.]', '', sdp_size)
                            if digits and digits != sdp_size:
                                sdp_full_map[digits] = sdp
                            if sdp_buy:
                                sdp_map[sdp_size] = sdp_buy
                                if digits and digits != sdp_size:
                                    sdp_map[digits] = sdp_buy

                    # ── 내 입찰 제외 함수 ──
                    def _exclude_my_bids_price(raw_price, pid, sz_name):
                        """
                        raw_price가 내 입찰가와 같으면 sell_bids에서 내 입찰 제거 후
                        남은 최저가를 반환. 경쟁자가 없으면 0.
                        """
                        if not raw_price or not my_bids_for_filter:
                            return raw_price

                        # 사이즈 키 매칭 (W215 ↔ 215)
                        sz_digits = re.sub(r'[^0-9.]', '', sz_name)
                        my_prices = []
                        for k_suffix in [sz_name, sz_digits, "ONE SIZE"]:
                            key = f"{pid}_{k_suffix}"
                            if key in my_bids_for_filter:
                                my_prices.extend(my_bids_for_filter[key])

                        if not my_prices:
                            return raw_price

                        # sell_bids에서 해당 사이즈의 모든 가격 수집
                        sell_bids_raw = kream.get("sell_bids", [])
                        sz_sell_prices = []
                        for sb in sell_bids_raw:
                            sb_sz = str(sb.get("size", "")).strip()
                            sb_digits = re.sub(r'[^0-9.]', '', sb_sz)
                            if sb_sz == sz_name or sb_digits == sz_digits or sz_name == "ONE SIZE":
                                for _ in range(sb.get("quantity", 1)):
                                    sz_sell_prices.append(sb["price"])

                        if not sz_sell_prices:
                            # sell_bids에 해당 사이즈 없으면 raw_price에서 직접 판단
                            if raw_price in my_prices:
                                add_log(tid, "warn",
                                    f"  사이즈 {sz_name}: 즉시구매가 {raw_price:,}원 = 내 입찰 → 경쟁자 없음")
                                return 0
                            return raw_price

                        # 내 입찰 수량만큼 제거
                        remaining = list(sz_sell_prices)
                        for mp in my_prices:
                            if mp in remaining:
                                remaining.remove(mp)

                        if not remaining:
                            add_log(tid, "info",
                                f"  사이즈 {sz_name}: 전체 판매입찰이 내 입찰 → 경쟁자 없음 (내가 최저가)")
                            return 0

                        competitor_low = min(remaining)
                        if competitor_low != raw_price:
                            add_log(tid, "info",
                                f"  사이즈 {sz_name}: 즉시구매가 {raw_price:,}원 → 내 입찰 제외 → 경쟁자 최저가 {competitor_low:,}원")
                        return competitor_low

                    # 전체 즉시구매가에서도 내 입찰 제외
                    if instant_buy and product_id:
                        item_sz = str(item.get("size", "ONE SIZE")).strip()
                        instant_buy = _exclude_my_bids_price(instant_buy, product_id, item_sz)

                    if input_sizes:
                        size_margins = []
                        for sz in input_sizes:
                            sz_cny = float(sz.get("cny_price", 0))
                            sz_name = str(sz["size"]).strip()
                            mi = calculate_margin_for_queue(
                                sz_cny, item["category"], item["shipping"]
                            )
                            # 사이즈별 즉시구매가 매칭 (해외배송 우선)
                            sz_sdp = sdp_full_map.get(sz_name, {})
                            sz_instant_buy = (
                                sz_sdp.get("buyOverseas")
                                or sz_sdp.get("buyPrice")
                                or sdp_map.get(sz_name, 0)
                            )
                            # ★ 내 입찰 제외
                            if sz_instant_buy and product_id:
                                sz_instant_buy = _exclude_my_bids_price(sz_instant_buy, product_id, sz_name)
                            add_log(tid, "info",
                                f"  사이즈 {sz_name}: 즉시구매가={sz_instant_buy or '매칭실패(경쟁자없음)'}"
                                f" (sdp_map keys: {list(sdp_map.keys())[:10]})")
                            sz_comp = {}
                            if sz_instant_buy and sz_sdp:
                                sz_comp = analyze_competitiveness(
                                    sz_instant_buy, item["category"], sz_sdp)
                            size_margins.append({
                                "size": sz_name,
                                "cny": sz_cny,
                                "totalCost": mi["total_cost"],
                                "margins": mi["margins"],
                                "instantBuyPrice": sz_instant_buy,
                                "comp": sz_comp,
                            })
                        # 대표 마진 (최저 CNY 기준)
                        min_cost = min(sm["totalCost"] for sm in size_margins)
                        max_cost = max(sm["totalCost"] for sm in size_margins)
                        rep_cny = min(float(sz.get("cny_price", 0)) for sz in input_sizes)
                        margin_info = calculate_margin_for_queue(
                            rep_cny, item["category"], item["shipping"]
                        )
                    else:
                        size_margins = []
                        margin_info = calculate_margin_for_queue(
                            item["cny"], item["category"], item["shipping"]
                        )
                        # 개별 사이즈 항목: 해당 사이즈의 즉시구매가 매칭
                        item_size = str(item.get("size", "")).strip()
                        if item_size and sdp_map:
                            sz_buy = sdp_map.get(item_size, 0)
                            if sz_buy:
                                # ★ 내 입찰 제외
                                sz_buy = _exclude_my_bids_price(sz_buy, product_id, item_size)
                                instant_buy = sz_buy
                                add_log(tid, "info",
                                    f"  사이즈 {item_size}: sdp_map에서 즉시구매가={sz_buy or '경쟁자없음'}원 매칭")
                            else:
                                add_log(tid, "info",
                                    f"  사이즈 {item_size}: sdp_map에 없음 (keys: {list(sdp_map.keys())[:10]})")
                        elif item_size:
                            add_log(tid, "info",
                                f"  사이즈 {item_size}: sdp_map 비어있음 (sizeDeliveryPrices 수집 실패)")

                    # ONE SIZE / 개별 사이즈 경쟁력 분석
                    one_size_comp = {}
                    if not size_margins and instant_buy:
                        os_key = item.get("size", "") or "ONE SIZE"
                        os_sdp = sdp_full_map.get(os_key, {})
                        if not os_sdp and sdp_full_map:
                            os_sdp = next(iter(sdp_full_map.values()), {})
                        if os_sdp:
                            one_size_comp = analyze_competitiveness(
                                instant_buy, item["category"], os_sdp)

                    # 시장 분류 계산
                    market_info = {"market_type": "데이터 부족", "market_color": "gray",
                                   "avg_margin_rate": None, "profitable_count": 0,
                                   "total_count": 0, "details": []}
                    if size_margins:
                        market_info = classify_market(size_margins)
                    elif instant_buy and margin_info["total_cost"]:
                        market_info = classify_market([{
                            "size": item.get("size", "ONE SIZE"),
                            "totalCost": margin_info["total_cost"],
                            "instantBuyPrice": instant_buy,
                        }])

                    item["result"] = {
                        "productId": product_id,
                        "nameKr": name_kr,
                        "nameEn": name_en,
                        "brand": kream.get("brand", ""),
                        "kreamSizes": kream.get("sizes", []),
                        "totalCost": margin_info["total_cost"],
                        "margins": margin_info["margins"],
                        "krwPrice": margin_info["krw_price"],
                        "customs": margin_info["customs"],
                        "importVat": margin_info["import_vat"],
                        "sizeMargins": size_margins,
                        "gosi": gosi,
                        # KREAM 가격 (명확히 구분)
                        "instantBuyPrice": instant_buy,       # 즉시구매가 = 현재 판매입찰 최저가
                        "instantSellPrice": kream.get("instant_sell_price"),  # 즉시판매가 = 현재 구매입찰 최고가
                        "recentTradePrice": recent_trade,     # 최근 체결가 (과거)
                        "totalTrades": kream.get("total_trades"),
                        "sellBids": kream.get("sell_bids", []),
                        "buyBids": kream.get("buy_bids", []),
                        "sizeDeliveryPrices": kream.get("size_delivery_prices", []),
                        # 수집 실패 플래그
                        "collectFailed": not bool(kream.get("size_delivery_prices")),
                        # 시장 분류
                        "marketType": market_info["market_type"],
                        "marketColor": market_info["market_color"],
                        "avgMarginRate": market_info["avg_margin_rate"],
                        "profitableCount": market_info["profitable_count"],
                        "marketTotalCount": market_info["total_count"],
                        "marketDetails": market_info["details"],
                        # 경쟁력 분석 (ONE SIZE용)
                        "comp": one_size_comp,
                    }
                    item["status"] = "완료"
                    cost_str = f"원가 {margin_info['total_cost']:,}원"
                    if input_sizes:
                        cost_str = f"{len(input_sizes)}사이즈, 원가 {min_cost:,}~{max_cost:,}원"
                    add_log(tid, "success",
                            f"{model}: {name_kr or name_en} → {cost_str}, 즉시구매가 {instant_buy or 0:,}원")

                except Exception as e:
                    item["status"] = "실패"
                    item["result"] = {"error": str(e)}
                    add_log(tid, "error", f"{model}: {e}")

            loop.close()

            done = sum(1 for q in pending if q["status"] == "완료")
            add_log(tid, "success", f"일괄 실행 완료: {done}/{len(pending)}건 성공")
            save_queue()

            # batch 히스토리 저장 (검색+마진 계산)
            batch_items = []
            for item in pending:
                r = item.get("result", {})
                margin_str = None
                if item["status"] == "완료" and r.get("totalCost"):
                    mg = (r.get("margins") or {}).get("margin_15", {})
                    if mg.get("profit"):
                        margin_str = f"+{mg['profit']:,.0f} ({mg.get('margin_rate',0):.1f}%)"
                batch_items.append({
                    "model": item["model"],
                    "name": r.get("nameKr") or r.get("nameEn") or "-",
                    "size": item.get("size", ""),
                    "bid_price": None,
                    "cost": r.get("totalCost"),
                    "instant_buy": r.get("instantBuyPrice"),
                    "margin": margin_str,
                    "status": item["status"],
                })
            save_batch_history("검색+마진계산", batch_items)

            finish_task(tid, result={"queue": product_queue})

        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/market-check", methods=["POST"])
def api_market_check():
    """모델번호 입력 → KREAM 시세 수집 → 得物 원가 비교 → 시장 분류 반환"""
    data = request.json or {}
    model = (data.get("model") or "").strip().upper()
    if not model:
        return jsonify({"error": "모델번호를 입력해주세요"}), 400

    # 1) 得物 가격 DB 조회 (없어도 계속 진행)
    dewu = get_dewu_prices(model)
    has_dewu = bool(dewu)

    # 2) KREAM 즉시구매가 조회 (큐에 완료 항목이 있으면 재사용)
    kream_prices = {}  # kr_size → sell_price
    kream_result = None
    for q in product_queue:
        if q.get("model", "").upper() == model and q.get("status") == "완료":
            kream_result = q.get("result", {})
            sdp = kream_result.get("sizeDeliveryPrices", [])
            for s in sdp:
                sz = str(s.get("size", "")).strip()
                bp = s.get("buyPrice") or s.get("buyNormal") or 0
                if sz and bp:
                    digits = re.sub(r'[^0-9.]', '', sz)
                    kream_prices[digits] = bp
                    kream_prices[sz] = bp
            if not kream_prices and kream_result.get("instantBuyPrice"):
                kream_prices["ALL"] = kream_result["instantBuyPrice"]
            break

    # 得物 데이터도 KREAM 데이터��� 없으면 에러
    if not has_dewu and not kream_prices:
        return jsonify({
            "error": f"분석 데이터가 없습니다: {model}",
            "hint": "큐에서 먼저 일괄 실행하여 KREAM 가격을 수집하거나, 得物 가격을 등록해주세요."
        }), 404

    # 3) 사이즈별 마진 계산
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    cny_rate = float(settings.get("cnyRate", 218.12))
    cny_margin = float(settings.get("cnyMargin", 1.03))

    size_data = []

    if has_dewu:
        # 得物 가격 기반 분석
        for key, info in dewu["sizes"].items():
            cny = info["cny"]
            kr_size = info["kr_size"] or info["eu_size"]
            krw_buy = round(cny * cny_rate * cny_margin)
            total_cost = krw_buy + 8000
            kream_sell = kream_prices.get(kr_size) or kream_prices.get(info["eu_size"]) or kream_prices.get("ALL") or 0
            size_data.append({
                "size": kr_size,
                "eu_size": info["eu_size"],
                "cny": cny,
                "totalCost": total_cost,
                "instantBuyPrice": kream_sell,
            })
    elif kream_result:
        # KREAM 데이터만으로 분석 (큐 결과 활용)
        sm = kream_result.get("sizeMargins", [])
        if sm:
            for s in sm:
                size_data.append({
                    "size": s.get("size", ""),
                    "eu_size": "",
                    "cny": 0,
                    "totalCost": s.get("totalCost", 0),
                    "instantBuyPrice": s.get("instantBuyPrice", 0),
                })
        elif kream_result.get("totalCost") and kream_result.get("instantBuyPrice"):
            size_data.append({
                "size": "ALL",
                "eu_size": "",
                "cny": 0,
                "totalCost": kream_result["totalCost"],
                "instantBuyPrice": kream_result["instantBuyPrice"],
            })

    # 4) 시장 분류
    market = classify_market(size_data) if size_data else {
        "market_type": "데이터 부족", "market_color": "gray",
        "avg_margin_rate": None, "profitable_count": 0,
        "total_count": 0, "details": []
    }

    # 5) 마진 양호 사이즈 목록
    good_sizes = [d["size"] for d in market["details"] if d.get("margin_rate") is not None and d["margin_rate"] >= 10]
    ok_sizes = [d["size"] for d in market["details"] if d.get("margin_rate") is not None and 0 <= d["margin_rate"] < 10]

    # 메시지 생성
    data_source = "得物+KREAM" if has_dewu else "KREAM"
    message = f"[{data_source} 기준] 이 상품은 {market['market_type']}입니다."
    if market["market_type"] == "혼합 시장" and good_sizes:
        message += f" {', '.join(good_sizes)} 사이즈만 마진이 충분합니다."
    elif market["market_type"] == "혼합 시장" and ok_sizes:
        message += f" {', '.join(ok_sizes)} 사이즈는 소량 마진이 남습니다."
    elif market["market_type"] == "비정상 시장":
        message += " 평균 마진율이 마이너스입니다. 입찰 비추천."
    elif market["market_type"] == "정상 시장":
        message += " 입찰 추천."
    if not has_dewu:
        message += " (得物 가격 미등록 — KREAM 시세로만 분석)"

    return jsonify({
        "model": model,
        "brand": dewu["brand"] if has_dewu else (kream_result or {}).get("brand", ""),
        "market_type": market["market_type"],
        "market_color": market["market_color"],
        "avg_margin_rate": market["avg_margin_rate"],
        "profitable_count": market["profitable_count"],
        "total_count": market["total_count"],
        "good_sizes": good_sizes,
        "ok_sizes": ok_sizes,
        "message": message,
        "details": market["details"],
        "has_kream_prices": bool(kream_prices),
        "has_dewu_prices": has_dewu,
    })


@app.route("/api/queue/auto-register", methods=["POST"])
def api_queue_auto_register():
    """큐에서 선택된 상품들을 자동으로 고시정보 등록 + 입찰 (Playwright)"""
    data = request.json or {}
    bid_items = data.get("items", [])
    # items: [{productId, model, size, price, quantity, gosi:{...}, category, gosiAlready}]

    if not bid_items:
        return jsonify({"error": "항목 없음"}), 400

    # CNY 필수 검증
    require_cny = True
    if SETTINGS_FILE.exists():
        try:
            _s = json.loads(SETTINGS_FILE.read_text())
            require_cny = _s.get("require_cny_on_bid", True)
        except Exception:
            pass
    if require_cny:
        missing_cny = [bi for bi in bid_items if not bi.get("cny_price") or float(bi.get("cny_price", 0)) <= 0]
        if missing_cny:
            models = set(bi.get("model", "?") for bi in missing_cny)
            return jsonify({"error": f"원가(CNY)는 필수입니다 ({len(missing_cny)}건: {', '.join(models)}). 설정에서 해제 가능"}), 400

    # 세션 파일 확인
    if not Path(STATE_FILE).exists():
        return jsonify({"error": "세션 없음. 먼저 python3 kream_bot.py --mode login 실행"}), 400

    tid = new_task()
    add_log(tid, "info", f"자동 입찰 시작: {len(bid_items)}건")

    with auto_bid_lock:
        auto_bid_control["state"] = "running"
        auto_bid_event.set()

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            results = []
            stopped = False
            gosi_done_pids = set()  # 이 배치에서 고시정보 등록 완료한 productId

            # ── 같은 productId를 그룹핑 ──
            # 순서를 유지하면서 productId별로 그룹핑
            from collections import OrderedDict
            pid_groups = OrderedDict()  # pid → [items]
            resolved_items = []  # pid가 확정된 아이템 목록

            # 1단계: pid 확인 (검색 필요하면 검색)
            for i, bi in enumerate(bid_items, 1):
                pid = bi.get("productId") or 0
                price = bi["price"]
                size = bi.get("size", "")
                model = bi.get("model", "")

                # Step 17-A: 사이즈 필수 검증 (신발 ONE SIZE 차단)
                is_valid, err_msg, _cat = validate_size_for_bid(model, size)
                if not is_valid:
                    add_log(tid, "error", f"[{i}] {model}: {err_msg}")
                    results.append({"productId": str(pid) if pid else "", "model": model,
                        "size": size, "price": price, "success": False, "error": err_msg})
                    continue

                # Step 17-E: 카테고리 미결정 차단 (큐 execute에서 0순위 DB로 채워졌어야 함)
                bi_category = str(bi.get("category", "")).strip()
                cat_ok, cat_err = validate_category_for_bid(model, bi_category)
                if not cat_ok:
                    add_log(tid, "error", f"[{i}] {model}: {cat_err}")
                    results.append({"productId": str(pid) if pid else "", "model": model,
                        "size": size, "price": price, "success": False,
                        "error": cat_err, "skip_reason": "category_undecided"})
                    continue

                # Step 17-E 보완: gosi 미결정 차단 (큐 execute에서 auto_fill_gosi 실패한 항목)
                bi_gosi = bi.get("gosi")
                gosi_ok, gosi_err = validate_gosi_for_bid(bi_gosi)
                if not gosi_ok:
                    add_log(tid, "error", f"[{i}] {model}: {gosi_err}")
                    results.append({"productId": str(pid) if pid else "", "model": model,
                        "size": size, "price": price, "success": False,
                        "error": gosi_err, "skip_reason": "gosi_undecided"})
                    continue

                if not pid or str(pid) == "0":
                    if model:
                        add_log(tid, "info", f"[준비] {model} 상품번호 검색 중...")
                        try:
                            search_results = loop.run_until_complete(search_by_model(model))
                            if search_results:
                                kream_data = search_results[0].get("kream", {})
                                pid = str(kream_data.get("product_id", ""))
                            if not pid or pid == "0":
                                add_log(tid, "error", f"{model}: 상품번호를 찾을 수 없음")
                                results.append({"productId": pid, "model": model,
                                    "size": size, "price": price, "success": False})
                                continue
                            add_log(tid, "info", f"{model} → #{pid}")
                        except Exception as e:
                            add_log(tid, "error", f"{model}: 검색 실패 — {e}")
                            results.append({"productId": 0, "model": model,
                                "size": size, "price": price, "success": False})
                            continue
                    else:
                        add_log(tid, "error", f"[{i}] productId와 model 모두 없음")
                        results.append({"productId": 0, "model": "",
                            "size": size, "price": price, "success": False})
                        continue

                bi["_resolved_pid"] = str(pid)
                pid_key = str(pid)
                if pid_key not in pid_groups:
                    pid_groups[pid_key] = []
                pid_groups[pid_key].append(bi)

            # 2단계: productId별로 고시정보 + 일괄 입찰 실행
            total_items = sum(len(g) for g in pid_groups.values())
            processed = 0
            for pid, group in pid_groups.items():
                # 일시정지/중단 체크
                if not auto_bid_event.is_set():
                    add_log(tid, "info", "⏸ 일시정지 중...")
                auto_bid_event.wait()
                with auto_bid_lock:
                    if auto_bid_control["state"] == "stopping":
                        add_log(tid, "info", f"⏹ 중단 — {processed}/{total_items}건 처리됨")
                        stopped = True
                        break

                first = group[0]
                model = first.get("model", "")
                gosi = first.get("gosi", {})
                gosi_already = first.get("gosiAlready", False)
                category = first.get("category", "가방")
                bid_days = int(first.get("bid_days", 30))

                sizes_str = ", ".join(bi.get("size", "?") for bi in group)
                add_log(tid, "info",
                        f"[#{pid}] {model} — {len(group)}사이즈: {sizes_str}")

                # 고시정보 등록 (중복 방지)
                if not gosi_already and pid not in gosi_done_pids:
                    add_log(tid, "info", f"  고시정보 등록 중... #{pid}")
                    try:
                        gosi_result = loop.run_until_complete(
                            _run_gosi_only(pid, gosi, category, tid)
                        )
                        if gosi_result:
                            gosi_done_pids.add(pid)
                            add_log(tid, "success", f"  고시정보 등록 완료")
                        else:
                            add_log(tid, "error", f"  고시정보 등록 실패")
                            for bi in group:
                                results.append({"productId": pid, "model": model,
                                    "size": bi.get("size"), "price": bi["price"], "success": False})
                                processed += 1
                            continue
                    except Exception as e:
                        add_log(tid, "error", f"  고시정보 오류: {e}")
                        for bi in group:
                            results.append({"productId": pid, "model": model,
                                "size": bi.get("size"), "price": bi["price"], "success": False})
                            processed += 1
                        continue
                else:
                    skip_reason = "gosiAlready" if gosi_already else "배치 내 이미 등록"
                    add_log(tid, "info", f"  고시정보 스킵 ({skip_reason})")

                # 입찰: 여러 사이즈면 일괄, 1사이즈면 기존 방식
                if len(group) > 1:
                    # 일괄 입찰
                    batch_bids = [{"size": bi.get("size", "ONE SIZE"),
                                   "price": bi["price"],
                                   "qty": bi.get("quantity", 1)} for bi in group]
                    add_log(tid, "info",
                            f"  일괄 입찰 {len(batch_bids)}사이즈 진행 중...")
                    try:
                        batch_result = loop.run_until_complete(
                            _run_batch_bid(pid, batch_bids, bid_days, tid)
                        )
                        ok = batch_result.get("success", 0)
                        fail = batch_result.get("fail", 0)
                        add_log(tid, "success" if ok > 0 else "error",
                                f"  일괄 입찰 결과: 성공 {ok}건, 실패 {fail}건")
                        for bi_result in batch_result.get("results", []):
                            matched_bi = next((b for b in group if b.get("size") == bi_result["size"]), group[0])
                            results.append({"productId": pid, "model": model,
                                "size": bi_result["size"], "price": bi_result["price"],
                                "success": bi_result.get("ok", False)})
                            processed += 1
                            if bi_result.get("ok"):
                                save_bid_local(pid, model=model, size=bi_result["size"],
                                              price=bi_result["price"], source="placed")
                                _cny = matched_bi.get("cny_price", 0)
                                try:
                                    saved = _save_bid_cost(
                                        order_id=bi_result.get("orderId") or f"{pid}_{bi_result['size']}",
                                        model=model, size=bi_result["size"],
                                        cny_price=float(_cny) if _cny else None,
                                        exchange_rate=float(matched_bi.get("exchange_rate", 0)),
                                        overseas_shipping=int(matched_bi.get("overseas_shipping", 8000)),
                                        cny_source=("manual" if _cny and float(_cny) > 0 else None),
                                    )
                                    if not saved:
                                        add_log(tid, "warn", f"  [{model} {bi_result['size']}] 원가 미저장 (식货 매칭 실패)")
                                except Exception as e:
                                    add_log(tid, "error", f"  bid_cost 저장 실패: {e}")
                    except Exception as e:
                        add_log(tid, "error", f"  일괄 입찰 오류: {e}")
                        for bi in group:
                            results.append({"productId": pid, "model": model,
                                "size": bi.get("size"), "price": bi["price"], "success": False})
                            processed += 1
                else:
                    # 단일 사이즈 — 기존 방식
                    bi = group[0]
                    price = bi["price"]
                    size = bi.get("size", "ONE SIZE")
                    qty = bi.get("quantity", 1)
                    add_log(tid, "info",
                            f"  입찰: {size} → {price:,}원 × {qty}개 ({bid_days}일)")
                    try:
                        result = loop.run_until_complete(
                            _run_bid_only(pid, price, size, qty, bid_days, tid, model)
                        )
                        ok = result.get("success", False)
                        results.append({"productId": pid, "model": model,
                            "size": size, "price": price, "success": ok})
                        processed += 1
                        if ok:
                            save_bid_local(pid, model=model, size=size,
                                          price=price, source="placed")
                            _cny = bi.get("cny_price", 0)
                            try:
                                saved = _save_bid_cost(
                                    order_id=result.get("orderId") or f"{pid}_{size}",
                                    model=model, size=size,
                                    cny_price=float(_cny) if _cny else None,
                                    exchange_rate=float(bi.get("exchange_rate", 0)),
                                    overseas_shipping=int(bi.get("overseas_shipping", 8000)),
                                    cny_source=("manual" if _cny and float(_cny) > 0 else None),
                                )
                                if not saved:
                                    add_log(tid, "warn", f"  [{model} {size}] 원가 미저장 (식货 매칭 실패)")
                            except Exception as e:
                                add_log(tid, "error", f"  bid_cost 저장 실패: {e}")
                    except Exception as e:
                        add_log(tid, "error", f"  입찰 오류: {e}")
                        results.append({"productId": pid, "model": model,
                            "size": size, "price": price, "success": False})
                        processed += 1

            loop.close()

            ok = sum(1 for r in results if r["success"])
            total_attempted = len(results)
            if stopped:
                add_log(tid, "info", f"중단됨: {ok}/{total_attempted}건 성공 (전체 {len(bid_items)}건 중 {total_attempted}건 처리)")
            else:
                add_log(tid, "success", f"완료: {ok}/{len(bid_items)}건 성공")

            # batch 히스토리 저장
            batch_items = []
            for bi, r in zip(bid_items[:total_attempted], results):
                margin_str = None
                if r["success"]:
                    cost = bi.get("cost", 0)
                    if cost and r.get("price"):
                        pi = _calc_profit_simple(r["price"], cost)
                        margin_str = pi
                batch_items.append({
                    "model": bi.get("model", ""),
                    "name": bi.get("nameEn", "") or bi.get("name", "-"),
                    "size": r.get("size", bi.get("size", "")),
                    "bid_price": r.get("price"),
                    "margin": margin_str,
                    "status": "입찰완료" if r["success"] else "실패",
                })
            if batch_items:
                save_batch_history("자동입찰", batch_items)

            finish_task(tid, result={"results": results, "success": ok, "total": len(bid_items), "stopped": stopped})

        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))
        finally:
            with auto_bid_lock:
                auto_bid_control["state"] = "idle"
                auto_bid_event.set()

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/auto-bid/pause", methods=["POST"])
def api_auto_bid_pause():
    with auto_bid_lock:
        if auto_bid_control["state"] != "running":
            return jsonify({"error": "실행 중이 아닙니다"}), 400
        auto_bid_control["state"] = "paused"
        auto_bid_event.clear()  # 다음 상품 전에 대기
    return jsonify({"ok": True, "state": "paused"})


@app.route("/api/auto-bid/resume", methods=["POST"])
def api_auto_bid_resume():
    with auto_bid_lock:
        if auto_bid_control["state"] != "paused":
            return jsonify({"error": "일시정지 상태가 아닙니다"}), 400
        auto_bid_control["state"] = "running"
        auto_bid_event.set()  # 대기 해제
    return jsonify({"ok": True, "state": "running"})


@app.route("/api/auto-bid/stop", methods=["POST"])
def api_auto_bid_stop():
    with auto_bid_lock:
        if auto_bid_control["state"] not in ("running", "paused"):
            return jsonify({"error": "실행 중이 아닙니다"}), 400
        auto_bid_control["state"] = "stopping"
        auto_bid_event.set()  # paused 상태에서도 깨움
    return jsonify({"ok": True, "state": "stopping"})


@app.route("/api/auto-bid/status")
def api_auto_bid_status():
    with auto_bid_lock:
        return jsonify({"state": auto_bid_control["state"]})


# ═══════════════════════════════════════════
# API: 실행 이력
# ═══════════════════════════════════════════

@app.route("/api/history")
def api_history():
    if HISTORY_FILE.exists():
        return jsonify(json.loads(HISTORY_FILE.read_text()))
    return jsonify([])


def save_history(task_type, product_id, price, qty, success):
    history = []
    if HISTORY_FILE.exists():
        try:
            history = json.loads(HISTORY_FILE.read_text())
        except Exception:
            pass

    history.insert(0, {
        "date": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "type": task_type,
        "productId": product_id,
        "price": price,
        "quantity": qty,
        "success": 1 if success else 0,
        "fail": 0 if success else 1,
    })
    history = history[:200]
    HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2))


def save_batch_history(batch_type, items_detail):
    """일괄 실행 단위(batch)로 히스토리 저장
    items_detail: [{model, name, bid_price, margin, status}, ...]
    """
    history = []
    if BATCH_HISTORY_FILE.exists():
        try:
            data = json.loads(BATCH_HISTORY_FILE.read_text())
            history = data.get("history", [])
        except Exception:
            pass

    total = len(items_detail)
    success = sum(1 for it in items_detail if "완료" in (it.get("status") or ""))
    failed = total - success

    entry = {
        "executed_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "type": batch_type,
        "total": total,
        "success": success,
        "failed": failed,
        "items": items_detail,
    }

    history.insert(0, entry)
    history = history[:30]  # 최근 30건만 유지

    BATCH_HISTORY_FILE.write_text(json.dumps(
        {"history": history}, ensure_ascii=False, indent=2
    ))


@app.route("/api/batch-history")
def api_batch_history():
    """일괄 실행 이력 조회 (최근 30건)"""
    if BATCH_HISTORY_FILE.exists():
        try:
            return jsonify(json.loads(BATCH_HISTORY_FILE.read_text()))
        except Exception:
            pass
    return jsonify({"history": []})


# ═══════════════════════════════════════════
# 내 입찰 현황 로컬 저장 헬퍼
# ═══════════════════════════════════════════

def load_my_bids_local():
    if MY_BIDS_FILE.exists():
        try:
            return json.loads(MY_BIDS_FILE.read_text())
        except Exception:
            pass
    return {"bids": [], "lastSync": None}


def save_bid_local(product_id, model="", size="ONE SIZE", price=0, source="placed", order_id=None):
    """입찰 성공 시 로컬 JSON에 기록 (동일 상품+사이즈는 덮어씀)"""
    data = load_my_bids_local()
    data["bids"] = [b for b in data["bids"]
                    if not (str(b.get("productId")) == str(product_id) and b.get("size") == size)]
    data["bids"].append({
        "productId": str(product_id),
        "model": model,
        "size": size,
        "price": price,
        "date": datetime.now().strftime("%Y/%m/%d %H:%M"),
        "source": source,
        "orderId": order_id,
    })
    MY_BIDS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))


@app.route("/api/my-bids/local")
def api_my_bids_local():
    """로컬 저장된 내 입찰 현황 조회"""
    return jsonify(load_my_bids_local())


@app.route("/api/my-bids/sync", methods=["POST"])
def api_my_bids_sync():
    """판매자센터에서 내 입찰 목록 동기화 → 로컬 JSON 저장"""
    tid = new_task()
    add_log(tid, "info", "내 입찰 현황 동기화 시작...")

    def run():
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            from kream_adjuster import collect_my_bids_via_menu
            bids = loop.run_until_complete(collect_my_bids_via_menu(headless=get_headless()))
            loop.close()

            data = {
                "bids": [{
                    "productId": str(b.get("productId", "")),
                    "model": b.get("model", ""),
                    "size": b.get("size", "ONE SIZE"),
                    "price": b.get("bidPrice", 0),
                    "rank": b.get("bidRank"),
                    "date": datetime.now().strftime("%Y/%m/%d %H:%M"),
                    "source": "sync",
                    "orderId": b.get("orderId", ""),
                    "nameKr": b.get("nameKr", ""),
                    "rawText": b.get("rawText", ""),
                } for b in bids],
                "lastSync": datetime.now().strftime("%Y/%m/%d %H:%M"),
            }
            MY_BIDS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))
            add_log(tid, "success", f"동기화 완료: {len(bids)}건 저장")
            finish_task(tid, result=data)
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# 입찰 순위 모니터링 + 가격 자동 조정
# ═══════════════════════════════════════════


def _get_next_monitor_time():
    """다음 모니터링 실행 시간 계산 (MONITOR_HOURS 기반)"""
    now = datetime.now()
    for h in MONITOR_HOURS:
        target = now.replace(hour=h, minute=0, second=0, microsecond=0)
        if target > now:
            return target
    tomorrow = now + timedelta(days=1)
    return tomorrow.replace(hour=MONITOR_HOURS[0], minute=0, second=0, microsecond=0)


def _schedule_next_monitor():
    """다음 모니터링 타이머 등록"""
    global _monitor_timer
    with _monitor_lock:
        if not monitor_state["running"]:
            return
        next_time = _get_next_monitor_time()
        delay = max(60, (next_time - datetime.now()).total_seconds())
        monitor_state["next_run"] = next_time.strftime("%Y-%m-%d %H:%M")
        _monitor_timer = threading.Timer(delay, _monitor_trigger)
        _monitor_timer.daemon = True
        _monitor_timer.start()
        print(f"[모니터] 다음 실행: {monitor_state['next_run']} ({delay:.0f}초 후)")


def _monitor_trigger():
    """타이머 콜백 → 모니터링 실행 + 다음 스케줄"""
    threading.Thread(target=_run_monitor_check, daemon=True).start()
    _schedule_next_monitor()


def _calc_settlement_for_monitor(sell_price):
    """판매가에 대한 정산액 계산"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    fee_rate = float(settings.get("feeRate", 0.06))
    fixed_fee = int(settings.get("fixedFee", 2500))
    vat_rate = float(settings.get("vatRate", 0.10))
    return round(sell_price * (1 - fee_rate * (1 + vat_rate)) - fixed_fee)


def _find_cost_for_bid(bid):
    """입찰의 원가(total_cost) 찾기 — bid_cost DB → 큐 메모리 순서로 조회"""
    order_id = bid.get("orderId") or ""
    model = (bid.get("model") or "").upper()
    size = bid.get("size") or ""

    # 1) bid_cost 테이블에서 조회 (order_id 매칭 또는 model+size 매칭)
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        row = None
        if order_id:
            c.execute("SELECT * FROM bid_cost WHERE order_id=?", (order_id,))
            row = c.fetchone()
        if not row and model:
            c.execute("SELECT * FROM bid_cost WHERE UPPER(model)=? AND size=? ORDER BY created_at DESC LIMIT 1",
                      (model, size))
            row = c.fetchone()
        conn.close()
        if row:
            r = dict(row)
            cny = r.get("cny_price", 0)
            rate = r.get("exchange_rate", 0)
            ship = r.get("overseas_shipping", 8000)
            other = r.get("other_costs", 0)
            if cny and rate:
                return round(cny * rate * 1.03 + ship + other)
    except Exception:
        pass

    # 2) 큐 메모리에서 조회 (기존 동작)
    for item in product_queue:
        if (item.get("model") or "").upper() != model:
            continue
        result = item.get("result")
        if result and result.get("total_cost"):
            return result["total_cost"]
    return None


def _log_bid_competition(bids, market):
    """각 입찰에 대해 bid_competition_log에 한 줄씩 기록"""
    conn = sqlite3.connect(str(PRICE_DB))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logged = 0

    for bid in bids:
        pid = bid.get("productId")
        my_price = bid.get("bidPrice", 0)
        if not my_price:
            continue

        mkt = market.get(pid, {})
        sell_bids = mkt.get("sell_bids", [])

        # 시장 최저가 (내 입찰가 제외)
        all_prices = sorted(set(s["price"] for s in sell_bids)) if sell_bids else []
        competitor_prices = [p for p in all_prices if p != my_price]
        market_lowest = competitor_prices[0] if competitor_prices else None
        am_i_lowest = 1 if (market_lowest is None or my_price < market_lowest) else 0

        # 경쟁자 수 (같은 사이즈의 판매 입찰 총 개수, 내 것 제외)
        bid_size = bid.get("size", "ONE SIZE")
        competitor_count = None
        if sell_bids:
            same_size_bids = [s for s in sell_bids if s.get("size") == bid_size]
            # 전체 개수에서 내 입찰 1개 차감
            competitor_count = max(0, len(same_size_bids) - 1)

        # 마진 계산: 정산액 - 원가
        my_margin = None
        try:
            total_cost = _find_cost_for_bid(bid)
            if total_cost is not None:
                settlement = _calc_settlement_for_monitor(my_price)
                my_margin = settlement - total_cost
        except Exception:
            pass  # 계산 실패 시 NULL

        conn.execute(
            """INSERT INTO bid_competition_log
               (product_id, model, size, my_price, market_lowest, am_i_lowest,
                my_margin, competitor_count, checked_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (pid, bid.get("model", ""), bid_size, my_price,
             market_lowest, am_i_lowest, my_margin, competitor_count, now)
        )
        logged += 1

    conn.commit()
    conn.close()
    print(f"[모니터] bid_competition_log: {logged}건 기록")


def _expire_old_pending():
    """24시간 경과한 pending/profit_low/deficit → expired 처리 (승인/거절 건은 보존)"""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute(
            """UPDATE price_adjustments SET status='expired', executed_at=?
               WHERE status IN ('pending','profit_low','deficit')
                 AND created_at < datetime('now','-24 hours')""",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),)
        )
        expired = c.rowcount
        conn.commit()
        conn.close()
        if expired:
            print(f"[모니터] 이전 대기 건 {expired}건 만료 처리 완료")
    except Exception as e:
        print(f"[모니터] 만료 처리 오류: {e}")


def _run_monitor_check():
    """모니터링: 순위 체크 → 가격 조정 계산 → DB 저장 → 이메일"""
    print(f"\n[모니터] ===== 순위 체크: {datetime.now().strftime('%m-%d %H:%M')} =====")
    _expire_old_pending()
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        from kream_adjuster import collect_my_bids, collect_market_data

        # 1) 내 입찰 수집
        print("[모니터] 내 입찰 수집 중...")
        bids = loop.run_until_complete(collect_my_bids(headless=True))
        if not bids:
            print("[모니터] 입찰 없음")
            with _monitor_lock:
                monitor_state["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                monitor_state["total_checks"] += 1
            loop.close()
            return

        # 2) 시장 데이터 수집
        pids = list(set(b["productId"] for b in bids if b.get("productId")))
        print(f"[모니터] 입찰 {len(bids)}건, 상품 {len(pids)}개 시세 수집")
        market = loop.run_until_complete(collect_market_data(pids, headless=True))
        loop.close()

        # 2.5) bid_competition_log 기록 (모든 입찰에 대해)
        try:
            _log_bid_competition(bids, market)
        except Exception as comp_e:
            print(f"[모니터] bid_competition_log 기록 오류: {comp_e}")

        # 3) 순위 분석 + 가격 조정 계산
        adjustments = []
        for bid in bids:
            pid = bid.get("productId")
            mkt = market.get(pid, {})
            sell_bids = mkt.get("sell_bids", [])
            if not sell_bids:
                continue

            my_price = bid.get("bidPrice", 0)
            if not my_price:
                continue

            # 판매입찰 가격 오름차순 (내 입찰가 제외)
            all_prices = sorted(set(s["price"] for s in sell_bids))
            competitor_prices = [p for p in all_prices if p != my_price]

            if not competitor_prices:
                # 내 입찰만 있고 경쟁자 없음 → 1위, 조정 불필요
                print(f"[모니터] {bid.get('model','')} {bid.get('size','')}: 경쟁자 없음 (내가 1위)")
                continue

            # 경쟁자 최저가
            competitor_low = competitor_prices[0]

            # 내가 이미 경쟁자보다 낮으면 패스
            if my_price <= competitor_low:
                continue

            # 새 가격 = 경쟁자 최저가 - 1,000원 (1,000원 단위 올림)
            new_price = int(math.ceil((competitor_low - 1000) / 1000) * 1000)
            if new_price <= 0:
                continue

            # 수익 계산
            total_cost = _find_cost_for_bid(bid)
            if total_cost is not None:
                expected_profit = _calc_settlement_for_monitor(new_price) - total_cost
            else:
                expected_profit = None

            # 상태 결정: 수익 마이너스면 deficit, 5,000원 미만이면 profit_low
            status = "pending"
            if expected_profit is not None:
                if expected_profit < 0:
                    status = "deficit"
                elif expected_profit < 5000:
                    status = "profit_low"

            adjustments.append({
                "order_id": bid.get("orderId", ""),
                "product_id": pid,
                "model": bid.get("model", ""),
                "name_kr": bid.get("nameKr", ""),
                "size": bid.get("size", "ONE SIZE"),
                "old_price": my_price,
                "competitor_price": competitor_low,
                "new_price": new_price,
                "expected_profit": expected_profit,
                "status": status,
            })

        # 4) DB 저장
        pending = [a for a in adjustments if a["status"] == "pending"]
        if adjustments:
            _save_adjustments(adjustments)

        # 5) 이메일 (pending 건만)
        if pending:
            _send_adjustment_email(pending)

        with _monitor_lock:
            monitor_state["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            monitor_state["total_checks"] += 1
            monitor_state["total_adjustments"] += len(pending)

        print(f"[모니터] 완료: {len(bids)}건 중 순위 밀림 {len(adjustments)}건, 조정 대상 {len(pending)}건")

        # 순위 변동 알림 추가
        if pending:
            for adj in pending[:5]:  # 최대 5건
                name = adj.get("name_kr") or adj.get("model", "")
                add_notification(
                    "rank_change",
                    f"{name} {adj.get('size','')}: 순위 변동",
                    f"현재가 {adj['old_price']:,}원 → 경쟁자 {adj['competitor_price']:,}원, 추천가 {adj['new_price']:,}원",
                    "/api/adjust/pending"
                )

        # 조건부 입찰 체크
        try:
            _check_conditional_bids()
        except Exception as ce:
            print(f"[모니터] 조건부 입찰 체크 오류: {ce}")

        # 만료 임박 체크
        try:
            _check_expiring_bids(bids)
        except Exception as ee:
            print(f"[모니터] 만료 임박 체크 오류: {ee}")

        # 모니터링 성공 → 연속 실패 카운터 리셋
        try:
            on_bid_monitor_success()
        except Exception:
            pass

        # 모니터링 완료 후 → 자동 가격 조정 실행 (설정 ON일 때만)
        try:
            aa_settings = _get_auto_adjust_settings()
            if aa_settings["enabled"]:
                print("[모니터] 자동 가격 조정 실행 중...")
                aa_result = auto_execute_approvals()
                print(f"[모니터] 자동 조정 결과: 수정 {aa_result['modified']}, "
                      f"건너뜀 {aa_result['skipped']['total']}, 실패 {aa_result['failed']}")
        except Exception as ae:
            print(f"[모니터] 자동 조정 오류: {ae}")
            traceback.print_exc()

        # 모니터링 완료 후 → 자동 입찰 정리 (설정 ON일 때만)
        try:
            cleanup_settings = _get_cleanup_settings()
            if cleanup_settings["enabled"]:
                print("[모니터] 입찰 정리 실행 중...")
                # 1) 먼저: 유예 지난 pending 처리
                exec_r = run_cleanup_execution()
                print(f"[모니터] 정리 실행: 삭제 {exec_r.get('executed',0)}, 실패 {exec_r.get('failed',0)}")
                # 2) 나중: 새 후보 탐지
                det_r = run_cleanup_detection()
                print(f"[모니터] 정리 탐지: {det_r.get('detected',0)}건 탐지, {det_r.get('saved',0)}건 등록")
        except Exception as cle:
            print(f"[모니터] 입찰 정리 오류: {cle}")
            traceback.print_exc()
    except Exception as e:
        print(f"[모니터] 오류: {e}")
        traceback.print_exc()
        # 모니터링 실패 → 연속 실패 카운터 증가
        try:
            on_bid_monitor_failure(str(e))
        except Exception:
            pass


def _check_expiring_bids(bids):
    """만료일 3일 이내 입찰 감지 → 알림"""
    from datetime import timedelta
    now = datetime.now()
    threshold = now + timedelta(days=3)
    expiring = []

    for bid in bids:
        deadline_str = bid.get("deadline")
        if not deadline_str:
            continue
        try:
            deadline = datetime.strptime(deadline_str, "%Y-%m-%d")
        except ValueError:
            continue
        if deadline <= threshold:
            days_left = (deadline - now).days
            expiring.append({
                "orderId": bid.get("orderId", ""),
                "productId": bid.get("productId", ""),
                "model": bid.get("model", ""),
                "nameKr": bid.get("nameKr", ""),
                "size": bid.get("size", "ONE SIZE"),
                "bidPrice": bid.get("bidPrice", 0),
                "deadline": deadline_str,
                "daysLeft": max(0, days_left),
            })

    if expiring:
        print(f"[모니터] 만료 임박 {len(expiring)}건 감지")
        # my_bids_local.json에 만료 정보 저장
        local_file = BASE_DIR / "my_bids_local.json"
        try:
            local_data = json.loads(local_file.read_text()) if local_file.exists() else {}
        except Exception:
            local_data = {}
        local_data["expiring"] = expiring
        local_file.write_text(json.dumps(local_data, ensure_ascii=False, indent=2))

        for eb in expiring[:5]:
            name = eb.get("nameKr") or eb.get("model", "")
            add_notification(
                "bid_expiry",
                f"입찰 만료 임박: {name} {eb['size']}",
                f"{eb['bidPrice']:,}원 입찰이 {eb['daysLeft']}일 후 만료됩니다",
                "/api/expiring-bids"
            )


def _save_adjustments(adjustments):
    """조정 대상 DB 저장 (같은 order_id로 pending이 있으면 스킵)"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    for adj in adjustments:
        c.execute(
            "SELECT id FROM price_adjustments WHERE order_id=? AND status='pending'",
            (adj["order_id"],)
        )
        if c.fetchone():
            continue
        c.execute(
            """INSERT INTO price_adjustments
            (order_id, product_id, model, name_kr, size, old_price,
             competitor_price, new_price, expected_profit, status, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (adj["order_id"], adj["product_id"], adj["model"],
             adj.get("name_kr", ""), adj["size"], adj["old_price"],
             adj["competitor_price"], adj["new_price"],
             adj["expected_profit"], adj["status"], now)
        )
    conn.commit()
    conn.close()


def _send_adjustment_email(pending):
    """가격 조정 알림 이메일 (Gmail SMTP)"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass

    app_password = settings.get("emailAppPassword", "")
    if not app_password:
        print("[이메일] 앱 비밀번호 미설정 (설정 → emailAppPassword)")
        return

    subject = f"[KREAM] 입찰 순위 변동 {len(pending)}건 - 가격 조정 필요"
    rows = ""
    for a in pending:
        name = a.get("name_kr") or a.get("model", "")
        profit = f"{a['expected_profit']:,}원" if a["expected_profit"] is not None else "미확인"
        rows += (
            f"<tr>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{name}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{a['size']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{a['old_price']:,}원</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{a['competitor_price']:,}원</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd;font-weight:700'>{a['new_price']:,}원</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{profit}</td>"
            f"</tr>"
        )

    body = f"""<html><body style="font-family:-apple-system,sans-serif">
<h2 style="color:#111">KREAM 입찰 순위 변동 알림</h2>
<p>{datetime.now().strftime('%Y-%m-%d %H:%M')} 기준, <b>{len(pending)}건</b>의 가격 조정이 필요합니다.</p>
<table style="border-collapse:collapse;width:100%;font-size:13px">
<thead><tr style="background:#f5f5f5">
<th style="padding:8px;border:1px solid #ddd">상품</th>
<th style="padding:8px;border:1px solid #ddd">사이즈</th>
<th style="padding:8px;border:1px solid #ddd">현재가</th>
<th style="padding:8px;border:1px solid #ddd">경쟁자</th>
<th style="padding:8px;border:1px solid #ddd">추천가</th>
<th style="padding:8px;border:1px solid #ddd">예상수익</th>
</tr></thead>
<tbody>{rows}</tbody>
</table>
<p style="margin-top:20px">
<a href="http://localhost:5001" style="background:#31b46e;color:#fff;padding:12px 28px;
text-decoration:none;border-radius:8px;font-weight:600">대시보드에서 승인</a>
</p>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_RECEIVER
    msg.attach(MIMEText(body, "html", "utf-8"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, app_password)
            server.send_message(msg)
        print(f"[이메일] 발송 완료: {len(pending)}건 알림")
    except Exception as e:
        print(f"[이메일] 발송 실패: {e}")


# ── 모니터링 API ──

@app.route("/api/monitor/status")
def api_monitor_status():
    """모니터링 상태 조회"""
    with _monitor_lock:
        return jsonify(monitor_state.copy())


@app.route("/api/monitor/start", methods=["POST"])
def api_monitor_start():
    """모니터링 시작"""
    with _monitor_lock:
        if monitor_state["running"]:
            return jsonify({"ok": True, "msg": "이미 실행 중"})
        monitor_state["running"] = True
    _schedule_next_monitor()
    return jsonify({"ok": True, "next_run": monitor_state.get("next_run")})


@app.route("/api/monitor/stop", methods=["POST"])
def api_monitor_stop():
    """모니터링 중지"""
    global _monitor_timer
    with _monitor_lock:
        monitor_state["running"] = False
        monitor_state["next_run"] = None
        if _monitor_timer:
            _monitor_timer.cancel()
            _monitor_timer = None
    return jsonify({"ok": True})


@app.route("/api/monitor/run-once", methods=["POST"])
def api_monitor_run_once():
    """수동 1회 모니터링"""
    tid = new_task()
    add_log(tid, "info", "수동 모니터링 시작...")

    def run():
        try:
            _run_monitor_check()
            add_log(tid, "success", "모니터링 완료")
            finish_task(tid, result={"ok": True})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ── 조정 대기/승인/거절 API ──

@app.route("/api/adjust/pending")
def api_adjust_pending():
    """pending/profit_low/deficit 상태의 조정 목록 — bid_cost JOIN으로 실시간 수익 재계산"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        """SELECT pa.*,
                  bc.cny_price, bc.exchange_rate, bc.overseas_shipping, bc.other_costs,
                  bc.cny_source AS bc_cny_source
           FROM price_adjustments pa
           LEFT JOIN bid_cost bc ON pa.order_id = bc.order_id
           WHERE pa.status IN ('pending', 'profit_low', 'deficit')
           ORDER BY pa.created_at DESC LIMIT 200"""
    )
    rows = []
    for r in c.fetchall():
        row = dict(r)
        # bid_cost 데이터가 있으면 expected_profit 실시간 재계산
        cny = row.pop("cny_price", None)
        rate = row.pop("exchange_rate", None)
        ship = row.pop("overseas_shipping", None)
        other = row.pop("other_costs", None)
        if cny and rate:
            total_cost = round(cny * rate * 1.03 + (ship or 8000) + (other or 0))
            settlement = _calc_settlement_for_monitor(row["new_price"])
            row["expected_profit"] = settlement - total_cost
            row["has_cost_data"] = True
            # 상태도 재계산 (deficit/profit_low/pending)
            ep = row["expected_profit"]
            if ep < 0:
                row["status"] = "deficit"
            elif ep < 5000:
                row["status"] = "profit_low"
            else:
                row["status"] = "pending"
        else:
            row["has_cost_data"] = False
        rows.append(row)
    conn.close()
    return jsonify({"adjustments": rows})


@app.route("/api/adjust/history-log")
def api_adjust_history_log():
    """조정 이력 (실행/거절/실패)"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        """SELECT * FROM price_adjustments
        WHERE status IN ('executed', 'rejected', 'failed')
        ORDER BY executed_at DESC LIMIT 100"""
    )
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({"history": rows})


@app.route("/api/adjust/approve", methods=["POST"])
def api_adjust_approve():
    """승인 → 가격 변경 실행"""
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "ids 필요"}), 400

    tid = new_task()
    add_log(tid, "info", f"가격 조정 승인: {len(ids)}건")

    def run():
        try:
            conn = sqlite3.connect(str(PRICE_DB))
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            placeholders = ",".join("?" * len(ids))
            c.execute(
                f"SELECT * FROM price_adjustments WHERE id IN ({placeholders}) AND status='pending'",
                ids
            )
            items = [dict(r) for r in c.fetchall()]
            conn.close()

            if not items:
                add_log(tid, "error", "승인 대상 없음 (적자/수익부족 항목은 승인 불가)")
                finish_task(tid, error="승인 대상 없음 (적자/수익부족 항목은 승인 불가)")
                return

            # 수익 5,000원 미만 항목 필터링
            valid_items = []
            for item in items:
                ep = item.get("expected_profit")
                if ep is not None and ep < 5000:
                    add_log(tid, "warn", f"{item['order_id']}: 수익 {ep:,}원 → 승인 불가 (최소 5,000원 필요)")
                else:
                    valid_items.append(item)
            items = valid_items

            if not items:
                add_log(tid, "error", "수익 조건 미달로 승인 가능 항목 없음")
                finish_task(tid, error="수익 조건 미달로 승인 가능 항목 없음")
                return

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            results = []
            for item in items:
                oid = item["order_id"]
                price = item["new_price"]
                add_log(tid, "info", f"{oid} → {price:,}원 수정 중...")

                ok = loop.run_until_complete(
                    modify_bid_price(oid, price, headless=get_headless())
                )
                results.append({"id": item["id"], "success": ok})

                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                conn2 = sqlite3.connect(str(PRICE_DB))
                conn2.execute(
                    "UPDATE price_adjustments SET status=?, executed_at=? WHERE id=?",
                    ("executed" if ok else "failed", now_str, item["id"])
                )
                conn2.commit()
                conn2.close()

                if ok:
                    add_log(tid, "success", f"{oid} 수정 완료")
                else:
                    add_log(tid, "error", f"{oid} 수정 실패")

            loop.close()

            success_cnt = sum(1 for r in results if r["success"])
            add_log(tid, "success", f"완료: {success_cnt}/{len(items)}건")
            finish_task(tid, result={"results": results, "success": success_cnt})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/adjust/reject", methods=["POST"])
def api_adjust_reject():
    """거절 처리"""
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "ids 필요"}), 400

    conn = sqlite3.connect(str(PRICE_DB))
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    placeholders = ",".join("?" * len(ids))
    conn.execute(
        f"UPDATE price_adjustments SET status='rejected', executed_at=? "
        f"WHERE id IN ({placeholders})",
        [now_str] + ids
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "count": len(ids)})


@app.route("/api/bid-cost/upsert", methods=["POST"])
def api_bid_cost_upsert():
    """원가 수동 입력/수정"""
    data = request.json or {}
    order_id = data.get("order_id", "").strip()
    model = data.get("model", "").strip()
    size = data.get("size", "").strip()
    cny_price = data.get("cny_price")
    exchange_rate = data.get("exchange_rate")

    if not order_id:
        return jsonify({"ok": False, "error": "order_id 필수"}), 400
    if not cny_price or float(cny_price) <= 0:
        return jsonify({"ok": False, "error": "CNY 가격 필수"}), 400

    # 환율 없으면 현재 설정에서 가져옴
    if not exchange_rate:
        settings = {}
        if SETTINGS_FILE.exists():
            settings = json.loads(SETTINGS_FILE.read_text())
        exchange_rate = settings.get("cnyRate", 215)

    cny_f = float(cny_price)
    rate_f = float(exchange_rate)
    ship_i = int(data.get("overseas_shipping", 8000))
    other_i = int(data.get("other_costs", 0))

    _save_bid_cost(
        order_id=order_id, model=model, size=size,
        cny_price=cny_f,
        exchange_rate=rate_f,
        overseas_shipping=ship_i,
        other_costs=other_i,
        cny_source="manual",
    )

    # pending 조정 건의 expected_profit + status 갱신
    total_cost = round(cny_f * rate_f * 1.03 + ship_i + other_i)
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute(
            "SELECT id, new_price FROM price_adjustments WHERE order_id=? AND status IN ('pending','profit_low','deficit')",
            (order_id,)
        )
        for pa in c.fetchall():
            settlement = _calc_settlement_for_monitor(pa["new_price"])
            ep = settlement - total_cost
            if ep < 0:
                new_status = "deficit"
            elif ep < 5000:
                new_status = "profit_low"
            else:
                new_status = "pending"
            c.execute(
                "UPDATE price_adjustments SET expected_profit=?, status=? WHERE id=?",
                (ep, new_status, pa["id"])
            )
        conn.commit()
        conn.close()
    except Exception:
        pass

    return jsonify({"ok": True})


@app.route("/api/bid-cost/get/<order_id>")
def api_bid_cost_get(order_id):
    """특정 order_id의 원가 조회"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM bid_cost WHERE order_id=?", (order_id,))
    row = c.fetchone()
    conn.close()
    if row:
        return jsonify({"ok": True, "cost": dict(row)})
    return jsonify({"ok": True, "cost": None})


@app.route("/api/bid-cost/missing")
def api_bid_cost_missing():
    """원가 없는 pending 조정건을 모델별 그룹화하여 반환"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        """SELECT pa.order_id, pa.product_id, pa.model, pa.name_kr, pa.size, pa.new_price
           FROM price_adjustments pa
           LEFT JOIN bid_cost bc ON pa.order_id = bc.order_id
           WHERE pa.status IN ('pending','profit_low','deficit')
             AND bc.order_id IS NULL
           ORDER BY pa.model, pa.size"""
    )
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    from collections import OrderedDict
    groups_dict = OrderedDict()
    for r in rows:
        key = r["model"] or r["product_id"] or "unknown"
        if key not in groups_dict:
            groups_dict[key] = {
                "model": r["model"] or "",
                "product_id": r["product_id"] or "",
                "display_name": r["name_kr"] or r["model"] or f"#{r['product_id']}",
                "bids": [],
            }
        groups_dict[key]["bids"].append({
            "order_id": r["order_id"],
            "size": r["size"],
            "current_price": r["new_price"],
        })

    return jsonify({"groups": list(groups_dict.values()), "total": len(rows)})


@app.route("/api/bid-cost/shihuo-diff")
def api_bid_cost_shihuo_diff():
    """등록된 bid_cost와 식货 활성 배치의 cny_price 차이 리포트.

    매칭: bc.model = sh.model AND CAST(bc.size AS INTEGER) = sh.kream_mm AND sh.active=1.
    가격 차이가 있는 행만 반환. ONE SIZE 등 캐스팅 불가 항목은 자동 제외(매칭 0).
    """
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    active_batch = c.execute(
        "SELECT batch_id FROM shihuo_prices WHERE active=1 ORDER BY imported_at DESC LIMIT 1"
    ).fetchone()
    active_batch_id = active_batch[0] if active_batch else None

    rows = c.execute("""
        SELECT bc.order_id, bc.model, bc.size,
               bc.cny_price        AS bc_cny,
               bc.cny_source       AS bc_source,
               bc.exchange_rate    AS bc_rate,
               sh.cny_price        AS sh_cny,
               sh.batch_id         AS sh_batch,
               sh.kream_mm         AS sh_kream_mm,
               sh.size_eu          AS sh_size_eu
          FROM bid_cost bc
          JOIN shihuo_prices sh
            ON sh.active=1
           AND sh.model = bc.model
           AND sh.kream_mm IS NOT NULL
           AND sh.kream_mm = CAST(bc.size AS INTEGER)
         WHERE sh.cny_price <> bc.cny_price
         ORDER BY ABS(sh.cny_price - bc.cny_price) DESC
    """).fetchall()
    conn.close()

    items = []
    for r in rows:
        bc_cny = float(r["bc_cny"] or 0)
        sh_cny = float(r["sh_cny"] or 0)
        diff = sh_cny - bc_cny
        diff_pct = (diff / bc_cny * 100.0) if bc_cny else None
        items.append({
            "order_id": r["order_id"],
            "model": r["model"],
            "size": r["size"],
            "bc_cny": bc_cny,
            "bc_source": r["bc_source"],
            "sh_cny": sh_cny,
            "diff_cny": round(diff, 2),
            "diff_pct": round(diff_pct, 2) if diff_pct is not None else None,
            "sh_batch": r["sh_batch"],
            "sh_size_eu": r["sh_size_eu"],
            "sh_kream_mm": r["sh_kream_mm"],
            "exchange_rate": r["bc_rate"],
        })

    return jsonify({
        "ok": True,
        "active_batch_id": active_batch_id,
        "count": len(items),
        "items": items,
    })


@app.route("/api/bid-cost/bulk-upsert", methods=["POST"])
def api_bid_cost_bulk_upsert():
    """여러 건 원가 한번에 저장 + price_adjustments 재계산"""
    data = request.json or {}
    entries = data.get("entries", [])
    if not entries:
        return jsonify({"ok": False, "error": "entries 필요"}), 400

    updated = 0
    failed = 0
    details = []

    # 환율 기본값
    default_rate = 215
    if SETTINGS_FILE.exists():
        try:
            s = json.loads(SETTINGS_FILE.read_text())
            default_rate = s.get("cnyRate", 215)
        except Exception:
            pass

    for entry in entries:
        oid = (entry.get("order_id") or "").strip()
        cny = entry.get("cny_price")
        if not oid or not cny or float(cny) <= 0:
            failed += 1
            details.append({"order_id": oid, "ok": False, "reason": "order_id 또는 cny_price 누락"})
            continue

        cny_f = float(cny)
        rate_f = float(entry.get("exchange_rate") or default_rate)
        ship_i = int(entry.get("overseas_shipping", 8000))
        other_i = int(entry.get("other_costs", 0))
        model = entry.get("model", "")
        size = entry.get("size", "")

        try:
            _save_bid_cost(
                order_id=oid, model=model, size=size,
                cny_price=cny_f, exchange_rate=rate_f,
                overseas_shipping=ship_i, other_costs=other_i,
                cny_source="manual",
            )

            # pending 건 expected_profit 재계산
            total_cost = round(cny_f * rate_f * 1.03 + ship_i + other_i)
            conn = sqlite3.connect(str(PRICE_DB))
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute(
                "SELECT id, new_price FROM price_adjustments WHERE order_id=? AND status IN ('pending','profit_low','deficit')",
                (oid,)
            )
            for pa in c.fetchall():
                settlement = _calc_settlement_for_monitor(pa["new_price"])
                ep = settlement - total_cost
                new_status = "deficit" if ep < 0 else ("profit_low" if ep < 5000 else "pending")
                c.execute("UPDATE price_adjustments SET expected_profit=?, status=? WHERE id=?",
                          (ep, new_status, pa["id"]))
            conn.commit()
            conn.close()

            updated += 1
            details.append({"order_id": oid, "ok": True})
        except Exception as e:
            failed += 1
            details.append({"order_id": oid, "ok": False, "reason": str(e)})

    return jsonify({"ok": True, "updated": updated, "failed": failed, "details": details})


@app.route("/api/email/test", methods=["POST"])
def api_email_test():
    """이메일 발송 테스트"""
    test_data = [{
        "name_kr": "[테스트] 나이키 에어포스 1",
        "model": "TEST-001",
        "size": "270",
        "old_price": 120000,
        "competitor_price": 115000,
        "new_price": 114000,
        "expected_profit": 8500,
    }]
    _send_adjustment_email(test_data)
    return jsonify({"ok": True, "msg": "테스트 이메일 발송 시도 완료"})


# ═══════════════════════════════════════════
# 조건부 입찰 API
# ═══════════════════════════════════════════

@app.route("/api/conditional-bids", methods=["GET"])
def api_conditional_bids_list():
    """조건부 입찰 목록"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM conditional_bids ORDER BY created_at DESC LIMIT 200"
    ).fetchall()
    conn.close()
    return jsonify({"bids": [dict(r) for r in rows]})


@app.route("/api/conditional-bids", methods=["POST"])
def api_conditional_bids_add():
    """조건부 입찰 추가"""
    data = request.json or {}
    required = ["product_id", "condition_type", "condition_value", "bid_price"]
    for f in required:
        if not data.get(f):
            return jsonify({"error": f"{f} 필요"}), 400
    if data["condition_type"] not in ("price_below", "competitor_above"):
        return jsonify({"error": "condition_type은 price_below 또는 competitor_above"}), 400

    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        """INSERT INTO conditional_bids
        (product_id, model, size, condition_type, condition_value, bid_price, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, 'active', ?)""",
        (
            str(data["product_id"]),
            data.get("model", ""),
            data.get("size", "ONE SIZE"),
            data["condition_type"],
            int(data["condition_value"]),
            int(data["bid_price"]),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/conditional-bids/<int:bid_id>", methods=["DELETE"])
def api_conditional_bids_delete(bid_id):
    """조건부 입찰 삭제"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute("DELETE FROM conditional_bids WHERE id=?", (bid_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/conditional-bids/<int:bid_id>/cancel", methods=["POST"])
def api_conditional_bids_cancel(bid_id):
    """조건부 입찰 비활성화"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute("UPDATE conditional_bids SET status='expired' WHERE id=?", (bid_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def _check_conditional_bids():
    """조건부 입찰 조건 체크 → 충족 시 자동 입찰 실행"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    active = conn.execute(
        "SELECT * FROM conditional_bids WHERE status='active'"
    ).fetchall()
    conn.close()

    if not active:
        return

    print(f"[조건부입찰] {len(active)}건 조건 체크 중...")

    for cb in active:
        cb = dict(cb)
        pid = cb["product_id"]
        ctype = cb["condition_type"]
        cval = cb["condition_value"]

        # price_history에서 최신 시세 조회
        conn2 = sqlite3.connect(str(PRICE_DB))
        conn2.row_factory = sqlite3.Row
        row = conn2.execute(
            "SELECT sell_now_price, recent_trade_price FROM price_history WHERE product_id=? ORDER BY collected_at DESC LIMIT 1",
            (pid,)
        ).fetchone()
        conn2.close()

        if not row:
            continue

        triggered = False
        if ctype == "price_below":
            # 즉시구매가가 X원 이하
            sell_price = row["sell_now_price"] or 0
            if sell_price > 0 and sell_price <= cval:
                triggered = True
                print(f"[조건부입찰] {cb['model']} {cb['size']}: 즉시구매가 {sell_price:,}원 <= {cval:,}원 → 조건 충족!")
        elif ctype == "competitor_above":
            # 경쟁자 최저가가 X원 이상
            sell_price = row["sell_now_price"] or 0
            if sell_price > 0 and sell_price >= cval:
                triggered = True
                print(f"[조건부입찰] {cb['model']} {cb['size']}: 경쟁자최저가 {sell_price:,}원 >= {cval:,}원 → 조건 충족!")

        if triggered:
            # 상태 업데이트
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn3 = sqlite3.connect(str(PRICE_DB))
            conn3.execute(
                "UPDATE conditional_bids SET status='triggered', triggered_at=? WHERE id=?",
                (now_str, cb["id"])
            )
            conn3.commit()
            conn3.close()

            # 알림 추가
            name = cb.get("model") or f"#{pid}"
            add_notification(
                "conditional_bid",
                f"조건부 입찰 조건 충족: {name} {cb['size']}",
                f"조건: {'즉시구매가 ≤' if ctype == 'price_below' else '경쟁자최저가 ≥'} {cval:,}원 → 입찰가 {cb['bid_price']:,}원",
                f"/api/conditional-bids"
            )

            # 자동 입찰 실행
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                from kream_bot import place_sell_bid
                ok = loop.run_until_complete(
                    place_sell_bid(pid, cb["bid_price"], cb["size"], 1, headless=True)
                )
                loop.close()
                if ok:
                    print(f"[조건부입찰] {name} {cb['size']} {cb['bid_price']:,}원 입찰 성공")
                else:
                    print(f"[조건부입찰] {name} {cb['size']} 입찰 실패")
            except Exception as e:
                print(f"[조건부입찰] 입찰 오류: {e}")


# ═══════════════════════════════════════════
# 수정 이력 API
# ═══════════════════════════════════════════

@app.route("/api/edit-log")
def api_edit_log():
    """수정 이력 조회"""
    date = request.args.get("date", "")
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    if date:
        rows = conn.execute(
            "SELECT * FROM edit_log WHERE edited_at LIKE ? ORDER BY edited_at DESC LIMIT 200",
            (date + "%",)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM edit_log ORDER BY edited_at DESC LIMIT 200"
        ).fetchall()
    conn.close()
    return jsonify({"logs": [dict(r) for r in rows]})


@app.route("/api/edit-log", methods=["POST"])
def api_edit_log_add():
    """수정 이력 추가 (프론트엔드에서 직접 기록)"""
    data = request.json or {}
    required = ["item_type", "item_id", "field_name", "old_value", "new_value"]
    for f in required:
        if f not in data:
            return jsonify({"error": f"{f} 필요"}), 400
    save_edit_log(data["item_type"], data["item_id"], data["field_name"],
                  data["old_value"], data["new_value"])
    return jsonify({"ok": True})


# ═══════════════════════════════════════════
# 입찰 만료 관련 API
# ═══════════════════════════════════════════

@app.route("/api/expiring-bids")
def api_expiring_bids():
    """만료 임박 입찰 목록"""
    local_file = BASE_DIR / "my_bids_local.json"
    try:
        data = json.loads(local_file.read_text()) if local_file.exists() else {}
    except Exception:
        data = {}
    expiring = data.get("expiring", [])
    return jsonify({"expiring": expiring, "count": len(expiring)})


@app.route("/api/expiring-bids/renew", methods=["POST"])
def api_renew_bids():
    """만료 임박 입찰 갱신 (동일 가격 재입찰)"""
    data = request.json or {}
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return jsonify({"error": "order_ids 필요"}), 400

    # 설정에서 자동 갱신 허용 여부 확인
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    if not settings.get("autoRenewBids", False):
        return jsonify({"error": "자동 갱신이 비활성화되어 있습니다. 설정에서 활성화해주세요."}), 400

    tid = new_task()
    add_log(tid, "info", f"입찰 갱신 시작: {len(order_ids)}건")

    def run():
        try:
            # my_bids_local에서 해당 입찰 정보 찾기
            local_file = BASE_DIR / "my_bids_local.json"
            local_data = json.loads(local_file.read_text()) if local_file.exists() else {}
            expiring = local_data.get("expiring", [])

            targets = [e for e in expiring if e.get("orderId") in order_ids]
            if not targets:
                add_log(tid, "error", "갱신 대상 없음")
                finish_task(tid, error="갱신 대상 없음")
                return

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            results = []
            for t in targets:
                pid = t["productId"]
                price = t["bidPrice"]
                size = t["size"]
                name = t.get("nameKr") or t.get("model", "")
                add_log(tid, "info", f"{name} {size} {price:,}원 재입찰 중...")

                try:
                    from kream_bot import run_bid
                    ok = loop.run_until_complete(run_bid(pid, price, size, 1, tid))
                    results.append({"orderId": t["orderId"], "success": bool(ok)})
                    if ok:
                        add_log(tid, "success", f"{name} {size} 갱신 완료")
                    else:
                        add_log(tid, "error", f"{name} {size} 갱신 실패")
                except Exception as e:
                    add_log(tid, "error", f"{name} {size} 오류: {e}")
                    results.append({"orderId": t["orderId"], "success": False})

            loop.close()
            success_cnt = sum(1 for r in results if r["success"])
            add_log(tid, "success", f"완료: {success_cnt}/{len(targets)}건")
            finish_task(tid, result={"results": results, "success": success_cnt})
        except Exception as e:
            traceback.print_exc()
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


# ═══════════════════════════════════════════
# 물류 관리 API
# ═══════════════════════════════════════════

# 업로드 폴더
UPLOADS_DIR = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)


@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(str(UPLOADS_DIR), filename)


@app.route("/api/logistics/suppliers")
def api_logistics_suppliers():
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM suppliers ORDER BY id").fetchall()
    conn.close()
    return jsonify({"suppliers": [dict(r) for r in rows]})


@app.route("/api/logistics/supplier", methods=["POST"])
def api_logistics_supplier_add():
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "이름 필요"}), 400
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        "INSERT INTO suppliers (name, contact, phone, wechat, notes, created_at) VALUES (?,?,?,?,?,?)",
        (name, data.get("contact", ""), data.get("phone", ""), data.get("wechat", ""),
         data.get("notes", ""), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/logistics/supplier/<int:sid>", methods=["DELETE"])
def api_logistics_supplier_delete(sid):
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute("DELETE FROM suppliers WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/logistics/pending")
def api_logistics_pending():
    """발송 대기: sales_history에 있는데 shipment_requests에 없는 건"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT sh.order_id, sh.product_id, sh.model, sh.size,
               sh.sale_price as amount, sh.trade_date as sold_at
        FROM sales_history sh
        LEFT JOIN shipment_requests sr ON sh.order_id = sr.order_id
        WHERE sr.id IS NULL
        ORDER BY sh.trade_date DESC
        LIMIT 200
    """).fetchall()
    conn.close()
    return jsonify({"pending": [dict(r) for r in rows]})


@app.route("/api/logistics/request", methods=["POST"])
def api_logistics_request_add():
    """발송 요청 생성"""
    # multipart/form-data 지원
    order_id = request.form.get("order_id", "") or (request.json or {}).get("order_id", "")
    product_id = request.form.get("product_id", "") or (request.json or {}).get("product_id", "")
    model = request.form.get("model", "") or (request.json or {}).get("model", "")
    size = request.form.get("size", "") or (request.json or {}).get("size", "")
    supplier_id = request.form.get("supplier_id") or (request.json or {}).get("supplier_id")
    hubnet_hbl = request.form.get("hubnet_hbl", "") or (request.json or {}).get("hubnet_hbl", "")
    notes = request.form.get("notes", "") or (request.json or {}).get("notes", "")

    if not supplier_id:
        return jsonify({"error": "협력사 선택 필요"}), 400

    # 증거 이미지 저장
    proof_filename = ""
    proof_file = request.files.get("proof_image")
    if proof_file and proof_file.filename:
        ext = proof_file.filename.rsplit(".", 1)[-1] if "." in proof_file.filename else "jpg"
        proof_filename = f"proof_{order_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        proof_file.save(str(UPLOADS_DIR / proof_filename))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        """INSERT INTO shipment_requests
        (order_id, product_id, model, size, supplier_id, hubnet_hbl, request_date,
         tracking_number, status, proof_image, notes, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (order_id, product_id, model, size, int(supplier_id), hubnet_hbl,
         now[:10], "", "요청", proof_filename, notes, now, now)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/logistics/request/<int:rid>", methods=["PUT"])
def api_logistics_request_update(rid):
    """발송 요청 업데이트 (트래킹/상태)"""
    data = request.json or {}
    conn = sqlite3.connect(str(PRICE_DB))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    updates = []
    params = []
    for field in ["tracking_number", "status", "hubnet_hbl", "notes"]:
        if field in data and data[field] is not None:
            updates.append(f"{field}=?")
            params.append(data[field])

    if not updates:
        conn.close()
        return jsonify({"ok": True})

    updates.append("updated_at=?")
    params.append(now)
    params.append(rid)

    # 상태 변경 이력 기록
    if "status" in data:
        old = conn.execute("SELECT status FROM shipment_requests WHERE id=?", (rid,)).fetchone()
        if old:
            save_edit_log("shipment", str(rid), "status", old[0], data["status"])

    # 트래킹 입력 시 자동 상태 변경
    if "tracking_number" in data and data["tracking_number"] and "status" not in data:
        old_status = conn.execute("SELECT status FROM shipment_requests WHERE id=?", (rid,)).fetchone()
        if old_status and old_status[0] == "요청":
            updates = [u for u in updates if not u.startswith("status")]
            updates.insert(0, "status=?")
            params.insert(0, "발송완료")
            save_edit_log("shipment", str(rid), "status", "요청", "발송완료")

    conn.execute(f"UPDATE shipment_requests SET {','.join(updates)} WHERE id=?", params)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/logistics/request/<int:rid>", methods=["DELETE"])
def api_logistics_request_delete(rid):
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute("DELETE FROM shipment_requests WHERE id=?", (rid,))
    conn.execute("DELETE FROM shipment_costs WHERE shipment_id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/logistics/requests")
def api_logistics_requests():
    """발송 요청 목록 (필터 가능)"""
    status = request.args.get("status", "")
    supplier_id = request.args.get("supplier_id", "")

    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    sql = "SELECT * FROM shipment_requests WHERE 1=1"
    params = []
    if status:
        sql += " AND status=?"
        params.append(status)
    if supplier_id:
        sql += " AND supplier_id=?"
        params.append(int(supplier_id))
    sql += " ORDER BY created_at DESC LIMIT 200"

    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify({"requests": [dict(r) for r in rows]})


@app.route("/api/logistics/stats")
def api_logistics_stats():
    """물류 현황 통계"""
    conn = sqlite3.connect(str(PRICE_DB))
    pending = conn.execute("SELECT COUNT(*) FROM sales_history sh LEFT JOIN shipment_requests sr ON sh.order_id=sr.order_id WHERE sr.id IS NULL").fetchone()[0]
    in_progress = conn.execute("SELECT COUNT(*) FROM shipment_requests WHERE status IN ('요청','발송완료','허브넷도착','통관중')").fetchone()[0]
    done = conn.execute("SELECT COUNT(*) FROM shipment_requests WHERE status='배송완료'").fetchone()[0]

    month_start = datetime.now().strftime("%Y-%m-01")
    month_cost = conn.execute(
        "SELECT COALESCE(SUM(CASE WHEN currency='KRW' THEN amount ELSE amount*218 END),0) FROM shipment_costs WHERE created_at>=?",
        (month_start,)
    ).fetchone()[0]
    conn.close()
    return jsonify({"pending": pending, "in_progress": in_progress, "done": done, "month_cost": int(month_cost)})


@app.route("/api/logistics/export")
def api_logistics_export():
    """엑셀 내보내기 (CSV)"""
    status = request.args.get("status", "")
    supplier_id = request.args.get("supplier_id", "")

    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    sql = "SELECT sr.*, s.name as supplier_name FROM shipment_requests sr LEFT JOIN suppliers s ON sr.supplier_id=s.id WHERE 1=1"
    params = []
    if status:
        sql += " AND sr.status=?"
        params.append(status)
    if supplier_id:
        sql += " AND sr.supplier_id=?"
        params.append(int(supplier_id))
    sql += " ORDER BY sr.created_at DESC"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    csv_data = "\ufeff주문번호,모델,사이즈,협력사,HBL,요청일,트래킹번호,상태,메모\n"
    for r in rows:
        csv_data += f"{r['order_id']},{r['model']},{r['size']},{r['supplier_name'] or ''},{r['hubnet_hbl'] or ''},{r['request_date'] or ''},{r['tracking_number'] or ''},{r['status']},{r['notes'] or ''}\n"

    return Response(csv_data, mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=shipments_{datetime.now().strftime('%Y%m%d')}.csv"})


@app.route("/api/logistics/import-tracking", methods=["POST"])
def api_logistics_import_tracking():
    """트래킹 번호 일괄 가져오기 (CSV)"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "파일 필요"}), 400

    import csv, io
    content = file.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    header = next(reader, [])

    # 열 찾기
    header_lower = [h.strip().lower() for h in header]
    order_col = -1
    hbl_col = -1
    tracking_col = -1

    for i, h in enumerate(header_lower):
        if "주문" in h or "order" in h:
            order_col = i
        if "hbl" in h:
            hbl_col = i
        if "트래킹" in h or "tracking" in h or "운송장" in h:
            tracking_col = i

    if tracking_col < 0:
        return jsonify({"error": "트래킹 열을 찾을 수 없습니다"}), 400
    if order_col < 0 and hbl_col < 0:
        return jsonify({"error": "주문번호 또는 HBL 열이 필요합니다"}), 400

    conn = sqlite3.connect(str(PRICE_DB))
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    updated = 0

    for row in reader:
        if len(row) <= tracking_col:
            continue
        tracking = row[tracking_col].strip()
        if not tracking:
            continue

        if order_col >= 0 and len(row) > order_col and row[order_col].strip():
            key_col = "order_id"
            key_val = row[order_col].strip()
        elif hbl_col >= 0 and len(row) > hbl_col and row[hbl_col].strip():
            key_col = "hubnet_hbl"
            key_val = row[hbl_col].strip()
        else:
            continue

        result = conn.execute(
            f"UPDATE shipment_requests SET tracking_number=?, status='발송완료', updated_at=? WHERE {key_col}=? AND (tracking_number IS NULL OR tracking_number='')",
            (tracking, now, key_val)
        )
        if result.rowcount > 0:
            updated += 1

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "updated": updated})


# ═══════════════════════════════════════════
# 판매 이력 수집 API + 스케줄러
# ═══════════════════════════════════════════

def _save_shipments_to_db(shipments):
    """발송관리 수집 결과를 DB에 저장, (새로 추가된 건수, 새 건 목록) 반환"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    new_count = 0
    new_items = []
    for s in shipments:
        try:
            c.execute(
                """INSERT OR IGNORE INTO sales_history
                (order_id, product_id, model, product_info, size,
                 sale_price, trade_date, ship_date, ship_status, collected_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (s.get("order_id", ""), s.get("product_id", ""),
                 s.get("model", ""), s.get("product_info", ""),
                 s.get("size", ""), s.get("sale_price", 0),
                 s.get("trade_date", ""), s.get("ship_date", ""),
                 s.get("ship_status", ""), now)
            )
            if c.rowcount > 0:
                new_count += 1
                new_items.append(s)
        except Exception:
            pass
    conn.commit()
    conn.close()
    return new_count, new_items


# 새 체결건 알림 (대시보드 폴링용)
_new_sales_alerts = []
_new_sales_lock = threading.Lock()
_hubnet_trigger_lock = threading.Lock()  # Step 10: 허브넷 PDF 트리거 동시실행 방지


def _run_sales_sync():
    """비동기 발송관리 수집 실행"""
    async def _do():
        try:
            settings = {}
            if SETTINGS_FILE.exists():
                settings = json.loads(SETTINGS_FILE.read_text())
            headless = settings.get("headless", True)

            async with async_playwright() as p:
                browser = await create_browser(p, headless=headless)
                context = await create_context(browser, STATE_FILE)
                page = await context.new_page()
                await apply_stealth(page)

                if not await ensure_logged_in(page, context):
                    await browser.close()
                    return {"ok": False, "error": "로그인 필요"}

                shipments = await collect_shipments(page, max_pages=10)
                await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
                await browser.close()

                new_count, new_items = _save_shipments_to_db(shipments)
                with _sales_lock:
                    sales_scheduler_state["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                    sales_scheduler_state["total_syncs"] += 1
                    sales_scheduler_state["last_new_count"] = new_count

                # 새 체결건 알림 추가
                if new_items:
                    with _new_sales_lock:
                        for item in new_items:
                            _new_sales_alerts.append({
                                **item,
                                "detected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            })

                print(f"[판매수집] 완료: 총 {len(shipments)}건 수집, 신규 {new_count}건")
                return {"ok": True, "total": len(shipments), "new_count": new_count, "new_items": new_items}
        except Exception as e:
            print(f"[판매수집] 오류: {e}")
            traceback.print_exc()
            return {"ok": False, "error": str(e)}

    loop = asyncio.new_event_loop()
    try:
        result = loop.run_until_complete(_do())
    finally:
        loop.close()

    # 새 판매 감지 시 자동 재입찰 시도 (sync 컨텍스트에서 호출)
    if result and result.get("ok") and result.get("new_items"):
        try:
            rebid_result = auto_rebid_after_sale(result["new_items"])
            print(f"[auto_rebid] 성공={rebid_result['success']} "
                  f"건너뜀={rebid_result['skipped']} "
                  f"실패={rebid_result['failed']}")
        except Exception as re:
            print(f"[auto_rebid] 예외: {re}")
            try:
                health_alerter.alert("auto_rebid_exception", str(re), cooldown_minutes=60)
            except Exception:
                pass

    # ───────── Step 10: 허브넷 자동 PDF 다운로드 트리거 ─────────
    # 격리 원칙: 어떤 예외도 외부로 던지지 않음. 판매 수집 결과(result)는 보존.
    try:
        hb_settings = {}
        if SETTINGS_FILE.exists():
            hb_settings = json.loads(SETTINGS_FILE.read_text())
        if hb_settings.get('hubnet_auto_pdf', False):
            acquired = _hubnet_trigger_lock.acquire(blocking=False)
            if not acquired:
                print("[HUBNET_AUTO] 이전 사이클 진행 중 — 이번 사이클 스킵")
            else:
                try:
                    from kream_hubnet_bot import download_pending_invoices
                    hb_result = download_pending_invoices(
                        limit=20,
                        triggered_by='scheduler'
                    )
                    print(f"[HUBNET_AUTO] {hb_result}")
                    if isinstance(hb_result, dict) and hb_result.get('failed', 0) > 0:
                        try:
                            health_alerter.alert(
                                'hubnet_pdf_failed',
                                f"허브넷 PDF 다운로드 {hb_result['failed']}건 실패",
                                cooldown_minutes=60
                            )
                        except Exception:
                            pass
                finally:
                    _hubnet_trigger_lock.release()
    except Exception as he:
        print(f"[HUBNET_AUTO_ERROR] {he}")
        try:
            health_alerter.alert(
                'hubnet_pdf_trigger_error',
                f"허브넷 자동 PDF 트리거 오류: {he}",
                cooldown_minutes=60
            )
        except Exception:
            pass
    # ─────────────────────────────────────────────────────────────

    return result


def _get_sales_sync_interval():
    """설정에서 판매 수집 간격(초) 조회 — 기본 30분 + ±5분 랜덤 지터"""
    interval_min = 30
    try:
        if SETTINGS_FILE.exists():
            settings = json.loads(SETTINGS_FILE.read_text())
            interval_min = settings.get("sales_sync_interval_minutes", 30)
    except Exception:
        pass
    base_seconds = max(5, interval_min) * 60
    jitter = random.randint(-300, 300)  # ±5분
    return max(300, base_seconds + jitter)  # 최소 5분


def _schedule_next_sales_sync():
    """다음 판매 수집 예약 (설정 간격 + 랜덤 지터)"""
    global _sales_timer
    with _sales_lock:
        if not sales_scheduler_state["running"]:
            return
    interval = _get_sales_sync_interval()
    _sales_timer = threading.Timer(interval, _sales_sync_tick)
    _sales_timer.daemon = True
    _sales_timer.start()
    with _sales_lock:
        next_time = (datetime.now() + timedelta(seconds=interval)).strftime("%Y-%m-%d %H:%M")
        sales_scheduler_state["next_run"] = next_time


def _sales_sync_tick():
    """스케줄러 틱 — 수집 실행 후 다음 예약"""
    _run_sales_sync()
    _schedule_next_sales_sync()


@app.route("/api/sales/recent")
def api_sales_recent():
    """최근 판매 내역"""
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        "SELECT * FROM sales_history ORDER BY trade_date DESC, id DESC LIMIT ? OFFSET ?",
        (limit, offset)
    )
    rows = [dict(r) for r in c.fetchall()]
    c.execute("SELECT COUNT(*) FROM sales_history")
    total = c.fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "sales": rows, "total": total})


@app.route("/api/sales/sync", methods=["POST"])
def api_sales_sync():
    """수동 판매 동기화"""
    def _bg():
        return _run_sales_sync()
    result = [None]
    def _run():
        result[0] = _bg()
    t = threading.Thread(target=_run)
    t.start()
    t.join(timeout=120)
    if result[0]:
        return jsonify(result[0])
    return jsonify({"ok": False, "error": "타임아웃 (120초)"})


@app.route("/api/sales/stats")
def api_sales_stats():
    """판매 통계"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # 총 판매 건수 & 금액
    c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(sale_price),0) as total_amount FROM sales_history")
    row = dict(c.fetchone())

    # 최근 7일 판매
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    c.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(sale_price),0) as total_amount FROM sales_history WHERE trade_date >= ?",
        (week_ago,)
    )
    weekly = dict(c.fetchone())

    # 모델별 판매 순위
    c.execute("""
        SELECT model, COUNT(*) as cnt, SUM(sale_price) as total_amount
        FROM sales_history WHERE model != ''
        GROUP BY model ORDER BY cnt DESC LIMIT 10
    """)
    top_models = [dict(r) for r in c.fetchall()]

    # 일별 판매 추이 (최근 30일)
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    c.execute("""
        SELECT trade_date, COUNT(*) as cnt, SUM(sale_price) as total_amount
        FROM sales_history WHERE trade_date >= ?
        GROUP BY trade_date ORDER BY trade_date
    """, (month_ago,))
    daily = [dict(r) for r in c.fetchall()]

    conn.close()
    return jsonify({
        "ok": True,
        "total_sales": row["cnt"],
        "total_amount": row["total_amount"],
        "weekly_sales": weekly["cnt"],
        "weekly_amount": weekly["total_amount"],
        "top_models": top_models,
        "daily_trend": daily,
    })


@app.route("/api/sales/dashboard")
def api_sales_dashboard():
    """판매 대시보드 — 요약/베스트셀러/최근 판매/시간대 분포"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    def _period_stats(date_from, date_to=None):
        if date_to:
            c.execute(
                "SELECT COUNT(*) as cnt, COALESCE(SUM(sale_price),0) as rev FROM sales_history WHERE trade_date >= ? AND trade_date <= ?",
                (date_from, date_to))
        else:
            c.execute(
                "SELECT COUNT(*) as cnt, COALESCE(SUM(sale_price),0) as rev FROM sales_history WHERE trade_date >= ?",
                (date_from,))
        r = c.fetchone()
        return {"count": r["cnt"], "total_revenue": r["rev"]}

    summary = {
        "today": _period_stats(today, today),
        "yesterday": _period_stats(yesterday, yesterday),
        "last_7days": _period_stats(week_ago),
        "last_30days": _period_stats(month_ago),
    }

    # 베스트셀러 TOP 10 (30일)
    c.execute("""
        SELECT model, COUNT(*) as count, COALESCE(SUM(sale_price),0) as total_revenue
        FROM sales_history WHERE model != '' AND trade_date >= ?
        GROUP BY model ORDER BY count DESC LIMIT 10
    """, (month_ago,))
    top_models = [dict(r) for r in c.fetchall()]

    # 최근 20건
    c.execute("SELECT * FROM sales_history ORDER BY trade_date DESC, id DESC LIMIT 20")
    recent_sales = []
    for r in c.fetchall():
        row = dict(r)
        row["margin_estimate"] = None  # 마진 추정 불가 시 NULL
        recent_sales.append(row)

    # 시간대별 분포 (30일, trade_date에서 시간 추출)
    c.execute("""
        SELECT CAST(SUBSTR(trade_date, 12, 2) AS INTEGER) as hour, COUNT(*) as count
        FROM sales_history WHERE trade_date >= ? AND LENGTH(trade_date) >= 13
        GROUP BY hour ORDER BY hour
    """, (month_ago,))
    hourly_raw = {r["hour"]: r["count"] for r in c.fetchall()}
    hourly_distribution = [{"hour": h, "count": hourly_raw.get(h, 0)} for h in range(24)]

    conn.close()
    return jsonify({
        "ok": True,
        "summary": summary,
        "top_models": top_models,
        "recent_sales": recent_sales,
        "hourly_distribution": hourly_distribution,
    })


@app.route("/api/sales/search")
def api_sales_search():
    """판매 이력 검색 — 모델/사이즈/기간 필터 + 페이지네이션"""
    model = request.args.get("model", "").strip()
    size = request.args.get("size", "").strip()
    from_date = request.args.get("from_date", "").strip()
    to_date = request.args.get("to_date", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    per_page = min(per_page, 100)

    conditions = []
    params = []
    if model:
        conditions.append("model LIKE ?")
        params.append(f"%{model}%")
    if size:
        conditions.append("size = ?")
        params.append(size)
    if from_date:
        conditions.append("trade_date >= ?")
        params.append(from_date)
    if to_date:
        conditions.append("trade_date <= ?")
        params.append(to_date + " 23:59:59")

    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""

    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute(f"SELECT COUNT(*) FROM sales_history{where}", params)
    total = c.fetchone()[0]

    offset = (page - 1) * per_page
    c.execute(
        f"SELECT * FROM sales_history{where} ORDER BY trade_date DESC, id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    )
    sales = [dict(r) for r in c.fetchall()]
    conn.close()

    return jsonify({
        "ok": True,
        "sales": sales,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": math.ceil(total / per_page) if per_page else 0,
    })


@app.route("/api/sales/by-model/<path:model>")
def api_sales_by_model(model):
    """모델별 판매 상세 — 회전율/평균가격/사이즈분포"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # 전체 판매 이력
    c.execute(
        "SELECT * FROM sales_history WHERE model = ? ORDER BY trade_date DESC",
        (model,))
    sales = [dict(r) for r in c.fetchall()]

    if not sales:
        conn.close()
        return jsonify({"ok": True, "model": model, "total_count": 0, "sales": [],
                        "avg_price": None, "size_distribution": [], "turnover_days": None})

    total_count = len(sales)
    prices = [s["sale_price"] for s in sales if s.get("sale_price")]
    avg_price = round(sum(prices) / len(prices)) if prices else None

    # 사이즈별 분포
    c.execute("""
        SELECT size, COUNT(*) as count, COALESCE(AVG(sale_price),0) as avg_price
        FROM sales_history WHERE model = ? AND size != ''
        GROUP BY size ORDER BY count DESC
    """, (model,))
    size_distribution = [dict(r) for r in c.fetchall()]

    # 회전율 (첫 판매 ~ 마지막 판매 사이 일수 / 판매 건수)
    c.execute(
        "SELECT MIN(trade_date) as first_sale, MAX(trade_date) as last_sale FROM sales_history WHERE model = ?",
        (model,))
    dates = c.fetchone()
    turnover_days = None
    if dates["first_sale"] and dates["last_sale"] and total_count > 1:
        try:
            d1 = datetime.strptime(dates["first_sale"][:10], "%Y-%m-%d")
            d2 = datetime.strptime(dates["last_sale"][:10], "%Y-%m-%d")
            span = (d2 - d1).days
            if span > 0:
                turnover_days = round(span / total_count, 1)
        except Exception:
            pass

    conn.close()
    return jsonify({
        "ok": True,
        "model": model,
        "total_count": total_count,
        "avg_price": avg_price,
        "size_distribution": size_distribution,
        "turnover_days": turnover_days,
        "sales": sales[:50],  # 최근 50건만
    })


@app.route("/api/sales/scheduler/status")
def api_sales_scheduler_status():
    """판매 수집 스케줄러 상태"""
    with _sales_lock:
        return jsonify({"ok": True, **sales_scheduler_state})


@app.route("/api/sales/scheduler/start", methods=["POST"])
def api_sales_scheduler_start():
    """판매 수집 스케줄러 시작"""
    with _sales_lock:
        if sales_scheduler_state["running"]:
            return jsonify({"ok": True, "msg": "이미 실행 중"})
        sales_scheduler_state["running"] = True
    _schedule_next_sales_sync()
    return jsonify({"ok": True, "msg": "스케줄러 시작됨 (30분 간격 + 지터)"})


@app.route("/api/sales/scheduler/stop", methods=["POST"])
def api_sales_scheduler_stop():
    """판매 수집 스케줄러 중지"""
    global _sales_timer
    with _sales_lock:
        sales_scheduler_state["running"] = False
        sales_scheduler_state["next_run"] = None
    if _sales_timer:
        _sales_timer.cancel()
        _sales_timer = None
    return jsonify({"ok": True, "msg": "스케줄러 중지됨"})


@app.route("/api/sales/alerts")
def api_sales_alerts():
    """새 체결건 알림 조회 (폴링용)"""
    with _new_sales_lock:
        alerts = list(_new_sales_alerts)
    return jsonify({"ok": True, "alerts": alerts, "count": len(alerts)})


@app.route("/api/sales/alerts/dismiss", methods=["POST"])
def api_sales_alerts_dismiss():
    """알림 전체 확인 (클리어)"""
    with _new_sales_lock:
        _new_sales_alerts.clear()
    return jsonify({"ok": True})


@app.route("/api/sales/rebid-recommendations")
def api_sales_rebid_recommendations():
    """재입찰 추천 목록 — 최근 판매 건 중 재입찰 가능한 항목"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    # 최근 30일 판매 중 모델번호가 있는 건
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    c.execute("""
        SELECT * FROM sales_history
        WHERE model != '' AND trade_date >= ?
        ORDER BY trade_date DESC LIMIT 50
    """, (month_ago,))
    sales = [dict(r) for r in c.fetchall()]
    conn.close()

    # 각 건에 대해 현재 시장 가격 정보 추가 (price_history에서 조회)
    recommendations = []
    for s in sales:
        rec = {
            "order_id": s["order_id"],
            "model": s["model"],
            "product_info": s["product_info"],
            "size": s["size"],
            "sale_price": s["sale_price"],
            "trade_date": s["trade_date"],
        }
        recommendations.append(rec)

    return jsonify({"ok": True, "recommendations": recommendations})


# ═══════════════════════════════════════════
# 헬스체크
# ════════════════════════════��══════════════

@app.route("/api/health")
def api_health():
    """시스템 헬스체크: 인증/스케줄러/데이터 신선도/DB/백업 상태"""
    try:
        now = datetime.now()
        result = {}

        # 인증 파일 상태
        for key, fname in [("auth_partner", "auth_state.json"), ("auth_kream", "auth_state_kream.json")]:
            fpath = BASE_DIR / fname
            if fpath.exists():
                mtime = datetime.fromtimestamp(fpath.stat().st_mtime)
                age_hours = round((now - mtime).total_seconds() / 3600, 1)
                # 파일 내용이 유효한지 확인 (빈 JSON이 아닌지)
                valid = False
                try:
                    data = json.loads(fpath.read_text())
                    valid = bool(data.get("cookies") or data.get("origins"))
                except Exception:
                    pass
                result[key] = {
                    "exists": True,
                    "last_modified": mtime.isoformat(),
                    "age_hours": age_hours,
                    "valid": valid,
                }
            else:
                result[key] = {
                    "exists": False,
                    "last_modified": None,
                    "age_hours": None,
                    "valid": False,
                }

        # 허브넷 세션 (Step 17-D Phase 2-A.1): PHP 세션 시스템 — PHPSESSID 1개만 있으면 정상
        # 정상 세션도 ~317byte. 빈 cookies 또는 JSON 파싱 실패 시 valid=false.
        hubnet_path = BASE_DIR / "auth_state_hubnet.json"
        auth_hubnet = {
            "exists": False,
            "last_modified": None,
            "age_hours": None,
            "valid": False,
        }
        if hubnet_path.exists():
            mtime = datetime.fromtimestamp(hubnet_path.stat().st_mtime)
            auth_hubnet["exists"] = True
            auth_hubnet["last_modified"] = mtime.isoformat()
            auth_hubnet["age_hours"] = round((now - mtime).total_seconds() / 3600, 1)
            try:
                data = json.loads(hubnet_path.read_text())
                cookies = data.get("cookies", [])
                has_phpsessid = any(c.get("name") == "PHPSESSID" for c in cookies)
                if has_phpsessid:
                    auth_hubnet["valid"] = True
            except Exception:
                auth_hubnet["valid"] = False
        result["auth_hubnet"] = auth_hubnet

        # 스케줄러 상태
        result["schedulers"] = {
            "monitor": "running" if monitor_state.get("running") else "stopped",
            "sales": "running" if sales_scheduler_state.get("running") else "stopped",
            "backup": "running" if backup_state.get("running") else "stopped",
        }

        # 세션 사전 갱신 스케줄러 (Step 17-D Phase 2-B)
        result["session_refresh"] = {
            "enabled": _session_refresh_status["enabled"],
            "last_run": _session_refresh_status["last_run"],
            "next_run": _session_refresh_status["next_run"],
        }

        # 마지막 판매 수집
        last_sale_collected = None
        last_sale_age_hours = None
        try:
            conn = sqlite3.connect(str(PRICE_DB))
            row = conn.execute("SELECT MAX(collected_at) FROM sales_history").fetchone()
            conn.close()
            if row and row[0]:
                last_sale_collected = row[0]
                try:
                    last_dt = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
                    last_sale_age_hours = round((now - last_dt).total_seconds() / 3600, 1)
                except Exception:
                    pass
        except Exception:
            pass
        result["last_sale_collected"] = last_sale_collected
        result["last_sale_age_hours"] = last_sale_age_hours

        # DB 크기
        db_size_mb = None
        try:
            db_size_mb = round(PRICE_DB.stat().st_size / (1024 * 1024), 2)
        except Exception:
            pass
        result["db_size_mb"] = db_size_mb

        # 마지막 백업
        backup_dir = Path.home() / "Desktop" / "kream_backups"
        last_backup = None
        last_backup_age_hours = None
        try:
            if backup_dir.exists():
                backups = sorted(backup_dir.glob("price_history_*.db"), key=lambda f: f.stat().st_mtime, reverse=True)
                if backups:
                    bmtime = datetime.fromtimestamp(backups[0].stat().st_mtime)
                    last_backup = bmtime.isoformat()
                    last_backup_age_hours = round((now - bmtime).total_seconds() / 3600, 1)
        except Exception:
            pass
        result["last_backup"] = last_backup
        result["last_backup_age_hours"] = last_backup_age_hours

        # ��합 상태 판정
        status = "healthy"

        # critical 조건
        if (not result["auth_partner"]["exists"] or not result["auth_partner"]["valid"]
                or (result["auth_partner"]["age_hours"] is not None and result["auth_partner"]["age_hours"] >= 24)
                or not result["auth_kream"]["exists"] or not result["auth_kream"]["valid"]
                or (result["auth_kream"]["age_hours"] is not None and result["auth_kream"]["age_hours"] >= 24)
                or (last_sale_age_hours is not None and last_sale_age_hours >= 24)):
            status = "critical"
        # warning 조건
        elif ((result["auth_partner"]["age_hours"] is not None and result["auth_partner"]["age_hours"] >= 12)
                or (result["auth_kream"]["age_hours"] is not None and result["auth_kream"]["age_hours"] >= 12)
                or (last_sale_age_hours is not None and last_sale_age_hours >= 12)
                or (last_backup_age_hours is not None and last_backup_age_hours >= 25)):
            status = "warning"

        result["status"] = status

        # 환경 정보 (Step 18-A) + 상세 (Step 18-B)
        try:
            settings_data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
            result["environment"] = settings_data.get("environment", "unknown")
            result["kream_main_accessible"] = settings_data.get("kream_main_accessible", None)
            result["env_detection_detail"] = settings_data.get("env_detection_detail", None)
            result["env_checked_at"] = settings_data.get("env_checked_at", None)
        except Exception:
            result["environment"] = "unknown"
            result["kream_main_accessible"] = None
            result["env_detection_detail"] = None
            result["env_checked_at"] = None

        return jsonify(result)

    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})


def detect_environment():
    """kream.co.kr 실제 HTTP 응답 확인. settings.json에 캐시.
    Step 33-D: timeout만으로 overseas 단정 X — 네이버 백업 체크로 IP차단/오프라인 분리.
    FORCE_ENV 환경변수가 설정되면 그 값을 우선 사용 (수동 우회).
    """
    try:
        settings = json.loads(SETTINGS_FILE.read_text(encoding="utf-8")) if SETTINGS_FILE.exists() else {}
    except Exception:
        settings = {}

    accessible = False
    detection_detail = "unknown"
    env_name = "unknown"

    forced = os.environ.get("FORCE_ENV", "").strip()
    if forced:
        env_name = forced
        accessible = forced in ("korea", "imac_kr")
        detection_detail = "forced_via_FORCE_ENV"
        settings["kream_main_accessible"] = accessible
        settings["environment"] = env_name
        settings["env_checked_at"] = datetime.now().isoformat()
        settings["env_detection_detail"] = detection_detail
        try:
            SETTINGS_FILE.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as _e:
            print(f"[ENV] settings 저장 실패: {_e}")
        print(f"[ENV] HTTP-check: accessible={accessible}, detail={detection_detail}, env={env_name}")
        return accessible

    try:
        try:
            import requests as _requests
        except ImportError:
            detection_detail = "requests_module_missing"
            settings["kream_main_accessible"] = False
            settings["environment"] = "offline"
            settings["env_checked_at"] = datetime.now().isoformat()
            settings["env_detection_detail"] = detection_detail
            try:
                SETTINGS_FILE.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception as _e:
                print(f"[ENV] settings 저장 실패: {_e}")
            print(f"[ENV] requests 모듈 없음 — fallback offline")
            return False

        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
        }
        kream_failed = False
        try:
            r = _requests.get("https://kream.co.kr", timeout=15, headers=headers, allow_redirects=True)
            if r.status_code == 200:
                body_lower = r.text.lower()[:5000]
                head_text = r.text[:5000]
                if "kream" in body_lower and ("한정판" in head_text or "application" in body_lower or "<title" in body_lower):
                    accessible = True
                    detection_detail = "http_200_kream_marker"
                else:
                    detection_detail = "http_200_but_no_marker"
                    kream_failed = True
            elif r.status_code in (403, 451):
                detection_detail = f"blocked_http_{r.status_code}"
                kream_failed = True
            else:
                detection_detail = f"http_{r.status_code}"
                kream_failed = True
        except _requests.exceptions.Timeout:
            detection_detail = "timeout"
            kream_failed = True
        except _requests.exceptions.ConnectionError:
            detection_detail = "connection_error"
            kream_failed = True
        except Exception as _e:
            detection_detail = f"error_{type(_e).__name__}"
            kream_failed = True

        if accessible:
            env_name = "korea"
        elif kream_failed:
            naver_ok = False
            try:
                rn = _requests.get("https://www.naver.com", timeout=5, headers=headers, allow_redirects=True)
                if rn.status_code == 200:
                    naver_ok = True
            except Exception:
                naver_ok = False
            if naver_ok:
                env_name = "overseas_blocked"
                detection_detail = f"{detection_detail}+naver_ok"
            else:
                env_name = "offline"
                detection_detail = f"{detection_detail}+naver_fail"
        else:
            env_name = "offline"
    except Exception as _e:
        detection_detail = f"outer_error_{type(_e).__name__}"
        env_name = "offline"

    settings["kream_main_accessible"] = accessible
    settings["environment"] = env_name
    settings["env_checked_at"] = datetime.now().isoformat()
    settings["env_detection_detail"] = detection_detail
    try:
        SETTINGS_FILE.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as _e:
        print(f"[ENV] settings 저장 실패: {_e}")
    print(f"[ENV] HTTP-check: accessible={accessible}, detail={detection_detail}, env={env_name}")
    return accessible


@app.route("/api/env/recheck", methods=["POST"])
def api_env_recheck():
    """환경 감지 수동 재실행 (VPN 토글 후 사용). Step 18-B."""
    try:
        accessible = detect_environment()
        try:
            settings = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            settings = {}
        return jsonify({
            "ok": True,
            "environment": settings.get("environment"),
            "accessible": accessible,
            "detail": settings.get("env_detection_detail"),
            "checked_at": settings.get("env_checked_at"),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/env/test-price-collection", methods=["POST"])
def api_test_price_collection():
    """실전 가격 수집 테스트 (JQ4110으로 1회). Step 18-B."""
    try:
        try:
            import requests as _rq
        except ImportError:
            return jsonify({"ok": False, "error": "requests module missing"}), 500

        try:
            r = _rq.post(
                "http://localhost:5001/api/search",
                json={"model": "JQ4110"},
                timeout=60,
            )
        except Exception as _e:
            return jsonify({
                "ok": False,
                "test_result": "request_error",
                "error": str(_e),
            })

        if r.status_code != 200:
            return jsonify({
                "ok": False,
                "test_result": "api_failed",
                "http_status": r.status_code,
            })

        try:
            d = r.json()
        except Exception:
            d = {}
        if not isinstance(d, dict):
            d = {"_raw_type": type(d).__name__}

        sizes = d.get("sizes", []) or d.get("size_prices", []) or []
        if not isinstance(sizes, list):
            sizes = []
        has_data = len(sizes) > 0 and any(
            (s.get("buy_price") or s.get("buyPrice")) for s in sizes if isinstance(s, dict)
        )

        cache_path = BASE_DIR / "kream_prices.json"
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}
        except Exception:
            cache = {}
        cache["_last_test"] = {
            "at": datetime.now().isoformat(),
            "model": "JQ4110",
            "has_data": has_data,
            "sizes_count": len(sizes),
        }
        try:
            cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as _e:
            print(f"[ENV-TEST] cache 저장 실패: {_e}")

        return jsonify({
            "ok": True,
            "test_result": "success" if has_data else "empty_result",
            "has_data": has_data,
            "sizes_count": len(sizes),
            "sample": sizes[:3] if sizes else [],
            "note": "환경 차단 시 sizes_count=0 또는 buy_price 없음",
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/daily-summary", methods=["GET"])
def api_daily_summary():
    """오늘 작업 요약: 입찰/삭제/판매/pending/auth 실패."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        today = datetime.now().strftime("%Y-%m-%d")

        bids_today = 0
        try:
            c.execute("""
                SELECT COUNT(*) FROM price_adjustments
                WHERE DATE(executed_at) = ? AND status = 'executed'
            """, (today,))
            bids_today = c.fetchone()[0] or 0
        except Exception:
            bids_today = 0

        auto_adjust_today = 0
        try:
            c.execute("""
                SELECT COUNT(*) FROM auto_adjust_log
                WHERE DATE(executed_at) = ?
            """, (today,))
            auto_adjust_today = c.fetchone()[0] or 0
        except Exception:
            auto_adjust_today = 0

        sales_today = 0
        try:
            c.execute("""
                SELECT COUNT(*) FROM sales_history
                WHERE DATE(trade_date) = ?
            """, (today,))
            sales_today = c.fetchone()[0] or 0
        except Exception:
            sales_today = 0

        pending_now = 0
        try:
            c.execute("SELECT COUNT(*) FROM price_adjustments WHERE status = 'pending'")
            pending_now = c.fetchone()[0] or 0
        except Exception:
            pending_now = 0

        auth_failures = 0
        try:
            c.execute("""
                SELECT COUNT(*) FROM notifications
                WHERE type = 'auth_failure'
                AND datetime(created_at) > datetime('now', '-24 hours')
                AND (dismissed IS NULL OR dismissed = 0)
            """)
            auth_failures = c.fetchone()[0] or 0
        except Exception:
            auth_failures = 0

        last_sale = None
        try:
            c.execute("SELECT MAX(trade_date) FROM sales_history")
            last_sale = c.fetchone()[0] or None
        except Exception:
            last_sale = None

        conn.close()

        # Step 18-B: 가격 수집 상태
        last_collection_at = None
        prices_collected_today = 0
        try:
            cache_path = BASE_DIR / "kream_prices.json"
            if cache_path.exists():
                try:
                    cache = json.loads(cache_path.read_text(encoding="utf-8"))
                except Exception:
                    cache = {}
                last_test = cache.get("_last_test", {}) if isinstance(cache, dict) else {}
                if isinstance(last_test, dict):
                    last_collection_at = last_test.get("at")
                    if last_collection_at and str(last_collection_at).startswith(today):
                        prices_collected_today = 1
                if isinstance(cache, dict):
                    for k, v in cache.items():
                        if str(k).startswith("_"):
                            continue
                        if isinstance(v, dict):
                            collected_at = v.get("collected_at") or v.get("updated_at")
                            if collected_at and str(collected_at).startswith(today):
                                prices_collected_today += 1
        except Exception:
            pass

        # Step 20: 의사결정 필요 항목 자동 추출
        decisions = []
        try:
            diag_resp = api_cleanup_diagnose()
            diag_data = None
            if hasattr(diag_resp, 'get_json'):
                try:
                    diag_data = diag_resp.get_json()
                except Exception:
                    diag_data = None
            if diag_data is None and hasattr(diag_resp, 'data'):
                try:
                    diag_data = json.loads(diag_resp.data)
                except Exception:
                    diag_data = None
            if diag_data and diag_data.get('ok'):
                stats = diag_data.get('stats', {}) or {}
                if stats.get('withdraw', 0) > 0:
                    decisions.append({
                        'type': 'cleanup_withdraw',
                        'priority': 'high',
                        'count': stats['withdraw'],
                        'message': f"회수 권장 입찰 {stats['withdraw']}건 — 적자 또는 조정 후 마진 미달",
                        'action_url': '#cleanup'
                    })
                if stats.get('need_cost', 0) > 0:
                    decisions.append({
                        'type': 'cleanup_need_cost',
                        'priority': 'medium',
                        'count': stats['need_cost'],
                        'message': f"원가 입력 필요 {stats['need_cost']}건",
                        'action_url': '#cleanup'
                    })
        except Exception:
            pass

        try:
            settings_data = json.loads((BASE_DIR / 'settings.json').read_text(encoding='utf-8'))
            if not settings_data.get('kream_main_accessible'):
                decisions.append({
                    'type': 'env_blocked',
                    'priority': 'medium',
                    'count': 1,
                    'message': 'kream.co.kr 접근 차단 — 가격수집/자동조정 제한',
                    'action_url': '#settings'
                })
        except Exception:
            pass

        try:
            for name in ['auth_state.json', 'auth_state_kream.json']:
                p = BASE_DIR / name
                if p.exists():
                    age_h = (datetime.now() - datetime.fromtimestamp(p.stat().st_mtime)).total_seconds() / 3600
                    if age_h > 18:
                        decisions.append({
                            'type': 'auth_aging',
                            'priority': 'high',
                            'count': 1,
                            'message': f"{name} {round(age_h)}시간 경과 — 곧 만료",
                            'action_url': None
                        })
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "date": today,
            "summary": {
                "bids_today": bids_today,
                "auto_adjust_today": auto_adjust_today,
                "sales_today": sales_today,
                "pending_now": pending_now,
                "auth_failures_24h": auth_failures,
                "last_sale_date": last_sale,
                "last_collection_at": last_collection_at,
                "prices_collected_today": prices_collected_today,
                "decisions_pending": decisions,
                "decisions_count": len(decisions),
            },
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/health/test-alert", methods=["POST"])
def api_health_test_alert():
    """테스트 알림 발송 (쿨다운 무시)"""
    data = request.json or {}
    key = data.get("key", "test_alert")
    message = data.get("message", "테스트 알림입니다")
    try:
        result = health_alerter.alert(key, message, force=True)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ═══════════════════════════════════════════
# 판매 패턴 분석
# ═══════════════════════════════════════════

@app.route("/api/sales/pattern-analysis")
def api_sales_pattern_analysis():
    """판매 패턴 분석: 모델별 판매 빈도, 시간대 분포, 추천 모니터링 간격"""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row

        # 모델별 판매 데이터
        rows = conn.execute("""
            SELECT model, trade_date, sale_price
            FROM sales_history
            WHERE model IS NOT NULL AND model != ''
            ORDER BY model, trade_date
        """).fetchall()

        if not rows:
            conn.close()
            return jsonify({
                "models": [],
                "hourly_distribution": [{"hour": h, "count": 0} for h in range(24)],
                "summary": {"total_models": 0, "models_recommended_30min": 0, "data_period_days": 0}
            })

        # 모델별 그룹핑
        from collections import defaultdict
        model_sales = defaultdict(list)
        all_dates = []
        hourly_counts = [0] * 24

        for r in rows:
            model = r["model"]
            td = r["trade_date"] or ""
            model_sales[model].append(td)
            all_dates.append(td)
            # 시간대 추출
            try:
                dt = datetime.strptime(td, "%Y-%m-%d %H:%M:%S")
                hourly_counts[dt.hour] += 1
            except Exception:
                try:
                    dt = datetime.strptime(td, "%Y-%m-%d")
                    hourly_counts[12] += 1  # 시간 정보 없으면 정오로
                except Exception:
                    pass

        # 전체 기간 계산
        valid_dates = []
        for d in all_dates:
            try:
                valid_dates.append(datetime.strptime(d[:10], "%Y-%m-%d"))
            except Exception:
                pass
        data_period_days = 0
        if valid_dates:
            data_period_days = max(1, (max(valid_dates) - min(valid_dates)).days)

        # 모델별 분석
        models_result = []
        models_30min = 0

        for model, dates in model_sales.items():
            count = len(dates)
            if count < 3:
                continue  # 3건 미만 제외

            # 날짜 파싱
            parsed = []
            for d in dates:
                try:
                    parsed.append(datetime.strptime(d, "%Y-%m-%d %H:%M:%S"))
                except Exception:
                    try:
                        parsed.append(datetime.strptime(d, "%Y-%m-%d"))
                    except Exception:
                        pass

            if len(parsed) < 3:
                continue

            parsed.sort()
            first_sale = parsed[0].strftime("%Y-%m-%d %H:%M")
            last_sale = parsed[-1].strftime("%Y-%m-%d %H:%M")
            span_days = max(1, (parsed[-1] - parsed[0]).days)

            # 평균 판매 간격 (시간)
            intervals = []
            for i in range(1, len(parsed)):
                diff_hours = (parsed[i] - parsed[i - 1]).total_seconds() / 3600
                intervals.append(diff_hours)
            avg_hours = sum(intervals) / len(intervals) if intervals else 999

            # 추천 모니터링 간격
            if avg_hours < 4:
                recommended = "30분"
                models_30min += 1
            elif avg_hours < 12:
                recommended = "1시간"
            else:
                recommended = "3시간"

            models_result.append({
                "model": model,
                "sales_count": count,
                "first_sale": first_sale,
                "last_sale": last_sale,
                "span_days": span_days,
                "avg_hours_between_sales": round(avg_hours, 1),
                "recommended_monitoring": recommended,
            })

        # 판매 수 내림차순 정렬
        models_result.sort(key=lambda x: x["sales_count"], reverse=True)

        conn.close()
        return jsonify({
            "models": models_result,
            "hourly_distribution": [{"hour": h, "count": hourly_counts[h]} for h in range(24)],
            "summary": {
                "total_models": len(models_result),
                "models_recommended_30min": models_30min,
                "data_period_days": data_period_days,
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ═══════════════════════════════════════════
# 경보 연동: 헬스체크 5분 모니터링 + 입찰 연속 실패 + 판매 무데이터
# ═══════════════════════════════════════════

_health_alert_timer = None
_bid_monitor_fail_count = 0  # 입찰 모니터링 연속 실패 카운트


def _health_alert_check():
    """5분마다 헬스체크 내부 호출 → critical 시 경보 발송"""
    global _health_alert_timer
    try:
        with app.test_request_context():
            resp = api_health()
            data = json.loads(resp.get_data())
            if data.get("status") == "critical":
                # critical 상세 이유 수집
                reasons = []
                if not data.get("auth_partner", {}).get("valid"):
                    reasons.append("판매자센터 인증 무효")
                if not data.get("auth_kream", {}).get("valid"):
                    reasons.append("KREAM 인증 무효")
                ah = data.get("auth_partner", {}).get("age_hours")
                if ah and ah >= 24:
                    reasons.append(f"판매자센터 인증 {ah}시간 경과")
                ah2 = data.get("auth_kream", {}).get("age_hours")
                if ah2 and ah2 >= 24:
                    reasons.append(f"KREAM 인증 {ah2}시간 경과")
                lsa = data.get("last_sale_age_hours")
                if lsa and lsa >= 24:
                    reasons.append(f"판매 수집 {lsa}시간 경과")
                msg = "시스템 상태 CRITICAL: " + ", ".join(reasons) if reasons else "시스템 상태 CRITICAL"
                health_alerter.alert("health_critical", msg, cooldown_minutes=60)

            # 판매 수집 12시간 무데이터 체크
            lsa = data.get("last_sale_age_hours")
            if lsa is not None and lsa >= 12:
                health_alerter.alert(
                    "sales_no_data_12h",
                    f"판매 수집 {lsa:.1f}시간 동안 새 데이터 없음. 스케줄러/세션 확인 필요.",
                    cooldown_minutes=120,
                )
    except Exception as e:
        print(f"[경보] 헬스체크 모니터링 오류: {e}")
    # 5분 후 재실행
    _health_alert_timer = threading.Timer(300, _health_alert_check)
    _health_alert_timer.daemon = True
    _health_alert_timer.start()


def on_bid_monitor_success():
    """입찰 모니터링 성공 시 카운터 리셋"""
    global _bid_monitor_fail_count
    _bid_monitor_fail_count = 0


def on_bid_monitor_failure(error_msg=""):
    """입찰 모니터링 실패 시 카운터 증가 → 3회 연속 시 경보"""
    global _bid_monitor_fail_count
    _bid_monitor_fail_count += 1
    if _bid_monitor_fail_count >= 3:
        try:
            health_alerter.alert(
                "bid_monitor_consecutive_fail",
                f"입찰 모니터링 {_bid_monitor_fail_count}회 연속 실패. 마지막 오류: {error_msg}",
                cooldown_minutes=60,
            )
        except Exception:
            pass


# ═══════════════════════════════════════════
# 자동 가격 조정 (언더컷 자동 방어)
# ═══════════════════════════════════════════


def _get_auto_adjust_settings():
    """자동 조정 관련 설정값 로드"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    return {
        "enabled": settings.get("auto_adjust_enabled", False),
        "daily_max": int(settings.get("auto_adjust_daily_max", 10)),
        "min_profit": int(settings.get("auto_adjust_min_profit", 4000)),
    }


def _auto_adjust_today_stats():
    """오늘 자동 실행 통계"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute(
        "SELECT action, COUNT(*) FROM auto_adjust_log WHERE date(executed_at)=? GROUP BY action",
        (today,)
    )
    stats = dict(c.fetchall())
    conn.close()
    return {
        "modified": stats.get("auto_modified", 0),
        "skipped_no_cost": stats.get("skipped_no_cost", 0),
        "skipped_profit_low": stats.get("skipped_profit_low", 0),
        "skipped_cooldown": stats.get("skipped_cooldown", 0),
        "skipped_daily_limit": stats.get("skipped_daily_limit", 0),
        "skipped_failure_rate": stats.get("skipped_failure_rate", 0),
        "skipped_stale_data": stats.get("skipped_stale_data", 0),
        "modify_failed": stats.get("modify_failed", 0),
    }


def _auto_adjust_failure_rate_1h():
    """최근 1시간 실행 실패율"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute(
        "SELECT action FROM auto_adjust_log WHERE action IN ('auto_modified','modify_failed') "
        "AND executed_at > datetime('now', '-1 hour')"
    )
    rows = [r[0] for r in c.fetchall()]
    conn.close()
    if not rows:
        return 0.0
    failed = sum(1 for r in rows if r == "modify_failed")
    return failed / len(rows)


def _log_auto_adjust(order_id, model, size, old_price, new_price, expected_profit, action, skip_reason=None, modify_result=None):
    """auto_adjust_log에 기록"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        """INSERT INTO auto_adjust_log (order_id, model, size, old_price, new_price,
           expected_profit, action, skip_reason, modify_result)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (order_id, model or "", size or "", old_price, new_price,
         expected_profit, action, skip_reason, modify_result)
    )
    conn.commit()
    conn.close()


def auto_execute_approvals(force=False):
    """자동 가격 조정 실행 — pending 건 중 조건 통과 건만 수정

    Args:
        force: True면 auto_adjust_enabled 설정 무시 (수동 1회 실행용)

    Returns:
        dict: {modified, skipped: {total, ...}, failed, details: [...]}
    """
    aa_settings = _get_auto_adjust_settings()
    if not force and not aa_settings["enabled"]:
        return {"modified": 0, "skipped": {"total": 0}, "failed": 0, "details": [], "reason": "disabled"}

    daily_max = aa_settings["daily_max"]
    min_profit = aa_settings["min_profit"]

    # 실패율 체크
    failure_rate = _auto_adjust_failure_rate_1h()
    if failure_rate > 0.2:
        # 자동 OFF
        if aa_settings["enabled"]:
            try:
                existing = json.loads(SETTINGS_FILE.read_text()) if SETTINGS_FILE.exists() else {}
                existing["auto_adjust_enabled"] = False
                existing["auto_adjust_disabled_reason"] = "failure_rate_exceeded"
                SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))
            except Exception:
                pass
            try:
                health_alerter.alert("auto_adjust_disabled",
                    f"자동 가격 조정이 비활성화되었습니다. 최근 1시간 실패율: {failure_rate*100:.0f}%")
            except Exception:
                pass
        return {"modified": 0, "skipped": {"total": 0}, "failed": 0, "details": [],
                "reason": f"failure_rate_exceeded ({failure_rate*100:.0f}%)"}

    # 오늘 통계
    today_stats = _auto_adjust_today_stats()
    today_modified = today_stats["modified"]

    # pending 건 조회 (bid_cost JOIN)
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        """SELECT pa.*, bc.cny_price, bc.exchange_rate, bc.overseas_shipping, bc.other_costs
           FROM price_adjustments pa
           LEFT JOIN bid_cost bc ON pa.order_id = bc.order_id
           WHERE pa.status IN ('pending','profit_low','deficit')
           ORDER BY pa.created_at ASC"""
    )
    pending_rows = [dict(r) for r in c.fetchall()]
    conn.close()

    modified = 0
    failed = 0
    skipped = {"total": 0, "no_cost": 0, "profit_low": 0, "cooldown": 0,
               "daily_limit": 0, "failure_rate": 0, "stale_data": 0}
    details = []
    start_time = datetime.now()
    profit_low_count = 0

    for row in pending_rows:
        # 5분 타임아웃
        if (datetime.now() - start_time).total_seconds() > 300:
            break

        oid = row["order_id"]
        model = row["model"] or ""
        size = row["size"] or ""
        old_price = row["old_price"]
        new_price = row["new_price"]

        # a) 원가 체크
        cny = row.get("cny_price")
        rate = row.get("exchange_rate")
        if not cny or not rate:
            skipped["no_cost"] += 1
            skipped["total"] += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, None,
                             "skipped_no_cost", "원가 미등록")
            details.append({"order_id": oid, "action": "skipped_no_cost"})
            continue

        # 실시간 수익 계산
        ship = row.get("overseas_shipping") or 8000
        other = row.get("other_costs") or 0
        total_cost = round(cny * rate * 1.03 + ship + other)
        settlement = _calc_settlement_for_monitor(new_price)
        expected_profit = settlement - total_cost

        # b) 마진 체크
        if expected_profit < min_profit:
            skipped["profit_low"] += 1
            skipped["total"] += 1
            profit_low_count += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "skipped_profit_low", f"마진 {expected_profit:,}원 < {min_profit:,}원")
            details.append({"order_id": oid, "action": "skipped_profit_low", "profit": expected_profit})
            continue

        # c) 쿨다운 체크 (24시간)
        conn2 = sqlite3.connect(str(PRICE_DB))
        c2 = conn2.cursor()
        c2.execute(
            "SELECT COUNT(*) FROM auto_adjust_log WHERE order_id=? AND action='auto_modified' "
            "AND executed_at > datetime('now', '-24 hours')", (oid,)
        )
        if c2.fetchone()[0] > 0:
            conn2.close()
            skipped["cooldown"] += 1
            skipped["total"] += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "skipped_cooldown", "24시간 쿨다운")
            details.append({"order_id": oid, "action": "skipped_cooldown"})
            continue
        conn2.close()

        # d) 하루 한도 체크
        if today_modified + modified >= daily_max:
            skipped["daily_limit"] += 1
            skipped["total"] += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "skipped_daily_limit", f"하루 한도 {daily_max}건 초과")
            details.append({"order_id": oid, "action": "skipped_daily_limit"})
            continue

        # e) 스테일 데이터 체크 — pending 상태 재확인
        conn3 = sqlite3.connect(str(PRICE_DB))
        c3 = conn3.cursor()
        c3.execute("SELECT status FROM price_adjustments WHERE id=?", (row["id"],))
        curr = c3.fetchone()
        conn3.close()
        if not curr or curr[0] not in ("pending",):
            skipped["stale_data"] += 1
            skipped["total"] += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "skipped_stale_data", f"상태 변경됨: {curr[0] if curr else 'deleted'}")
            details.append({"order_id": oid, "action": "skipped_stale_data"})
            continue

        # f) 실행: modify_bid_price
        print(f"[자동조정] {oid} {model} {size}: {old_price:,} → {new_price:,}원 (수익 {expected_profit:,}원)")
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            ok = loop.run_until_complete(modify_bid_price(oid, new_price, headless=True))
            loop.close()
        except Exception as e:
            ok = False
            print(f"[자동조정] 오류: {e}")

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if ok:
            # 성공 → price_adjustments 상태 업데이트
            conn4 = sqlite3.connect(str(PRICE_DB))
            conn4.execute(
                "UPDATE price_adjustments SET status='executed', executed_at=? WHERE id=?",
                (now_str, row["id"])
            )
            conn4.commit()
            conn4.close()
            modified += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "auto_modified", modify_result="success")
            details.append({"order_id": oid, "action": "auto_modified", "profit": expected_profit})
            print(f"[자동조정] ✓ 성공")
        else:
            # 실패 → price_adjustments는 건드리지 않음
            conn4 = sqlite3.connect(str(PRICE_DB))
            conn4.execute(
                "UPDATE price_adjustments SET status='failed', executed_at=? WHERE id=?",
                (now_str, row["id"])
            )
            conn4.commit()
            conn4.close()
            failed += 1
            _log_auto_adjust(oid, model, size, old_price, new_price, expected_profit,
                             "modify_failed", modify_result="playwright_error")
            details.append({"order_id": oid, "action": "modify_failed"})
            print(f"[자동조정] ✗ 실패")

    # 마진 부족 5건 이상 → 알림
    if profit_low_count >= 5:
        try:
            health_alerter.alert("auto_adjust_low_margin",
                f"마진 부족 건 {profit_low_count}건 누적", cooldown_minutes=1440)
        except Exception:
            pass

    # 하루 한도 초과 → 알림
    if skipped["daily_limit"] > 0:
        try:
            health_alerter.alert("auto_adjust_daily_limit",
                f"하루 자동 조정 한도 {daily_max}건 초과", cooldown_minutes=1440)
        except Exception:
            pass

    result = {"modified": modified, "skipped": skipped, "failed": failed, "details": details}
    print(f"[자동조정] 완료: 수정 {modified}, 건너뜀 {skipped['total']}, 실패 {failed}")
    return result


# ── 자동 조정 API ──

@app.route("/api/auto-adjust/status")
def api_auto_adjust_status():
    """자동 조정 상태"""
    aa = _get_auto_adjust_settings()
    today_stats = _auto_adjust_today_stats()
    failure_rate = _auto_adjust_failure_rate_1h()

    # 마지막 실행 시각
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute("SELECT MAX(executed_at) FROM auto_adjust_log")
    last_row = c.fetchone()
    last_run = last_row[0] if last_row and last_row[0] else None

    # 최근 50건 성공률
    c.execute(
        "SELECT action FROM auto_adjust_log WHERE action IN ('auto_modified','modify_failed') "
        "ORDER BY executed_at DESC LIMIT 50"
    )
    recent = [r[0] for r in c.fetchall()]
    conn.close()

    success_rate_50 = None
    if recent:
        success_rate_50 = round(sum(1 for r in recent if r == "auto_modified") / len(recent) * 100, 1)

    # disabled_reason
    disabled_reason = None
    if SETTINGS_FILE.exists():
        try:
            s = json.loads(SETTINGS_FILE.read_text())
            disabled_reason = s.get("auto_adjust_disabled_reason")
        except Exception:
            pass

    return jsonify({
        "enabled": aa["enabled"],
        "daily_max": aa["daily_max"],
        "min_profit": aa["min_profit"],
        "today_modified": today_stats["modified"],
        "today_skipped": {
            "total": sum(v for k, v in today_stats.items() if k.startswith("skipped_")),
            "no_cost": today_stats["skipped_no_cost"],
            "profit_low": today_stats["skipped_profit_low"],
            "cooldown": today_stats["skipped_cooldown"],
            "daily_limit": today_stats["skipped_daily_limit"],
            "failure_rate": today_stats["skipped_failure_rate"],
            "stale_data": today_stats["skipped_stale_data"],
        },
        "today_failed": today_stats["modify_failed"],
        "last_run": last_run,
        "failure_rate_1h": round(failure_rate * 100, 1),
        "success_rate_50": success_rate_50,
        "disabled_reason": disabled_reason,
    })


@app.route("/api/auto-adjust/toggle", methods=["POST"])
def api_auto_adjust_toggle():
    """자동 조정 ON/OFF 토글"""
    data = request.json or {}
    enabled = bool(data.get("enabled", False))
    existing = {}
    if SETTINGS_FILE.exists():
        try:
            existing = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    existing["auto_adjust_enabled"] = enabled
    if enabled:
        existing.pop("auto_adjust_disabled_reason", None)
    SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))
    return jsonify({"ok": True, "enabled": enabled})


@app.route("/api/auto-adjust/run-once", methods=["POST"])
def api_auto_adjust_run_once():
    """수동 1회 실행 (auto_adjust_enabled 무관)"""
    tid = new_task()
    add_log(tid, "info", "자동 조정 수동 실행 시작...")

    def run():
        try:
            result = auto_execute_approvals(force=True)
            add_log(tid, "success",
                    f"완료: 수정 {result['modified']}, 건너뜀 {result['skipped']['total']}, 실패 {result['failed']}")
            finish_task(tid, result=result)
        except Exception as e:
            traceback.print_exc()
            add_log(tid, "error", f"오류: {e}")
            finish_task(tid, error=str(e))

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"taskId": tid})


@app.route("/api/auto-adjust/history")
def api_auto_adjust_history():
    """자동 조정 이력"""
    limit = int(request.args.get("limit", 50))
    action_filter = request.args.get("filter", "all")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")

    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row

    # Step 17-C: competitor_price (price_adjustments 최근 1건) + bc_cny_source (bid_cost) JOIN
    query = (
        "SELECT aal.*, "
        "(SELECT pa.competitor_price FROM price_adjustments pa "
        "  WHERE pa.order_id = aal.order_id "
        "  ORDER BY pa.created_at DESC LIMIT 1) AS competitor_price, "
        "bc.cny_source AS bc_cny_source "
        "FROM auto_adjust_log aal "
        "LEFT JOIN bid_cost bc ON bc.order_id = aal.order_id "
        "WHERE 1=1"
    )
    params = []

    if action_filter == "modified":
        query += " AND aal.action='auto_modified'"
    elif action_filter == "skipped":
        query += " AND aal.action LIKE 'skipped_%'"
    elif action_filter == "failed":
        query += " AND aal.action='modify_failed'"

    if from_date:
        query += " AND date(aal.executed_at) >= ?"
        params.append(from_date)
    if to_date:
        query += " AND date(aal.executed_at) <= ?"
        params.append(to_date)

    query += " ORDER BY aal.executed_at DESC LIMIT ?"
    params.append(limit)

    c = conn.cursor()
    c.execute(query, params)
    items = [dict(r) for r in c.fetchall()]

    # 전체 건수
    count_query = "SELECT COUNT(*) FROM auto_adjust_log WHERE 1=1"
    count_params = []
    if action_filter == "modified":
        count_query += " AND action='auto_modified'"
    elif action_filter == "skipped":
        count_query += " AND action LIKE 'skipped_%'"
    elif action_filter == "failed":
        count_query += " AND action='modify_failed'"
    if from_date:
        count_query += " AND date(executed_at) >= ?"
        count_params.append(from_date)
    if to_date:
        count_query += " AND date(executed_at) <= ?"
        count_params.append(to_date)

    c.execute(count_query, count_params)
    total = c.fetchone()[0]
    conn.close()

    return jsonify({"items": items, "total": total})


# ═══════════════════════════════════════════
# 자동 재입찰 시스템
# ═══════════════════════════════════════════


def _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                    expected_profit, action, skip_reason=None, new_order_id=None):
    """auto_rebid_log에 기록"""
    conn = sqlite3.connect(str(PRICE_DB))
    conn.execute(
        """INSERT INTO auto_rebid_log
        (original_order_id, model, size, sold_price, new_bid_price,
         expected_profit, action, skip_reason, new_order_id)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (order_id, model or "", size or "", sold_price, new_bid_price,
         expected_profit, action, skip_reason, new_order_id)
    )
    conn.commit()
    conn.close()


def _count_recent_rebids(model, size, hours=24):
    """같은 모델+사이즈의 최근 N시간 성공 재입찰 횟수"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute(
        "SELECT COUNT(*) FROM auto_rebid_log "
        "WHERE model=? AND size=? AND action='auto_rebid_success' "
        "AND executed_at > datetime('now', ?)",
        (model, size, f'-{hours} hours')
    )
    count = c.fetchone()[0]
    conn.close()
    return count


def _count_today_rebid_success():
    """오늘 성공 재입찰 총 건수"""
    conn = sqlite3.connect(str(PRICE_DB))
    c = conn.cursor()
    c.execute(
        "SELECT COUNT(*) FROM auto_rebid_log "
        "WHERE action='auto_rebid_success' "
        "AND date(executed_at)=date('now', 'localtime')"
    )
    count = c.fetchone()[0]
    conn.close()
    return count


def _get_my_other_bids(model, size, exclude_order_id):
    """내 입찰 중 해당 모델+사이즈의 다른 입찰들 (자기 입찰 제외용)"""
    try:
        bids_file = BASE_DIR / "my_bids_local.json"
        if not bids_file.exists():
            return []
        data = json.loads(bids_file.read_text())
        bids = data.get("bids", [])
        return [b for b in bids
                if (b.get("model") or "").upper() == (model or "").upper()
                and str(b.get("size")) == str(size)
                and str(b.get("orderId")) != str(exclude_order_id)]
    except Exception:
        return []


async def _fetch_kream_prices_for_model(model):
    """모델번호로 KREAM 사이즈별 즉시구매가 수집.
    Returns: {size: buy_price} dict
    """
    results = await search_by_model(model)
    if not results:
        return {}

    kream = results[0].get("kream", {})
    sizes = kream.get("sizes", [])
    price_map = {}
    for s in sizes:
        sz = str(s.get("size", ""))
        bp = s.get("buy_price") or s.get("buyPrice") or 0
        if sz and bp:
            price_map[sz] = bp
    return price_map


async def _execute_rebid(product_id, model, size, price, cny_price):
    """Playwright로 실제 입찰 실행"""
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await create_browser(p, headless=get_headless())
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page, context):
            await browser.close()
            return {"success": False, "error": "로그인 필요"}

        bid_data = {
            "product_id": str(product_id),
            "model": model,
            "사이즈": size,
            "입찰가격": price,
            "수량": 1,
            "bid_days": 30,
        }

        try:
            success = await place_bid(page, bid_data, delay=2.0)
        except Exception as e:
            await browser.close()
            return {"success": False, "error": str(e)}

        if success:
            await save_state_with_localstorage(page, context, STATE_FILE, PARTNER_URL)
            settings = {}
            if SETTINGS_FILE.exists():
                try:
                    settings = json.loads(SETTINGS_FILE.read_text())
                except Exception:
                    pass
            rate = settings.get("cnyRate", 215)
            try:
                _save_bid_cost(
                    order_id=f"{product_id}_{size}_rebid",  # v3 확정: timestamp 제거 → UPSERT 보존
                    model=model, size=size,
                    cny_price=float(cny_price) if cny_price else None,
                    exchange_rate=float(rate),
                    overseas_shipping=8000,
                    cny_source=("manual" if cny_price and float(cny_price) > 0 else None),
                )
            except Exception as e:
                print(f"[auto_rebid] bid_cost 실패: {e}")

        await browser.close()
        return {"success": success}


def auto_rebid_after_sale(sale_records):
    """판매 감지 시 자동 재입찰 실행.
    Args:
        sale_records: list of dict [{order_id, model, size, sale_price, product_id}, ...]
    Returns:
        dict: {success, skipped, failed, details}
    """
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass

    if not settings.get("auto_rebid_enabled", False):
        return {"success": 0, "skipped": len(sale_records), "failed": 0,
                "details": [{"reason": "skipped_disabled"} for _ in sale_records]}

    daily_max = int(settings.get("auto_rebid_daily_max", 20))
    blacklist = set(settings.get("auto_rebid_blacklist", []))
    min_profit = int(settings.get("auto_adjust_min_profit", 4000))
    undercut = int(settings.get("undercutAmount", 1000))

    results = {"success": 0, "skipped": 0, "failed": 0, "details": []}

    # 모델별 그룹핑 (KREAM 가격 1회만 수집)
    model_groups = {}
    for sale in sale_records:
        m = sale.get("model") or ""
        model_groups.setdefault(m, []).append(sale)

    async def _process():
        for model, sales_for_model in model_groups.items():
            # 블랙리스트 체크
            if model in blacklist:
                for sale in sales_for_model:
                    _log_auto_rebid(sale.get("order_id"), model, sale.get("size"),
                                    sale.get("sale_price"), None, None,
                                    "skipped_blacklist", f"Model {model} in blacklist")
                    results["skipped"] += 1
                    results["details"].append({"order_id": sale.get("order_id"), "action": "skipped_blacklist"})
                continue

            # 모델별 KREAM 가격 수집
            try:
                kream_prices = await _fetch_kream_prices_for_model(model)
            except Exception as e:
                print(f"[auto_rebid] KREAM 수집 실패 {model}: {e}")
                for sale in sales_for_model:
                    _log_auto_rebid(sale.get("order_id"), model, sale.get("size"),
                                    sale.get("sale_price"), None, None,
                                    "rebid_failed", f"KREAM fetch failed: {e}")
                    results["failed"] += 1
                continue

            for sale in sales_for_model:
                order_id = sale.get("order_id")
                size = str(sale.get("size", ""))
                sold_price = sale.get("sale_price", 0)
                product_id = sale.get("product_id", "")

                # 하루 한도
                today_count = _count_today_rebid_success()
                if today_count + results["success"] >= daily_max:
                    _log_auto_rebid(order_id, model, size, sold_price, None, None,
                                    "skipped_daily_limit", f"Today: {today_count}/{daily_max}")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "skipped_daily_limit"})
                    continue

                # 원가 체크
                conn = sqlite3.connect(str(PRICE_DB))
                conn.row_factory = sqlite3.Row
                c = conn.cursor()
                c.execute("SELECT * FROM bid_cost WHERE order_id=?", (order_id,))
                cost_row = c.fetchone()
                if not cost_row:
                    # model+size로 재시도
                    c.execute(
                        "SELECT * FROM bid_cost WHERE UPPER(model)=? AND size=? ORDER BY created_at DESC LIMIT 1",
                        ((model or "").upper(), size)
                    )
                    cost_row = c.fetchone()
                conn.close()

                if not cost_row:
                    _log_auto_rebid(order_id, model, size, sold_price, None, None,
                                    "skipped_no_cost", "bid_cost not found")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "skipped_no_cost"})
                    continue

                # 루프 가드 (24시간 내 5회)
                recent_count = _count_recent_rebids(model, size, hours=24)
                if recent_count >= 5:
                    _log_auto_rebid(order_id, model, size, sold_price, None, None,
                                    "skipped_loop_guard", f"{recent_count} rebids in 24h")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "skipped_loop_guard"})
                    try:
                        health_alerter.alert("auto_rebid_loop_guard",
                            f"{model} {size} 24시간 내 {recent_count}회 재입찰 - 수동 확인 필요",
                            cooldown_minutes=1440)
                    except Exception:
                        pass
                    continue

                # 재입찰가 계산
                competitor_price = kream_prices.get(size) or kream_prices.get(str(size))
                if not competitor_price:
                    _log_auto_rebid(order_id, model, size, sold_price, None, None,
                                    "rebid_failed", f"Size {size} not in KREAM data")
                    results["failed"] += 1
                    continue

                # 자기 입찰 제외
                my_others = _get_my_other_bids(model, size, order_id)
                if my_others:
                    my_lowest = min((b.get("price", 0) for b in my_others), default=0)
                    if my_lowest and my_lowest <= competitor_price:
                        _log_auto_rebid(order_id, model, size, sold_price, None, None,
                                        "skipped_margin_low", f"My own bid is lowest ({my_lowest})")
                        results["skipped"] += 1
                        continue

                new_bid_price = int(math.ceil((competitor_price - undercut) / 1000) * 1000)
                if new_bid_price <= 0:
                    _log_auto_rebid(order_id, model, size, sold_price, new_bid_price, None,
                                    "rebid_failed", "Calculated price <= 0")
                    results["failed"] += 1
                    continue

                # 가격 급변 체크 (±10%)
                if new_bid_price < sold_price * 0.9 or new_bid_price > sold_price * 1.1:
                    _log_auto_rebid(order_id, model, size, sold_price, new_bid_price, None,
                                    "skipped_price_shift",
                                    f"Sold: {sold_price}, New: {new_bid_price}")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "skipped_price_shift",
                                               "new_bid_price": new_bid_price})
                    try:
                        health_alerter.alert("auto_rebid_price_shift",
                            f"{model} {size} 가격 급변 - 판매가 {sold_price:,} → 재입찰가 {new_bid_price:,}",
                            cooldown_minutes=1440)
                    except Exception:
                        pass
                    continue

                # 예상 수익 계산
                cny = dict(cost_row).get("cny_price", 0)
                rate = dict(cost_row).get("exchange_rate", 0)
                ship = dict(cost_row).get("overseas_shipping", 8000)
                other = dict(cost_row).get("other_costs", 0)
                total_cost = round(cny * rate * 1.03 + ship + other) if cny and rate else None
                expected_profit = None
                if total_cost:
                    settlement = _calc_settlement_for_monitor(new_bid_price)
                    expected_profit = settlement - total_cost

                # 마진 하한
                if expected_profit is not None and expected_profit < min_profit:
                    _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                                    expected_profit, "skipped_margin_low",
                                    f"Profit {expected_profit} < {min_profit}")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "skipped_margin_low",
                                               "expected_profit": expected_profit})
                    continue

                # Step 13: dry-run 모드 (실제 입찰 박지 않음, 의사결정만 로그)
                if settings.get("auto_rebid_dry_run", False):
                    _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                                    expected_profit, "auto_rebid_dry_run",
                                    f"DRY-RUN: would rebid {new_bid_price}")
                    results["skipped"] += 1
                    results["details"].append({"order_id": order_id, "action": "auto_rebid_dry_run",
                                               "new_bid_price": new_bid_price,
                                               "expected_profit": expected_profit})
                    print(f"[auto_rebid] DRY-RUN {model} {size}: sold {sold_price:,} → would rebid {new_bid_price:,}")
                    continue

                # 실제 입찰 실행
                print(f"[auto_rebid] {model} {size}: sold {sold_price:,} → rebid {new_bid_price:,}")
                try:
                    bid_result = await _execute_rebid(
                        product_id=product_id, model=model, size=size,
                        price=new_bid_price, cny_price=cny,
                    )
                    if bid_result.get("success"):
                        _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                                        expected_profit, "auto_rebid_success")
                        results["success"] += 1
                        results["details"].append({"order_id": order_id, "action": "auto_rebid_success",
                                                   "new_bid_price": new_bid_price, "expected_profit": expected_profit})
                        print(f"[auto_rebid] ✓ 성공")
                    else:
                        _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                                        expected_profit, "rebid_failed",
                                        bid_result.get("error", "unknown"))
                        results["failed"] += 1
                        print(f"[auto_rebid] ✗ 실패: {bid_result.get('error')}")
                except Exception as e:
                    _log_auto_rebid(order_id, model, size, sold_price, new_bid_price,
                                    expected_profit, "rebid_failed", str(e))
                    results["failed"] += 1
                    print(f"[auto_rebid] ✗ 예외: {e}")

    # async 실행
    loop = asyncio.new_event_loop()
    try:
        loop.run_until_complete(_process())
    finally:
        loop.close()

    print(f"[auto_rebid] 완료: 성공 {results['success']}, 건너뜀 {results['skipped']}, 실패 {results['failed']}")
    return results


# ── 자동 재입찰 API ──

@app.route("/api/price-book/list", methods=["GET"])
def api_price_book_list():
    """단가표 전체 조회 (Step 37). bulk_only=1 → is_bulk_item=1만."""
    try:
        from services.price_book import list_all
        bulk_only = request.args.get("bulk_only", "0") in ("1", "true", "True")
        items = list_all(bulk_only=bulk_only)
        return jsonify({"ok": True, "items": items, "count": len(items)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/price-book/lookup", methods=["GET"])
def api_price_book_lookup():
    """단가표 조회 (model+size). size 생략 시 size IS NULL 매칭."""
    try:
        from services.price_book import lookup_price
        model = request.args.get("model", "").strip()
        size = request.args.get("size")
        if not model:
            return jsonify({"ok": False, "error": "model required"}), 400
        if size is not None:
            size = size.strip() or None
        row = lookup_price(model, size)
        return jsonify({"ok": True, "matched": bool(row), "row": row})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/price-book/upsert", methods=["POST"])
def api_price_book_upsert():
    """단가표 등록/수정. body: {model, size, cny_price, category, brand, is_bulk_item, notes, source}."""
    try:
        from services.price_book import upsert_price, lookup_price
        body = request.get_json(silent=True) or {}
        model = (body.get("model") or "").strip()
        cny_price = body.get("cny_price")
        if not model or cny_price in (None, ""):
            return jsonify({"ok": False, "error": "model and cny_price required"}), 400
        size = body.get("size")
        if isinstance(size, str):
            size = size.strip() or None
        upsert_price(
            model, size, float(cny_price),
            category=body.get("category"),
            brand=body.get("brand"),
            is_bulk_item=int(body.get("is_bulk_item", 0) or 0),
            notes=body.get("notes"),
            source=body.get("source", "사장님 직접 입력"),
        )
        return jsonify({"ok": True, "row": lookup_price(model, size)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-rebid/no-cost-analysis", methods=["GET"])
def api_auto_rebid_no_cost_analysis():
    """NO_COST 후보 분류 (Step 36 진단). 어떤 원가를 보강해야 할지 알려줌."""
    try:
        from services.auto_rebid import get_rebid_candidates

        hours = int(request.args.get("hours", 720))
        candidates = get_rebid_candidates(hours=hours)

        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row

        result = {
            "model_size_missing": [],
            "size_mismatch": [],
            "matched": [],
        }

        for c in candidates:
            model = c["model"]
            size = str(c["size"])

            model_cnt = conn.execute(
                "SELECT COUNT(*) FROM bid_cost WHERE model = ?", (model,)
            ).fetchone()[0]

            if model_cnt == 0:
                result["model_size_missing"].append({
                    "model": model, "size": size,
                    "sale_price": c["sale_price"], "order_id": c["order_id"],
                })
                continue

            exact_cnt = conn.execute(
                "SELECT COUNT(*) FROM bid_cost WHERE model = ? AND size = ?",
                (model, size),
            ).fetchone()[0]

            if exact_cnt == 0:
                sizes = [r[0] for r in conn.execute(
                    "SELECT DISTINCT size FROM bid_cost WHERE model = ?", (model,)
                ).fetchall()]
                result["size_mismatch"].append({
                    "model": model, "sale_size": size,
                    "registered_sizes": sizes,
                    "sale_price": c["sale_price"], "order_id": c["order_id"],
                })
            else:
                result["matched"].append({"model": model, "size": size})

        conn.close()

        return jsonify({
            "ok": True,
            "summary": {
                "total_candidates": len(candidates),
                "model_missing_count": len(result["model_size_missing"]),
                "size_mismatch_count": len(result["size_mismatch"]),
                "matched_count": len(result["matched"]),
            },
            "details": result,
        })
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/auto-rebid/dry-run", methods=["POST"])
def api_auto_rebid_dry_run():
    """판매 후 N시간 내 후보 시뮬레이션 (실제 입찰 X). Step 35."""
    try:
        from services.auto_rebid import run_dry_run, format_dry_run_for_discord

        settings = {}
        if SETTINGS_FILE.exists():
            try:
                settings = json.loads(SETTINGS_FILE.read_text())
            except Exception:
                pass

        body = request.get_json(silent=True) or {}
        hours = int(body.get("hours", 24))
        result = run_dry_run(settings, hours=hours)

        try:
            msg = format_dry_run_for_discord(result)
            safe_send_alert(subject="자동 재입찰 dry-run", body=msg, alert_type="info")
        except Exception as de:
            result["discord_error"] = str(de)

        return jsonify({"ok": True, "result": result})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/auto-rebid/status")
def api_auto_rebid_status():
    """자동 재입찰 상태"""
    try:
        settings = {}
        if SETTINGS_FILE.exists():
            try:
                settings = json.loads(SETTINGS_FILE.read_text())
            except Exception:
                pass

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        today_success = c.execute(
            "SELECT COUNT(*) FROM auto_rebid_log WHERE action='auto_rebid_success' "
            "AND date(executed_at)=date('now','localtime')").fetchone()[0]
        today_skipped = c.execute(
            "SELECT COUNT(*) FROM auto_rebid_log WHERE action LIKE 'skipped_%' "
            "AND date(executed_at)=date('now','localtime')").fetchone()[0]
        today_failed = c.execute(
            "SELECT COUNT(*) FROM auto_rebid_log WHERE action='rebid_failed' "
            "AND date(executed_at)=date('now','localtime')").fetchone()[0]
        today_dry_run = c.execute(
            "SELECT COUNT(*) FROM auto_rebid_log WHERE action LIKE 'dry_run_%' "
            "AND date(executed_at)=date('now','localtime')").fetchone()[0]
        today_dry_run_go = c.execute(
            "SELECT COUNT(*) FROM auto_rebid_log WHERE action='dry_run_GO' "
            "AND date(executed_at)=date('now','localtime')").fetchone()[0]
        last_sale = c.execute("SELECT MAX(collected_at) FROM sales_history").fetchone()[0]
        conn.close()

        return jsonify({
            "ok": True,
            "enabled": settings.get("auto_rebid_enabled", False),
            "dry_run": settings.get("auto_rebid_dry_run", True),
            "daily_max": settings.get("auto_rebid_daily_max", 20),
            "min_profit": settings.get("auto_rebid_min_profit", 4000),
            "blacklist": settings.get("auto_rebid_blacklist", []),
            "today": {
                "success": today_success,
                "skipped": today_skipped,
                "failed": today_failed,
                "dry_run": today_dry_run,
                "dry_run_go": today_dry_run_go,
            },
            "last_sale": last_sale,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-rebid/toggle", methods=["POST"])
def api_auto_rebid_toggle():
    """자동 재입찰 ON/OFF"""
    try:
        data = request.json or {}
        enabled = bool(data.get("enabled", False))
        existing = {}
        if SETTINGS_FILE.exists():
            try:
                existing = json.loads(SETTINGS_FILE.read_text())
            except Exception:
                pass
        existing["auto_rebid_enabled"] = enabled
        SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))
        return jsonify({"ok": True, "enabled": enabled})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-rebid/run-once", methods=["POST"])
def api_auto_rebid_run_once():
    """수동 1회 실행 (enabled 무관). 최근 1시간 내 sales_history 대상."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT order_id, product_id, model, size, sale_price "
            "FROM sales_history WHERE collected_at > datetime('now', '-1 hour') "
            "ORDER BY collected_at DESC"
        ).fetchall()
        sales = [dict(r) for r in rows]
        conn.close()

        if not sales:
            return jsonify({"ok": True, "message": "최근 1시간 내 판매 없음",
                            "success": 0, "skipped": 0, "failed": 0})

        # 일시적으로 enabled=true로 설정
        existing = {}
        if SETTINGS_FILE.exists():
            try:
                existing = json.loads(SETTINGS_FILE.read_text())
            except Exception:
                pass
        original_enabled = existing.get("auto_rebid_enabled", False)
        existing["auto_rebid_enabled"] = True
        SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))

        try:
            result = auto_rebid_after_sale(sales)
        finally:
            # 원래 상태 복원
            existing2 = {}
            if SETTINGS_FILE.exists():
                try:
                    existing2 = json.loads(SETTINGS_FILE.read_text())
                except Exception:
                    pass
            existing2["auto_rebid_enabled"] = original_enabled
            SETTINGS_FILE.write_text(json.dumps(existing2, ensure_ascii=False, indent=2))

        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-rebid/history")
def api_auto_rebid_history():
    """자동 재입찰 이력"""
    try:
        limit = int(request.args.get("limit", 50))
        filter_type = request.args.get("filter", "all")
        from_date = request.args.get("from_date", "")
        to_date = request.args.get("to_date", "")

        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row

        query = "SELECT * FROM auto_rebid_log WHERE 1=1"
        params = []
        if filter_type == "success":
            query += " AND action='auto_rebid_success'"
        elif filter_type == "skipped":
            query += " AND action LIKE 'skipped_%'"
        elif filter_type == "failed":
            query += " AND action='rebid_failed'"
        elif filter_type == "dry_run":
            query += " AND action LIKE 'dry_run_%'"
        elif filter_type == "real":
            query += " AND action NOT LIKE 'dry_run_%'"
        if from_date:
            query += " AND date(executed_at) >= ?"
            params.append(from_date)
        if to_date:
            query += " AND date(executed_at) <= ?"
            params.append(to_date)
        query += " ORDER BY executed_at DESC LIMIT ?"
        params.append(limit)

        rows = conn.execute(query, params).fetchall()
        history = [dict(r) for r in rows]
        conn.close()

        return jsonify({"ok": True, "history": history, "count": len(history)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ═══════════════════════════════════════════
# 입찰 정리 도구 (Step 5)
# ═══════════════════════════════════════════

def _get_cleanup_settings():
    """정리 설정 조회"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    return {
        "enabled": settings.get("auto_cleanup_enabled", False),
        "types": settings.get("auto_cleanup_types", {
            "duplicate_price": False, "expiring_soon": False,
            "margin_low": False, "low_rank_duplicate": False
        }),
        "daily_max": settings.get("auto_cleanup_daily_max", 30),
        "grace_minutes": settings.get("auto_cleanup_grace_minutes", 60),
    }


def _get_sold_order_ids():
    """판매 완료 order_id 집합 (정리 제외용)"""
    sold = set()
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        rows = conn.execute("SELECT order_id FROM sales_history WHERE order_id IS NOT NULL").fetchall()
        sold = {r[0] for r in rows}
        conn.close()
    except Exception:
        pass
    return sold


def _get_today_cleanup_count():
    """오늘 삭제/pending_delete 건수"""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        row = conn.execute(
            "SELECT COUNT(*) FROM bid_cleanup_log WHERE date(detected_at) = date('now') AND status IN ('pending_delete','deleted')"
        ).fetchone()
        conn.close()
        return row[0] if row else 0
    except Exception:
        return 0


def _get_pending_cleanup_order_ids():
    """현재 pending_delete인 order_id 집합"""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        rows = conn.execute("SELECT order_id FROM bid_cleanup_log WHERE status='pending_delete'").fetchall()
        conn.close()
        return {r[0] for r in rows}
    except Exception:
        return set()


def _load_bids_for_cleanup():
    """my_bids_local.json에서 현재 입찰 로드"""
    try:
        with open("my_bids_local.json", "r") as f:
            data = json.load(f)
        return data.get("bids", [])
    except Exception:
        return []


def _detect_duplicate_price(bids, sold_ids, pending_ids):
    """같은 품번+사이즈 3건 이상 중 비싼 것부터 탐지 (2건은 확정전략으로 유지)"""
    from collections import defaultdict
    groups = defaultdict(list)
    for b in bids:
        oid = b.get("orderId", "")
        if oid in sold_ids or oid in pending_ids:
            continue
        key = (b.get("model", "").upper(), b.get("size", ""))
        if key[0]:
            groups[key].append(b)

    candidates = []
    for (model, size), group in groups.items():
        if len(group) < 3:
            continue
        # 비싼 순 정렬, 2건은 유지하고 나머지 탐지
        sorted_bids = sorted(group, key=lambda x: -(x.get("price") or x.get("bidPrice") or 0))
        for b in sorted_bids[:-2]:  # 가장 싼 2건은 유지
            price = b.get("price") or b.get("bidPrice") or 0
            candidates.append({
                "order_id": b.get("orderId", ""),
                "model": b.get("model", ""),
                "size": b.get("size", ""),
                "price": price,
                "cleanup_type": "duplicate_price",
                "reason": f"같은 {model} {size} {len(group)}건 중 고가 (가격: {price:,}원)",
            })
    return candidates


def _detect_expiring_soon(bids, sold_ids, pending_ids):
    """만료 24시간 이내 탐지 — 만료 필드 없으면 빈 리스트"""
    # 현재 입찰 데이터에 만료일 필드 없음 → 스텁
    print("[정리] expiring_soon: 만료 필드 없음 → 스킵 (향후 KREAM API 지원 시 활성화)")
    return []


def _detect_margin_low(bids, sold_ids, pending_ids):
    """bid_cost 기준 현재 마진 4,000원 미만 탐지 (원가 없으면 스킵 — 가짜값 금지)"""
    candidates = []
    for b in bids:
        oid = b.get("orderId", "")
        if oid in sold_ids or oid in pending_ids:
            continue
        price = b.get("price") or b.get("bidPrice") or 0
        if not price:
            continue
        total_cost = _find_cost_for_bid(b)
        if total_cost is None:
            continue  # 원가 없으면 스킵 (가짜값 금지)
        settlement = _calc_settlement_for_monitor(price)
        margin = settlement - total_cost
        if margin < 4000:
            candidates.append({
                "order_id": oid,
                "model": b.get("model", ""),
                "size": b.get("size", ""),
                "price": price,
                "cleanup_type": "margin_low",
                "reason": f"마진 {margin:,}원 (정산 {settlement:,} - 원가 {total_cost:,})",
            })
    return candidates


def _detect_low_rank_duplicate(bids, sold_ids, pending_ids):
    """내 입찰 중 같은 품번+사이즈에서 순위 1위 아닌 중복 탐지"""
    from collections import defaultdict
    groups = defaultdict(list)
    for b in bids:
        oid = b.get("orderId", "")
        if oid in sold_ids or oid in pending_ids:
            continue
        key = (b.get("model", "").upper(), b.get("size", ""))
        if key[0]:
            groups[key].append(b)

    candidates = []
    for (model, size), group in groups.items():
        if len(group) < 2:
            continue
        # 순위 오름차순 (1위 = 최상), rank 없으면 가격 오름차순
        sorted_bids = sorted(group, key=lambda x: (
            x.get("rank") or x.get("bidRank") or 999,
            x.get("price") or x.get("bidPrice") or 0
        ))
        # 1위 유지, 나머지 탐지
        for b in sorted_bids[1:]:
            rank = b.get("rank") or b.get("bidRank") or "?"
            price = b.get("price") or b.get("bidPrice") or 0
            candidates.append({
                "order_id": b.get("orderId", ""),
                "model": b.get("model", ""),
                "size": b.get("size", ""),
                "price": price,
                "cleanup_type": "low_rank_duplicate",
                "reason": f"{model} {size} 순위 {rank}위 (1위 아닌 중복)",
            })
    return candidates


def detect_cleanup_candidates():
    """4가지 유형 탐지 통합"""
    cs = _get_cleanup_settings()
    bids = _load_bids_for_cleanup()
    if not bids:
        return []

    sold_ids = _get_sold_order_ids()
    pending_ids = _get_pending_cleanup_order_ids()

    # my_bids_local의 status가 있으면 판매완료 추가 체크
    for b in bids:
        st = (b.get("status") or "").lower()
        if "완료" in st or "sold" in st:
            sold_ids.add(b.get("orderId", ""))

    all_candidates = []
    type_map = {
        "duplicate_price": _detect_duplicate_price,
        "expiring_soon": _detect_expiring_soon,
        "margin_low": _detect_margin_low,
        "low_rank_duplicate": _detect_low_rank_duplicate,
    }

    for tname, func in type_map.items():
        if cs["types"].get(tname, False):
            try:
                results = func(bids, sold_ids, pending_ids)
                all_candidates.extend(results)
            except Exception as e:
                print(f"[정리] {tname} 탐지 오류: {e}")

    return all_candidates


def run_cleanup_detection():
    """탐지 + pending_delete 기록 + 이메일 알림"""
    cs = _get_cleanup_settings()
    today_count = _get_today_cleanup_count()
    remaining = cs["daily_max"] - today_count

    if remaining <= 0:
        print(f"[정리] 하루 한도 도달 ({cs['daily_max']}건)")
        return {"detected": 0, "saved": 0, "reason": "daily_max_reached"}

    candidates = detect_cleanup_candidates()
    if not candidates:
        return {"detected": 0, "saved": 0}

    # 하루 한도 적용
    candidates = candidates[:remaining]

    grace = cs["grace_minutes"]
    now = datetime.now()
    scheduled = now + timedelta(minutes=grace)
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    sched_str = scheduled.strftime("%Y-%m-%d %H:%M:%S")

    saved = 0
    conn = sqlite3.connect(str(PRICE_DB))
    for c_item in candidates:
        snapshot = json.dumps(c_item, ensure_ascii=False)
        try:
            conn.execute(
                """INSERT INTO bid_cleanup_log
                   (order_id, model, size, price, cleanup_type, reason, status, detected_at, scheduled_delete_at, snapshot)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (c_item["order_id"], c_item["model"], c_item["size"], c_item["price"],
                 c_item["cleanup_type"], c_item["reason"], "pending_delete", now_str, sched_str, snapshot)
            )
            saved += 1
        except Exception as e:
            print(f"[정리] DB 저장 오류: {e}")
    conn.commit()
    conn.close()

    print(f"[정리] 탐지 {len(candidates)}건 → pending_delete {saved}건 (삭제 예정: {sched_str})")

    # 이메일 알림
    if saved > 0:
        try:
            _send_cleanup_email(candidates[:saved])
        except Exception as e:
            print(f"[정리] 이메일 발송 오류: {e}")

    return {"detected": len(candidates), "saved": saved}


def _send_cleanup_email(items):
    """정리 대상 이메일 알림"""
    settings = {}
    if SETTINGS_FILE.exists():
        try:
            settings = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass
    app_password = settings.get("gmail_app_password") or settings.get("emailAppPassword", "")
    if not app_password:
        return

    type_labels = {
        "duplicate_price": "중복(고가)",
        "expiring_soon": "만료 임박",
        "margin_low": "저마진",
        "low_rank_duplicate": "순위 밀림 중복",
    }

    rows = ""
    for it in items:
        rows += (
            f"<tr>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{it['model']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{it['size']}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{it['price']:,}원</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{type_labels.get(it['cleanup_type'], it['cleanup_type'])}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ddd'>{it['reason']}</td>"
            f"</tr>"
        )

    grace = _get_cleanup_settings()["grace_minutes"]
    body = f"""<html><body style="font-family:-apple-system,sans-serif">
<h2 style="color:#111">🧹 입찰 정리 대상 알림</h2>
<p>{datetime.now().strftime('%Y-%m-%d %H:%M')} 기준, <b>{len(items)}건</b>의 입찰이 정리 대상입니다.</p>
<p style="color:#e65100;font-weight:600">⏰ {grace}분 유예 후 자동 삭제됩니다. 대시보드에서 취소 가능합니다.</p>
<table style="border-collapse:collapse;width:100%;font-size:13px">
<thead><tr style="background:#f5f5f5">
<th style="padding:8px;border:1px solid #ddd">모델</th>
<th style="padding:8px;border:1px solid #ddd">사이즈</th>
<th style="padding:8px;border:1px solid #ddd">입찰가</th>
<th style="padding:8px;border:1px solid #ddd">유형</th>
<th style="padding:8px;border:1px solid #ddd">사유</th>
</tr></thead>
<tbody>{rows}</tbody>
</table>
<p style="margin-top:20px">
<a href="http://localhost:5001" style="background:#e65100;color:#fff;padding:12px 28px;
text-decoration:none;border-radius:8px;font-weight:600">대시보드에서 확인</a>
</p>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[KREAM] 입찰 정리 대상 {len(items)}건 ({grace}분 유예)"
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_RECEIVER
    msg.attach(MIMEText(body, "html", "utf-8"))
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, app_password)
            server.send_message(msg)
        print(f"[정리] 이메일 발송 완료: {len(items)}건")
    except Exception as e:
        print(f"[정리] 이메일 발송 실패: {e}")


def run_cleanup_execution():
    """scheduled_delete_at 지난 pending_delete 건 실제 삭제"""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(str(PRICE_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM bid_cleanup_log WHERE status='pending_delete' AND scheduled_delete_at <= ?",
        (now_str,)
    ).fetchall()
    conn.close()

    if not rows:
        return {"executed": 0, "failed": 0}

    sold_ids = _get_sold_order_ids()
    executed = 0
    failed = 0
    stale = 0

    for row in rows:
        row_id = row["id"]
        order_id = row["order_id"]

        # 5중 안전장치: 판매 완료 건 재확인 (스테일 체크)
        if order_id in sold_ids:
            conn = sqlite3.connect(str(PRICE_DB))
            conn.execute(
                "UPDATE bid_cleanup_log SET status='stale_skipped', executed_at=? WHERE id=?",
                (now_str, row_id)
            )
            conn.commit()
            conn.close()
            stale += 1
            print(f"[정리] {order_id}: 판매 완료 확인 → 스킵")
            continue

        # 스테일 체크: 현재 입찰 목록에 아직 존재하는지 확인
        current_bids = _load_bids_for_cleanup()
        current_oids = {b.get("orderId", "") for b in current_bids}
        if order_id not in current_oids:
            conn = sqlite3.connect(str(PRICE_DB))
            conn.execute(
                "UPDATE bid_cleanup_log SET status='stale_skipped', executed_at=?, cancel_reason='입찰 목록에 없음' WHERE id=?",
                (now_str, row_id)
            )
            conn.commit()
            conn.close()
            stale += 1
            print(f"[정리] {order_id}: 입찰 목록에 없음 → 스킵")
            continue

        # 실제 삭제 실행 (기존 delete_bids 로직 재사용)
        success = _execute_cleanup_delete(order_id)
        conn = sqlite3.connect(str(PRICE_DB))
        if success:
            conn.execute(
                "UPDATE bid_cleanup_log SET status='deleted', executed_at=? WHERE id=?",
                (now_str, row_id)
            )
            executed += 1
            print(f"[정리] {order_id}: 삭제 완료")
        else:
            conn.execute(
                "UPDATE bid_cleanup_log SET status='delete_failed', executed_at=? WHERE id=?",
                (now_str, row_id)
            )
            failed += 1
            print(f"[정리] {order_id}: 삭제 실패")
        conn.commit()
        conn.close()

    print(f"[정리] 실행 완료: 삭제 {executed}, 실패 {failed}, 스테일 {stale}")
    return {"executed": executed, "failed": failed, "stale": stale}


def _execute_cleanup_delete(order_id):
    """기존 delete_bids 로직 재사용하여 단건 삭제"""
    try:
        tid = new_task()
        add_log(tid, "info", f"[정리] {order_id} 삭제")
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        result = loop.run_until_complete(delete_bids([order_id], tid))
        loop.close()
        finish_task(tid, result=result)
        return result.get("success", 0) > 0
    except Exception as e:
        print(f"[정리] 삭제 실행 오류 ({order_id}): {e}")
        return False


# ── 입찰 정리 API 6개 ──

@app.route("/api/auto-cleanup/status")
def api_auto_cleanup_status():
    """정리 상태 조회"""
    cs = _get_cleanup_settings()
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        pending_count = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE status='pending_delete'").fetchone()[0]
        today_deleted = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE status='deleted' AND date(executed_at)=date('now')").fetchone()[0]
        today_detected = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE date(detected_at)=date('now')").fetchone()[0]
        total_deleted = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE status='deleted'").fetchone()[0]
        total_cancelled = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE status='cancelled'").fetchone()[0]
        total_failed = conn.execute("SELECT COUNT(*) FROM bid_cleanup_log WHERE status='delete_failed'").fetchone()[0]
        conn.close()
    except Exception:
        pending_count = today_deleted = today_detected = total_deleted = total_cancelled = total_failed = 0

    return jsonify({
        "ok": True,
        "enabled": cs["enabled"],
        "types": cs["types"],
        "daily_max": cs["daily_max"],
        "grace_minutes": cs["grace_minutes"],
        "stats": {
            "pending": pending_count,
            "today_deleted": today_deleted,
            "today_detected": today_detected,
            "total_deleted": total_deleted,
            "total_cancelled": total_cancelled,
            "total_failed": total_failed,
        }
    })


@app.route("/api/auto-cleanup/toggle", methods=["POST"])
def api_auto_cleanup_toggle():
    """정리 ON/OFF 토글 (enabled 및 types 부분 업데이트 지원)"""
    data = request.json or {}
    existing = {}
    if SETTINGS_FILE.exists():
        try:
            existing = json.loads(SETTINGS_FILE.read_text())
        except Exception:
            pass

    if "enabled" in data:
        existing["auto_cleanup_enabled"] = bool(data["enabled"])
    if "types" in data and isinstance(data["types"], dict):
        cur_types = existing.get("auto_cleanup_types", {
            "duplicate_price": False, "expiring_soon": False,
            "margin_low": False, "low_rank_duplicate": False
        })
        for k, v in data["types"].items():
            if k in cur_types:
                cur_types[k] = bool(v)
        existing["auto_cleanup_types"] = cur_types
    if "daily_max" in data:
        existing["auto_cleanup_daily_max"] = int(data["daily_max"])
    if "grace_minutes" in data:
        existing["auto_cleanup_grace_minutes"] = int(data["grace_minutes"])

    SETTINGS_FILE.write_text(json.dumps(existing, ensure_ascii=False, indent=2))
    return jsonify({
        "ok": True,
        "enabled": existing.get("auto_cleanup_enabled", False),
        "types": existing.get("auto_cleanup_types", {}),
    })


@app.route("/api/auto-cleanup/pending")
def api_auto_cleanup_pending():
    """pending_delete 목록 조회"""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM bid_cleanup_log WHERE status='pending_delete' ORDER BY scheduled_delete_at ASC"
        ).fetchall()
        conn.close()
        return jsonify({"ok": True, "items": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-cleanup/cancel", methods=["POST"])
def api_auto_cleanup_cancel():
    """pending_delete 취소 (Undo)"""
    data = request.json or {}
    ids = data.get("ids", [])
    reason = data.get("reason", "사용자 취소")
    if not ids:
        return jsonify({"error": "ids 필요"}), 400

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cancelled = 0
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        for cid in ids:
            result = conn.execute(
                "UPDATE bid_cleanup_log SET status='cancelled', executed_at=?, cancel_reason=? WHERE id=? AND status='pending_delete'",
                (now_str, reason, cid)
            )
            if result.rowcount > 0:
                cancelled += 1
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True, "cancelled": cancelled})


@app.route("/api/auto-cleanup/run-once", methods=["POST"])
def api_auto_cleanup_run_once():
    """수동 1회 실행 (탐지 + 실행)"""
    try:
        exec_result = run_cleanup_execution()
        detect_result = run_cleanup_detection()
        return jsonify({
            "ok": True,
            "execution": exec_result,
            "detection": detect_result,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/auto-cleanup/history")
def api_auto_cleanup_history():
    """정리 이력 조회"""
    limit = request.args.get("limit", 50, type=int)
    filter_type = request.args.get("filter", "")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")

    try:
        conn = sqlite3.connect(str(PRICE_DB))
        conn.row_factory = sqlite3.Row
        query = "SELECT * FROM bid_cleanup_log WHERE 1=1"
        params = []
        if filter_type == "deleted":
            query += " AND status='deleted'"
        elif filter_type == "cancelled":
            query += " AND status='cancelled'"
        elif filter_type == "failed":
            query += " AND status='delete_failed'"
        elif filter_type == "pending":
            query += " AND status='pending_delete'"
        if from_date:
            query += " AND date(detected_at) >= ?"
            params.append(from_date)
        if to_date:
            query += " AND date(detected_at) <= ?"
            params.append(to_date)
        query += " ORDER BY detected_at DESC LIMIT ?"
        params.append(limit)

        rows = conn.execute(query, params).fetchall()
        conn.close()
        return jsonify({"ok": True, "history": [dict(r) for r in rows], "count": len(rows)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ═══════════════════════════════════════════
# 허브넷 통합 API (Step 8 — 작업지시서 §4.1)
# 응답 표준: {"success": bool, "data": ...} 또는 {"success": false, "error": "..."}
# ═══════════════════════════════════════════

def _hubnet_session_meta() -> dict:
    """auth_state_hubnet.json 세션 메타 + 만료 추정 체크."""
    meta = {"valid": False, "saved_at": None, "expires_estimate": None}
    try:
        with open(BASE_DIR / "settings.json", "r", encoding="utf-8") as f:
            settings = json.load(f)
    except OSError:
        return meta
    session_path = settings.get("hubnet_session_path") or "auth_state_hubnet.json"
    p = Path(session_path)
    if not p.is_absolute():
        p = BASE_DIR / p
    if not p.exists():
        return meta
    try:
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        return meta
    meta["saved_at"] = data.get("saved_at")
    meta["expires_estimate"] = data.get("expires_estimate")
    meta["valid"] = bool(data.get("cookies"))
    if meta["expires_estimate"]:
        try:
            from datetime import timezone as _tz
            exp = datetime.fromisoformat(meta["expires_estimate"])
            if exp.tzinfo is None:
                exp = exp.replace(tzinfo=_tz.utc)
            if exp <= datetime.now(_tz.utc):
                meta["valid"] = False
        except (ValueError, TypeError):
            pass
    return meta


def _hubnet_today_stats() -> dict:
    """hubnet_pdf_log에서 오늘(localtime) 상태별 카운트 + 평균 duration + 자동토글."""
    stats = {
        "success": 0, "failed": 0, "skipped": 0, "matching_failed": 0,
        "avg_duration_ms": None,
        "auto_pdf_enabled": False,
    }
    # 자동 토글 값
    try:
        with open(BASE_DIR / "settings.json", "r", encoding="utf-8") as f:
            settings = json.load(f)
        stats["auto_pdf_enabled"] = bool(settings.get("hubnet_auto_pdf", False))
    except OSError:
        pass
    # DB 통계
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT status, COUNT(*) FROM hubnet_pdf_log "
                "WHERE date(created_at) = date('now', 'localtime') "
                "GROUP BY status"
            )
            for status, count in cur.fetchall():
                if status in stats:
                    stats[status] = count
            cur.execute(
                "SELECT AVG(duration_ms) FROM hubnet_pdf_log "
                "WHERE date(created_at) = date('now', 'localtime') "
                "  AND duration_ms IS NOT NULL"
            )
            row = cur.fetchone()
            if row and row[0] is not None:
                stats["avg_duration_ms"] = int(row[0])
        finally:
            conn.close()
    except sqlite3.Error:
        pass
    return stats


@app.route("/api/hubnet/status")
def api_hubnet_status():
    try:
        return jsonify({
            "success": True,
            "data": {
                "session": _hubnet_session_meta(),
                "today": _hubnet_today_stats(),
            },
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/login", methods=["POST"])
def api_hubnet_login():
    try:
        with open(BASE_DIR / "settings.json", "r", encoding="utf-8") as f:
            settings = json.load(f)
        email = settings.get("hubnet_email")
        password = settings.get("hubnet_password")
        if not email or not password:
            return jsonify({
                "success": False,
                "error": "settings.json에 hubnet_email/password 없음",
            }), 400
        sess = hubnet_login(email, password)
        session_path = settings.get("hubnet_session_path") or "auth_state_hubnet.json"
        p = Path(session_path)
        if not p.is_absolute():
            p = BASE_DIR / p
        save_hubnet_session(sess, str(p))
        meta = _hubnet_session_meta()
        return jsonify({
            "success": True,
            "data": {
                "saved_at": meta.get("saved_at"),
                "session_path": str(p),
            },
        })
    except RuntimeError as e:
        return jsonify({"success": False, "error": str(e)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/sync", methods=["POST"])
def api_hubnet_sync():
    try:
        body = request.get_json(silent=True) or {}
        start_date = (body.get("start_date") or "").strip()
        end_date = (body.get("end_date") or "").strip()
        date_re = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        if not date_re.match(start_date):
            return jsonify({
                "success": False,
                "error": "start_date 형식 오류 (YYYY-MM-DD 필수)",
            }), 400
        if not date_re.match(end_date):
            return jsonify({
                "success": False,
                "error": "end_date 형식 오류 (YYYY-MM-DD 필수)",
            }), 400
        if start_date > end_date:
            return jsonify({
                "success": False,
                "error": "start_date가 end_date보다 늦음",
            }), 400

        sess = ensure_hubnet_logged_in()
        orders = fetch_hubnet_orders(sess, start_date=start_date, end_date=end_date)
        raw_orders = [o['raw'] for o in orders if isinstance(o, dict) and 'raw' in o]
        upsert_result = upsert_hubnet_orders(raw_orders)
        match_result = match_all_unmatched()
        return jsonify({
            "success": True,
            "data": {
                "fetched": len(orders),
                "upserted": upsert_result.get('total', 0),
                "matched": match_result.get('matched', 0),
            },
        })
    except RuntimeError as e:
        return jsonify({"success": False, "error": str(e)}), 502
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/pdf/download", methods=["POST"])
def api_hubnet_pdf_download():
    try:
        body = request.get_json(silent=True) or {}
        hbl = (body.get("hbl_number") or "").strip()
        order_id = (body.get("order_id") or "").strip() or None
        if not hbl:
            return jsonify({"success": False, "error": "hbl_number 필수"}), 400
        result = download_invoice_pdf(
            hbl, kream_order_id=order_id, triggered_by="manual",
        )
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/pdf/batch", methods=["POST"])
def api_hubnet_pdf_batch():
    try:
        body = request.get_json(silent=True) or {}
        limit = body.get("limit")
        if limit is not None:
            try:
                limit = int(limit)
            except (TypeError, ValueError):
                return jsonify({
                    "success": False,
                    "error": "limit은 정수여야 함",
                }), 400
            if limit <= 0:
                limit = None
        result = download_pending_invoices(limit=limit, triggered_by="manual")
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/pdf/log")
def api_hubnet_pdf_log():
    try:
        try:
            limit = int(request.args.get("limit", "50"))
        except (TypeError, ValueError):
            limit = 50
        if limit < 1:
            limit = 50
        if limit > 500:
            limit = 500  # clamp
        status = (request.args.get("status") or "all").strip().lower()
        valid_statuses = {"all", "success", "failed", "skipped", "matching_failed"}
        if status not in valid_statuses:
            return jsonify({
                "success": False,
                "error": f"status는 {sorted(valid_statuses)} 중 하나",
            }), 400

        conn = sqlite3.connect(str(PRICE_DB))
        try:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            base_cols = (
                "id, hbl_number, kream_order_id, pdf_path, file_size, "
                "status, error_message, duration_ms, triggered_by, created_at"
            )
            if status == "all":
                cur.execute(
                    f"SELECT {base_cols} FROM hubnet_pdf_log "
                    f"ORDER BY id DESC LIMIT ?",
                    (limit,),
                )
                items = [dict(r) for r in cur.fetchall()]
                cur.execute("SELECT COUNT(*) FROM hubnet_pdf_log")
                total = cur.fetchone()[0]
            else:
                cur.execute(
                    f"SELECT {base_cols} FROM hubnet_pdf_log "
                    f"WHERE status = ? ORDER BY id DESC LIMIT ?",
                    (status, limit),
                )
                items = [dict(r) for r in cur.fetchall()]
                cur.execute(
                    "SELECT COUNT(*) FROM hubnet_pdf_log WHERE status = ?",
                    (status,),
                )
                total = cur.fetchone()[0]
        finally:
            conn.close()

        return jsonify({
            "success": True,
            "data": {"items": items, "total": total},
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/hubnet/auto-toggle", methods=["POST"])
def api_hubnet_auto_toggle():
    """settings.json hubnet_auto_pdf 토글. body={"enabled": bool}"""
    try:
        body = request.get_json(silent=True) or {}
        if "enabled" not in body or not isinstance(body["enabled"], bool):
            return jsonify({
                "success": False,
                "error": "enabled (bool) 필수",
            }), 400
        new_val = body["enabled"]
        settings_path = BASE_DIR / "settings.json"
        with open(settings_path, "r", encoding="utf-8") as f:
            settings = json.load(f)
        previous = bool(settings.get("hubnet_auto_pdf", False))
        settings["hubnet_auto_pdf"] = new_val
        # atomic write
        tmp = settings_path.with_suffix(".json.tmp")
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
        tmp.replace(settings_path)
        return jsonify({
            "success": True,
            "data": {"enabled": new_val, "previous": previous},
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/labels/<path:filename>")
def serve_label_pdf(filename):
    """labels 폴더 PDF 정적 서빙. settings.json hubnet_pdf_dir 기준.

    보안: '..' 또는 절대경로 차단 (path traversal 방지).
    Cloudflare Tunnel 호환을 위해 file:// 대신 사용.
    """
    if ".." in filename or filename.startswith("/"):
        return jsonify({"success": False, "error": "잘못된 경로"}), 400
    try:
        with open(BASE_DIR / "settings.json", "r", encoding="utf-8") as f:
            settings = json.load(f)
        labels_dir = settings.get("hubnet_pdf_dir") or str(BASE_DIR / "labels")
    except OSError:
        labels_dir = str(BASE_DIR / "labels")
    full_path = Path(labels_dir) / filename
    if not full_path.exists() or not full_path.is_file():
        return jsonify({"success": False, "error": "PDF not found"}), 404
    return send_from_directory(labels_dir, filename, mimetype="application/pdf")


# ═══════════════════════════════════════════
# ═══════════════════════════════════════════
# Step 12: 사이즈 변환 시스템 API
# ═══════════════════════════════════════════

@app.route("/api/size-charts/import", methods=["POST"])
def api_size_charts_import():
    """엑셀 업로드 → size_charts 임포트."""
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "file 필드 없음"}), 400
    f = request.files["file"]
    tmp_path = BASE_DIR / f"_tmp_size_chart_{int(time.time())}.xlsx"
    try:
        f.save(str(tmp_path))
        from size_converter import import_size_chart_from_xlsx
        result = import_size_chart_from_xlsx(str(tmp_path))
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


@app.route("/api/size-charts/list")
def api_size_charts_list():
    """저장된 사이즈표 목록 (그룹화)."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        rows = conn.execute("""
            SELECT brand, gender, category, COUNT(*) AS size_count,
                   MIN(kream_mm) AS min_mm, MAX(kream_mm) AS max_mm
            FROM size_charts
            GROUP BY brand, gender, category
            ORDER BY brand, gender, category
        """).fetchall()
    return jsonify({
        "ok": True,
        "charts": [
            {"brand": r[0], "gender": r[1], "category": r[2],
             "size_count": r[3], "min_mm": r[4], "max_mm": r[5]}
            for r in rows
        ]
    })


@app.route("/api/size-charts/test", methods=["POST"])
def api_size_charts_test():
    """변환 테스트 (디버깅용). body={brand, gender, category, model, size, model_sizes:[...]}"""
    data = request.get_json() or {}
    from size_converter import convert_to_kream_mm
    mm, used_norm = convert_to_kream_mm(
        brand=data.get("brand", "ADIDAS"),
        gender=data.get("gender", "M"),
        category=data.get("category", "shoes"),
        model=data.get("model", "TEST"),
        size_str=data.get("size", ""),
        model_sizes_set=set(data.get("model_sizes", [])),
        purchase_country=data.get("purchase_country", "ALL"),
        log=False,
    )
    return jsonify({"ok": True, "kream_mm": mm, "size_normalized": used_norm})


# ═══════════════════════════════════════════
# Step 15: 識货 시장가 임포트 API
# ═══════════════════════════════════════════

def normalize_brand(brand_raw, supplier_raw):
    """品牌名称/供应商 → 정규화 브랜드.

    1. 品牌名称에 'adidas' 포함 (대소문자 무관) → 'ADIDAS'
    2. 品牌名称가 명확한 ADIDAS 표기 ('三叶草' 단독, 'ADIDAS' 단독) → 'ADIDAS'
    3. 品牌名称가 '无品牌' 또는 비어있음 → 供应商名称에서 추정
       - supplier에 'adidas' 포함 → 'ADIDAS'
       - supplier에 '三叶草' + '官方旗舰店' → 'ADIDAS'
    4. 그 외 → 'unknown'
    """
    b = (brand_raw or "").strip()
    s = (supplier_raw or "").strip()
    b_lower = b.lower()

    if "adidas" in b_lower:
        return "ADIDAS"
    if b in ("三叶草", "ADIDAS", "Adidas", "adidas"):
        return "ADIDAS"

    if not b or b == "无品牌":
        s_lower = s.lower()
        if "adidas" in s_lower:
            return "ADIDAS"
        if "三叶草" in s and "官方旗舰店" in s:
            return "ADIDAS"
        return "unknown"

    return "unknown"


SHIHUO_REQUIRED_HEADERS = [
    "产品编号", "产品尺寸欧码", "产品价格", "创建时间",
    "品牌名称", "产品分类", "供应商名称", "平台名称",
]


def _shihuo_category_to_internal(cat_raw):
    """产品分类 → size_charts.category (shoes/bags/...)."""
    c = (cat_raw or "").strip()
    if "鞋" in c:
        return "shoes"
    if "包" in c:
        return "bags"
    if "凉" in c or "拖" in c:
        return "sandals"
    return "shoes"


def _shihuo_is_no_size_category(cat_raw):
    """가방/액세서리 등 사이즈 차원이 없는 카테고리 판별."""
    c = (cat_raw or "").strip()
    return "包" in c


@app.route("/api/shihuo/import", methods=["POST"])
def api_shihuo_import():
    """識货 엑셀 업로드 → shihuo_prices 임포트.

    응답: {ok, batch_id, total_rows, mapped, no_size, mapping_failed,
           unknown_brand, models_count, errors[:10]}
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "file 필드 없음"}), 400

    f = request.files["file"]
    tmp_path = BASE_DIR / f"_tmp_shihuo_{int(time.time())}.xlsx"
    try:
        f.save(str(tmp_path))
        from openpyxl import load_workbook
        wb = load_workbook(str(tmp_path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"ok": False, "error": "데이터 없음"}), 400

        header = [str(c).strip() if c else "" for c in rows[0]]
        missing = [h for h in SHIHUO_REQUIRED_HEADERS if h not in header]
        if missing:
            return jsonify({
                "ok": False,
                "error": f"필수 헤더 누락: {missing}",
                "header_found": header,
            }), 400

        col = {h: header.index(h) for h in SHIHUO_REQUIRED_HEADERS}
        col_color = header.index("产品颜色") if "产品颜色" in header else None

        # 1) (model, size_eu) 그룹별 MIN(cny_price) 추출
        grouped = {}
        for row in rows[1:]:
            if row is None or all(v is None for v in row):
                continue
            model_v = row[col["产品编号"]]
            if model_v is None or str(model_v).strip() == "":
                continue
            try:
                price_v = row[col["产品价格"]]
                if price_v is None or str(price_v).strip() == "":
                    continue
                price = float(str(price_v).strip())
            except (ValueError, TypeError):
                continue

            model = str(model_v).strip()
            size_v = row[col["产品尺寸欧码"]]
            size_eu = "" if size_v is None else str(size_v).strip()
            key = (model, size_eu)

            entry = {
                "model": model,
                "size_eu": size_eu,
                "cny_price": price,
                "brand_raw": str(row[col["品牌名称"]]).strip() if row[col["品牌名称"]] is not None else "",
                "category_raw": str(row[col["产品分类"]]).strip() if row[col["产品分类"]] is not None else "",
                "supplier": str(row[col["供应商名称"]]).strip() if row[col["供应商名称"]] is not None else "",
                "platform": str(row[col["平台名称"]]).strip() if row[col["平台名称"]] is not None else "",
                "source_created_at": row[col["创建时间"]],
                "color": (str(row[col_color]).strip() if col_color is not None and row[col_color] is not None else ""),
            }

            if key not in grouped or entry["cny_price"] < grouped[key]["cny_price"]:
                grouped[key] = entry

        # 2) 모델별 사이즈 집합 (분수 매핑 시 모델 내 정수 EU 존재 여부 판단)
        model_sizes_map = {}
        for (model, size_eu) in grouped.keys():
            model_sizes_map.setdefault(model, set()).add(size_eu)

        # 3) batch_id 생성
        batch_id = f"shihuo_{datetime.now():%Y%m%d_%H%M%S}"

        # 4) 정규화/매핑/INSERT — 단일 트랜잭션
        from size_converter import convert_to_kream_mm
        inserted = 0
        mapped = 0
        no_size = 0
        mapping_failed = 0
        unknown_brand = 0
        errors = []

        conn = sqlite3.connect(str(PRICE_DB))
        try:
            conn.execute("BEGIN")
            cur = conn.cursor()

            for key, entry in grouped.items():
                try:
                    brand_norm = normalize_brand(entry["brand_raw"], entry["supplier"])
                    category = _shihuo_category_to_internal(entry["category_raw"])
                    is_no_size = _shihuo_is_no_size_category(entry["category_raw"]) or entry["size_eu"] == ""

                    size_normalized = None
                    kream_mm = None
                    mapping_status = None
                    mapping_note = None

                    if is_no_size:
                        mapping_status = "no_size"
                        mapping_note = "가방/사이즈없음"
                        no_size += 1
                    elif brand_norm == "unknown":
                        mapping_status = "unknown_brand"
                        mapping_note = f"brand_raw={entry['brand_raw']!r}, supplier={entry['supplier']!r}"
                        unknown_brand += 1
                    else:
                        from size_converter import normalize_size as _ns
                        ns_result, ns_rule = _ns(entry["size_eu"], model_sizes_map.get(entry["model"], set()))
                        if ns_result is None:
                            # 분수 정규화 자체 실패 (excluded_int_exists / parse_failed 등)
                            mapping_status = "mapping_failed"
                            mapping_note = f"size={entry['size_eu']!r}, rule={ns_rule}"
                            mapping_failed += 1
                        else:
                            kream_mm, used_norm = convert_to_kream_mm(
                                brand=brand_norm,
                                gender="M",
                                category=category,
                                model=entry["model"],
                                size_str=entry["size_eu"],
                                model_sizes_set=model_sizes_map.get(entry["model"], set()),
                                purchase_country="ALL",
                                log=False,
                            )
                            # Step 15d: 폴백 발생 시 used_norm이 정수로 바뀜
                            size_normalized = used_norm if used_norm is not None else ns_result
                            if kream_mm is None:
                                # size_charts에 해당 EU가 없음
                                mapping_status = "no_size_chart"
                                mapping_note = f"size={entry['size_eu']!r}→{ns_result}, brand={brand_norm}"
                                mapping_failed += 1
                            else:
                                mapping_status = "mapped"
                                mapped += 1

                    src_dt = entry["source_created_at"]
                    src_str = None
                    if src_dt is not None:
                        try:
                            src_str = src_dt.strftime("%Y-%m-%d %H:%M:%S") if hasattr(src_dt, "strftime") else str(src_dt)
                        except Exception:
                            src_str = str(src_dt)

                    cur.execute("""
                        INSERT OR REPLACE INTO shihuo_prices
                        (batch_id, active, brand_raw, brand_normalized, category,
                         model, color, size_eu, size_normalized, kream_mm,
                         cny_price, supplier, platform, source_created_at,
                         mapping_status, mapping_note)
                        VALUES (?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        batch_id, entry["brand_raw"], brand_norm, category,
                        entry["model"], entry["color"], entry["size_eu"],
                        size_normalized, kream_mm,
                        entry["cny_price"], entry["supplier"], entry["platform"],
                        src_str, mapping_status, mapping_note,
                    ))
                    inserted += 1
                except Exception as e:
                    errors.append(f"{key}: {e}")

            # 5) 모든 INSERT 성공 후에만 옛날 batch active=0
            cur.execute("UPDATE shihuo_prices SET active=0 WHERE batch_id != ?", (batch_id,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            return jsonify({"ok": False, "error": f"트랜잭션 실패: {e}", "errors": errors[:10]}), 500
        finally:
            conn.close()

        return jsonify({
            "ok": True,
            "batch_id": batch_id,
            "total_rows": inserted,
            "mapped": mapped,
            "no_size": no_size,
            "mapping_failed": mapping_failed,
            "unknown_brand": unknown_brand,
            "models_count": len(model_sizes_map),
            "errors": errors[:10],
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass


@app.route("/api/shihuo/latest")
def api_shihuo_latest():
    """현재 활성 batch 요약."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        row = conn.execute("""
            SELECT batch_id, MIN(imported_at), COUNT(*)
            FROM shihuo_prices WHERE active=1
            GROUP BY batch_id LIMIT 1
        """).fetchone()
        if not row:
            return jsonify({"ok": True, "batch_id": None, "total_count": 0})
        batch_id, imported_at, total = row
        by_status = dict(conn.execute("""
            SELECT mapping_status, COUNT(*) FROM shihuo_prices
            WHERE active=1 GROUP BY mapping_status
        """).fetchall())
        by_model = dict(conn.execute("""
            SELECT model, COUNT(*) FROM shihuo_prices
            WHERE active=1 GROUP BY model
        """).fetchall())
    return jsonify({
        "ok": True,
        "batch_id": batch_id,
        "imported_at": imported_at,
        "total_count": total,
        "by_status": by_status,
        "by_model": by_model,
    })


@app.route("/api/shihuo/by-model/<path:model>")
def api_shihuo_by_model(model):
    """특정 모델의 활성 시장가 (사이즈별)."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        rows = conn.execute("""
            SELECT size_eu, size_normalized, kream_mm, cny_price,
                   supplier, platform, mapping_status, mapping_note
            FROM shihuo_prices
            WHERE active=1 AND model=?
            ORDER BY CASE WHEN kream_mm IS NULL THEN 1 ELSE 0 END, kream_mm
        """, (model,)).fetchall()
    return jsonify({
        "ok": True,
        "model": model,
        "items": [
            {
                "size_eu": r[0], "size_normalized": r[1], "kream_mm": r[2],
                "cny_price": r[3], "supplier": r[4], "platform": r[5],
                "mapping_status": r[6], "mapping_note": r[7],
            } for r in rows
        ]
    })


@app.route("/api/shihuo/unmapped")
def api_shihuo_unmapped():
    """매핑 실패/미지정 건 (사람 검토용)."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        rows = conn.execute("""
            SELECT model, size_eu, brand_raw, brand_normalized,
                   mapping_status, mapping_note, cny_price, supplier
            FROM shihuo_prices
            WHERE active=1 AND mapping_status IN ('no_size_chart','mapping_failed','unknown_brand')
            ORDER BY mapping_status, model, size_eu
        """).fetchall()
    return jsonify({
        "ok": True,
        "count": len(rows),
        "items": [
            {
                "model": r[0], "size_eu": r[1], "brand_raw": r[2],
                "brand_normalized": r[3], "mapping_status": r[4],
                "mapping_note": r[5], "cny_price": r[6], "supplier": r[7],
            } for r in rows
        ]
    })


@app.route("/api/shihuo/activate/<batch_id>", methods=["POST"])
def api_shihuo_activate(batch_id):
    """지정 batch_id를 active=1로, 그 외 active=0으로 전환."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        existing = conn.execute(
            "SELECT COUNT(*) FROM shihuo_prices WHERE batch_id=?", (batch_id,)
        ).fetchone()[0]
        if existing == 0:
            return jsonify({"ok": False, "error": f"batch_id {batch_id} 존재하지 않음"}), 404
        conn.execute("UPDATE shihuo_prices SET active=0 WHERE batch_id != ?", (batch_id,))
        conn.execute("UPDATE shihuo_prices SET active=1 WHERE batch_id = ?", (batch_id,))
        conn.commit()
    return jsonify({"ok": True, "batch_id": batch_id, "activated": existing})


@app.route("/api/shihuo/deactivate", methods=["POST"])
def api_shihuo_deactivate():
    """현재 active 배치를 모두 끔 — 진짜 비활성화."""
    with sqlite3.connect(str(PRICE_DB)) as conn:
        cur = conn.execute("UPDATE shihuo_prices SET active=0 WHERE active=1")
        conn.commit()
        cnt = cur.rowcount
    return jsonify({"ok": True, "deactivated": cnt})


# 백워드 호환 — 한 분기 유지 후 v17에서 제거 예정
@app.route("/api/shihuo/rollback/<batch_id>", methods=["POST"])
def api_shihuo_rollback(batch_id):
    """[DEPRECATED] /api/shihuo/activate/<batch_id> 사용 권장."""
    return api_shihuo_activate(batch_id)


# ═══════════════════════════════════════════
# In-App 도움말 API
# ═══════════════════════════════════════════

@app.route('/api/help/<tab_id>', methods=['GET'])
def api_help(tab_id):
    """탭별 In-App 도움말 콘텐츠 반환."""
    try:
        from pathlib import Path
        help_path = Path(__file__).parent / 'help_content.json'
        if not help_path.exists():
            return jsonify({'ok': False, 'error': 'help_content.json 없음'}), 404
        data = json.loads(help_path.read_text(encoding='utf-8'))
        if tab_id not in data:
            return jsonify({'ok': False, 'error': f'tab_id={tab_id} 없음'}), 404
        return jsonify({'ok': True, 'help': data[tab_id]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 19: 밀린 입찰 진단 + 회수 전략 도구
# ═══════════════════════════════════════════

@app.route('/api/cleanup/diagnose', methods=['GET'])
def api_cleanup_diagnose():
    """rank 밀린 입찰 + 원가 + 마진 분석 → 회수 전략 추천."""
    try:
        from collections import defaultdict

        local_path = Path(__file__).parent / 'my_bids_local.json'
        if not local_path.exists():
            return jsonify({'ok': True, 'total': 0, 'items': [], 'stats': {},
                            'by_model': {}, 'note': 'local cache 없음'})

        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids = local.get('bids', []) if isinstance(local, dict) else []

        # 설정값
        try:
            settings = json.loads(Path(__file__).parent.joinpath('settings.json').read_text(encoding='utf-8'))
        except Exception:
            settings = {}
        fee_rate = settings.get('commission_rate', 6) / 100
        fixed_fee = 2500
        min_margin = settings.get('min_margin', 4000)
        undercut = settings.get('undercut_amount', 1000)
        overseas_ship_default = settings.get('overseas_shipping', 8000)

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        # 판매 완료된 order_id (절대 건드리지 않음)
        try:
            c.execute("SELECT DISTINCT order_id FROM sales_history WHERE order_id IS NOT NULL")
            sold_ids = {row[0] for row in c.fetchall()}
        except Exception:
            sold_ids = set()

        # 같은 model+size 입찰 카운트 (마지막 재고 안전장치용)
        by_size_count = defaultdict(int)
        for b in bids:
            by_size_count[(b.get('model'), b.get('size'))] += 1

        items = []
        for b in bids:
            rank = b.get('rank')
            if not rank or rank == 1:
                continue

            order_id = b.get('orderId')
            if order_id in sold_ids:
                continue

            model = b.get('model', '-')
            size = b.get('size', '-')
            price = b.get('price') or 0

            cny_price = exchange_rate = overseas_ship = None
            try:
                c.execute("""
                    SELECT cny_price, exchange_rate, overseas_shipping, other_costs
                    FROM bid_cost WHERE order_id = ?
                """, (order_id,))
                row = c.fetchone()
                if row:
                    cny_price, exchange_rate, overseas_ship, _other_costs = row
            except Exception:
                pass

            cost = None
            margin = None
            margin_status = 'no_cost'

            if cny_price is not None and exchange_rate is not None:
                try:
                    ship = overseas_ship if overseas_ship is not None else overseas_ship_default
                    cost = round(float(cny_price) * float(exchange_rate) * 1.03 + float(ship))
                    settlement = price * (1 - fee_rate * 1.1) - fixed_fee
                    margin = round(settlement - cost)

                    if margin >= min_margin:
                        margin_status = 'ok'
                    elif margin >= 0:
                        margin_status = 'low'
                    else:
                        margin_status = 'deficit'
                except Exception:
                    pass

            recommendation = 'hold'
            recommendation_reason = ''

            if margin_status == 'no_cost':
                recommendation = 'need_cost'
                recommendation_reason = '원가 미등록 → CNY 입력 후 재진단'
            elif margin_status == 'deficit':
                recommendation = 'withdraw'
                recommendation_reason = f'적자 ({margin:,}원) → 회수 권장'
            elif margin_status == 'low':
                hypothetical_price = price - undercut
                hyp_settlement = hypothetical_price * (1 - fee_rate * 1.1) - fixed_fee
                hyp_margin = hyp_settlement - cost
                if hyp_margin >= min_margin:
                    recommendation = 'adjust'
                    recommendation_reason = f'-{undercut}원 조정 시 마진 {round(hyp_margin):,}원 (충분)'
                else:
                    recommendation = 'withdraw'
                    recommendation_reason = f'조정해도 마진 {round(hyp_margin):,}원 (미달) → 회수 권장'
            elif margin_status == 'ok':
                recommendation = 'adjust'
                recommendation_reason = f'마진 {margin:,}원 → 가격수집 복원 후 자동조정'

            same_size_total = by_size_count[(model, size)]
            is_last_in_size = (same_size_total <= 1)
            if recommendation == 'withdraw' and is_last_in_size:
                recommendation = 'withdraw_blocked'
                recommendation_reason += ' (단 마지막 재고라 안전장치 발동, 강제 회수만 가능)'

            items.append({
                'orderId': order_id,
                'model': model,
                'size': size,
                'price': price,
                'rank': rank,
                'cny_price': cny_price,
                'exchange_rate': exchange_rate,
                'cost': cost,
                'margin': margin,
                'margin_status': margin_status,
                'recommendation': recommendation,
                'reason': recommendation_reason,
                'is_last_in_size': is_last_in_size,
            })

        conn.close()

        stats = defaultdict(int)
        for it in items:
            stats[it['recommendation']] += 1
            stats[f"margin_{it['margin_status']}"] += 1

        by_model = defaultdict(list)
        for it in items:
            by_model[it['model']].append(it)

        return jsonify({
            'ok': True,
            'total': len(items),
            'stats': dict(stats),
            'items': items,
            'by_model': dict(by_model),
            'settings_used': {
                'min_margin': min_margin,
                'undercut': undercut,
                'fee_rate': fee_rate,
                'overseas_ship_default': overseas_ship_default,
            }
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/cleanup/bulk-withdraw', methods=['POST'])
def api_cleanup_bulk_withdraw():
    """선택한 order_id 일괄 회수. force 미지정 시 마지막 재고 안전장치 발동."""
    data = request.get_json() or {}
    order_ids = data.get('orderIds', [])
    force = data.get('force', False)

    if not order_ids:
        return jsonify({'ok': False, 'error': 'orderIds required'}), 400

    try:
        diag_resp = api_cleanup_diagnose()
        if hasattr(diag_resp, 'get_json'):
            diag = diag_resp.get_json()
        elif hasattr(diag_resp, 'data'):
            diag = json.loads(diag_resp.data)
        else:
            diag = diag_resp

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        try:
            c.execute("SELECT DISTINCT order_id FROM sales_history WHERE order_id IS NOT NULL")
            sold_ids = {row[0] for row in c.fetchall()}
        except Exception:
            sold_ids = set()
        conn.close()

        items_map = {it['orderId']: it for it in (diag.get('items') or [])}
        approved_ids = []
        blocked = []

        for oid in order_ids:
            if oid in sold_ids:
                blocked.append({'orderId': oid, 'reason': '판매 완료 건 (보호)'})
                continue

            it = items_map.get(oid)
            if not it:
                blocked.append({'orderId': oid, 'reason': '진단 대상 아님 (1위거나 알 수 없음)'})
                continue

            if it.get('is_last_in_size') and not force:
                blocked.append({'orderId': oid, 'reason': '같은 사이즈 마지막 재고 (force=true 필요)'})
                continue

            approved_ids.append(oid)

        delete_result = None
        if approved_ids:
            try:
                import requests as rq
                r = rq.post('http://localhost:5001/api/my-bids/delete',
                            json={'orderIds': approved_ids}, timeout=10)
                try:
                    delete_result = r.json()
                except Exception:
                    delete_result = {'status_code': r.status_code, 'text': r.text[:300]}
            except Exception as e:
                delete_result = {'error': str(e)}

        return jsonify({
            'ok': True,
            'requested': len(order_ids),
            'approved': len(approved_ids),
            'blocked': blocked,
            'delete_task': delete_result,
            'note': 'task 완료 후 5분 대기 → /api/my-bids/verify-deleted로 확인 권장'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/cleanup/bulk-adjust', methods=['POST'])
def api_cleanup_bulk_adjust():
    """선택한 order_id 일괄 가격 -N원 조정. 마진 사전 체크."""
    data = request.get_json() or {}
    order_ids = data.get('orderIds', [])
    decrement = data.get('decrement', 1000)

    if not order_ids:
        return jsonify({'ok': False, 'error': 'orderIds required'}), 400

    try:
        diag_resp = api_cleanup_diagnose()
        if hasattr(diag_resp, 'get_json'):
            diag = diag_resp.get_json()
        elif hasattr(diag_resp, 'data'):
            diag = json.loads(diag_resp.data)
        else:
            diag = diag_resp

        items_map = {it['orderId']: it for it in (diag.get('items') or [])}

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        try:
            c.execute("SELECT DISTINCT order_id FROM sales_history WHERE order_id IS NOT NULL")
            sold_ids = {row[0] for row in c.fetchall()}
        except Exception:
            sold_ids = set()
        conn.close()

        try:
            settings = json.loads(Path(__file__).parent.joinpath('settings.json').read_text(encoding='utf-8'))
        except Exception:
            settings = {}
        fee_rate = settings.get('commission_rate', 6) / 100
        fixed_fee = 2500
        min_margin = settings.get('min_margin', 4000)

        approved = []
        blocked = []

        for oid in order_ids:
            if oid in sold_ids:
                blocked.append({'orderId': oid, 'reason': '판매 완료 (보호)'})
                continue

            it = items_map.get(oid)
            if not it:
                blocked.append({'orderId': oid, 'reason': '진단 결과에 없음'})
                continue

            cost = it.get('cost')
            if cost is None:
                blocked.append({'orderId': oid, 'reason': '원가 미등록 → 조정 불가'})
                continue

            new_price = it['price'] - decrement
            new_price = math.ceil(new_price / 1000) * 1000

            settlement = new_price * (1 - fee_rate * 1.1) - fixed_fee
            new_margin = settlement - cost

            if new_margin < min_margin:
                blocked.append({
                    'orderId': oid,
                    'reason': f'조정 후 마진 {round(new_margin):,}원 < {min_margin:,} (미달)'
                })
                continue

            approved.append({
                'orderId': oid,
                'old_price': it['price'],
                'new_price': new_price,
                'expected_margin': round(new_margin)
            })

        modify_results = []
        if approved:
            try:
                import requests as rq
                for app_item in approved:
                    try:
                        r = rq.post('http://localhost:5001/api/my-bids/modify',
                                    json={'orderId': app_item['orderId'],
                                          'newPrice': app_item['new_price']},
                                    timeout=10)
                        try:
                            resp_json = r.json()
                        except Exception:
                            resp_json = None
                        modify_results.append({
                            'orderId': app_item['orderId'],
                            'status': r.status_code,
                            'response': resp_json
                        })
                    except Exception as e:
                        modify_results.append({'orderId': app_item['orderId'], 'error': str(e)})
            except Exception:
                pass

        return jsonify({
            'ok': True,
            'requested': len(order_ids),
            'approved': approved,
            'blocked': blocked,
            'modify_results': modify_results,
            'note': '가격 수정은 5분 후 sync에 반영됨'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 22: 구매대행 모델 반영 + 정리 자동화 + 체결률 KPI
# ═══════════════════════════════════════════

# ── Step 23: 진단 라우트 ──
@app.route('/api/diagnostics/sync-page-dump', methods=['POST'])
def api_sync_page_dump():
    """판매자센터 입찰 페이지를 직접 열어서 HTML + 스크린샷 저장.
    sync가 0건 반환할 때 페이지 상태를 사장이 직접 확인."""
    try:
        import asyncio
        from pathlib import Path as _Path
        from datetime import datetime as _dt

        ts = _dt.now().strftime('%Y%m%d_%H%M%S')
        dump_dir = _Path(__file__).parent / 'diagnostics'
        dump_dir.mkdir(exist_ok=True)

        html_path = dump_dir / f'sync_page_{ts}.html'
        png_path = dump_dir / f'sync_page_{ts}.png'

        async def dump():
            from playwright.async_api import async_playwright
            from kream_bot import create_browser, create_context, ensure_logged_in, dismiss_popups

            async with async_playwright() as p:
                browser = await create_browser(p, headless=True)
                context = await create_context(browser, storage='auth_state.json')
                page = await context.new_page()

                await page.goto('https://partner.kream.co.kr/c2c/sell/bid', wait_until='domcontentloaded', timeout=30000)
                await page.wait_for_timeout(3000)

                logged_in = await ensure_logged_in(page, context)

                try:
                    await dismiss_popups(page)
                except Exception:
                    pass

                await page.wait_for_timeout(2000)

                html = await page.content()
                html_path.write_text(html, encoding='utf-8')

                await page.screenshot(path=str(png_path), full_page=True)

                count_info = {}
                for selector_desc, selector in [
                    ('table_rows', 'table tbody tr'),
                    ('list_items', '.bid-item, .list-item, [class*="bid"]'),
                    ('total_text', '[class*="total"], [class*="count"]'),
                ]:
                    try:
                        elements = await page.query_selector_all(selector)
                        count_info[selector_desc] = len(elements)
                    except Exception:
                        count_info[selector_desc] = -1

                final_url = page.url
                title = await page.title()

                await browser.close()

                return {
                    'logged_in': bool(logged_in),
                    'final_url': final_url,
                    'title': title,
                    'count_info': count_info,
                    'html_size': len(html),
                }

        result = asyncio.run(dump())

        return jsonify({
            'ok': True,
            'timestamp': ts,
            'html_path': str(html_path),
            'screenshot_path': str(png_path),
            **result,
            'note': '스크린샷 + HTML 저장됨. 직접 열어서 입찰 보이는지 확인'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/diagnostics/explore-menu', methods=['POST'])
def api_explore_menu():
    """판매자센터 메인에서 입찰 관련 링크 자동 탐색."""
    try:
        import asyncio

        async def explore():
            from playwright.async_api import async_playwright
            from kream_bot import create_browser, create_context, ensure_logged_in, dismiss_popups

            async with async_playwright() as p:
                browser = await create_browser(p, headless=True)
                context = await create_context(browser, storage='auth_state.json')
                page = await context.new_page()

                await page.goto('https://partner.kream.co.kr/c2c', wait_until='domcontentloaded', timeout=30000)
                await page.wait_for_timeout(3000)

                logged_in = await ensure_logged_in(page, context)
                try:
                    await dismiss_popups(page)
                except Exception:
                    pass

                await page.wait_for_timeout(2000)

                links = await page.evaluate("""
                    () => {
                        const allLinks = Array.from(document.querySelectorAll('a, [role="link"], button'));
                        return allLinks.map(el => ({
                            text: (el.textContent || '').trim().slice(0, 50),
                            href: el.href || el.getAttribute('data-href') || '',
                            classes: el.className || ''
                        })).filter(l => l.text.length > 0).slice(0, 100);
                    }
                """)

                bid_keywords = ['입찰', '판매', 'bid', 'sell', '내 입찰', '입찰 관리', '판매 관리', 'C2C', 'P2P']
                bid_links = []
                for link in links:
                    text_lower = link['text'].lower()
                    href_lower = link['href'].lower()
                    if any(kw.lower() in text_lower for kw in bid_keywords) or \
                       any(kw in href_lower for kw in ['bid', 'sell', 'c2c']):
                        bid_links.append(link)

                page_info = {
                    'url': page.url,
                    'title': await page.title(),
                    'logged_in': logged_in,
                }

                from datetime import datetime as _dt2
                from pathlib import Path as _Path2
                ts = _dt2.now().strftime('%Y%m%d_%H%M%S')
                dump_dir = _Path2(__file__).parent / 'diagnostics'
                dump_dir.mkdir(exist_ok=True)
                screenshot_path = dump_dir / f'menu_explore_{ts}.png'
                await page.screenshot(path=str(screenshot_path), full_page=True)

                await browser.close()

                return {
                    'page_info': page_info,
                    'all_links_count': len(links),
                    'bid_related_links': bid_links,
                    'all_links_sample': links[:30],
                    'screenshot': str(screenshot_path)
                }

        result = asyncio.run(explore())
        return jsonify({'ok': True, **result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/diagnostics/list-dumps', methods=['GET'])
def api_diagnostics_list_dumps():
    """저장된 진단 덤프 목록."""
    try:
        from pathlib import Path as _Path
        dump_dir = _Path(__file__).parent / 'diagnostics'
        if not dump_dir.exists():
            return jsonify({'ok': True, 'dumps': []})

        dumps = []
        for f in sorted(dump_dir.glob('sync_page_*.png'), reverse=True)[:20]:
            html_f = f.with_suffix('.html')
            dumps.append({
                'timestamp': f.stem.replace('sync_page_', ''),
                'screenshot': f.name,
                'html': html_f.name if html_f.exists() else None,
                'size_mb': round(f.stat().st_size / 1024 / 1024, 2),
            })

        return jsonify({'ok': True, 'dumps': dumps})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/diagnostics/<path:filename>', methods=['GET'])
def serve_diagnostics(filename):
    """진단 파일 (스크린샷, HTML) 직접 접근."""
    from pathlib import Path as _Path
    from flask import send_from_directory
    dump_dir = _Path(__file__).parent / 'diagnostics'
    return send_from_directory(str(dump_dir), filename)


@app.route('/api/real-margin', methods=['GET'])
def api_real_margin():
    """판매 체결된 건의 실제 bid_cost join → 진짜 마진 (추정 제거)."""
    try:
        from datetime import datetime, timedelta
        days = request.args.get('days', 30, type=int)
        since = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        try:
            settings = json.loads((BASE_DIR / 'settings.json').read_text(encoding='utf-8'))
        except Exception:
            settings = {}
        fee_rate = settings.get('commission_rate', 6) / 100
        fixed_fee = 2500
        overseas_ship_default = settings.get('overseas_shipping', 8000)

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        c.execute("""
            SELECT s.order_id, s.model, s.size, s.sale_price, s.trade_date,
                   b.cny_price, b.exchange_rate, b.overseas_shipping
            FROM sales_history s
            LEFT JOIN bid_cost b ON s.order_id = b.order_id
            WHERE DATE(s.trade_date) >= ?
            ORDER BY s.trade_date DESC
        """, (since,))
        rows = c.fetchall()

        items = []
        confirmed_revenue = 0
        confirmed_cost = 0
        confirmed_margin = 0
        confirmed_count = 0
        unknown_cost_count = 0
        unknown_revenue = 0

        for r in rows:
            order_id, model, size, sale_price, trade_date, cny, fx, ship = r
            sale_price = sale_price or 0

            if cny is not None and fx is not None:
                ship = ship if ship is not None else overseas_ship_default
                cost = float(cny) * float(fx) * 1.03 + float(ship)
                settlement = sale_price * (1 - fee_rate * 1.1) - fixed_fee
                margin = settlement - cost
                items.append({
                    'order_id': order_id,
                    'model': model,
                    'size': size,
                    'sale_price': sale_price,
                    'trade_date': trade_date,
                    'cost': round(cost),
                    'margin': round(margin),
                    'confirmed': True
                })
                confirmed_revenue += sale_price
                confirmed_cost += cost
                confirmed_margin += margin
                confirmed_count += 1
            else:
                items.append({
                    'order_id': order_id,
                    'model': model,
                    'size': size,
                    'sale_price': sale_price,
                    'trade_date': trade_date,
                    'cost': None,
                    'margin': None,
                    'confirmed': False
                })
                unknown_cost_count += 1
                unknown_revenue += sale_price

        # Step 23: fuzzy 매칭 보강 (model+size → model)
        fuzzy_matched_count = 0
        estimated_revenue = 0.0
        estimated_cost = 0.0
        estimated_margin = 0.0

        for item in items:
            if item.get('confirmed') or item.get('cost') is not None:
                continue
            sp = item.get('sale_price') or 0
            mdl = item.get('model')
            sz = item.get('size')
            if not mdl:
                continue
            est_cost = None
            est_source = None
            try:
                c.execute(
                    """
                    SELECT AVG(cny_price), AVG(exchange_rate), AVG(COALESCE(overseas_shipping, ?))
                    FROM bid_cost
                    WHERE model = ? AND (size = ? OR size = ?)
                    """,
                    (overseas_ship_default, mdl, sz, 'ONE SIZE' if not sz else sz)
                )
                row = c.fetchone()
                if row and row[0] is not None:
                    cny_a, fx_a, ship_a = row
                    est_cost = float(cny_a) * float(fx_a) * 1.03 + float(ship_a)
                    est_source = 'fuzzy_model_size'
                else:
                    c.execute(
                        """
                        SELECT AVG(cny_price), AVG(exchange_rate), AVG(COALESCE(overseas_shipping, ?))
                        FROM bid_cost
                        WHERE model = ?
                        """,
                        (overseas_ship_default, mdl)
                    )
                    row2 = c.fetchone()
                    if row2 and row2[0] is not None:
                        cny_a, fx_a, ship_a = row2
                        est_cost = float(cny_a) * float(fx_a) * 1.03 + float(ship_a)
                        est_source = 'fuzzy_model_only'
            except Exception:
                est_cost = None

            if est_cost is not None:
                settlement = sp * (1 - fee_rate * 1.1) - fixed_fee
                est_margin_v = settlement - est_cost
                item['cost'] = round(est_cost)
                item['margin'] = round(est_margin_v)
                item['estimation_source'] = est_source
                fuzzy_matched_count += 1
                estimated_revenue += sp
                estimated_cost += est_cost
                estimated_margin += est_margin_v
                # unknown_cost 카운트에서 차감 (estimated로 분류)
                unknown_cost_count = max(0, unknown_cost_count - 1)
                unknown_revenue = max(0.0, unknown_revenue - sp)

        conn.close()

        return jsonify({
            'ok': True,
            'period_days': days,
            'total_sales': len(items),
            'confirmed': {
                'count': confirmed_count,
                'revenue': round(confirmed_revenue),
                'cost': round(confirmed_cost),
                'margin': round(confirmed_margin),
                'avg_margin': round(confirmed_margin / confirmed_count) if confirmed_count else 0,
                'margin_rate_pct': round((confirmed_margin / confirmed_revenue * 100) if confirmed_revenue else 0, 1),
            },
            'estimated': {
                'count': fuzzy_matched_count,
                'revenue': round(estimated_revenue),
                'cost': round(estimated_cost),
                'margin': round(estimated_margin),
                'avg_margin': round(estimated_margin / fuzzy_matched_count) if fuzzy_matched_count else 0,
                'margin_rate_pct': round((estimated_margin / estimated_revenue * 100) if estimated_revenue else 0, 1),
                'note': 'bid_cost 평균치 기반 추정 (model+size 또는 model). estimation_source 키 참조'
            },
            'unknown_cost': {
                'count': unknown_cost_count,
                'revenue': round(unknown_revenue),
                'note': 'bid_cost 데이터 없어서 마진 계산 불가'
            },
            'items': items[:50],
            'note': '확정값(bid_cost order_id 매칭) + 추정값(model+size/model 평균) 분리'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/conversion-rate', methods=['GET'])
def api_conversion_rate():
    """체결률: 활성 입찰 → 판매로 가는 비율 추이 (구매대행 모델 핵심 KPI)."""
    try:
        from datetime import datetime, timedelta
        days = request.args.get('days', 30, type=int)
        since_date = datetime.now() - timedelta(days=days)
        since = since_date.strftime('%Y-%m-%d')

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        c.execute("""
            SELECT DATE(trade_date) as d, COUNT(*) as cnt
            FROM sales_history
            WHERE DATE(trade_date) >= ?
            GROUP BY DATE(trade_date)
            ORDER BY d
        """, (since,))
        daily_sales = {r[0]: r[1] for r in c.fetchall()}

        local_path = BASE_DIR / 'my_bids_local.json'
        active_bids = 0
        if local_path.exists():
            try:
                local = json.loads(local_path.read_text(encoding='utf-8'))
                active_bids = len(local.get('bids', []) if isinstance(local, dict) else [])
            except Exception:
                pass

        total_sales = sum(daily_sales.values())
        total_pool = active_bids + total_sales
        conversion_pct = (total_sales / total_pool * 100) if total_pool else 0

        items = []
        cur = since_date
        end = datetime.now()
        while cur <= end:
            d = cur.strftime('%Y-%m-%d')
            sales = daily_sales.get(d, 0)
            items.append({'date': d, 'sales': sales})
            cur += timedelta(days=1)

        conn.close()

        return jsonify({
            'ok': True,
            'period_days': days,
            'active_bids_now': active_bids,
            'total_sales_period': total_sales,
            'conversion_pct': round(conversion_pct, 1),
            'daily': items,
            'note': '체결률 = 기간 내 체결 / (현재 활성 + 기간 내 체결). 구매대행 모델의 핵심 KPI.'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/cleanup/auto-execute', methods=['POST'])
def api_cleanup_auto_execute():
    """회수 권장 자동 정리. 안전장치는 bulk-withdraw 그대로 사용 (force=False)."""
    data = request.get_json() or {}
    dry_run = data.get('dry_run', False)

    try:
        diag_resp = api_cleanup_diagnose()
        if hasattr(diag_resp, 'get_json'):
            diag = diag_resp.get_json()
        elif hasattr(diag_resp, 'data'):
            diag = json.loads(diag_resp.data)
        else:
            diag = diag_resp

        if not diag or not diag.get('ok'):
            return jsonify({'ok': False, 'error': 'diagnose 실패'}), 500

        target_ids = [
            it['orderId'] for it in (diag.get('items') or [])
            if it.get('recommendation') == 'withdraw'
        ]

        if not target_ids:
            return jsonify({
                'ok': True,
                'dry_run': dry_run,
                'targets': [],
                'count': 0,
                'note': '회수 권장 건 없음'
            })

        if dry_run:
            return jsonify({
                'ok': True,
                'dry_run': True,
                'targets': target_ids,
                'count': len(target_ids),
                'note': '실행하려면 dry_run=false로 다시 호출'
            })

        cap_before_resp = api_capital_status()
        if hasattr(cap_before_resp, 'get_json'):
            cap_before_data = cap_before_resp.get_json()
        elif hasattr(cap_before_resp, 'data'):
            cap_before_data = json.loads(cap_before_resp.data)
        else:
            cap_before_data = cap_before_resp
        capital_before = cap_before_data.get('tied_total', 0) if cap_before_data and cap_before_data.get('ok') else 0

        import requests as rq
        r = rq.post('http://localhost:5001/api/cleanup/bulk-withdraw',
                    json={'orderIds': target_ids, 'force': False},
                    timeout=30)
        try:
            result = r.json()
        except Exception:
            result = {'status_code': r.status_code, 'text': r.text[:300]}

        return jsonify({
            'ok': True,
            'dry_run': False,
            'requested': len(target_ids),
            'count': len(target_ids),
            'targets': target_ids,
            'result': result,
            'capital_before': capital_before,
            'note': '5분 후 /api/cleanup/effect-report로 효과 확인'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/cleanup/effect-report', methods=['GET'])
def api_cleanup_effect():
    """capital_history에서 최근 변화 자동 분석 (정리 전후 비교)."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            SELECT timestamp, tied_total, tied_count, recoverable, recoverable_count
            FROM capital_history
            ORDER BY timestamp DESC LIMIT 10
        """)
        rows = c.fetchall()
        conn.close()

        if len(rows) < 2:
            return jsonify({
                'ok': True,
                'sufficient_data': False,
                'snapshots_count': len(rows),
                'note': 'capital_history 데이터 부족. 1시간 후 재시도.'
            })

        latest = rows[0]
        oldest = rows[-1]

        return jsonify({
            'ok': True,
            'sufficient_data': True,
            'before': {
                'timestamp': oldest[0], 'tied_total': oldest[1],
                'tied_count': oldest[2], 'recoverable': oldest[3]
            },
            'after': {
                'timestamp': latest[0], 'tied_total': latest[1],
                'tied_count': latest[2], 'recoverable': latest[3]
            },
            'delta': {
                'tied_total': latest[1] - oldest[1],
                'tied_count': latest[2] - oldest[2],
                'recoverable': latest[3] - oldest[3],
            },
            'snapshots_count': len(rows)
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


# ═══════════════════════════════════════════
# Step 18-C: 환경 비의존 가치 작업 4건
# ═══════════════════════════════════════════

@app.route('/api/my-bids/rank-changes', methods=['GET'])
def api_rank_changes():
    """내 입찰 중 rank가 1이 아닌 건 + 최근 변동 표시."""
    try:
        from pathlib import Path
        local_path = Path(__file__).parent / 'my_bids_local.json'
        if not local_path.exists():
            return jsonify({'ok': True, 'total_bids': 0, 'rank_1_count': 0,
                            'rank_lost_count': 0, 'unknown_count': 0,
                            'rank_lost_by_model': {}, 'last_sync': None,
                            'note': 'local cache 없음'})

        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids = local.get('bids', []) if isinstance(local, dict) else []

        not_first = [b for b in bids if b.get('rank') and b.get('rank') > 1]
        unknown = [b for b in bids if not b.get('rank')]

        from collections import defaultdict
        by_model = defaultdict(list)
        for b in not_first:
            by_model[b.get('model', '?')].append(b)

        return jsonify({
            'ok': True,
            'total_bids': len(bids),
            'rank_1_count': sum(1 for b in bids if b.get('rank') == 1),
            'rank_lost_count': len(not_first),
            'unknown_count': len(unknown),
            'rank_lost_by_model': dict(by_model),
            'last_sync': local.get('last_sync') or local.get('updated_at'),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/sales/analytics', methods=['GET'])
def api_sales_analytics():
    """판매 분석: 회전율, 매출 추이, 모델별 통계."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        c.execute("""
            SELECT DATE(trade_date) as d, COUNT(*) as cnt, SUM(sale_price) as total
            FROM sales_history
            WHERE DATE(trade_date) > DATE('now', '-7 days')
            GROUP BY DATE(trade_date)
            ORDER BY d
        """)
        daily_7d = [{'date': r[0], 'count': r[1], 'revenue': r[2] or 0} for r in c.fetchall()]

        c.execute("""
            SELECT DATE(trade_date) as d, COUNT(*) as cnt, SUM(sale_price) as total
            FROM sales_history
            WHERE DATE(trade_date) > DATE('now', '-30 days')
            GROUP BY DATE(trade_date)
            ORDER BY d
        """)
        daily_30d = [{'date': r[0], 'count': r[1], 'revenue': r[2] or 0} for r in c.fetchall()]

        c.execute("""
            SELECT model, COUNT(*) as cnt, AVG(sale_price) as avg_price, SUM(sale_price) as total
            FROM sales_history
            GROUP BY model
            ORDER BY cnt DESC
            LIMIT 10
        """)
        top_models = [{'model': r[0], 'count': r[1], 'avg_price': r[2] or 0, 'revenue': r[3] or 0} for r in c.fetchall()]

        c.execute("""
            SELECT size, COUNT(*) as cnt
            FROM sales_history
            GROUP BY size
            ORDER BY cnt DESC
            LIMIT 15
        """)
        size_freq = [{'size': r[0] or '-', 'count': r[1]} for r in c.fetchall()]

        c.execute("SELECT COUNT(*), SUM(sale_price) FROM sales_history")
        total_row = c.fetchone()

        conn.close()

        return jsonify({
            'ok': True,
            'total_count': total_row[0] or 0,
            'total_revenue': total_row[1] or 0,
            'daily_7d': daily_7d,
            'daily_30d': daily_30d,
            'top_models': top_models,
            'size_freq': size_freq,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


def _api_daily_log(date):
    """특정 날짜의 작업 일지 마크다운 반환."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        try:
            c.execute("""
                SELECT model, size, new_price, expected_profit
                FROM price_adjustments
                WHERE DATE(executed_at) = ? AND status = 'executed'
                ORDER BY executed_at
            """, (date,))
            bids = c.fetchall()
        except Exception:
            bids = []

        try:
            c.execute("""
                SELECT model, size, old_price, new_price, action
                FROM auto_adjust_log
                WHERE DATE(executed_at) = ?
                ORDER BY executed_at
            """, (date,))
            adjusts = c.fetchall()
        except Exception:
            adjusts = []

        try:
            c.execute("""
                SELECT model, size, sale_price
                FROM sales_history
                WHERE DATE(trade_date) = ?
                ORDER BY trade_date
            """, (date,))
            sales = c.fetchall()
        except Exception:
            sales = []

        try:
            c.execute("""
                SELECT title, message, created_at
                FROM notifications
                WHERE type = 'auth_failure' AND DATE(created_at) = ?
                ORDER BY created_at
            """, (date,))
            auth_fails = c.fetchall()
        except Exception:
            auth_fails = []

        conn.close()

        md = f"# 작업 일지 — {date}\n\n"
        md += f"## 📊 요약\n\n"
        md += f"- 입찰 실행: **{len(bids)}건**\n"
        md += f"- 자동 가격조정: **{len(adjusts)}건**\n"
        md += f"- 판매 체결: **{len(sales)}건**\n"
        md += f"- 인증 실패: **{len(auth_fails)}건**\n\n"

        if sales:
            total_rev = sum((s[2] or 0) for s in sales)
            md += f"### 💰 매출\n\n총 {total_rev:,}원 ({len(sales)}건)\n\n"
            md += "| 모델 | 사이즈 | 판매가 |\n|---|---|---|\n"
            for s in sales:
                md += f"| {s[0]} | {s[1] or '-'} | {(s[2] or 0):,}원 |\n"
            md += "\n"

        if bids:
            md += f"### 📦 입찰 실행 ({len(bids)}건)\n\n"
            md += "| 모델 | 사이즈 | 가격 | 예상수익 |\n|---|---|---|---|\n"
            for b in bids:
                profit = b[3] if b[3] is not None else '-'
                profit_str = profit if isinstance(profit, str) else f'{profit:,}원'
                md += f"| {b[0]} | {b[1] or '-'} | {(b[2] or 0):,}원 | {profit_str} |\n"
            md += "\n"

        if adjusts:
            md += f"### 🎯 자동 가격 조정 ({len(adjusts)}건)\n\n"
            for a in adjusts:
                md += f"- {a[0]} {a[1] or '-'}: {(a[2] or 0):,} → {(a[3] or 0):,}원 [{a[4]}]\n"
            md += "\n"

        if auth_fails:
            md += f"### ⚠️ 인증 실패 ({len(auth_fails)}건)\n\n"
            for af in auth_fails:
                md += f"- [{af[2]}] {af[0]}\n"
            md += "\n"

        if not (bids or adjusts or sales or auth_fails):
            md += "_이 날짜에 기록된 작업 없음_\n"

        return jsonify({'ok': True, 'date': date, 'markdown': md})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/daily-log/today', methods=['GET'])
def api_daily_log_today():
    return _api_daily_log(datetime.now().strftime('%Y-%m-%d'))


@app.route('/api/daily-log/<date>', methods=['GET'])
def api_daily_log_by_date(date):
    return _api_daily_log(date)


@app.route('/api/daily-log/save-today', methods=['POST'])
def api_daily_log_save():
    """오늘 일지를 daily_log/YYYY-MM-DD.md 파일로 저장."""
    try:
        from pathlib import Path
        date = datetime.now().strftime('%Y-%m-%d')
        result = _api_daily_log(date)
        try:
            data = result.get_json()
        except Exception:
            data = json.loads(result.data)
        if not data.get('ok'):
            return jsonify({'ok': False, 'error': 'log generation failed'}), 500

        log_dir = Path(__file__).parent / 'daily_log'
        log_dir.mkdir(exist_ok=True)
        log_path = log_dir / f'{date}.md'
        log_path.write_text(data['markdown'], encoding='utf-8')

        return jsonify({'ok': True, 'saved_to': str(log_path), 'date': date})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 18-D: 일일 자동화 + 운영 가시성 4건
# ═══════════════════════════════════════════

# ── 작업 #1: 작업 일지 자동 저장 (매일 23:55) ──
def _schedule_daily_log_save():
    """매일 23:55에 오늘 일지를 daily_log/YYYY-MM-DD.md로 저장."""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        with app.app_context():
            result = _api_daily_log(today)
            try:
                data = result.get_json()
            except Exception:
                data = json.loads(result.data) if hasattr(result, 'data') else result

        if data.get('ok'):
            log_dir = Path(__file__).parent / 'daily_log'
            log_dir.mkdir(exist_ok=True)
            log_path = log_dir / f'{today}.md'
            log_path.write_text(data['markdown'], encoding='utf-8')
            print(f"[DAILY-LOG] 저장 완료: {log_path}")
        else:
            print(f"[DAILY-LOG] 생성 실패: {data.get('error')}")
    except Exception as e:
        print(f"[DAILY-LOG] 스케줄러 에러: {e}")
        import sys as _sys
        _sys.stderr.write(f"daily-log scheduler error: {e}\n")


# ── 작업 #2: 내 입찰 자동 동기화 + rank 변동 알림 (30분) ──
_last_rank_snapshot = {}


def _schedule_my_bids_sync_with_alert():
    """30분마다 내 입찰 sync + rank 변동 감지 (1순위 → 다른 순위)."""
    global _last_rank_snapshot
    try:
        try:
            import requests as rq
            sync_resp = rq.post('http://localhost:5001/api/my-bids/sync', timeout=120)
            sync_data = sync_resp.json() if sync_resp.status_code == 200 else {}
            sync_task_id = sync_data.get('taskId') or sync_data.get('task_id')

            if sync_task_id:
                for _ in range(20):
                    time.sleep(3)
                    task_resp = rq.get(
                        f'http://localhost:5001/api/task/{sync_task_id}', timeout=5
                    )
                    if task_resp.status_code == 200:
                        status = task_resp.json().get('status')
                        if status in ('done', 'completed', 'success'):
                            break
                        if status in ('failed', 'error'):
                            print("[BIDS-MONITOR] sync 실패")
                            return
        except Exception as e:
            print(f"[BIDS-MONITOR] sync 호출 에러: {e}")
            return

        local_path = Path(__file__).parent / 'my_bids_local.json'
        if not local_path.exists():
            return
        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids = local.get('bids', [])

        current = {b.get('orderId'): b.get('rank') for b in bids if b.get('orderId')}

        dropped = []
        for oid, rank in current.items():
            prev = _last_rank_snapshot.get(oid)
            if prev == 1 and rank and rank > 1:
                bid_info = next((b for b in bids if b.get('orderId') == oid), {})
                dropped.append({
                    'orderId': oid,
                    'model': bid_info.get('model', '-'),
                    'size': bid_info.get('size', '-'),
                    'price': bid_info.get('price'),
                    'old_rank': prev,
                    'new_rank': rank,
                })

        if dropped:
            try:
                lines = [
                    f"- {d['model']} {d['size']} {d.get('price','-')}원: rank 1 → {d['new_rank']}"
                    for d in dropped
                ]
                body = (
                    f"내 입찰 중 {len(dropped)}건이 1위에서 밀렸습니다.\n\n"
                    + "\n".join(lines)
                )
                try:
                    safe_send_alert(
                        subject=f"[KREAM] 입찰 순위 변동 {len(dropped)}건",
                        body=body,
                        alert_type='rank_drop',
                    )
                except NameError:
                    print(f"[BIDS-MONITOR] {body}")
            except Exception as e:
                print(f"[BIDS-MONITOR] 알림 에러: {e}")

        _last_rank_snapshot = current
        print(f"[BIDS-MONITOR] sync 완료: {len(bids)}건, dropped {len(dropped)}")
    except Exception as e:
        print(f"[BIDS-MONITOR] 에러: {e}")


@app.route('/api/scheduler/bids-monitor/toggle', methods=['POST'])
def api_bids_monitor_toggle():
    """내 입찰 모니터 스케줄러 ON/OFF 토글."""
    data = request.get_json(silent=True) or {}
    enabled = data.get('enabled', True)
    if scheduler is None:
        return jsonify({'ok': False, 'error': 'scheduler unavailable'}), 500
    try:
        if enabled:
            try:
                scheduler.resume_job('my_bids_sync_monitor')
                return jsonify({'ok': True, 'enabled': True})
            except Exception:
                scheduler.add_job(
                    _schedule_my_bids_sync_with_alert,
                    'interval', minutes=30,
                    id='my_bids_sync_monitor',
                    replace_existing=True,
                    misfire_grace_time=300,
                )
                return jsonify({'ok': True, 'enabled': True, 'created': True})
        else:
            scheduler.pause_job('my_bids_sync_monitor')
            return jsonify({'ok': True, 'enabled': False})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── 작업 #3: 주간 리포트 ──
@app.route('/api/weekly-report', methods=['GET'])
def api_weekly_report():
    """지난 7일 종합 리포트 (마크다운)."""
    try:
        end = datetime.now()
        start = end - timedelta(days=7)
        start_str = start.strftime('%Y-%m-%d')
        end_str = end.strftime('%Y-%m-%d')

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        try:
            c.execute("""
                SELECT COUNT(*), SUM(expected_profit) FROM price_adjustments
                WHERE DATE(executed_at) BETWEEN ? AND ? AND status='executed'
            """, (start_str, end_str))
            bid_row = c.fetchone() or (0, 0)
        except Exception:
            bid_row = (0, 0)

        try:
            c.execute("""
                SELECT COUNT(*) FROM auto_adjust_log
                WHERE DATE(executed_at) BETWEEN ? AND ?
            """, (start_str, end_str))
            adjust_count = c.fetchone()[0] or 0
        except Exception:
            adjust_count = 0

        try:
            c.execute("""
                SELECT COUNT(*), SUM(sale_price) FROM sales_history
                WHERE DATE(trade_date) BETWEEN ? AND ?
            """, (start_str, end_str))
            sales_row = c.fetchone() or (0, 0)
        except Exception:
            sales_row = (0, 0)

        try:
            c.execute("""
                SELECT DATE(trade_date), COUNT(*), SUM(sale_price)
                FROM sales_history
                WHERE DATE(trade_date) BETWEEN ? AND ?
                GROUP BY DATE(trade_date) ORDER BY DATE(trade_date)
            """, (start_str, end_str))
            daily = c.fetchall()
        except Exception:
            daily = []

        try:
            c.execute("""
                SELECT model, COUNT(*), SUM(sale_price)
                FROM sales_history
                WHERE DATE(trade_date) BETWEEN ? AND ?
                GROUP BY model ORDER BY COUNT(*) DESC LIMIT 5
            """, (start_str, end_str))
            top_models = c.fetchall()
        except Exception:
            top_models = []

        try:
            c.execute("""
                SELECT COUNT(*) FROM notifications
                WHERE type='auth_failure' AND DATE(created_at) BETWEEN ? AND ?
            """, (start_str, end_str))
            auth_fails = c.fetchone()[0] or 0
        except Exception:
            auth_fails = 0

        conn.close()

        md = f"# 주간 리포트 — {start_str} ~ {end_str}\n\n"
        md += "## 📊 요약\n\n"
        md += f"- 입찰 실행: **{bid_row[0] or 0}건**\n"
        md += f"- 자동 가격조정: **{adjust_count}건**\n"
        md += f"- 판매 체결: **{sales_row[0] or 0}건** ({(sales_row[1] or 0):,}원)\n"
        md += f"- 인증 실패: **{auth_fails}건**\n\n"

        if daily:
            md += "## 📈 일별 매출\n\n| 날짜 | 건수 | 매출 |\n|---|---|---|\n"
            for d in daily:
                md += f"| {d[0]} | {d[1]} | {(d[2] or 0):,}원 |\n"
            md += "\n"

        if top_models:
            md += "## 🏆 모델 TOP 5\n\n| 모델 | 건수 | 매출 |\n|---|---|---|\n"
            for m in top_models:
                md += f"| {m[0]} | {m[1]} | {(m[2] or 0):,}원 |\n"
            md += "\n"

        if not (daily or top_models):
            md += "_지난 7일 판매 데이터 없음_\n"

        return jsonify({
            'ok': True,
            'period': {'start': start_str, 'end': end_str},
            'summary': {
                'bids_executed': bid_row[0] or 0,
                'adjustments': adjust_count,
                'sales_count': sales_row[0] or 0,
                'sales_revenue': sales_row[1] or 0,
                'auth_failures': auth_fails,
            },
            'markdown': md,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/weekly-report/save', methods=['POST'])
def api_weekly_report_save():
    """주간 리포트를 weekly_report/YYYY-WW.md로 저장."""
    try:
        result = api_weekly_report()
        try:
            data = result.get_json()
        except Exception:
            data = json.loads(result.data) if hasattr(result, 'data') else result
        if not data.get('ok'):
            return jsonify({'ok': False, 'error': 'report generation failed'}), 500

        now = datetime.now()
        week = now.isocalendar()[1]
        filename = f"{now.year}-W{week:02d}.md"
        rep_dir = Path(__file__).parent / 'weekly_report'
        rep_dir.mkdir(exist_ok=True)
        rep_path = rep_dir / filename
        rep_path.write_text(data['markdown'], encoding='utf-8')

        return jsonify({'ok': True, 'saved_to': str(rep_path), 'filename': filename})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


def _schedule_weekly_report_save():
    """매주 월 0:05 자동 저장."""
    try:
        with app.app_context():
            api_weekly_report_save()
        print("[WEEKLY-REPORT] 자동 저장 트리거")
    except Exception as e:
        print(f"[WEEKLY-REPORT] 에러: {e}")


# ── 작업 #4: 시스템 상태 종합 페이지 ──
@app.route('/admin/status', methods=['GET'])
def admin_status_page():
    """시스템 상태 종합 페이지 (HTML 직접 렌더)."""
    try:
        settings_path = Path(__file__).parent / 'settings.json'
        try:
            settings = json.loads(settings_path.read_text(encoding='utf-8'))
        except Exception:
            settings = {}

        auth_files = {}
        for name in ['auth_state.json', 'auth_state_kream.json']:
            p = Path(__file__).parent / name
            if p.exists():
                mtime = datetime.fromtimestamp(p.stat().st_mtime)
                age_h = (datetime.now() - mtime).total_seconds() / 3600
                auth_files[name] = {
                    'exists': True,
                    'modified': mtime.strftime('%Y-%m-%d %H:%M'),
                    'age_hours': round(age_h, 1),
                }
            else:
                auth_files[name] = {'exists': False}

        scheduler_jobs = []
        try:
            if scheduler is not None:
                for job in scheduler.get_jobs():
                    scheduler_jobs.append({
                        'id': job.id,
                        'next_run': str(job.next_run_time) if job.next_run_time else 'paused',
                    })
        except Exception:
            pass

        try:
            conn = sqlite3.connect(str(PRICE_DB))
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM price_adjustments WHERE status='pending'")
            pa_pending = c.fetchone()[0] or 0
            c.execute("SELECT COUNT(*) FROM sales_history")
            sales_total = c.fetchone()[0] or 0
            c.execute("SELECT MAX(trade_date) FROM sales_history")
            last_sale = c.fetchone()[0] or '-'
            conn.close()
        except Exception:
            pa_pending = 0
            sales_total = 0
            last_sale = '-'

        toggles = {
            '자동 입찰': settings.get('auto_bid_enabled', False),
            '자동 가격조정': settings.get('auto_adjust_enabled', False),
            '자동 재입찰': settings.get('auto_rebid_enabled', False),
            '자동 정리': settings.get('auto_cleanup_enabled', False),
            '허브넷 자동 PDF': settings.get('hubnet_auto_pdf', False),
            '사전 갱신': settings.get('session_refresh_enabled', True),
        }

        env = settings.get('environment', 'unknown')
        env_detail = settings.get('env_detection_detail', '-') or '-'
        env_checked_at = (settings.get('env_checked_at') or '-')[:16]

        html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>시스템 상태 — KREAM 자동화</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,sans-serif;background:#f9fafb;margin:0;padding:24px;color:#111}}
.container{{max-width:1100px;margin:0 auto}}
h1{{margin:0 0 24px 0}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;margin-bottom:24px}}
.card{{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:16px}}
.card h2{{margin:0 0 12px 0;font-size:15px;color:#374151}}
.row{{display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #f3f4f6;font-size:13px}}
.row:last-child{{border:none}}
.k{{color:#6b7280}}
.v{{font-weight:600}}
.ok{{color:#059669}}
.warn{{color:#d97706}}
.err{{color:#dc2626}}
.muted{{color:#9ca3af}}
.refresh{{padding:6px 14px;background:#2563eb;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px}}
table{{width:100%;font-size:12px;border-collapse:collapse}}
table td{{padding:6px;border-bottom:1px solid #f3f4f6}}
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px}}
</style>
</head>
<body>
<div class="container">
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
  <h1>🛠 시스템 상태</h1>
  <button class="refresh" onclick="location.reload()">새로고침</button>
</div>
<div style="font-size:12px;color:#6b7280;margin-bottom:16px">최종 갱신: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>

<div class="grid">

  <div class="card">
    <h2>🌐 환경</h2>
    <div class="row"><span class="k">environment</span><span class="v {('ok' if env in ('korea','imac_kr') else 'warn')}">{env}</span></div>
    <div class="row"><span class="k">detail</span><span class="v">{env_detail}</span></div>
    <div class="row"><span class="k">checked_at</span><span class="v muted">{env_checked_at}</span></div>
  </div>

  <div class="card">
    <h2>🔐 인증</h2>'''

        for name, info in auth_files.items():
            if info['exists']:
                cls = 'ok' if info['age_hours'] < 12 else ('warn' if info['age_hours'] < 24 else 'err')
                html += f'<div class="row"><span class="k">{name}</span><span class="v {cls}">{info["age_hours"]}h 전</span></div>'
            else:
                html += f'<div class="row"><span class="k">{name}</span><span class="v err">없음</span></div>'

        html += '''
  </div>

  <div class="card">
    <h2>⚙️ 자동 토글</h2>'''
        for name, val in toggles.items():
            cls = 'ok' if val else 'muted'
            badge = 'ON' if val else 'OFF'
            html += f'<div class="row"><span class="k">{name}</span><span class="badge {cls}">{badge}</span></div>'
        html += f'''
  </div>

  <div class="card">
    <h2>📅 스케줄러 ({len(scheduler_jobs)}개)</h2>'''
        if scheduler_jobs:
            for job in scheduler_jobs:
                next_run = str(job['next_run'])[:16] if job['next_run'] != 'paused' else 'paused'
                html += f'<div class="row"><span class="k">{job["id"]}</span><span class="v muted" style="font-size:11px">{next_run}</span></div>'
        else:
            html += '<div class="muted" style="font-size:13px">등록된 작업 없음</div>'
        html += f'''
  </div>

  <div class="card">
    <h2>📦 DB 현황</h2>
    <div class="row"><span class="k">조정 대기 (pending)</span><span class="v {('warn' if pa_pending > 10 else 'ok')}">{pa_pending}</span></div>
    <div class="row"><span class="k">판매 누적</span><span class="v">{sales_total}</span></div>
    <div class="row"><span class="k">최근 판매</span><span class="v muted">{last_sale}</span></div>
  </div>

  <div class="card">
    <h2>🔗 빠른 링크</h2>
    <div style="font-size:13px;line-height:1.8">
      <div><a href="/" style="color:#2563eb">📊 메인 대시보드</a></div>
      <div><a href="/api/health" style="color:#2563eb">/api/health</a></div>
      <div><a href="/api/daily-summary" style="color:#2563eb">/api/daily-summary</a></div>
      <div><a href="/api/daily-log/today" style="color:#2563eb">/api/daily-log/today</a></div>
      <div><a href="/api/weekly-report" style="color:#2563eb">/api/weekly-report</a></div>
    </div>
  </div>

</div>
</div>
</body>
</html>'''

        return Response(html, mimetype='text/html')
    except Exception as e:
        return Response(f"<pre>Error: {e}</pre>", status=500, mimetype='text/html')


# ── Step 23: sync 0건 자동 경고 ──
def _check_sync_health():
    """sync 결과가 0건이면 알림."""
    try:
        from pathlib import Path as _Path
        local_path = _Path(__file__).parent / 'my_bids_local.json'
        if not local_path.exists():
            return

        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids_count = len(local.get('bids', []) if isinstance(local, dict) else [])
        last_sync = local.get('last_sync') or local.get('updated_at') if isinstance(local, dict) else None

        if bids_count == 0:
            from datetime import datetime as _dt, timedelta as _td
            try:
                if last_sync:
                    if '/' in str(last_sync):
                        last_sync_dt = _dt.strptime(last_sync, '%Y/%m/%d %H:%M')
                    else:
                        last_sync_dt = _dt.fromisoformat(last_sync)
                    if _dt.now() - last_sync_dt < _td(hours=1):
                        try:
                            safe_send_alert(
                                subject='[KREAM] sync 0건 경고',
                                body='판매자센터 sync 결과 0건. 페이지 파싱 깨졌을 가능성.\n\n/api/diagnostics/sync-page-dump 호출하여 확인 필요.',
                                alert_type='sync_zero'
                            )
                        except Exception:
                            pass
            except Exception:
                pass
    except Exception:
        pass


# ── 스케줄러 작업 등록 (멱등성: replace_existing=True) ──
def _register_step18d_jobs():
    """Step 18-D 신규 스케줄러 작업 등록 (이미 있으면 교체)."""
    if scheduler is None:
        print("[SCHEDULER] APScheduler 사용 불가 — Step 18-D 작업 등록 스킵")
        return

    try:
        scheduler.add_job(
            _schedule_daily_log_save,
            'cron',
            hour=23, minute=55,
            id='daily_log_save',
            replace_existing=True,
            misfire_grace_time=600,
        )
        print("[SCHEDULER] daily_log_save 등록 (매일 23:55)")
    except Exception as e:
        print(f"[SCHEDULER] daily_log_save 등록 실패: {e}")

    try:
        scheduler.add_job(
            _schedule_my_bids_sync_with_alert,
            'interval',
            minutes=30,
            id='my_bids_sync_monitor',
            replace_existing=True,
            misfire_grace_time=300,
        )
        print("[SCHEDULER] my_bids_sync_monitor 등록 (30분 간격)")
    except Exception as e:
        print(f"[SCHEDULER] my_bids_sync_monitor 등록 실패: {e}")

    try:
        scheduler.add_job(
            _schedule_weekly_report_save,
            'cron',
            day_of_week='mon', hour=0, minute=5,
            id='weekly_report_save',
            replace_existing=True,
            misfire_grace_time=3600,
        )
        print("[SCHEDULER] weekly_report_save 등록 (매주 월 0:05)")
    except Exception as e:
        print(f"[SCHEDULER] weekly_report_save 등록 실패: {e}")

    # Step 23: sync 0건 자동 경고 (35분 간격)
    try:
        scheduler.add_job(
            _check_sync_health,
            'interval', minutes=35,
            id='sync_health_check',
            replace_existing=True,
            misfire_grace_time=300,
        )
        print("[SCHEDULER] sync_health_check 등록 (35분 간격)")
    except Exception as e:
        print(f"[SCHEDULER] sync_health_check 등록 실패: {e}")


# ═══════════════════════════════════════════
# Step 20: 자본 + 우회 + 모델분석
# ═══════════════════════════════════════════

@app.route('/api/capital-status', methods=['GET'])
def api_capital_status():
    """현재 입찰에 묶인 자본 + 회수 가능 자본 분석."""
    try:
        from collections import defaultdict

        local_path = BASE_DIR / 'my_bids_local.json'
        if not local_path.exists():
            return jsonify({'ok': True, 'tied_total': 0, 'tied_count': 0,
                            'unknown_cost_count': 0, 'recoverable': 0,
                            'recoverable_count': 0, 'top_models': [],
                            'avg_cost_estimate': 0, 'note': 'my_bids_local 없음'})

        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids = local.get('bids', []) if isinstance(local, dict) else []

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        try:
            c.execute("SELECT DISTINCT order_id FROM sales_history WHERE order_id IS NOT NULL")
            sold_ids = {row[0] for row in c.fetchall()}
        except Exception:
            sold_ids = set()

        active_bids = [b for b in bids if b.get('orderId') not in sold_ids]

        try:
            settings = json.loads((BASE_DIR / 'settings.json').read_text(encoding='utf-8'))
        except Exception:
            settings = {}
        fee_rate = settings.get('commission_rate', 6) / 100
        fixed_fee = 2500
        min_margin = settings.get('min_margin', 4000)
        undercut = settings.get('undercut_amount', 1000)
        overseas_ship_default = settings.get('overseas_shipping', 8000)

        try:
            c.execute(
                "SELECT AVG(cny_price * exchange_rate * 1.03 + COALESCE(overseas_shipping, ?)) "
                "FROM bid_cost",
                (overseas_ship_default,)
            )
            avg_cost_row = c.fetchone()
            avg_cost = avg_cost_row[0] if avg_cost_row and avg_cost_row[0] else 50000
        except Exception:
            avg_cost = 50000

        tied_total = 0
        tied_by_model = defaultdict(lambda: {'count': 0, 'capital': 0})
        recoverable = 0
        recoverable_count = 0
        unknown_cost_count = 0

        for b in active_bids:
            order_id = b.get('orderId')
            model = b.get('model', '-')
            price = b.get('price') or 0
            rank = b.get('rank')

            cost_known = False
            cost = avg_cost
            try:
                c.execute(
                    "SELECT cny_price, exchange_rate, overseas_shipping FROM bid_cost WHERE order_id = ?",
                    (order_id,)
                )
                row = c.fetchone()
                if row and row[0] is not None and row[1] is not None:
                    ship = row[2] if row[2] is not None else overseas_ship_default
                    cost = float(row[0]) * float(row[1]) * 1.03 + float(ship)
                    cost_known = True
            except Exception:
                pass

            if not cost_known:
                unknown_cost_count += 1

            tied_total += cost
            tied_by_model[model]['count'] += 1
            tied_by_model[model]['capital'] += cost

            if rank and rank > 1 and cost_known:
                try:
                    hyp_settlement = (price - undercut) * (1 - fee_rate * 1.1) - fixed_fee
                    hyp_margin = hyp_settlement - cost
                    if hyp_margin < min_margin:
                        recoverable += cost
                        recoverable_count += 1
                except Exception:
                    pass

        conn.close()

        sorted_models = sorted(tied_by_model.items(), key=lambda x: -x[1]['capital'])[:10]

        return jsonify({
            'ok': True,
            'tied_total': round(tied_total),
            'tied_count': len(active_bids),
            'unknown_cost_count': unknown_cost_count,
            'recoverable': round(recoverable),
            'recoverable_count': recoverable_count,
            'top_models': [
                {'model': m, 'count': v['count'], 'capital': round(v['capital'])}
                for m, v in sorted_models
            ],
            'avg_cost_estimate': round(avg_cost),
            'labels': {
                'tied_total': '활성 입찰 노출액',
                'tied_count': '활성 입찰',
                'recoverable': '정리 가능 노출액',
                'recoverable_count': '정리 가능 건',
                'business_model': 'consignment_purchase',
                'note': '구매대행: 체결 시점에 매입, 입찰만으로는 자본 미지출'
            }
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/market-prices/from-bids', methods=['GET'])
def api_market_prices_from_bids():
    """my_bids_local에서 추출한 경쟁가 (가격수집 우회 데이터)."""
    try:
        from collections import defaultdict

        local_path = BASE_DIR / 'my_bids_local.json'
        if not local_path.exists():
            return jsonify({'ok': True, 'count': 0, 'items': [], 'by_model': {},
                            'note': 'local cache 없음'})

        local = json.loads(local_path.read_text(encoding='utf-8'))
        bids = local.get('bids', []) if isinstance(local, dict) else []

        items = []
        for b in bids:
            mlp = b.get('market_lowest_price') or b.get('marketLowestPrice')
            if mlp:
                items.append({
                    'orderId': b.get('orderId'),
                    'model': b.get('model'),
                    'size': b.get('size'),
                    'my_price': b.get('price'),
                    'market_lowest': mlp,
                    'rank': b.get('rank'),
                    'gap': mlp - (b.get('price') or 0),
                })

        by_model = defaultdict(list)
        for it in items:
            by_model[it['model']].append(it)

        return jsonify({
            'ok': True,
            'count': len(items),
            'items': items,
            'by_model': dict(by_model),
            'last_sync': local.get('last_sync') or local.get('updated_at') if isinstance(local, dict) else None,
            'note': '판매자센터에서 sync 시 추출한 경쟁가. KREAM 일반사이트 수집 우회'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/model/<path:model>/deep-analysis', methods=['GET'])
def api_model_deep(model):
    """특정 모델 종합 분석: 입찰/판매이력/마진/추이 + 추천 액션."""
    try:
        from collections import defaultdict

        local_path = BASE_DIR / 'my_bids_local.json'
        bids = []
        if local_path.exists():
            try:
                local = json.loads(local_path.read_text(encoding='utf-8'))
                bids = [b for b in local.get('bids', []) if b.get('model') == model]
            except Exception:
                bids = []

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        sales = []
        try:
            c.execute("""
                SELECT size, sale_price, trade_date, ship_status
                FROM sales_history
                WHERE model = ?
                ORDER BY trade_date DESC
            """, (model,))
            sales = [
                {'size': r[0], 'price': r[1], 'date': r[2], 'status': r[3]}
                for r in c.fetchall()
            ]
        except Exception:
            pass

        adjustments = []
        try:
            c.execute("""
                SELECT old_price, new_price, expected_profit, status, executed_at
                FROM price_adjustments
                WHERE model = ?
                ORDER BY created_at DESC LIMIT 20
            """, (model,))
            adjustments = [
                {'old': r[0], 'new': r[1], 'profit': r[2], 'status': r[3], 'at': r[4]}
                for r in c.fetchall()
            ]
        except Exception:
            pass

        costs = []
        try:
            c.execute("""
                SELECT order_id, size, cny_price, exchange_rate, overseas_shipping
                FROM bid_cost
                WHERE model = ?
            """, (model,))
            costs = [
                {'order_id': r[0], 'size': r[1], 'cny': r[2], 'fx': r[3], 'ship': r[4]}
                for r in c.fetchall()
            ]
        except Exception:
            pass

        conn.close()

        total_sales = len(sales)
        total_revenue = sum((s.get('price') or 0) for s in sales)
        avg_sale_price = total_revenue / total_sales if total_sales else 0

        active_count = len(bids)
        rank_1_count = sum(1 for b in bids if b.get('rank') == 1)

        size_freq = defaultdict(int)
        for s in sales:
            size_freq[s.get('size') or '-'] += 1

        recommendation = 'monitor'
        rec_reason = ''
        if total_sales == 0:
            recommendation = 'no_data'
            rec_reason = '판매 이력 없음 — 충분히 누적 후 재평가'
        elif active_count == 0:
            recommendation = 'restock' if total_sales > 2 else 'consider'
            rec_reason = f'입찰 없음, 판매 {total_sales}건 → 재입찰 고려'
        elif rank_1_count == 0 and active_count > 0:
            recommendation = 'review_pricing'
            rec_reason = f'활성 입찰 {active_count}건 모두 1위 아님 → 가격 재검토'
        elif total_sales > 5 and rank_1_count >= active_count * 0.5:
            recommendation = 'expand'
            rec_reason = f'판매 {total_sales}건, 1위 비율 양호 → 사이즈/수량 확대 검토'

        return jsonify({
            'ok': True,
            'model': model,
            'summary': {
                'active_bids': active_count,
                'rank_1_bids': rank_1_count,
                'total_sales': total_sales,
                'total_revenue': total_revenue,
                'avg_sale_price': round(avg_sale_price),
                'has_cost_data': len(costs) > 0,
            },
            'bids': bids,
            'sales': sales[:30],
            'adjustments': adjustments,
            'costs': costs,
            'size_frequency': dict(size_freq),
            'recommendation': recommendation,
            'recommendation_reason': rec_reason,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


# ═══════════════════════════════════════════
# Step 21: 효과 측정 인프라 (자본 추이 + 포트폴리오 + ROI)
# ═══════════════════════════════════════════

def _migrate_capital_history():
    """capital_history 테이블 확인/생성 (CREATE IF NOT EXISTS)."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS capital_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                tied_total INTEGER,
                tied_count INTEGER,
                recoverable INTEGER,
                recoverable_count INTEGER,
                unknown_cost_count INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_capital_history_ts ON capital_history(timestamp)")
        conn.commit()
        conn.close()
        print("[MIGRATE] capital_history 테이블 확인/생성")
    except Exception as e:
        print(f"[MIGRATE] capital_history 에러: {e}")


def _snapshot_capital():
    """자본 현황 스냅샷을 capital_history에 저장."""
    try:
        with app.app_context():
            resp = api_capital_status()
            if hasattr(resp, 'get_json'):
                data = resp.get_json()
            elif hasattr(resp, 'data'):
                data = json.loads(resp.data)
            else:
                data = None

        if not data or not data.get('ok'):
            print(f"[CAPITAL-SNAPSHOT] 실패: {data}")
            return

        from datetime import datetime
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            INSERT INTO capital_history
            (timestamp, tied_total, tied_count, recoverable, recoverable_count, unknown_cost_count)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            datetime.now().isoformat(),
            data.get('tied_total', 0),
            data.get('tied_count', 0),
            data.get('recoverable', 0),
            data.get('recoverable_count', 0),
            data.get('unknown_cost_count', 0),
        ))
        conn.commit()
        conn.close()
        print(f"[CAPITAL-SNAPSHOT] 기록: tied={data.get('tied_total')}")
    except Exception as e:
        print(f"[CAPITAL-SNAPSHOT] 에러: {e}")


@app.route('/api/capital-history', methods=['GET'])
def api_capital_history():
    """자본 추이 조회. ?hours=24 (기본) ?days=7 등."""
    try:
        from datetime import datetime, timedelta
        hours = request.args.get('hours', type=int)
        days = request.args.get('days', type=int)

        if days:
            since = (datetime.now() - timedelta(days=days)).isoformat()
        elif hours:
            since = (datetime.now() - timedelta(hours=hours)).isoformat()
        else:
            since = (datetime.now() - timedelta(hours=24)).isoformat()

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            SELECT timestamp, tied_total, tied_count, recoverable, recoverable_count
            FROM capital_history
            WHERE timestamp >= ?
            ORDER BY timestamp ASC
        """, (since,))
        rows = c.fetchall()
        conn.close()

        items = [
            {'timestamp': r[0], 'tied_total': r[1], 'tied_count': r[2],
             'recoverable': r[3], 'recoverable_count': r[4]}
            for r in rows
        ]

        change = None
        if len(items) >= 2:
            try:
                change = {
                    'tied_delta': (items[-1]['tied_total'] or 0) - (items[0]['tied_total'] or 0),
                    'recoverable_delta': (items[-1]['recoverable'] or 0) - (items[0]['recoverable'] or 0),
                    'period_hours': round(
                        (datetime.fromisoformat(items[-1]['timestamp']) -
                         datetime.fromisoformat(items[0]['timestamp'])).total_seconds() / 3600, 1
                    )
                }
            except Exception:
                change = None

        return jsonify({
            'ok': True,
            'count': len(items),
            'items': items,
            'change': change,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/portfolio/overview', methods=['GET'])
def api_portfolio_overview():
    """모든 활성 모델 자동 분류 (expand/restock/review_pricing/monitor/no_data/archive)."""
    try:
        from collections import defaultdict

        local_path = BASE_DIR / 'my_bids_local.json'
        bids = []
        if local_path.exists():
            try:
                local = json.loads(local_path.read_text(encoding='utf-8'))
                bids = local.get('bids', []) if isinstance(local, dict) else []
            except Exception:
                bids = []

        active_models = set(b.get('model') for b in bids if b.get('model'))

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        try:
            c.execute("SELECT DISTINCT model FROM sales_history WHERE model IS NOT NULL")
            sold_models = set(r[0] for r in c.fetchall())
        except Exception:
            sold_models = set()

        all_models = active_models | sold_models

        models_info = []
        for model in all_models:
            mb = [b for b in bids if b.get('model') == model]
            active_count = len(mb)
            rank_1 = sum(1 for b in mb if b.get('rank') == 1)

            try:
                c.execute(
                    "SELECT COUNT(*), SUM(sale_price), MAX(trade_date) FROM sales_history WHERE model = ?",
                    (model,)
                )
                row = c.fetchone()
                sales_count = row[0] or 0
                revenue = row[1] or 0
                last_sale = row[2]
            except Exception:
                sales_count, revenue, last_sale = 0, 0, None

            rec = 'monitor'
            rec_priority = 3
            if sales_count == 0 and active_count > 0:
                rec = 'no_data'
                rec_priority = 3
            elif active_count == 0 and sales_count > 2:
                rec = 'restock'
                rec_priority = 1
            elif active_count > 0 and rank_1 == 0:
                rec = 'review_pricing'
                rec_priority = 2
            elif sales_count > 5 and active_count > 0 and rank_1 >= active_count * 0.5:
                rec = 'expand'
                rec_priority = 1
            elif sales_count == 0 and active_count == 0:
                rec = 'archive'
                rec_priority = 3

            models_info.append({
                'model': model,
                'active_bids': active_count,
                'rank_1_bids': rank_1,
                'sales_count': sales_count,
                'revenue': revenue,
                'last_sale': last_sale,
                'recommendation': rec,
                'priority': rec_priority,
            })

        conn.close()

        by_rec = defaultdict(list)
        for m in models_info:
            by_rec[m['recommendation']].append(m)

        models_info.sort(key=lambda x: (x['priority'], -x['sales_count']))

        return jsonify({
            'ok': True,
            'total_models': len(models_info),
            'by_recommendation': {k: v for k, v in by_rec.items()},
            'models': models_info,
            'stats': {
                'expand': len(by_rec.get('expand', [])),
                'restock': len(by_rec.get('restock', [])),
                'review_pricing': len(by_rec.get('review_pricing', [])),
                'monitor': len(by_rec.get('monitor', [])),
                'no_data': len(by_rec.get('no_data', [])),
                'archive': len(by_rec.get('archive', [])),
            }
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/capital-efficiency', methods=['GET'])
def api_capital_efficiency():
    """30일 ROI 추정 + 모델별 효율."""
    try:
        from datetime import datetime, timedelta

        thirty_days_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()

        try:
            c.execute("""
                SELECT model, COUNT(*) as cnt, SUM(sale_price) as revenue
                FROM sales_history
                WHERE DATE(trade_date) >= ?
                GROUP BY model
            """, (thirty_days_ago,))
            sales_by_model = {r[0]: {'count': r[1], 'revenue': r[2] or 0} for r in c.fetchall()}
        except Exception:
            sales_by_model = {}

        try:
            c.execute("""
                SELECT AVG(tied_total) FROM capital_history
                WHERE DATE(timestamp) >= ?
            """, (thirty_days_ago,))
            avg_row = c.fetchone()
            avg_capital = avg_row[0] if avg_row and avg_row[0] else None
        except Exception:
            avg_capital = None

        if not avg_capital:
            try:
                with app.app_context():
                    cap_resp = api_capital_status()
                    cap_data = cap_resp.get_json() if hasattr(cap_resp, 'get_json') else (
                        json.loads(cap_resp.data) if hasattr(cap_resp, 'data') else {}
                    )
                avg_capital = cap_data.get('tied_total', 1) if cap_data and cap_data.get('ok') else 1
            except Exception:
                avg_capital = 1
            data_source = 'current_only (capital_history 누적 부족)'
        else:
            data_source = '30day_avg'

        if not avg_capital or avg_capital <= 0:
            avg_capital = 1

        total_sales = sum(s['count'] for s in sales_by_model.values())
        total_revenue = sum(s['revenue'] for s in sales_by_model.values())

        try:
            c.execute("""
                SELECT AVG(cny_price * exchange_rate * 1.03 + COALESCE(overseas_shipping, 8000))
                FROM bid_cost
            """)
            avg_cost_row = c.fetchone()
            avg_cost = avg_cost_row[0] if avg_cost_row and avg_cost_row[0] else 50000
        except Exception:
            avg_cost = 50000

        estimated_total_cost = avg_cost * total_sales
        gross_profit = total_revenue - estimated_total_cost

        roi_30d = (gross_profit / avg_capital) if avg_capital else 0

        local_path = BASE_DIR / 'my_bids_local.json'
        try:
            local = json.loads(local_path.read_text(encoding='utf-8')) if local_path.exists() else {}
            all_bids = local.get('bids', []) if isinstance(local, dict) else []
        except Exception:
            all_bids = []

        model_roi = []
        for model, sales in sales_by_model.items():
            est_cost = avg_cost * sales['count']
            profit = sales['revenue'] - est_cost
            model_bids = sum(1 for b in all_bids if b.get('model') == model)
            model_capital_est = max(avg_cost * model_bids, 1)
            model_roi.append({
                'model': model,
                'sales_count': sales['count'],
                'revenue': sales['revenue'],
                'est_profit': round(profit),
                'est_capital': round(model_capital_est),
                'roi_estimate': round(profit / model_capital_est, 3),
            })

        model_roi.sort(key=lambda x: -x['roi_estimate'])

        conn.close()

        return jsonify({
            'ok': True,
            'period_days': 30,
            'avg_capital': round(avg_capital),
            'data_source': data_source,
            'total_sales': total_sales,
            'total_revenue': round(total_revenue),
            'estimated_cost': round(estimated_total_cost),
            'gross_profit': round(gross_profit),
            'roi_30d': round(roi_30d, 3),
            'roi_30d_pct': round(roi_30d * 100, 1),
            'top_models_by_roi': model_roi[:10],
            'note': '원가는 bid_cost 평균치 기반 추정. 정확도는 bid_cost 입력률에 의존.'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


# ═══════════════════════════════════════════
# Step 31: 신규 입찰 / 판매 마진 / 시장 모니터링
# ═══════════════════════════════════════════

def _calc_margin_status_msg(status, margin, min_margin):
    if status == 'LOW':
        return f'마진 {round(margin):,}원 < {min_margin:,} (단가 협상 필요)'
    if status == 'DEFICIT':
        return f'적자 {round(margin):,}원 (입찰 불가)'
    return ''


@app.route('/api/new-bid/calc-batch', methods=['POST'])
def api_new_bid_calc_batch():
    """신규 입찰 일괄 마진 계산. 모델/사이즈/판매가/CNY 리스트 → 원가/정산/마진/GO|LOW|DEFICIT.

    입력: {"items": [{"model": "JQ4110", "size": "260", "sale_price": 150000, "cny": 350}, ...]}
    """
    try:
        data = request.get_json() or {}
        items = data.get('items', [])
        if not items:
            return jsonify({'ok': False, 'error': 'items required'}), 400

        try:
            settings = json.loads(Path(__file__).parent.joinpath('settings.json').read_text(encoding='utf-8'))
        except Exception:
            settings = {}
        fee_rate = settings.get('commission_rate', 6) / 100.0
        fixed_fee = 2500
        min_margin = settings.get('min_margin', 4000)
        overseas_ship = settings.get('overseas_shipping', 8000)
        undercut = settings.get('undercut_amount', 1000)

        try:
            import requests as _rq
            fx_resp = _rq.get('http://localhost:5001/api/exchange-rate', timeout=5)
            fx = fx_resp.json().get('rate', settings.get('exchange_rate_cny', 216))
        except Exception:
            fx = settings.get('exchange_rate_cny', 216)

        results = []
        for item in items:
            model = (item.get('model') or '').strip()
            size = item.get('size', '')
            try:
                sale_price = float(item.get('sale_price', 0) or 0)
                cny = float(item.get('cny', 0) or 0)
            except Exception:
                sale_price = 0
                cny = 0

            if not model or cny <= 0 or sale_price <= 0:
                results.append({**item, 'status': 'INVALID', 'reason': '모델/CNY/판매가 누락'})
                continue

            bid_price = math.ceil((sale_price - undercut) / 1000.0) * 1000
            cost = cny * fx * 1.03 + overseas_ship
            settlement = bid_price * (1 - fee_rate * 1.1) - fixed_fee
            margin = settlement - cost
            status = 'GO' if margin >= min_margin else ('LOW' if margin >= 0 else 'DEFICIT')

            results.append({
                **item,
                'fx': round(fx, 2),
                'bid_price': int(bid_price),
                'cost': round(cost),
                'settlement': round(settlement),
                'margin': round(margin),
                'margin_pct': round((margin / bid_price * 100) if bid_price else 0, 1),
                'status': status,
                'reason': '마진 OK' if status == 'GO' else _calc_margin_status_msg(status, margin, min_margin),
            })

        go_count = sum(1 for r in results if r.get('status') == 'GO')
        low_count = sum(1 for r in results if r.get('status') == 'LOW')
        deficit_count = sum(1 for r in results if r.get('status') == 'DEFICIT')
        invalid_count = sum(1 for r in results if r.get('status') == 'INVALID')

        return jsonify({
            'ok': True,
            'total': len(results),
            'go': go_count,
            'low': low_count,
            'deficit': deficit_count,
            'invalid': invalid_count,
            'items': results,
            'settings_used': {
                'fx': round(fx, 2),
                'fee_rate': fee_rate,
                'min_margin': min_margin,
                'overseas_ship': overseas_ship,
                'undercut': undercut,
            },
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/new-bid/auto-fetch-prices', methods=['POST'])
def api_new_bid_auto_fetch_prices():
    """모델 리스트 받아서 /api/search 호출 → 사이즈별 시장가 수집."""
    try:
        data = request.get_json() or {}
        models = data.get('models', [])
        if not models:
            return jsonify({'ok': False, 'error': 'models required'}), 400

        import requests as _rq
        results = []
        for model in models[:20]:
            model = (model or '').strip()
            if not model:
                continue
            try:
                r = _rq.post('http://localhost:5001/api/search', json={'model': model}, timeout=60)
                if r.status_code == 200:
                    d = r.json()
                    sizes = d.get('sizes') or d.get('size_prices') or []
                    results.append({'model': model, 'ok': True, 'sizes': sizes})
                else:
                    results.append({'model': model, 'ok': False, 'error': f'HTTP {r.status_code}'})
            except Exception as e:
                results.append({'model': model, 'ok': False, 'error': str(e)})

        return jsonify({'ok': True, 'count': len(results), 'results': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/realized-margin/cumulative', methods=['GET'])
def api_realized_margin_cumulative():
    """월별 누적 + 모델별 ROI (real-margin 365일 호출)."""
    try:
        import requests as _rq
        from collections import defaultdict
        try:
            r = _rq.get('http://localhost:5001/api/real-margin?days=365', timeout=30)
            rm = r.json()
        except Exception as e:
            return jsonify({'ok': False, 'error': f'real-margin 호출 실패: {e}'}), 500
        if not rm.get('ok'):
            return jsonify({'ok': False, 'error': 'real-margin ok=false'}), 500

        items = rm.get('items', []) or []
        monthly = defaultdict(lambda: {'count': 0, 'revenue': 0, 'margin': 0, 'unknown_count': 0})
        by_model = defaultdict(lambda: {'count': 0, 'revenue': 0, 'margin': 0, 'has_cost': 0})

        for it in items:
            d = it.get('trade_date') or ''
            if not d:
                continue
            month = d[:7]
            sp = it.get('sale_price') or 0
            mg = it.get('margin')
            monthly[month]['count'] += 1
            monthly[month]['revenue'] += sp
            if mg is not None:
                monthly[month]['margin'] += mg
            else:
                monthly[month]['unknown_count'] += 1

            model = it.get('model') or '?'
            by_model[model]['count'] += 1
            by_model[model]['revenue'] += sp
            if mg is not None:
                by_model[model]['margin'] += mg
                by_model[model]['has_cost'] += 1

        return jsonify({
            'ok': True,
            'monthly': sorted([{'month': k, **v} for k, v in monthly.items()], key=lambda x: x['month']),
            'top_models': sorted(
                [{'model': k, **v,
                  'avg_margin': round(v['margin'] / v['has_cost']) if v['has_cost'] else None}
                 for k, v in by_model.items()],
                key=lambda x: -x['count'],
            )[:10],
            'total_count': len(items),
            'total_revenue': sum((it.get('sale_price') or 0) for it in items),
            'total_confirmed_margin': sum((it.get('margin') or 0) for it in items if it.get('margin') is not None),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


def _migrate_market_price_history():
    """market_price_history 테이블 보장 (CREATE IF NOT EXISTS)."""
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS market_price_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model TEXT NOT NULL,
                size TEXT,
                buy_price INTEGER,
                recent_price INTEGER,
                collected_at TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_mph_model_collected ON market_price_history(model, collected_at)")
        conn.commit()
        conn.close()
        print("[MIGRATE] market_price_history 확인/생성")
    except Exception as e:
        print(f"[MIGRATE] market_price_history 에러: {e}")


def _collect_active_models_market_prices():
    """활성 모델(my_bids_local 또는 sales_history 최근)의 사이즈별 시장가 수집 → DB 저장."""
    try:
        local_path = BASE_DIR / 'my_bids_local.json'
        active_models = []
        if local_path.exists():
            try:
                local = json.loads(local_path.read_text(encoding='utf-8'))
                active_models = list({(b.get('model') or '').strip()
                                      for b in (local.get('bids') or [])
                                      if b.get('model')})
            except Exception:
                active_models = []

        if not active_models:
            try:
                conn = sqlite3.connect(str(PRICE_DB))
                c = conn.cursor()
                c.execute("SELECT DISTINCT model FROM sales_history WHERE model IS NOT NULL ORDER BY trade_date DESC LIMIT 10")
                active_models = [r[0] for r in c.fetchall() if r[0]]
                conn.close()
            except Exception:
                active_models = []

        if not active_models:
            print("[MARKET-COLLECT] 수집할 모델 없음")
            return

        import requests as _rq
        now_iso = datetime.now().isoformat()
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        success = 0
        for model in active_models[:30]:
            try:
                r = _rq.post('http://localhost:5001/api/search', json={'model': model}, timeout=60)
                if r.status_code != 200:
                    continue
                d = r.json()
                sizes = d.get('sizes') or d.get('size_prices') or []
                for sz in sizes:
                    c.execute("""
                        INSERT INTO market_price_history
                        (model, size, buy_price, recent_price, collected_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        model,
                        sz.get('size'),
                        sz.get('buy_price') or sz.get('buyPrice'),
                        sz.get('recent_price') or sz.get('recentPrice'),
                        now_iso,
                    ))
                success += 1
            except Exception as e:
                print(f"[MARKET-COLLECT] {model} 실패: {e}")
        conn.commit()
        conn.close()
        print(f"[MARKET-COLLECT] {success}/{len(active_models)} 모델 수집")
    except Exception as e:
        print(f"[MARKET-COLLECT] 에러: {e}")


@app.route('/api/market/history/<path:model>', methods=['GET'])
def api_market_history(model):
    """모델별 시장가 추이 조회."""
    try:
        days = int(request.args.get('days', 30))
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            SELECT model, size, buy_price, recent_price, collected_at
            FROM market_price_history
            WHERE model = ?
              AND datetime(collected_at) >= datetime('now', ?)
            ORDER BY collected_at ASC
        """, (model, f'-{days} days'))
        rows = c.fetchall()
        conn.close()

        history = [
            {'model': r[0], 'size': r[1], 'buy_price': r[2], 'recent_price': r[3], 'collected_at': r[4]}
            for r in rows
        ]
        return jsonify({'ok': True, 'model': model, 'days': days, 'count': len(history), 'history': history})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/market/alerts', methods=['GET'])
def api_market_alerts():
    """최근 24h 대비 직전 24h buy_price 변동률 ±10% 이상 알림."""
    try:
        threshold = float(request.args.get('threshold', 10))
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("""
            SELECT model, size, buy_price, collected_at FROM market_price_history
            WHERE datetime(collected_at) >= datetime('now', '-2 days')
              AND buy_price IS NOT NULL
            ORDER BY model, size, collected_at ASC
        """)
        rows = c.fetchall()
        conn.close()

        from collections import defaultdict
        groups = defaultdict(list)
        for r in rows:
            groups[(r[0], r[1] or '')].append((r[2], r[3]))

        alerts = []
        for (model, size), pts in groups.items():
            if len(pts) < 2:
                continue
            first_price = pts[0][0]
            last_price = pts[-1][0]
            if not first_price or not last_price:
                continue
            pct = (last_price - first_price) / first_price * 100.0
            if abs(pct) >= threshold:
                alerts.append({
                    'model': model,
                    'size': size,
                    'from_price': first_price,
                    'to_price': last_price,
                    'change_pct': round(pct, 2),
                    'from_at': pts[0][1],
                    'to_at': pts[-1][1],
                    'direction': 'UP' if pct > 0 else 'DOWN',
                })
        alerts.sort(key=lambda x: -abs(x['change_pct']))
        return jsonify({'ok': True, 'threshold_pct': threshold, 'count': len(alerts), 'alerts': alerts})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/market/collect-now', methods=['POST'])
def api_market_collect_now():
    """수동 즉시 수집 트리거 (백그라운드 실행)."""
    try:
        threading.Thread(target=_collect_active_models_market_prices, daemon=True).start()
        return jsonify({'ok': True, 'message': '백그라운드 수집 시작'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 33-A: 자동 재로그인 + 알림 통계 API
# ─────────────────────────────────────────────
@app.route('/api/auth/relogin-status', methods=['GET'])
def api_relogin_status():
    from pathlib import Path
    state_file = Path(__file__).parent / '.relogin_state.json'
    state = {}
    if state_file.exists():
        try:
            state = json.loads(state_file.read_text())
        except Exception:
            pass
    auth_path = Path(__file__).parent / 'auth_state.json'
    auth_mtime = None
    if auth_path.exists():
        from datetime import datetime
        auth_mtime = datetime.fromtimestamp(auth_path.stat().st_mtime).isoformat()
    return jsonify({
        'ok': True,
        'auth_state_updated_at': auth_mtime,
        'last_attempt': state.get('last_attempt'),
        'last_success': state.get('last_success'),
        'last_failure': state.get('last_failure'),
        'last_failure_reason': state.get('last_failure_reason'),
    })


@app.route('/api/auth/relogin-now', methods=['POST'])
def api_relogin_now():
    import threading as _th
    def run():
        from pathlib import Path
        state_file = Path(__file__).parent / '.relogin_state.json'
        if state_file.exists():
            try:
                state = json.loads(state_file.read_text())
                state.pop('last_attempt', None)
                state_file.write_text(json.dumps(state))
            except Exception:
                pass
        _check_session_and_relogin()
    _th.Thread(target=run, daemon=True).start()
    return jsonify({'ok': True, 'note': '백그라운드 재로그인 시작'})


@app.route('/api/notifications/stats', methods=['GET'])
def api_notifications_stats():
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("SELECT type, COUNT(*) as cnt, MAX(created_at) as latest FROM notifications WHERE datetime(created_at) > datetime('now', '-7 days') GROUP BY type ORDER BY cnt DESC")
        rows = c.fetchall()
        conn.close()
        return jsonify({'ok': True, 'period_days': 7, 'by_type': [{'type': r[0], 'count': r[1], 'latest': r[2]} for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/notifications/cleanup-old', methods=['POST'])
def api_notifications_cleanup():
    try:
        conn = sqlite3.connect(str(PRICE_DB))
        c = conn.cursor()
        c.execute("DELETE FROM notifications WHERE datetime(created_at) < datetime('now', '-30 days')")
        deleted = c.rowcount
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ============================================================
# Step 42: 송금 환율 시스템 API
# ============================================================

@app.route('/api/remittance/add', methods=['POST'])
def api_remittance_add():
    """송금 이력 등록.
    body: {remittance_date, amount_cny, amount_krw, supplier?, wechat_id?, fee_krw?, notes?}
    """
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        required = ['remittance_date', 'amount_cny', 'amount_krw']
        for k in required:
            if k not in data:
                return jsonify({'success': False, 'error': f'missing {k}'}), 400

        result = remittance_svc.add_remittance(
            remittance_date=data['remittance_date'],
            amount_cny=float(data['amount_cny']),
            amount_krw=float(data['amount_krw']),
            supplier=data.get('supplier'),
            wechat_id=data.get('wechat_id'),
            fee_krw=float(data.get('fee_krw', 0)),
            notes=data.get('notes'),
        )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/list', methods=['GET'])
def api_remittance_list():
    """송금 이력 목록.
    query: ?limit=50&status=active|depleted|cancelled
    """
    try:
        from services import remittance as remittance_svc
        limit = int(request.args.get('limit', 50))
        status = request.args.get('status')
        items = remittance_svc.list_remittances(limit=limit, status=status)
        summary = remittance_svc.get_summary()
        return jsonify({'success': True, 'items': items, 'summary': summary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/match', methods=['POST'])
def api_remittance_match():
    """매칭 실행.
    body:
      - 수동: {remittance_id, bid_cost_id, order_id?, allocated_cny?}
      - FIFO 자동: {auto_fifo: true, max_matches?}
      - 협력사 인지 (Step 43-3): {auto_supplier: true, supplier_id?: int, max_matches?}
    """
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        if data.get('auto_fifo'):
            result = remittance_svc.auto_match_fifo(
                max_matches=int(data.get('max_matches', 100))
            )
        elif data.get('auto_supplier'):
            result = remittance_svc.auto_match_supplier_aware(
                supplier_id=int(data['supplier_id']) if data.get('supplier_id') else None,
                max_matches=int(data.get('max_matches', 100))
            )
        else:
            if 'remittance_id' not in data or 'bid_cost_id' not in data:
                return jsonify({'success': False,
                                'error': 'remittance_id and bid_cost_id required'}), 400
            result = remittance_svc.match_bid_to_remittance(
                remittance_id=int(data['remittance_id']),
                bid_cost_id=int(data['bid_cost_id']),
                order_id=data.get('order_id'),
                allocated_cny=float(data['allocated_cny']) if data.get('allocated_cny') else None,
                method='manual',
            )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/unmatched-bids', methods=['GET'])
def api_remittance_unmatched_bids():
    """매칭 안 된 bid_cost 목록."""
    try:
        from services import remittance as remittance_svc
        items = remittance_svc.get_unmatched_bids()
        return jsonify({
            'success': True,
            'items': items,
            'count': len(items),
            'total_cny': round(sum(b['cny_price'] - b['matched_cny'] for b in items), 2),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 43-2: 환율 손익 대시보드 (FX P&L)
# ═══════════════════════════════════════════

@app.route('/api/fx-pnl/portfolio', methods=['GET'])
def api_fx_pnl_portfolio():
    """전체 포트폴리오 환율 손익."""
    try:
        from services import fx_pnl as fx_pnl_svc
        result = fx_pnl_svc.calculate_portfolio_fx_pnl()
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fx-pnl/bid/<order_id>', methods=['GET'])
def api_fx_pnl_bid(order_id):
    """특정 입찰의 환율 손익."""
    try:
        from services import fx_pnl as fx_pnl_svc
        bid_cost_id = int(request.args.get('bid_cost_id', 0))
        result = fx_pnl_svc.calculate_fx_pnl_for_bid(bid_cost_id, order_id)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fx-pnl/supplier-comparison', methods=['GET'])
def api_fx_pnl_supplier_comparison():
    """협력사별 평균 환율 비교."""
    try:
        from services import fx_pnl as fx_pnl_svc
        items = fx_pnl_svc.supplier_fx_comparison()
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fx-pnl/monthly', methods=['GET'])
def api_fx_pnl_monthly():
    """월별 송금 통계 (Step 43-6)."""
    try:
        from services import fx_pnl as fx_pnl_svc
        items = fx_pnl_svc.monthly_remittance_stats()
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fx-pnl/trends', methods=['GET'])
def api_fx_pnl_trends():
    """최근 N일 송금 추세 (Step 43-6)."""
    try:
        from services import fx_pnl as fx_pnl_svc
        days = int(request.args.get('days', 90))
        result = fx_pnl_svc.remittance_trends(days)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 43-7: model_price_book CSV 일괄 입력
# ═══════════════════════════════════════════

@app.route('/api/price-book/bulk-upload', methods=['POST'])
def api_price_book_bulk_upload():
    """CSV 일괄 업로드. body: {csv_text} 또는 multipart file."""
    try:
        from services import price_book as price_book_svc
        if 'file' in request.files:
            csv_text = request.files['file'].read().decode('utf-8-sig')
        else:
            data = request.get_json() or {}
            csv_text = data.get('csv_text', '')

        if not csv_text:
            return jsonify({'success': False, 'error': 'csv_text or file required'}), 400

        result = price_book_svc.bulk_upsert_from_csv(csv_text)
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 43-8: bid_cost 단가 불일치 감지
# ═══════════════════════════════════════════

@app.route('/api/price-book/anomalies', methods=['GET'])
def api_price_book_anomalies():
    """bid_cost와 model_price_book 단가 차이 ±N% 이상 탐지."""
    try:
        from services import price_book as price_book_svc
        threshold = float(request.args.get('threshold', 20))
        items = price_book_svc.detect_bid_cost_anomalies(threshold)
        return jsonify({'success': True, 'items': items, 'count': len(items), 'threshold_pct': threshold})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 43-4: 인보이스번호 추적
# ═══════════════════════════════════════════

@app.route('/api/remittance/<int:rid>/invoice', methods=['POST'])
def api_remittance_link_invoice(rid):
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        if 'invoice_no' not in data:
            return jsonify({'success': False, 'error': 'invoice_no required'}), 400
        result = remittance_svc.link_invoice(
            rid,
            invoice_no=data['invoice_no'],
            invoice_date=data.get('invoice_date'),
            invoice_amount_usd=float(data['invoice_amount_usd']) if data.get('invoice_amount_usd') else None,
            invoice_amount_cny=float(data['invoice_amount_cny']) if data.get('invoice_amount_cny') else None,
            description=data.get('description'),
        )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/invoices', methods=['GET'])
def api_remittance_invoices(rid):
    try:
        from services import remittance as remittance_svc
        items = remittance_svc.list_invoices(rid)
        return jsonify({'success': True, 'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/invoice/search', methods=['GET'])
def api_invoice_search():
    """?q=JPUSD-202604"""
    try:
        from services import remittance as remittance_svc
        q = request.args.get('q', '').strip()
        if not q:
            return jsonify({'success': False, 'error': 'query required'}), 400
        items = remittance_svc.find_by_invoice(q)
        return jsonify({'success': True, 'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 43-5: 매칭 해제 / 송금 취소 API
# ═══════════════════════════════════════════

@app.route('/api/remittance/match/<int:match_id>', methods=['DELETE'])
def api_unmatch(match_id):
    try:
        from services import remittance as remittance_svc
        result = remittance_svc.unmatch(match_id)
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/cancel', methods=['POST'])
def api_cancel_remittance(rid):
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        reason = data.get('reason', '')
        result = remittance_svc.cancel_remittance(rid, reason)
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/matches', methods=['GET'])
def api_remittance_matches(rid):
    try:
        from services import remittance as remittance_svc
        items = remittance_svc.list_matches(rid)
        return jsonify({'success': True, 'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# Step 42-Phase 2.5: 영수증 + USD/CNY 분리 + 협력사
# ═══════════════════════════════════════════

@app.route('/api/remittance/upload-receipt', methods=['POST'])
def api_remittance_upload_receipt():
    """
    영수증 파일 업로드 (multipart/form-data).
    files: receipt
    form: transaction_no (optional, 파일명 prefix용)
    응답: {success, path, sha256, size_bytes, original_name}
    """
    try:
        from services import remittance as remittance_svc
        if 'receipt' not in request.files:
            return jsonify({'success': False, 'error': 'no file uploaded'}), 400

        f = request.files['receipt']
        if not f.filename:
            return jsonify({'success': False, 'error': 'empty filename'}), 400

        # 확장자 화이트리스트
        allowed_ext = {'.png', '.jpg', '.jpeg', '.pdf', '.heic', '.webp'}
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in allowed_ext:
            return jsonify({'success': False,
                            'error': f'extension not allowed: {ext}'}), 400

        # 크기 제한 (10MB)
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size > 10 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'file too large (max 10MB)'}), 400

        # 임시 저장 후 services로 위임
        original_name = secure_filename(f.filename) or 'receipt'
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        transaction_no = request.form.get('transaction_no')
        try:
            result = remittance_svc.save_receipt_file(tmp_path, original_name, transaction_no)
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

        return jsonify(result), (200 if result['success'] else 500)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/add-v2', methods=['POST'])
def api_remittance_add_v2():
    """
    USD/CNY 분리 + 영수증 메타 포함 v2 등록.
    body: {
      remittance_date, received_cny, amount_krw,
      send_currency?, send_amount?, send_fx_rate?,
      sender_service?, transaction_no?, supplier_id?,
      supplier?, wechat_id?, fee_krw?, notes?,
      receipt_path?, receipt_original_name?, receipt_sha256?
    }
    """
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        for k in ('remittance_date', 'received_cny', 'amount_krw'):
            if k not in data:
                return jsonify({'success': False, 'error': f'missing {k}'}), 400

        result = remittance_svc.add_remittance_v2(
            remittance_date=data['remittance_date'],
            received_cny=float(data['received_cny']),
            amount_krw=float(data['amount_krw']),
            send_currency=data.get('send_currency', 'CNY'),
            send_amount=float(data['send_amount']) if data.get('send_amount') else None,
            send_fx_rate=float(data['send_fx_rate']) if data.get('send_fx_rate') else None,
            sender_service=data.get('sender_service'),
            transaction_no=data.get('transaction_no'),
            supplier_id=int(data['supplier_id']) if data.get('supplier_id') else None,
            supplier=data.get('supplier'),
            wechat_id=data.get('wechat_id'),
            fee_krw=float(data.get('fee_krw', 0)),
            notes=data.get('notes'),
            receipt_path=data.get('receipt_path'),
            receipt_original_name=data.get('receipt_original_name'),
            receipt_sha256=data.get('receipt_sha256'),
        )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/update-cny', methods=['POST'])
def api_remittance_update_cny(rid):
    """USD 송금 후 협력사 입금 CNY 확인 시 호출. body: {received_cny}"""
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        if 'received_cny' not in data:
            return jsonify({'success': False, 'error': 'received_cny required'}), 400
        result = remittance_svc.update_received_cny(rid, float(data['received_cny']))
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/receipt', methods=['GET'])
def api_remittance_get_receipt(rid):
    """저장된 영수증 파일 다운로드/표시."""
    try:
        from services import remittance as remittance_svc
        rem = remittance_svc.get_remittance(rid)
        if not rem:
            return jsonify({'success': False, 'error': 'not found'}), 404
        if not rem.get('receipt_path'):
            return jsonify({'success': False, 'error': 'no receipt'}), 404

        full_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            rem['receipt_path']
        )
        if not os.path.exists(full_path):
            return jsonify({'success': False, 'error': 'file missing on disk'}), 404

        return send_file(full_path, as_attachment=False)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/verify', methods=['GET'])
def api_remittance_verify(rid):
    """영수증 무결성(SHA256) 검증."""
    try:
        from services import remittance as remittance_svc
        result = remittance_svc.verify_receipt_integrity(rid)
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/receipts', methods=['GET'])
def api_remittance_receipts(rid):
    """송금에 첨부된 모든 영수증 목록."""
    try:
        from services import remittance as remittance_svc
        items = remittance_svc.list_receipts(rid)
        return jsonify({'success': True, 'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/remittance/<int:rid>/attach-receipt', methods=['POST'])
def api_remittance_attach_receipt(rid):
    """
    이미 업로드된 영수증을 송금에 추가 첨부.
    body: {receipt_path, original_name?, sha256?, size_bytes?, receipt_type?, description?}

    워크플로우:
    1. /api/remittance/upload-receipt 로 파일 업로드 → path/sha256 받음
    2. 이 API로 receipt_type + description과 함께 첨부
    """
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        if 'receipt_path' not in data:
            return jsonify({'success': False, 'error': 'receipt_path required'}), 400

        result = remittance_svc.attach_receipt(
            remittance_id=rid,
            receipt_path=data['receipt_path'],
            original_name=data.get('original_name'),
            sha256=data.get('sha256'),
            size_bytes=int(data['size_bytes']) if data.get('size_bytes') else None,
            receipt_type=data.get('receipt_type', 'other'),
            description=data.get('description'),
        )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/receipt/<int:receipt_id>', methods=['GET'])
def api_receipt_get(receipt_id):
    """영수증 파일 다운로드 (ID 기반)."""
    try:
        from services import remittance as remittance_svc
        r = remittance_svc.get_receipt(receipt_id)
        if not r:
            return jsonify({'success': False, 'error': 'not found'}), 404

        full_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            r['receipt_path']
        )
        if not os.path.exists(full_path):
            return jsonify({'success': False, 'error': 'file missing on disk'}), 404

        from flask import send_file
        return send_file(full_path, as_attachment=False)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/receipt/<int:receipt_id>', methods=['DELETE'])
def api_receipt_delete(receipt_id):
    """영수증 메타 삭제 (파일은 보존)."""
    try:
        from services import remittance as remittance_svc
        result = remittance_svc.delete_receipt(receipt_id)
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/supplier/list', methods=['GET'])
def api_supplier_list():
    try:
        from services import remittance as remittance_svc
        return jsonify({'success': True, 'items': remittance_svc.list_suppliers()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auto-rebid/realtime-stats', methods=['GET'])
def api_auto_rebid_realtime():
    try:
        from services import rebid_monitor as rebid_monitor_svc
        hours = int(request.args.get('hours', 24))
        result = rebid_monitor_svc.realtime_stats(hours)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auto-rebid/model-stats', methods=['GET'])
def api_auto_rebid_model_stats():
    try:
        from services import rebid_monitor as rebid_monitor_svc
        hours = int(request.args.get('hours', 168))
        items = rebid_monitor_svc.model_stats(hours)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auto-rebid/skip-reasons', methods=['GET'])
def api_auto_rebid_skip_reasons():
    try:
        from services import rebid_monitor as rebid_monitor_svc
        hours = int(request.args.get('hours', 24))
        items = rebid_monitor_svc.skip_reasons(hours)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/auto-rebid/recent-executions', methods=['GET'])
def api_auto_rebid_recent():
    try:
        from services import rebid_monitor as rebid_monitor_svc
        limit = int(request.args.get('limit', 50))
        items = rebid_monitor_svc.recent_executions(limit)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/supplier/add', methods=['POST'])
def api_supplier_add():
    try:
        from services import remittance as remittance_svc
        data = request.get_json() or {}
        if 'name' not in data:
            return jsonify({'success': False, 'error': 'name required'}), 400
        result = remittance_svc.add_supplier(
            name=data['name'],
            name_en=data.get('name_en'),
            wechat_id=data.get('wechat_id'),
            bank_account=data.get('bank_account'),
            default_currency=data.get('default_currency', 'CNY'),
            notes=data.get('notes'),
        )
        return jsonify(result), (200 if result['success'] else 400)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════
# 실행
# ═══════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 50)
    print("  KREAM 판매자 대시보드 서버")
    print("  http://localhost:5001")
    print(f"  모니터링 스케줄: 매일 {MONITOR_HOURS}시")
    print(f"  판매 수집: 30분 간격 + ±5분 지터")
    print("=" * 50)
    # 환경 자동 감지 (Step 18-A): kream.co.kr 접근 가능 여부 1회 체크
    try:
        detect_environment()
    except Exception as _e:
        print(f"[ENV] detect_environment 실패: {_e}")
    # 서버 시작 시 환율 자동 조회 (백그라운드)
    threading.Thread(target=fetch_exchange_rates, daemon=True).start()
    # 모니터링 자동 시작
    monitor_state["running"] = True
    _schedule_next_monitor()
    # 판매 수집 스케줄러 자동 시작
    sales_scheduler_state["running"] = True
    _schedule_next_sales_sync()
    # 헬스체크 경보 모니터링 (5분 간격)
    _health_alert_timer = threading.Timer(60, _health_alert_check)  # 서버 시작 1분 후 첫 실행
    _health_alert_timer.daemon = True
    _health_alert_timer.start()
    print("  경보 모니터링: 5분 간격")
    # 자동 백업 스케줄러 (24h 주기, 첫 실행 60초 후 — 단기 시뮬)
    backup_state["running"] = True
    _backup_timer = threading.Timer(60, _backup_tick)
    _backup_timer.daemon = True
    _backup_timer.start()
    print("  자동 백업: 60초 후 첫 실행 → 이후 24시간 주기")
    # 세션 사전 갱신 스케줄러 (Step 17-D Phase 2-B)
    start_session_refresh_scheduler()
    # Step 18-D: APScheduler 작업 등록 + 시작
    _register_step18d_jobs()
    # Step 21: capital_history 마이그레이션 + 스냅샷 스케줄러 등록
    _migrate_capital_history()
    if scheduler is not None:
        try:
            scheduler.add_job(
                _snapshot_capital,
                'interval',
                hours=1,
                id='capital_snapshot',
                replace_existing=True,
                misfire_grace_time=600,
            )
            print("[SCHEDULER] capital_snapshot 등록 (1시간 간격)")
        except Exception as _ce:
            print(f"[SCHEDULER] capital_snapshot 등록 실패: {_ce}")
    # 첫 스냅샷 즉시 실행 (서버 재시작 시 데이터 누적)
    try:
        _snapshot_capital()
    except Exception as _se:
        print(f"[CAPITAL-SNAPSHOT] 초기 실행 실패: {_se}")
    # Step 31: 시장가 자동 수집 스케줄러 + 마이그레이션
    try:
        _migrate_market_price_history()
    except Exception as _me:
        print(f"[MIGRATE] market_price_history 호출 실패: {_me}")
    if scheduler is not None:
        try:
            scheduler.add_job(
                _collect_active_models_market_prices,
                'interval', hours=2,
                id='market_price_collect',
                replace_existing=True,
                misfire_grace_time=600,
            )
            print("[SCHEDULER] market_price_collect 등록 (2h)")
        except Exception as _me:
            print(f"[SCHEDULER] market_price_collect 실패: {_me}")
    # Step 33-A: 자동 재로그인 체크 (30분 간격)
    if scheduler is not None:
        try:
            scheduler.add_job(
                _check_session_and_relogin,
                'interval', minutes=30,
                id='auto_relogin_check',
                replace_existing=True,
                misfire_grace_time=300,
            )
            print("[SCHEDULER] auto_relogin_check 등록 (30분)")
        except Exception as e:
            print(f"[SCHEDULER] auto_relogin_check 실패: {e}")
    if scheduler is not None:
        try:
            scheduler.start()
            print("[SCHEDULER] APScheduler 시작 (Step 18-D)")
        except Exception as _se:
            print(f"[SCHEDULER] 시작 실패: {_se}")
    app.run(host="0.0.0.0", port=5001, debug=False)

