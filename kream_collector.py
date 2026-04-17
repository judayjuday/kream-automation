"""
KREAM 가격 수집 스크립트
- kream.co.kr 상품 페이지에서 가격 정보 수집
- partner.kream.co.kr 판매자센터에서 입찰 현황 수집
- Playwright (async) + Stealth 기반
- channel="chrome", auth_state.json 세션 사용

사용법:
  python3 kream_collector.py --products 299954
  python3 kream_collector.py --products 299954,300123,301456
  python3 kream_collector.py --products 299954 --save-excel
"""

import asyncio
import argparse
import json
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, Page
from playwright_stealth import Stealth

STATE_FILE_PARTNER = "auth_state.json"
STATE_FILE_KREAM = "auth_state_kream.json"
DB_PATH = Path(__file__).parent / "price_history.db"


def _init_db():
    """가격 이력 DB 초기화 (테이블 생성)"""
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS price_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id TEXT NOT NULL,
        model TEXT,
        size TEXT,
        delivery_type TEXT,
        buy_price INTEGER,
        sell_price INTEGER,
        recent_trade_price INTEGER,
        bid_count INTEGER,
        collected_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS my_bids_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        product_id TEXT,
        model TEXT,
        size TEXT,
        price INTEGER,
        rank INTEGER,
        status TEXT DEFAULT '입찰중',
        recorded_at TEXT NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS competitor_info (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id TEXT NOT NULL,
        size TEXT,
        delivery_type TEXT,
        price INTEGER,
        seller_name TEXT,
        first_seen_at TEXT NOT NULL,
        last_seen_at TEXT NOT NULL,
        price_changes TEXT DEFAULT '[]'
    )""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ph_pid ON price_history(product_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ph_collected ON price_history(collected_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_mb_pid ON my_bids_history(product_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ci_pid ON competitor_info(product_id)")
    conn.commit()
    conn.close()


def save_prices_to_db(product_id, model, size_delivery_prices, recent_trade=None):
    """사이즈×배송타입별 가격을 DB에 저장"""
    if not size_delivery_prices:
        return
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    now = datetime.now().isoformat()
    for sdp in size_delivery_prices:
        size = sdp.get("size", "")
        for dtype, buy_key, sell_key in [
            ("빠른배송", "buyFast", "sellFast"),
            ("일반배송", "buyNormal", "sellNormal"),
            ("해외배송", "buyOverseas", "sellOverseas"),
        ]:
            bp = sdp.get(buy_key)
            sp = sdp.get(sell_key)
            if bp or sp:
                c.execute(
                    "INSERT INTO price_history "
                    "(product_id,model,size,delivery_type,buy_price,sell_price,recent_trade_price,collected_at) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (product_id, model or "", size, dtype, bp, sp, recent_trade, now)
                )
        # 전체 최저가도 별도 저장
        bp_all = sdp.get("buyPrice")
        sp_all = sdp.get("sellPrice")
        if bp_all or sp_all:
            c.execute(
                "INSERT INTO price_history "
                "(product_id,model,size,delivery_type,buy_price,sell_price,recent_trade_price,collected_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (product_id, model or "", size, "최저가", bp_all, sp_all, recent_trade, now)
            )
    conn.commit()
    conn.close()


def save_my_bids_to_db(bids):
    """내 입찰 현황을 DB에 저장"""
    if not bids:
        return
    conn = sqlite3.connect(str(DB_PATH))
    c = conn.cursor()
    now = datetime.now().isoformat()
    for b in bids:
        c.execute(
            "INSERT INTO my_bids_history "
            "(order_id,product_id,model,size,price,rank,status,recorded_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (b.get("orderId"), b.get("productId"), b.get("model", ""),
             b.get("size", ""), b.get("bidPrice"), b.get("bidRank"),
             "입찰중", now)
        )
    conn.commit()
    conn.close()


# DB 초기화 (모듈 로드 시 실행)
_init_db()
EXCEL_OUTPUT = "kream_price_data.xlsx"
KREAM_URL = "https://kream.co.kr"
PARTNER_URL = "https://partner.kream.co.kr"


# ═══════════════════════════════════════════
# 브라우저 설정
# ═══════════════════════════════════════════

async def create_browser(playwright, headless=False):
    return await playwright.chromium.launch(
        channel="chrome",
        headless=headless,
        args=[
            '--disable-blink-features=AutomationControlled',
            '--no-sandbox',
            '--disable-dev-shm-usage',
        ]
    )


async def create_context(browser, storage=None):
    return await browser.new_context(
        storage_state=storage if storage and Path(storage).exists() else None,
        viewport={"width": 1440, "height": 900},
        locale="ko-KR",
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
    )


async def save_state_with_localstorage(page, context, path, origin_url):
    """storage_state에 localStorage 데이터를 병합하여 저장"""
    try:
        local_storage_data = await page.evaluate('() => JSON.stringify(localStorage)')
        ls_items = json.loads(local_storage_data) if local_storage_data else {}
        ls_entries = [{"name": k, "value": v} for k, v in ls_items.items()]
    except Exception:
        ls_entries = []

    state = await context.storage_state()

    if ls_entries:
        origin_found = False
        for origin in state.get("origins", []):
            if origin.get("origin") == origin_url:
                origin["localStorage"] = ls_entries
                origin_found = True
                break
        if not origin_found:
            state.setdefault("origins", []).append({
                "origin": origin_url,
                "localStorage": ls_entries,
            })

    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


async def apply_stealth(page):
    await Stealth().apply_stealth_async(page)


# ═══════════════════════════════════════════
# KREAM 공개 페이지 가격 수집
# ═══════════════════════════════════════════

async def collect_from_kream(page: Page, product_id: str) -> dict:
    """
    kream.co.kr/products/{product_id} 에서 가격 정보 수집
    - 상품명, 영문명, 모델번호
    - 즉시구매가, 즉시판매가, 최근거래가
    - 체결 거래 내역
    - 판매입찰/구매입찰 현황 (로그인 시)
    """
    url = f"{KREAM_URL}/products/{product_id}"
    print(f"\n{'='*50}")
    print(f"  KREAM 수집: 상품 #{product_id}")
    print(f"  URL: {url}")
    print(f"{'='*50}")

    result = {
        "product_id": product_id,
        "product_name": None,
        "product_name_en": None,
        "model_number": None,
        "display_price": None,        # 표시 가격 (메인)
        "instant_buy_price": None,    # 즉시구매가
        "instant_sell_price": None,   # 즉시판매가
        "recent_trade_price": None,   # 최근거래가
        "price_change": None,         # 가격 변동
        "total_trades": None,         # 총 거래수
        "trade_history": [],          # 체결 거래 내역
        "sell_bids": [],              # 판매입찰 현황
        "buy_bids": [],               # 구매입찰 현황
        "collected_at": datetime.now().isoformat(),
        "source": "kream.co.kr",
    }

    # ── API 응답 리스너를 페이지 로드 전에 등록 (reload 없이 캡처) ──
    _api_pre_captured: dict = {}

    async def _on_api_resp(response):
        if 'api.kream.co.kr/api/p/options/display' in response.url:
            try:
                body = await response.json()
                key = 'sell' if 'picker_type=sell' in response.url else 'buy'
                _api_pre_captured[key] = body
                print(f"  [API 사전캡처] picker_type={key}")
            except Exception:
                pass

    page.on('response', _on_api_resp)

    try:
        # 페이지 로드 (재시도 최대 2회)
        for attempt in range(3):
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=15000)
                # 스마트 대기: 상품 콘텐츠 렌더 확인 (최대 3.5초)
                try:
                    await page.wait_for_selector(
                        'p.bold.pc-bold, [class*="product_title"], #Product',
                        timeout=3000
                    )
                    await page.wait_for_timeout(500)
                except Exception:
                    await page.wait_for_timeout(2500)
                break
            except Exception as e:
                if attempt < 2:
                    print(f"  페이지 로드 재시도 ({attempt+1}/2): {e}")
                    await page.wait_for_timeout(1500)
                else:
                    raise

        # 세션 만료 감지
        if "/login" in page.url:
            print(f"  ** KREAM 세션 만료! 재로그인 필요 (kream_bot.py --mode login-kream) **")
            result["session_expired"] = "kream"
            return result

        # ── 1) JSON-LD 스키마에서 상품 기본정보 추출 ──
        schema_data = await page.evaluate("""() => {
            const el = document.getElementById('Product');
            if (el) {
                try { return JSON.parse(el.textContent); }
                catch(e) {}
            }
            return null;
        }""")

        if schema_data:
            result["product_name_en"] = schema_data.get("name")
            desc = schema_data.get("description", "")
            # description 형태: "IX7693 아디다스 올웨이즈 오리지널 데님 숄더백 블루 ..."
            if desc:
                parts = desc.split(" ", 1)
                if len(parts) == 2 and len(parts[0]) <= 20:
                    result["model_number"] = parts[0]
            offers = schema_data.get("offers", {})
            if isinstance(offers, dict):
                schema_price = offers.get("price")
                if schema_price:
                    result["display_price"] = int(schema_price)
                    # JSON-LD offers.price = KREAM 즉시구매가 (판매입찰 최저가, 없으면 최근거래가)
                    # → 이것이 가장 안정적인 즉시구매가 초기값 (API/DOM 실패해도 보장)
                    result["instant_buy_price"] = int(schema_price)
                    print(f"  JSON-LD offers.price → 즉시구매가 초기값: {result['instant_buy_price']}원")
            print(f"  JSON-LD: {result['product_name_en']}")

        # ── 2) DOM에서 상세 정보 추출 ──
        dom_data = await page.evaluate("""() => {
            const r = {};
            const body = document.body.innerText;

            // 메인 가격 (p.bold.pc-bold 첫 번째)
            const priceEls = document.querySelectorAll('p.bold.pc-bold');
            if (priceEls.length > 0) {
                const priceText = priceEls[0].innerText.trim();
                const m = priceText.match(/([0-9,]+)/);
                if (m) r.display_price = parseInt(m[1].replace(/,/g, ''));
            }

            // 상품명 (한글) — 여러 셀렉터 시도
            const nameSelectors = [
                'p.text-element',
                'p.detail_product_title',
                '[class*="product_title"]',
                '[class*="product-title"]',
                'h1',
            ];
            for (const sel of nameSelectors) {
                const els = document.querySelectorAll(sel);
                for (const p of els) {
                    const t = p.innerText.trim();
                    if (/[\uAC00-\uD7AF]/.test(t) && t.length > 4
                        && !t.includes('\uBC30\uC1A1') && !t.includes('KREAM')
                        && !t.includes('\uAC70\uB798') && !t.includes('\uB9AC\uBDF0')
                        && !t.includes('\uACE0\uAC1D')
                        && !t.includes('\uBC1C\uB9E4\uAC00')
                        && !t.includes('\uC218\uC218\uB8CC')
                        && !t.includes('\uC0C1\uD488')
                        && !/^[\d,]+\uC6D0$/.test(t)) {  // "129,000원" 같은 순수 가격 텍스트 제외
                        r.product_name = t;
                        break;
                    }
                }
                if (r.product_name) break;
            }

            // 모델번호
            const modelMatch = body.match(/\uBAA8\uB378\uBC88\uD638\s*([A-Z0-9-]+)/i);
            if (modelMatch) r.model_number = modelMatch[1];

            // 거래수
            const tradeMatch = body.match(/\uAC70\uB798\s*([0-9,]+)/);
            if (tradeMatch) r.total_trades = parseInt(tradeMatch[1].replace(/,/g, ''));

            // 가격변동 (▼/▲ 문자)
            const changeMatch = body.match(/([\u25BC\u25B2\u2193\u2191])\s*([0-9,]+)\uC6D0\(([0-9.]+)%\)/);
            if (changeMatch) {
                r.price_change = {
                    direction: (changeMatch[1] === '\u25BC' || changeMatch[1] === '\u2193') ? 'down' : 'up',
                    amount: parseInt(changeMatch[2].replace(/,/g, '')),
                    percent: parseFloat(changeMatch[3])
                };
            }

            return r;
        }""")

        if dom_data:
            if dom_data.get("display_price"):
                dom_price = dom_data["display_price"]
                result["display_price"] = dom_price
                # p.bold.pc-bold 는 즉시구매가(판매입찰 최저가) 또는 최근거래가를 표시
                # JSON-LD offers.price 와 동일하거나 클 경우에만 즉시구매가로 인정
                # (과거 체결가가 현재 판매입찰보다 낮은 경우 잘못된 값이 들어오는 것 방지)
                json_ld_price = result.get("instant_buy_price")
                if json_ld_price is None:
                    result["instant_buy_price"] = dom_price
                # recent_trade_price 는 체결거래 내역에서 가져오므로 DOM 메인가격과 구분
                # → 여기서는 display_price 만 업데이트, recent_trade_price 는 체결거래 탭 데이터 사용
            if dom_data.get("product_name"):
                pname = dom_data["product_name"]
                # "발매가 $65 (약 96,700원)" 등 발매가 문자열 필터링
                import re as _re
                pname = _re.sub(r'\s*발매가\s*\$?\d[\d,]*\s*(\(약\s*[\d,]+원\))?\s*', '', pname).strip()
                if pname:
                    result["product_name"] = pname
            if dom_data.get("model_number") and not result["model_number"]:
                result["model_number"] = dom_data["model_number"]
            if dom_data.get("total_trades"):
                result["total_trades"] = dom_data["total_trades"]
            if dom_data.get("price_change"):
                result["price_change"] = dom_data["price_change"]

        print(f"  상품명: {result.get('product_name', 'N/A')}")
        print(f"  모델번호: {result.get('model_number', 'N/A')}")
        print(f"  표시가격: {result.get('display_price', 'N/A')}원")
        print(f"  총거래: {result.get('total_trades', 'N/A')}건")

        # ── 3) 체결 거래 내역 수집 ──
        # 먼저 "체결 거래" 탭이 활성화되어 있는지 확인 (기본 활성)
        trade_data = await page.evaluate(r"""() => {
            const trades = [];
            const seen = new Set();

            // visible한 가격 요소만 수집 (체결 거래 탭)
            const priceEls = document.querySelectorAll(
                '.transaction_history_summary__content__item_price'
            );

            for (const el of priceEls) {
                if (!el || el.offsetParent === null) continue;
                const priceText = (el.innerText || '').trim();
                if (!priceText) continue;

                const m = priceText.match(/([0-9,]+)/);
                if (!m) continue;

                const entry = {price: parseInt(m[1].replace(/,/g, ''))};

                const row = el.parentElement;
                if (row) {
                    const children = row.children;
                    for (const child of children) {
                        if (child === el) continue;
                        const t = (child.innerText || '').trim();
                        if (!t) continue;
                        if (/^\d{2}\/\d{2}\/\d{2}$/.test(t)) {
                            entry.date = t;
                        }
                        else if (!/[0-9,]+\uC6D0/.test(t) && t.length < 20) {
                            entry.size = t;
                        }
                    }
                }

                const key = entry.price + '_' + (entry.date || '') + '_' + (entry.size || '');
                if (!seen.has(key)) {
                    seen.add(key);
                    trades.push(entry);
                }
            }

            return trades;
        }""")

        if trade_data:
            result["trade_history"] = trade_data
            print(f"  체결거래: {len(trade_data)}건")
            # 체결거래 최근가 = 첫 번째 항목의 가격 (최신 순 정렬 가정)
            if trade_data and trade_data[0].get("price"):
                result["recent_trade_price"] = trade_data[0]["price"]
                print(f"  최근거래가 (체결거래): {result['recent_trade_price']}원")

        # ── 4) 판매입찰 탭 클릭 → 수집 ──
        # 탭이 2세트 렌더링됨 (hidden + visible). .last로 visible 탭 선택
        try:
            sell_tab = page.locator('a.item_link:has-text("판매 입찰")').last
            await page.evaluate(
                "el => el.scrollIntoView({block:'center'})",
                await sell_tab.element_handle()
            )
            await page.wait_for_timeout(500)
            await sell_tab.click(timeout=5000)
            await page.wait_for_timeout(1500)

            sell_data = await parse_bid_section(page)
            result["sell_bids"] = sell_data
            print(f"  판매입찰: {len(sell_data)}건")

            # 판매입찰 최저가 = 즉시구매가
            if sell_data:
                prices = [b["price"] for b in sell_data if b.get("price")]
                if prices:
                    result["instant_buy_price"] = min(prices)
                    print(f"  즉시구매가 (최저 판매입찰): {result['instant_buy_price']}원")
        except Exception as e:
            print(f"  판매입찰 탭 실패: {e}")

        # ── 5) 구매입찰 탭 클릭 → 수집 ──
        try:
            buy_tab = page.locator('a.item_link:has-text("구매 입찰")').last
            await buy_tab.click(timeout=5000)
            await page.wait_for_timeout(1500)

            buy_data = await parse_bid_section(page)
            result["buy_bids"] = buy_data
            print(f"  구매입찰: {len(buy_data)}건")

            # 구매입찰 최고가 = 즉시판매가
            if buy_data:
                prices = [b["price"] for b in buy_data if b.get("price")]
                if prices:
                    result["instant_sell_price"] = max(prices)
                    print(f"  즉시판매가 (최고 구매입찰): {result['instant_sell_price']}원")
        except Exception as e:
            print(f"  구매입찰 탭 실패: {e}")

        # ── 6) 사이즈별 배송타입별 가격 (KREAM API) ──
        try:
            api_sizes = await collect_size_prices_via_api(
                page, product_id, pre_captured=_api_pre_captured
            )
            if api_sizes:
                result["size_delivery_prices"] = api_sizes
                print(f"  사이즈×배송타입: {len(api_sizes)}개 (API)")

                # API 데이터로 즉시구매가/즉시판매가/판매입찰/구매입찰 재설정
                # picker_type=buy  → buyPrice  = 각 사이즈의 최저 판매입찰가 = 즉시구매가
                # picker_type=sell → sellPrice = 각 사이즈의 최고 구매입찰가 = 즉시판매가
                buy_prices = [s["buyPrice"] for s in api_sizes if s.get("buyPrice") is not None]
                sell_prices = [s["sellPrice"] for s in api_sizes if s.get("sellPrice") is not None]
                if buy_prices:
                    result["instant_buy_price"] = min(buy_prices)
                    print(f"  즉시구매가 (API 재설정): {result['instant_buy_price']}원")
                else:
                    # API에서 buyPrice가 None → 판매입찰 탭 직접 클릭해서 최저가 수집
                    # 체결거래 가격을 즉시구매가로 사용하지 않기 위해 DOM fallback 대신 탭 재클릭
                    print(f"  API buyPrice 없음 → 판매입찰 탭 DOM 재수집")
                    try:
                        sell_tab2 = page.locator('a.item_link:has-text("판매 입찰")').last
                        await page.evaluate(
                            "el => el.scrollIntoView({block:'center'})",
                            await sell_tab2.element_handle()
                        )
                        await page.wait_for_timeout(500)
                        await sell_tab2.click(timeout=5000)
                        await page.wait_for_timeout(1500)

                        sell_data2 = await parse_bid_section(page)
                        if sell_data2:
                            prices2 = [b["price"] for b in sell_data2 if b.get("price")]
                            if prices2:
                                result["instant_buy_price"] = min(prices2)
                                result["sell_bids"] = sell_data2
                                print(f"  즉시구매가 (판매입찰 탭 재수집): {result['instant_buy_price']}원")
                            else:
                                result["instant_buy_price"] = None
                                print(f"  즉시구매 불가 (판매입찰 가격 없음)")
                        else:
                            result["instant_buy_price"] = None
                            print(f"  즉시구매 불가 (판매입찰 없음)")
                    except Exception as e2:
                        print(f"  판매입찰 탭 재수집 실패: {e2}")

                if sell_prices:
                    result["instant_sell_price"] = max(sell_prices)
                    print(f"  즉시판매가 (API 재설정): {result['instant_sell_price']}원")

                # sell_bids / buy_bids 도 API 데이터로 재설정
                # DOM 스크래핑이 체결거래 클래스를 잘못 읽는 문제 방지
                # (사이즈별 최저가 1건씩 → market_low·rank 계산 기준값으로 사용)
                api_sell_bids = [
                    {"price": s["buyPrice"], "size": s["size"], "quantity": 1}
                    for s in api_sizes if s.get("buyPrice") is not None
                ]
                if api_sell_bids:
                    result["sell_bids"] = api_sell_bids
                    print(f"  판매입찰 (API 재설정): {len(api_sell_bids)}건")

                api_buy_bids = [
                    {"price": s["sellPrice"], "size": s["size"], "quantity": 1}
                    for s in api_sizes if s.get("sellPrice") is not None
                ]
                if api_buy_bids:
                    result["buy_bids"] = api_buy_bids
                    print(f"  구매입찰 (API 재설정): {len(api_buy_bids)}건")
        except Exception as e:
            print(f"  API 사이즈 수집 실패: {e}")

        # ── 7) 사이즈별 가격 계산 (sell_bids/buy_bids에서 파생) ──
        sell_bids = result.get("sell_bids", [])
        buy_bids = result.get("buy_bids", [])

        if sell_bids or buy_bids:
            size_map = {}
            for b in sell_bids:
                sz = b.get("size", "ONE SIZE")
                if sz not in size_map:
                    size_map[sz] = {"size": sz, "buyNowPrice": None, "sellNowPrice": None,
                                    "sellBidCount": 0, "buyBidCount": 0}
                size_map[sz]["sellBidCount"] += b.get("quantity", 1)
                p = b["price"]
                if size_map[sz]["buyNowPrice"] is None or p < size_map[sz]["buyNowPrice"]:
                    size_map[sz]["buyNowPrice"] = p

            for b in buy_bids:
                sz = b.get("size", "ONE SIZE")
                if sz not in size_map:
                    size_map[sz] = {"size": sz, "buyNowPrice": None, "sellNowPrice": None,
                                    "sellBidCount": 0, "buyBidCount": 0}
                size_map[sz]["buyBidCount"] += b.get("quantity", 1)
                p = b["price"]
                if size_map[sz]["sellNowPrice"] is None or p > size_map[sz]["sellNowPrice"]:
                    size_map[sz]["sellNowPrice"] = p

            # 사이즈 정렬
            def size_sort_key(s):
                try:
                    return float(s)
                except ValueError:
                    return 0
            result["size_prices"] = sorted(size_map.values(), key=lambda x: size_sort_key(x["size"]))
            print(f"  사이즈별: {len(result['size_prices'])}개")

        # 로그인 필요 메시지 감지
        login_needed = await page.locator('text="모든 시세는 로그인 후 확인 가능합니다"').count()
        if login_needed > 0:
            print(f"  ** 입찰 상세는 kream.co.kr 로그인 필요 **")

        # ── 최종 가격 검증 및 로그 ──
        ibp = result.get("instant_buy_price")
        isp = result.get("instant_sell_price")
        rtp = result.get("recent_trade_price")
        dp  = result.get("display_price")
        print(f"\n  ── 최종 가격 요약 ──")
        print(f"  즉시구매가 : {ibp}원  (출처: JSON-LD→API→DOM 우선순위)")
        print(f"  즉시판매가 : {isp}원")
        print(f"  최근거래가 : {rtp}원")
        print(f"  display    : {dp}원")
        print(f"  판매입찰   : {[b['price'] for b in result.get('sell_bids', [])]}")
        print(f"  구매입찰   : {[b['price'] for b in result.get('buy_bids', [])]}")

    except Exception as e:
        print(f"  ERROR: {e}")
    finally:
        page.remove_listener('response', _on_api_resp)

    # 가격 이력 DB 저장
    try:
        sdp = result.get("size_delivery_prices", [])
        if sdp:
            save_prices_to_db(
                product_id,
                result.get("model_number", ""),
                sdp,
                recent_trade=result.get("recent_trade_price"),
            )
            print(f"  💾 가격 이력 DB 저장: {len(sdp)}사이즈")
    except Exception as e:
        print(f"  ⚠ DB 저장 실패: {e}")

    return result


async def collect_size_prices_via_api(page: Page, product_id: str, pre_captured: dict = None) -> list:
    """KREAM 내부 API로 사이즈별 배송타입별 가격 수집 (빠름)
    - pre_captured: 이미 캡처된 dict가 있으면 reload 생략
    - 없으면 페이지 reload로 API 재호출 (fallback)
    """
    result = []
    captured = {}

    # pre_captured가 있으면 reload 없이 바로 사용
    if pre_captured is not None:
        captured = pre_captured
        if not captured:
            print(f"  [API] 사전캡처 없음 → reload fallback")
        else:
            print(f"  [API] 사전캡처 사용 (reload 생략): {list(captured.keys())}")

    if not captured:
        # fallback: reload로 API 재캡처 (wait 단축: 4000ms → 2500ms)
        async def on_resp(response):
            if 'api.kream.co.kr/api/p/options/display' in response.url:
                try:
                    body = await response.json()
                    key = 'sell' if 'picker_type=sell' in response.url else 'buy'
                    captured[key] = body
                    print(f"  [API 캡처] picker_type={key}")
                except Exception:
                    pass

        page.on('response', on_resp)
        await page.reload(wait_until="domcontentloaded")
        await page.wait_for_timeout(2500)
        page.remove_listener('response', on_resp)

    if not captured:
        print(f"  [API] API 캡처 실패 (URL 불일치 또는 미호출)")

    for picker_type in ["buy", "sell"]:
        try:
            resp = captured.get(picker_type)
            if not resp:
                continue

            items = resp.get("content", {}).get("items", [])
            print(f"  [API] picker_type={picker_type} → {len(items)}개 사이즈")

            for item in items:
                size_id = str(item.get("id", ""))
                # 사이즈 텍스트
                size_text = size_id
                ti = item.get("title_item", {}).get("text_element", {})
                dv = ti.get("default_variation", {})
                if isinstance(dv, dict) and dv.get("text"):
                    size_text = dv["text"]

                # 전체 가격 (description_item)
                di = item.get("description_item", {}).get("text_element", {})
                ddv = di.get("default_variation", {})
                overall_price = None
                if isinstance(ddv, dict) and ddv.get("text"):
                    pt = ddv["text"].replace(",", "").replace("원", "").strip()
                    if pt.isdigit():
                        overall_price = int(pt)

                # 배송타입별 가격 (associated_item)
                ai = item.get("associated_item", {})
                ai_items = ai.get("items", [])

                delivery_prices = {}
                for sub in ai_items:
                    sub_items = sub.get("items", [])
                    for si in sub_items:
                        dt_text = ""
                        sti = si.get("title_item", {}).get("text_element", {})
                        sdv = sti.get("default_variation", {})
                        if isinstance(sdv, dict):
                            dt_text = sdv.get("text", "")

                        price_text = ""
                        sdi = si.get("description_item", {}).get("text_element", {})
                        sddv = sdi.get("default_variation", {})
                        if isinstance(sddv, dict):
                            price_text = sddv.get("text", "")

                        price_val = None
                        if price_text:
                            pt = price_text.replace(",", "").replace("원", "").strip()
                            if pt.isdigit():
                                price_val = int(pt)

                        if "빠른" in dt_text:
                            delivery_prices["fast"] = price_val
                        elif "일반" in dt_text:
                            delivery_prices["normal"] = price_val
                        elif "해외" in dt_text:
                            delivery_prices["overseas"] = price_val

                # 기존 result에 merge
                existing = next((r for r in result if r["size"] == size_text), None)
                if not existing:
                    existing = {
                        "size": size_text,
                        "buyPrice": None, "sellPrice": None,
                        "buyFast": None, "buyNormal": None, "buyOverseas": None,
                        "sellFast": None, "sellNormal": None, "sellOverseas": None,
                    }
                    result.append(existing)

                if picker_type == "buy":
                    # buyPrice: 모든 배송타입(빠른/일반/해외) 중 최저가
                    all_prices = [p for p in [
                        delivery_prices.get("fast"),
                        delivery_prices.get("normal"),
                        delivery_prices.get("overseas"),
                    ] if p]
                    if all_prices:
                        existing["buyPrice"] = min(all_prices)
                    elif overall_price is not None:
                        existing["buyPrice"] = overall_price
                    existing["buyFast"] = delivery_prices.get("fast")
                    existing["buyNormal"] = delivery_prices.get("normal")
                    existing["buyOverseas"] = delivery_prices.get("overseas")
                else:
                    existing["sellPrice"] = overall_price
                    existing["sellFast"] = delivery_prices.get("fast")
                    existing["sellNormal"] = delivery_prices.get("normal")
                    existing["sellOverseas"] = delivery_prices.get("overseas")

        except Exception as e:
            print(f"  API 수집 오류 ({picker_type}): {e}")

    # 정렬
    def sort_key(s):
        try:
            return float(s["size"])
        except (ValueError, TypeError):
            return 0
    result.sort(key=sort_key)

    if result:
        print(f"  [API] 사이즈별 가격: {[(r['size'], r['buyPrice'], r['sellPrice']) for r in result]}")

    return result


async def parse_bid_section(page: Page) -> list:
    """현재 활성화된 입찰 탭에서 가격/수량 파싱
    테이블 구조: 옵션(사이즈) | 판매희망가/구매희망가 | 수량

    주의: transaction_history_summary__content__item_price 클래스는
    체결 거래 섹션과 공유되므로 사용 불가. 현재 활성 탭의 가시 영역만 스캔.
    """
    bids = []

    try:
        data = await page.evaluate(r"""() => {
            const results = [];

            // 현재 뷰포트에 실제로 렌더링된 가격 요소를 찾는다.
            // getBoundingClientRect()로 실제 가시성 확인 (opacity/transform 숨김도 처리)
            const priceEls = document.querySelectorAll(
                '[class*="transaction_history_summary__content__item_price"]'
            );

            for (const el of priceEls) {
                if (!el) continue;

                // offsetParent 대신 getBoundingClientRect로 실제 가시성 확인
                const rect = el.getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;

                // 조상 중에 display:none 또는 visibility:hidden인 요소가 있으면 스킵
                let hidden = false;
                let node = el;
                while (node && node !== document.body) {
                    const st = window.getComputedStyle(node);
                    if (st.display === 'none' || st.visibility === 'hidden' || st.opacity === '0') {
                        hidden = true;
                        break;
                    }
                    node = node.parentElement;
                }
                if (hidden) continue;

                const priceText = (el.innerText || '').trim();
                if (!priceText) continue;

                const m = priceText.match(/([0-9,]+)/);
                if (!m) continue;

                const entry = {price: parseInt(m[1].replace(/,/g, ''))};

                // 부모 행에서 사이즈/수량 추출
                const row = el.parentElement;
                if (row) {
                    const children = Array.from(row.children).filter(c => c !== el);
                    const texts = children.map(c => (c.innerText || '').trim()).filter(t => t);

                    // 체결거래 행 판별: 날짜(YY/MM/DD) sibling이 있으면 체결거래 → 제외
                let hasDate = false;
                for (let i = 0; i < texts.length; i++) {
                        const t = texts[i];
                        if (/^\d{2}\/\d{2}\/\d{2}$/.test(t)) {
                            hasDate = true;
                            break;
                        }
                        if (/[0-9,]+\uC6D0/.test(t)) continue; // 가격 스킵

                        if (!entry.size && (t === 'ONE SIZE' || /^\d{3}(\.\d)?$/.test(t) || /^[A-Z0-9]{1,5}$/.test(t))) {
                            entry.size = t;
                        } else if (/^\d{1,3}$/.test(t) && parseInt(t) < 100) {
                            entry.quantity = parseInt(t);
                        }
                    }
                if (hasDate) continue; // 체결거래 행 제외
                }

                results.push(entry);
            }

            return results;
        }""")

        if data:
            bids = data
    except Exception as e:
        print(f"  입찰 파싱 오류: {e}")

    return bids


# ═══════════════════════════════════════════
# 판매자센터 가격 수집
# ═══════════════════════════════════════════

async def collect_from_partner(page: Page, product_id: str) -> dict:
    """partner.kream.co.kr에서 상품 검색 및 내 입찰 현황 수집"""
    print(f"\n  판매자센터 수집: 상품 #{product_id}")

    result = {
        "product_id": product_id,
        "partner_products": [],
        "my_bids": [],
        "source": "partner.kream.co.kr",
        "collected_at": datetime.now().isoformat(),
    }

    try:
        # 상품 목록에서 검색
        url = (
            f"{PARTNER_URL}/business/products"
            f"?page=1&perPage=10"
            f"&startDate=&endDate="
            f"&keyword={product_id}"
        )
        await page.goto(url, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)

        # 로그인 확인
        if "/sign-in" in page.url:
            print(f"  ** 판매자센터 로그인 필요 (auth_state.json 만료?) **")
            return result

        table_data = await page.evaluate("""() => {
            const results = [];
            const rows = document.querySelectorAll('table tbody tr');
            for (const row of rows) {
                const cells = row.querySelectorAll('td');
                const texts = Array.from(cells).map(c => c.innerText.trim());
                if (texts.length > 0 && texts.some(t => t.length > 0)) {
                    results.push(texts);
                }
            }
            return results;
        }""")

        if table_data:
            result["partner_products"] = table_data
            print(f"  상품목록: {len(table_data)}행")

        # 입찰 현황 페이지
        bid_url = (
            f"{PARTNER_URL}/business/bids"
            f"?page=1&perPage=10"
            f"&keyword={product_id}"
        )
        await page.goto(bid_url, wait_until="domcontentloaded")
        await page.wait_for_timeout(3000)

        bid_data = await page.evaluate("""() => {
            const results = [];
            const rows = document.querySelectorAll('table tbody tr');
            for (const row of rows) {
                const cells = row.querySelectorAll('td');
                const texts = Array.from(cells).map(c => c.innerText.trim());
                if (texts.length > 0 && texts.some(t => t.length > 0)) {
                    results.push(texts);
                }
            }
            return results;
        }""")

        if bid_data:
            result["my_bids"] = bid_data
            print(f"  내 입찰: {len(bid_data)}건")

    except Exception as e:
        print(f"  판매자센터 수집 오류: {e}")

    return result


# ═══════════════════════════════════════════
# 유틸리티
# ═══════════════════════════════════════════

def parse_price(text: str):
    """가격 텍스트에서 숫자만 추출 (예: "79,000원" → 79000)"""
    if not text:
        return None
    cleaned = ''.join(c for c in text if c.isdigit())
    return int(cleaned) if cleaned else None


def save_to_json(data: list, path: str = "kream_prices.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n  JSON 저장: {path}")


def save_to_excel(data: list, path: str = EXCEL_OUTPUT):
    wb = openpyxl.Workbook()

    # 가격 요약
    ws = wb.active
    ws.title = "가격요약"
    ws.append([
        "상품번호", "상품명", "영문명", "모델번호",
        "표시가격", "즉시구매가", "즉시판매가", "최근거래가",
        "총거래수", "판매입찰수", "구매입찰수", "수집시간"
    ])
    for item in data:
        k = item.get("kream", {})
        ws.append([
            k.get("product_id"),
            k.get("product_name"),
            k.get("product_name_en"),
            k.get("model_number"),
            k.get("display_price"),
            k.get("instant_buy_price"),
            k.get("instant_sell_price"),
            k.get("recent_trade_price"),
            k.get("total_trades"),
            len(k.get("sell_bids", [])),
            len(k.get("buy_bids", [])),
            k.get("collected_at"),
        ])

    # 체결 거래
    ws2 = wb.create_sheet("체결거래")
    ws2.append(["상품번호", "사이즈", "가격", "거래일"])
    for item in data:
        k = item.get("kream", {})
        for t in k.get("trade_history", []):
            ws2.append([
                k.get("product_id"),
                t.get("size"),
                t.get("price"),
                t.get("date"),
            ])

    # 판매입찰
    ws3 = wb.create_sheet("판매입찰현황")
    ws3.append(["상품번호", "사이즈", "가격", "수량"])
    for item in data:
        k = item.get("kream", {})
        for b in k.get("sell_bids", []):
            ws3.append([k.get("product_id"), b.get("size"), b.get("price"), b.get("quantity")])

    # 구매입찰
    ws4 = wb.create_sheet("구매입찰현황")
    ws4.append(["상품번호", "사이즈", "가격", "수량"])
    for item in data:
        k = item.get("kream", {})
        for b in k.get("buy_bids", []):
            ws4.append([k.get("product_id"), b.get("size"), b.get("price"), b.get("quantity")])

    wb.save(path)
    print(f"  엑셀 저장: {path}")


# ═══════════════════════════════════════════
# 메인 수집 함수
# ═══════════════════════════════════════════

async def collect_prices(product_ids: list, headless=False, save_excel=False,
                         include_partner=True) -> list:
    """여러 상품의 가격을 한번에 수집"""
    all_results = []

    kream_session = STATE_FILE_KREAM if Path(STATE_FILE_KREAM).exists() else None
    partner_session = STATE_FILE_PARTNER if Path(STATE_FILE_PARTNER).exists() else None

    if not kream_session:
        print("  ** auth_state_kream.json 없음 → 먼저 python3 kream_bot.py --mode login-kream 실행 **")

    async with async_playwright() as p:
        browser = await create_browser(p, headless=headless)

        # KREAM 수집용 컨텍스트 (kream.co.kr 세션)
        kream_context = await create_context(browser, kream_session)
        kream_page = await kream_context.new_page()
        await apply_stealth(kream_page)

        # 판매자센터용 컨텍스트 (partner.kream.co.kr 세션)
        partner_page = None
        partner_context = None
        if include_partner and partner_session:
            partner_context = await create_context(browser, partner_session)
            partner_page = await partner_context.new_page()
            await apply_stealth(partner_page)

        print(f"\n{'='*60}")
        print(f"  KREAM 가격 수집 시작")
        print(f"  대상: {', '.join(product_ids)}")
        print(f"  KREAM 세션: {'있음' if kream_session else '없음 (비로그인)'}")
        print(f"  판매자센터 세션: {'있음' if partner_session else '없음'}")
        print(f"  시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")

        for i, pid in enumerate(product_ids, 1):
            print(f"\n[{i}/{len(product_ids)}] 상품 #{pid}")
            item = {"product_id": pid}

            # KREAM 공개 페이지 (kream.co.kr 세션 사용)
            item["kream"] = await collect_from_kream(kream_page, pid)

            # 판매자센터 (partner.kream.co.kr 세션 사용)
            if include_partner and partner_page:
                item["partner"] = await collect_from_partner(partner_page, pid)

            all_results.append(item)

            if i < len(product_ids):
                await kream_page.wait_for_timeout(2000)

        # 세션 갱신 저장 (localStorage 포함)
        if kream_session:
            await save_state_with_localstorage(kream_page, kream_context, STATE_FILE_KREAM, KREAM_URL)
        if partner_context:
            await save_state_with_localstorage(partner_page, partner_context, STATE_FILE_PARTNER, PARTNER_URL)
        await browser.close()

    # 요약 출력
    print(f"\n{'='*60}")
    print(f"  수집 완료! {len(all_results)}건")
    print(f"{'='*60}")

    for item in all_results:
        k = item.get("kream", {})
        print(f"\n  #{k.get('product_id')}: {k.get('product_name', 'N/A')}")
        print(f"    표시가격:   {k.get('display_price', '-')}원")
        print(f"    즉시구매가: {k.get('instant_buy_price', '-')}원")
        print(f"    즉시판매가: {k.get('instant_sell_price', '-')}원")
        print(f"    최근거래가: {k.get('recent_trade_price', '-')}원")
        print(f"    거래내역:   {len(k.get('trade_history', []))}건")
        print(f"    판매입찰:   {len(k.get('sell_bids', []))}건")
        print(f"    구매입찰:   {len(k.get('buy_bids', []))}건")

    save_to_json(all_results)
    if save_excel:
        save_to_excel(all_results)

    return all_results


# ═══════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════

async def main():
    parser = argparse.ArgumentParser(description="KREAM 가격 수집")
    parser.add_argument("--products", required=True,
                        help="상품번호 (쉼표 구분)")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--save-excel", action="store_true")
    parser.add_argument("--no-partner", action="store_true",
                        help="판매자센터 수집 제외")
    args = parser.parse_args()

    product_ids = [pid.strip() for pid in args.products.split(",") if pid.strip()]

    results = await collect_prices(
        product_ids=product_ids,
        headless=args.headless,
        save_excel=args.save_excel,
        include_partner=not args.no_partner,
    )

    print("\n--- JSON ---")
    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    asyncio.run(main())
