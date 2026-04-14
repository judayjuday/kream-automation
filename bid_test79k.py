"""
79000원 2개 입찰 테스트 + 입찰 내역 확인
"""
import asyncio
import os
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

PARTNER_URL = "https://partner.kream.co.kr"
STATE_FILE = "auth_state.json"
EXCEL_PATH = "kream_data_template.xlsx"
SS_DIR = "debug_screenshots"
os.makedirs(SS_DIR, exist_ok=True)
step = 0

async def ss(page, name):
    global step
    step += 1
    f = f"{SS_DIR}/{step:02d}_{name}.png"
    await page.screenshot(path=f, full_page=False)
    print(f"  📸 [{step}] {name}")


async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["입찰데이터"]
    headers = [cell.value for cell in ws[1]]
    bid = dict(zip(headers, [cell.value for cell in ws[2]]))
    wb.close()

    product_id = str(bid["product_id"])
    price = str(int(bid["입찰가격"]))
    qty = int(bid["수량"])
    print(f"📋 입찰: #{product_id}, {price}원, {qty}개")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            channel="chrome", headless=False,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        context = await browser.new_context(
            storage_state=STATE_FILE if Path(STATE_FILE).exists() else None,
            viewport={"width": 1440, "height": 900}, locale="ko-KR",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await context.new_page()
        await Stealth().apply_stealth_async(page)

        # 로그인 확인
        await page.goto(f"{PARTNER_URL}/c2c")
        await page.wait_for_timeout(3000)
        if "/sign-in" in page.url:
            print("❌ 로그인 필요!")
            await browser.close()
            return
        print("✅ 로그인 OK")

        # 팝업 닫기
        for _ in range(5):
            await page.wait_for_timeout(500)
            try:
                cb = page.locator('text="다시 보지 않기"')
                if await cb.is_visible(timeout=300): await cb.click()
            except: pass
            try:
                btn = page.locator('button:has-text("확인")').first
                if await btn.is_visible(timeout=300): await btn.click(); continue
            except: pass
            try:
                x = page.locator('button:has-text("✕"), button:has-text("×")').first
                if await x.is_visible(timeout=300): await x.click(); continue
            except: pass
            break

        # ── 1. 상품 검색 ──
        print("\n[1] 상품 검색...")
        await page.goto(
            f"{PARTNER_URL}/business/products?page=1&perPage=10"
            f"&startDate=&endDate=&keyword={product_id}"
            f"&categoryId=&brandId=&productId=&sort="
        )
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        try:
            all_tab = page.locator('button:has-text("전체")').first
            if await all_tab.is_visible(timeout=1000):
                await all_tab.click()
                await page.wait_for_timeout(1500)
        except: pass

        # ── 2. "판매 입찰하기" 버튼 ──
        print("[2] '판매 입찰하기' 클릭...")
        bid_buttons = page.locator('button:has-text("판매 입찰하기")')
        if await bid_buttons.count() == 0:
            print("  ❌ 버튼 없음")
            await browser.close()
            return
        await bid_buttons.first.click()
        await page.wait_for_timeout(2000)

        try:
            if await page.locator('text="상품 정보 입력 필요"').is_visible(timeout=1000):
                print("  ⚠ 고시정보 필요")
                await browser.close()
                return
        except: pass

        # ── 3. 옵션/수량 모달 ──
        print("[3] 옵션/수량 모달...")
        try:
            await page.locator('text="옵션/수량 선택"').wait_for(timeout=5000)
        except:
            print("  ❌ 모달 안 열림")
            await ss(page, "modal_fail")
            await browser.close()
            return

        # + 버튼으로 수량 설정 (qty번 클릭)
        plus_btn = page.locator('[class*="Counter_plus"]')
        if await plus_btn.count() == 0:
            plus_btn = page.locator('button:has-text("+")').first
        for i in range(qty):
            await plus_btn.first.click()
            await page.wait_for_timeout(300)
        print(f"  ✅ 수량 {qty}개")
        await ss(page, "quantity_set")

        # ── 4. "판매 입찰 계속" ──
        print("[4] '판매 입찰 계속'...")
        await page.locator('button:has-text("판매 입찰 계속")').click(timeout=3000)
        await page.wait_for_timeout(2000)
        await ss(page, "after_continue")

        # ── 5. 가격 입력 ──
        print("[5] 가격 입력...")
        price_inputs = page.locator('input[placeholder*="판매 희망가"]')
        if await price_inputs.count() == 0:
            price_inputs = page.locator('input[placeholder*="희망가"]')

        input_count = await price_inputs.count()
        print(f"  가격 입력 필드: {input_count}개")

        for i in range(input_count):
            inp = price_inputs.nth(i)
            await inp.click()
            await page.keyboard.press("Meta+a")
            await page.keyboard.press("Backspace")
            await inp.type(price, delay=50)
            await page.wait_for_timeout(300)

        # blur로 React 값 반영
        await page.keyboard.press("Tab")
        await page.wait_for_timeout(1000)
        print(f"  ✅ {price}원 × {input_count}개 입력")
        await ss(page, "after_price")

        # ── 6. 체크박스 (부모 클릭 방식) ──
        print("[6] 체크박스 선택...")
        checkboxes = page.locator('input[type="checkbox"]')
        cnt = await checkboxes.count()
        checked_count = 0

        for i in range(cnt):
            cb = checkboxes.nth(i)
            if not await cb.is_visible():
                continue
            if await cb.is_checked():
                checked_count += 1
                continue
            parent = page.locator(f'input[type="checkbox"] >> nth={i} >> xpath=..')
            try:
                await parent.click(timeout=2000)
                await page.wait_for_timeout(300)
                if await cb.is_checked():
                    checked_count += 1
            except:
                try:
                    await cb.click(force=True)
                    await page.wait_for_timeout(300)
                    if await cb.is_checked():
                        checked_count += 1
                except: pass

        await page.wait_for_timeout(500)
        print(f"  ✅ 체크됨: {checked_count}개")

        # 총 N건 확인
        try:
            total = await page.locator('text=/총.*건/').first.inner_text()
            print(f"  📊 {total}")
        except: pass
        await ss(page, "after_checkbox")

        # ── 7. "판매 입찰하기" 최종 ──
        print("[7] '판매 입찰하기' 최종 클릭...")
        final_btn = page.locator('button:has-text("판매 입찰하기")').last
        enabled = await final_btn.is_enabled()
        print(f"  enabled={enabled}")
        await final_btn.click(timeout=3000)
        await page.wait_for_timeout(2000)
        await ss(page, "after_final_click")

        # ── 8. 확인 팝업 ──
        print("[8] 확인 팝업...")
        await page.wait_for_timeout(1000)

        # 팝업 텍스트 확인
        try:
            body = await page.locator('body').inner_text()
            for line in body.split('\n'):
                l = line.strip()
                if '총' in l and '건' in l and '입찰' in l:
                    print(f"  팝업: '{l}'")
        except: pass

        await ss(page, "confirm_popup")

        confirm_btns = page.locator('button:has-text("확인")')
        ccnt = await confirm_btns.count()
        if ccnt > 0:
            for i in range(ccnt):
                if await confirm_btns.nth(i).is_visible():
                    await confirm_btns.nth(i).click()
                    print("  ✅ 확인 클릭")
                    break
            await page.wait_for_timeout(3000)
        else:
            print("  ❌ 확인 버튼 없음")

        await ss(page, "after_confirm")

        # ── 9. 결과 ──
        print("[9] 결과...")
        try:
            await page.locator('text="입찰 신청 결과"').wait_for(timeout=5000)
            result_text = await page.locator('text=/성공.*건/').first.inner_text()
            print(f"  🎉 {result_text}")
            # 결과 팝업 닫기
            ok = page.locator('button:has-text("확인")')
            if await ok.count() > 0:
                for i in range(await ok.count()):
                    if await ok.nth(i).is_visible():
                        await ok.nth(i).click()
                        break
                await page.wait_for_timeout(1000)
        except:
            print("  ⚠ 결과 팝업 없음")

        await ss(page, "result")

        # ── 10. 입찰 내역 확인 ──
        print("\n[10] 입찰 내역 확인 (partner.kream.co.kr/business/asks)...")
        await page.goto(f"{PARTNER_URL}/business/asks")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        await ss(page, "asks_page")

        # 79,000원 검색
        price_els = page.locator('text="79,000원"')
        pcnt = await price_els.count()
        print(f"  '79,000원' 발견: {pcnt}건")

        if pcnt >= 2:
            print(f"  🎉🎉🎉 79,000원 {pcnt}건 확인 완료! 성공!")
        elif pcnt == 1:
            print(f"  ⚠ 79,000원 1건만 보임 (2건 기대)")
        else:
            print(f"  ❌ 79,000원 입찰 안 보임")
            # 79000 형태도 확인
            p2 = await page.locator('text="79000"').count()
            print(f"  '79000' 텍스트: {p2}건")

        # 페이지 스크롤해서 더 확인
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        await page.wait_for_timeout(1000)
        await ss(page, "asks_page_scrolled")

        await context.storage_state(path=STATE_FILE)
        print("\n✅ 테스트 완료")
        await page.wait_for_timeout(5000)
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())
