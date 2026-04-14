"""
판매 입찰 디버그 스크립트 - 각 단계마다 스크린샷 촬영
"""
import asyncio
import os
from pathlib import Path
from datetime import datetime

import openpyxl
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

PARTNER_URL = "https://partner.kream.co.kr"
STATE_FILE = "auth_state.json"
EXCEL_PATH = "kream_data_template.xlsx"
SCREENSHOT_DIR = "debug_screenshots"

os.makedirs(SCREENSHOT_DIR, exist_ok=True)

step_num = 0

async def screenshot(page, name):
    global step_num
    step_num += 1
    fname = f"{SCREENSHOT_DIR}/{step_num:02d}_{name}.png"
    await page.screenshot(path=fname, full_page=False)
    print(f"  📸 [{step_num}] {name} → {fname}")
    return fname


async def main():
    global step_num
    step_num = 0

    # 엑셀에서 입찰 데이터 로드
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["입찰데이터"]
    headers = [cell.value for cell in ws[1]]
    bid = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            bid = dict(zip(headers, row))
            break
    wb.close()

    if not bid:
        print("❌ 입찰 데이터 없음")
        return

    product_id = str(bid["product_id"])
    price = str(int(bid.get("입찰가격", 0)))
    qty = int(bid.get("수량", 1))
    print(f"📋 입찰 대상: 상품 #{product_id}, {price}원, {qty}개")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            channel="chrome",
            headless=False,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        context = await browser.new_context(
            storage_state=STATE_FILE if Path(STATE_FILE).exists() else None,
            viewport={"width": 1440, "height": 900},
            locale="ko-KR",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        )
        page = await context.new_page()
        await Stealth().apply_stealth_async(page)

        # ── 0. 로그인 확인 ──
        print("\n[STEP 0] 로그인 확인...")
        await page.goto(f"{PARTNER_URL}/c2c")
        await page.wait_for_timeout(3000)
        await screenshot(page, "login_check")

        if "/sign-in" in page.url:
            print("❌ 로그인 필요! 먼저 python3 kream_bot.py --mode login 실행")
            await browser.close()
            return
        print("  ✅ 로그인 OK")

        # 팝업 닫기
        for i in range(5):
            await page.wait_for_timeout(500)
            try:
                cb = page.locator('text="다시 보지 않기"')
                if await cb.is_visible(timeout=300):
                    await cb.click()
            except: pass
            try:
                btn = page.locator('button:has-text("확인")').first
                if await btn.is_visible(timeout=300):
                    await btn.click()
                    continue
            except: pass
            try:
                x = page.locator('button:has-text("✕"), button:has-text("×")').first
                if await x.is_visible(timeout=300):
                    await x.click()
                    continue
            except: pass
            break

        # ── 1. 상품 검색 페이지 이동 ──
        print("\n[STEP 1] 상품 검색 페이지 이동...")
        bid_url = (
            f"{PARTNER_URL}/business/products"
            f"?page=1&perPage=10&startDate=&endDate="
            f"&keyword={product_id}"
            f"&categoryId=&brandId=&productId=&sort="
        )
        await page.goto(bid_url)
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        await screenshot(page, "product_search_page")

        # "전체" 기간 탭
        try:
            all_tab = page.locator('button:has-text("전체")').first
            if await all_tab.is_visible(timeout=1000):
                await all_tab.click()
                await page.wait_for_timeout(1500)
                await screenshot(page, "after_all_tab_click")
        except:
            pass

        # ── 2. "판매 입찰하기" 버튼 찾기 ──
        print("\n[STEP 2] '판매 입찰하기' 버튼 찾기...")
        bid_buttons = page.locator('button:has-text("판매 입찰하기")')
        count = await bid_buttons.count()
        print(f"  발견된 버튼 수: {count}")
        await screenshot(page, "before_bid_button_click")

        if count == 0:
            print("  ❌ '판매 입찰하기' 버튼 없음!")
            # 페이지에 있는 모든 버튼 텍스트 확인
            all_buttons = page.locator('button')
            btn_count = await all_buttons.count()
            print(f"  페이지 버튼 총 {btn_count}개:")
            for i in range(min(btn_count, 20)):
                try:
                    txt = await all_buttons.nth(i).inner_text()
                    print(f"    [{i}] '{txt.strip()[:50]}'")
                except: pass
            await browser.close()
            return

        # 버튼 클릭
        print("  → 버튼 클릭...")
        await bid_buttons.first.click()
        await page.wait_for_timeout(2000)
        await screenshot(page, "after_bid_button_click")

        # 상품 정보 입력 필요 팝업 체크
        try:
            info_needed = page.locator('text="상품 정보 입력 필요"')
            if await info_needed.is_visible(timeout=1000):
                print("  ⚠ 고시정보 미입력 팝업 감지!")
                await screenshot(page, "product_info_needed_popup")
                await page.locator('button:has-text("취소")').first.click()
                await browser.close()
                return
        except: pass

        # ── 3. 옵션/수량 선택 모달 ──
        print("\n[STEP 3] 옵션/수량 선택 모달...")
        try:
            await page.locator('text="옵션/수량 선택"').wait_for(timeout=5000)
            print("  ✅ 모달 열림")
        except:
            print("  ❌ 모달 안 열림!")
            await screenshot(page, "modal_not_opened")
            # 현재 보이는 모든 텍스트 확인
            body_text = await page.locator('body').inner_text()
            print(f"  페이지 텍스트 (처음 500자): {body_text[:500]}")
            await browser.close()
            return

        await screenshot(page, "option_quantity_modal")

        # + 버튼으로 수량 설정
        print(f"  수량 {qty}개 설정 중...")
        plus_btn = page.locator('[class*="Counter_plus"]')
        pcnt = await plus_btn.count()
        print(f"  Counter_plus 버튼: {pcnt}개")
        if pcnt == 0:
            plus_btn = page.locator('button:has-text("+")').first
            pcnt = await plus_btn.count()
            print(f"  + 텍스트 버튼: {pcnt}개")

        for i in range(qty):
            await plus_btn.first.click()
            await page.wait_for_timeout(300)
        await screenshot(page, "after_quantity_set")

        # ── 4. "판매 입찰 계속" 클릭 ──
        print("\n[STEP 4] '판매 입찰 계속' 클릭...")
        continue_btn = page.locator('button:has-text("판매 입찰 계속")')
        ccnt = await continue_btn.count()
        print(f"  '판매 입찰 계속' 버튼: {ccnt}개")
        await screenshot(page, "before_continue_click")

        if ccnt == 0:
            print("  ❌ '판매 입찰 계속' 버튼 없음!")
            all_buttons = page.locator('button')
            btn_count = await all_buttons.count()
            for i in range(min(btn_count, 20)):
                try:
                    txt = await all_buttons.nth(i).inner_text()
                    vis = await all_buttons.nth(i).is_visible()
                    print(f"    [{i}] '{txt.strip()[:50]}' visible={vis}")
                except: pass
            await browser.close()
            return

        try:
            await continue_btn.click(timeout=3000)
            await page.wait_for_timeout(2000)
            print("  ✅ 클릭 성공")
        except Exception as e:
            print(f"  ❌ 클릭 실패: {e}")
            await screenshot(page, "continue_click_failed")
            await browser.close()
            return

        await screenshot(page, "after_continue_click")

        # ── 5. 판매 희망가 입력 ──
        print("\n[STEP 5] 판매 희망가 입력...")
        price_inputs = page.locator('input[placeholder*="판매 희망가"]')
        input_count = await price_inputs.count()
        print(f"  '판매 희망가' placeholder 입력 필드: {input_count}개")

        if input_count == 0:
            price_inputs = page.locator('input[placeholder*="희망가"]')
            input_count = await price_inputs.count()
            print(f"  '희망가' placeholder 입력 필드: {input_count}개")

        if input_count == 0:
            # 모든 input 필드 탐색
            all_inputs = page.locator('input')
            acnt = await all_inputs.count()
            print(f"  페이지 총 input: {acnt}개")
            for i in range(min(acnt, 20)):
                try:
                    ph = await all_inputs.nth(i).get_attribute("placeholder")
                    tp = await all_inputs.nth(i).get_attribute("type")
                    nm = await all_inputs.nth(i).get_attribute("name")
                    vis = await all_inputs.nth(i).is_visible()
                    print(f"    [{i}] type={tp} name={nm} placeholder={ph} visible={vis}")
                except: pass
            await screenshot(page, "no_price_input")
            await browser.close()
            return

        await screenshot(page, "before_price_input")

        try:
            for i in range(input_count):
                inp = price_inputs.nth(i)
                await inp.click()
                await page.keyboard.press("Meta+a")
                await page.keyboard.press("Backspace")
                await inp.type(price, delay=50)
                await page.wait_for_timeout(300)
            print(f"  ✅ {price}원 입력 완료 ({input_count}개 필드)")
        except Exception as e:
            print(f"  ❌ 가격 입력 실패: {e}")
            await screenshot(page, "price_input_failed")
            await browser.close()
            return

        await screenshot(page, "after_price_input")

        # ── 6. 체크박스 선택 ──
        print("\n[STEP 6] 체크박스 선택...")

        # 다양한 체크박스 locator 시도
        cb_locators = [
            ('[data-sentry-element="BaseCheckbox"]', 'BaseCheckbox'),
            ('input[type="checkbox"]', 'input checkbox'),
            ('[class*="checkbox" i]', 'class contains checkbox'),
            ('[class*="Checkbox" i]', 'class contains Checkbox'),
            ('[role="checkbox"]', 'role checkbox'),
        ]

        found_cb = None
        for selector, label in cb_locators:
            loc = page.locator(selector)
            cnt = await loc.count()
            print(f"  {label} ({selector}): {cnt}개")
            if cnt > 0 and found_cb is None:
                found_cb = (loc, cnt, label)

        await screenshot(page, "before_checkbox")

        if found_cb:
            loc, cnt, label = found_cb
            print(f"  → {label} 사용 ({cnt}개)")
            for i in range(cnt):
                cb = loc.nth(i)
                try:
                    # 방법 1: JS evaluate로 React state 토글
                    await cb.evaluate("""(el) => {
                        const input = el.tagName === 'INPUT' ? el : el.querySelector('input[type="checkbox"]');
                        if (!input) return;
                        const nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'checked'
                        ).set;
                        nativeInputValueSetter.call(input, true);
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new MouseEvent('click', { bubbles: true }));
                    }""")
                    await page.wait_for_timeout(300)
                except Exception as e:
                    print(f"    JS 방법 실패, 직접 클릭 시도: {e}")
                    try:
                        await cb.click(force=True)
                        await page.wait_for_timeout(300)
                    except Exception as e2:
                        print(f"    클릭도 실패: {e2}")

            await page.wait_for_timeout(500)
            await screenshot(page, "after_checkbox")

            # 체크박스 상태 확인
            for i in range(cnt):
                cb = loc.nth(i)
                try:
                    checked = await cb.evaluate("""(el) => {
                        const input = el.tagName === 'INPUT' ? el : el.querySelector('input[type="checkbox"]');
                        return input ? input.checked : 'no input found';
                    }""")
                    print(f"    체크박스[{i}] checked={checked}")
                except Exception as e:
                    print(f"    체크박스[{i}] 상태 확인 실패: {e}")
        else:
            print("  ⚠ 체크박스를 찾지 못함")
            await screenshot(page, "no_checkbox_found")

        # ── 7. "판매 입찰하기" 최종 버튼 ──
        print("\n[STEP 7] '판매 입찰하기' 최종 버튼 클릭...")
        final_btns = page.locator('button:has-text("판매 입찰하기")')
        fcnt = await final_btns.count()
        print(f"  '판매 입찰하기' 버튼: {fcnt}개")
        for i in range(fcnt):
            try:
                txt = await final_btns.nth(i).inner_text()
                vis = await final_btns.nth(i).is_visible()
                enabled = await final_btns.nth(i).is_enabled()
                cls = await final_btns.nth(i).get_attribute("class")
                print(f"    [{i}] text='{txt.strip()}' visible={vis} enabled={enabled} class={cls[:80] if cls else 'None'}")
            except: pass

        await screenshot(page, "before_final_bid_click")

        if fcnt == 0:
            print("  ❌ 최종 버튼 없음!")
            await browser.close()
            return

        final_btn = final_btns.last
        try:
            is_enabled = await final_btn.is_enabled()
            print(f"  최종 버튼 enabled={is_enabled}")
            if not is_enabled:
                print("  ❌ 버튼이 비활성화 상태!")
                await screenshot(page, "final_btn_disabled")
                # 왜 비활성화인지 확인 - 체크박스 다시 확인
                print("  체크박스 재확인...")
                cbs_all = page.locator('input[type="checkbox"]')
                for i in range(await cbs_all.count()):
                    chk = await cbs_all.nth(i).is_checked()
                    print(f"    checkbox[{i}] checked={chk}")

            await final_btn.click(timeout=3000)
            await page.wait_for_timeout(2000)
            print("  ✅ 클릭 성공")
        except Exception as e:
            print(f"  ❌ 클릭 실패: {e}")
            await screenshot(page, "final_bid_click_failed")
            await browser.close()
            return

        await screenshot(page, "after_final_bid_click")

        # ── 8. 확인 팝업 ──
        print("\n[STEP 8] 확인 팝업 ('총 N건의 판매 입찰하기')...")
        await page.wait_for_timeout(1000)

        # 팝업/모달 텍스트 확인
        try:
            body_text = await page.locator('body').inner_text()
            # "총" 포함 텍스트 찾기
            for line in body_text.split('\n'):
                if '총' in line or '입찰' in line or '확인' in line:
                    print(f"  텍스트: '{line.strip()[:80]}'")
        except: pass

        await screenshot(page, "confirmation_popup")

        confirm_btn = page.locator('button:has-text("확인")')
        ccnt = await confirm_btn.count()
        print(f"  '확인' 버튼: {ccnt}개")

        if ccnt == 0:
            print("  ❌ 확인 버튼 없음! 팝업이 나타나지 않았을 수 있음")
            await browser.close()
            return

        for i in range(ccnt):
            try:
                txt = await confirm_btn.nth(i).inner_text()
                vis = await confirm_btn.nth(i).is_visible()
                print(f"    [{i}] text='{txt.strip()}' visible={vis}")
            except: pass

        try:
            # visible한 확인 버튼 찾기
            for i in range(ccnt):
                if await confirm_btn.nth(i).is_visible():
                    await confirm_btn.nth(i).click()
                    print(f"  ✅ 확인 버튼[{i}] 클릭 완료")
                    break
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"  ❌ 확인 클릭 실패: {e}")

        await screenshot(page, "after_confirm_click")

        # ── 9. 결과 확인 ──
        print("\n[STEP 9] 입찰 결과 확인...")
        try:
            result = page.locator('text="입찰 신청 결과"')
            await result.wait_for(timeout=5000)
            print("  ✅ '입찰 신청 결과' 텍스트 발견")
            result_text = await page.locator('text=/성공.*건/').first.inner_text()
            print(f"  🎉 결과: {result_text}")
        except:
            print("  ⚠ '입찰 신청 결과' 없음")

        await screenshot(page, "final_result")

        # 현재 페이지 URL 확인
        print(f"\n  최종 URL: {page.url}")

        # ── 10. 입찰 내역 페이지에서 확인 ──
        print("\n[STEP 10] 입찰 내역 페이지에서 78000원 확인...")
        await page.goto(f"{PARTNER_URL}/business/asks")
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(3000)
        await screenshot(page, "asks_page")

        # 78000 텍스트 검색
        try:
            price_text = page.locator('text="78,000"')
            pcnt = await price_text.count()
            print(f"  '78,000' 텍스트: {pcnt}개 발견")
            if pcnt > 0:
                print("  🎉🎉🎉 78,000원 입찰 등록 확인됨!")
            else:
                print("  ❌ 78,000원 입찰이 보이지 않음")
                # "78000" 형태도 확인
                p2 = page.locator('text="78000"')
                p2cnt = await p2.count()
                print(f"  '78000' 텍스트: {p2cnt}개")
        except Exception as e:
            print(f"  확인 실패: {e}")

        # 세션 저장
        await context.storage_state(path=STATE_FILE)
        print(f"\n📸 스크린샷 {step_num}장 저장: {SCREENSHOT_DIR}/")
        print("✅ 디버그 완료")

        # 브라우저 10초 대기 후 닫기 (결과 확인용)
        await page.wait_for_timeout(10000)
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
