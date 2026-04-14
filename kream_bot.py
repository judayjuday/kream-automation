"""
KREAM 판매자센터 자동화 스크립트 v2
- 상품 고시정보 자동 입력
- 판매 입찰 자동 등록
- Stealth 모드 (봇 감지 우회)
- Playwright (async) 기반

사용법:
  pip3 install playwright openpyxl
  python3 -m playwright install chromium
  
  python3 kream_bot.py --mode login       # 첫 로그인 (세션 저장)
  python3 kream_bot.py --mode product     # 상품 고시정보 입력
  python3 kream_bot.py --mode bid         # 판매 입찰
  python3 kream_bot.py --mode all         # 고시정보 + 입찰 한번에
"""

import asyncio
import argparse
import re
import sys
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, Page, BrowserContext
from playwright_stealth import Stealth

EXCEL_PATH = "kream_data_template.xlsx"
PARTNER_URL = "https://partner.kream.co.kr"
KREAM_URL = "https://kream.co.kr"
STATE_FILE = "auth_state.json"
STATE_FILE_KREAM = "auth_state_kream.json"


# ═══════════════════════════════════════════
# 엑셀 데이터 로드
# ═══════════════════════════════════════════

def load_product_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["상품정보"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def load_bid_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["입찰데이터"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def load_settings(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["설정"]
    settings = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            settings[row[0]] = row[1]
    wb.close()
    return settings


# ═══════════════════════════════════════════
# Stealth 모드 (봇 감지 우회)
# ═══════════════════════════════════════════

async def apply_stealth(page):
    """playwright-stealth 패키지로 봇 감지 우회"""
    await Stealth().apply_stealth_async(page)


# ═══════════════════════════════════════════
# React Input 헬퍼
# ═══════════════════════════════════════════

async def react_clear_and_fill(page, selector, value):
    if value is None or str(value).strip() == "":
        return
    value = str(value).strip()
    el = page.locator(selector)
    await el.click()
    await page.keyboard.press("Meta+a")
    await page.keyboard.press("Backspace")
    await page.wait_for_timeout(100)
    await el.type(value, delay=30)
    await page.wait_for_timeout(200)


async def select_dropdown(page, button_selector, option_text):
    if not option_text or str(option_text).strip() == "":
        return
    option_text = str(option_text).strip()

    btn = page.locator(button_selector)
    try:
        await btn.wait_for(state="attached", timeout=10000)
    except Exception:
        print(f"  ⚠ 드롭다운 버튼 '{button_selector}' DOM에 없음")
        return

    try:
        await btn.scroll_into_view_if_needed()
        await page.wait_for_timeout(300)
        await btn.click(timeout=5000)
    except Exception:
        # JS 클릭 폴백
        try:
            await btn.evaluate("el => el.click()")
        except Exception:
            print(f"  ⚠ 드롭다운 버튼 '{button_selector}' 클릭 실패")
            return

    await page.wait_for_timeout(500)
    option = page.get_by_text(option_text, exact=False).first
    try:
        await option.click(timeout=3000)
    except Exception:
        print(f"  ⚠ 드롭다운 옵션 '{option_text}' 찾기 실패")
        await page.keyboard.press("Escape")


# ═══════════════════════════════════════════
# 브라우저 & 로그인
# ═══════════════════════════════════════════

async def create_browser(playwright, headless=False):
    browser = await playwright.chromium.launch(channel="chrome",
        headless=headless,
        args=[
            '--disable-blink-features=AutomationControlled',
            '--no-sandbox',
            '--disable-dev-shm-usage',
        ]
    )
    return browser


async def create_context(browser, storage=None):
    context = await browser.new_context(
        storage_state=storage if storage and Path(storage).exists() else None,
        viewport={"width": 1440, "height": 900},
        locale="ko-KR",
        user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    )
    return context


async def login_manual(playwright):
    print("🔐 판매자센터 로그인 모드")
    print("  브라우저가 열리면 직접 로그인해주세요.")
    print("  로그인 완료 후 터미널에서 Enter를 눌러주세요.\n")

    browser = await create_browser(playwright, headless=False)
    context = await create_context(browser)
    page = await context.new_page()
    await apply_stealth(page)
    await page.goto(f"{PARTNER_URL}/sign-in")
    await page.wait_for_load_state("networkidle")

    input("\n✅ 로그인 완료되면 Enter를 누르세요...")

    await context.storage_state(path=STATE_FILE)
    print(f"✅ 로그인 상태 저장 완료 → {STATE_FILE}")
    await browser.close()


async def login_kream(playwright):
    print("🔐 KREAM 일반 사이트 로그인 모드")
    print("  브라우저에서 직접 로그인해주세요.")
    print("  로그인 완료 후 터미널에서 Enter를 눌러주세요.\n")

    browser = await create_browser(playwright, headless=False)
    context = await create_context(browser)
    page = await context.new_page()
    await apply_stealth(page)
    await page.goto(KREAM_URL)
    await page.wait_for_load_state("domcontentloaded")
    await page.wait_for_timeout(2000)

    input("\n✅ 로그인 완료되면 Enter를 누르세요...")

    await context.storage_state(path=STATE_FILE_KREAM)
    print(f"✅ KREAM 로그인 상태 저장 완료 → {STATE_FILE_KREAM}")
    await browser.close()


async def ensure_logged_in(page):
    print(f"🔐 로그인 상태 확인 중... → {PARTNER_URL}/c2c")
    try:
        await page.goto(f"{PARTNER_URL}/c2c", wait_until="domcontentloaded")
        await page.wait_for_load_state("networkidle", timeout=15000)
    except Exception as e:
        print(f"  ⚠ 페이지 로드 지연: {e}")
    await page.wait_for_timeout(2000)
    current_url = page.url
    print(f"  → 현재 URL: {current_url}")
    if "/sign-in" in current_url:
        print("❌ 로그인이 필요합니다. 먼저 --mode login 으로 로그인해주세요.")
        return False
    print("✅ 로그인 확인 완료")
    await dismiss_popups(page)
    return True


async def dismiss_popups(page):
    """KREAM 팝업/모달 자동 닫기 (신규 가입자 혜택, 공지사항 등)"""
    print("🔄 팝업 확인 중...")
    max_attempts = 10  # 최대 10개까지 처리

    for i in range(max_attempts):
        await page.wait_for_timeout(300)
        closed = False

        # "다시 보지 않기" 체크박스가 있으면 체크
        try:
            checkbox = page.locator('text="다시 보지 않기"')
            if await checkbox.is_visible(timeout=500):
                await checkbox.click()
        except Exception:
            pass

        # "확인" 버튼 클릭
        try:
            confirm_btn = page.locator('button:has-text("확인")').first
            if await confirm_btn.is_visible(timeout=500):
                await confirm_btn.click()
                await page.wait_for_timeout(500)
                closed = True
                print(f"  ✓ 팝업 {i+1} 닫음 (확인)")
                continue
        except Exception:
            pass

        # X 버튼 클릭 (모달 닫기)
        try:
            close_x = page.locator('button:has-text("✕"), button:has-text("×")').first
            if await close_x.is_visible(timeout=500):
                await close_x.click()
                await page.wait_for_timeout(500)
                closed = True
                print(f"  ✓ 팝업 {i+1} 닫음 (X)")
                continue
        except Exception:
            pass

        # "오늘 하루 안 보기", "오늘 안 봄" 등
        try:
            today_btn = page.locator('button:has-text("오늘"), text="오늘 하루 안 보기"').first
            if await today_btn.is_visible(timeout=500):
                await today_btn.click()
                await page.wait_for_timeout(500)
                closed = True
                print(f"  ✓ 팝업 {i+1} 닫음 (오늘 안 봄)")
                continue
        except Exception:
            pass

        # 더 이상 팝업이 없으면 종료
        if not closed:
            break

    print("  ✓ 팝업 처리 완료")


# ═══════════════════════════════════════════
# 상품 고시정보 입력
# ═══════════════════════════════════════════

async def fill_product_info(page, product, delay=2.0):
    product_id = product["product_id"]
    url = f"{PARTNER_URL}/business/my/products/{product_id}"

    print(f"\n📦 상품 #{product_id} 고시정보 입력 중...")
    print(f"  → URL: {url}")
    await page.goto(url)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(1500)

    # 현재 URL 확인 (리다이렉트 감지)
    current_url = page.url
    print(f"  → 현재 URL: {current_url}")
    if "/sign-in" in current_url:
        print(f"  ❌ 세션 만료 — 로그인 페이지로 리다이렉트됨")
        raise Exception("세션 만료: 로그인 페이지로 리다이렉트됨")

    # 고시정보 섹션이 보일 때까지 스크롤 & 대기
    gosi_section = page.locator('text=상품 고시정보')
    try:
        await gosi_section.wait_for(state="visible", timeout=10000)
        await gosi_section.scroll_into_view_if_needed()
        await page.wait_for_timeout(500)
    except Exception:
        print(f"  ⚠ '상품 고시정보' 섹션을 찾을 수 없음 — 스크롤 시도")
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.5)")
        await page.wait_for_timeout(1000)

    # 페이지에 상품 편집 폼이 있는지 확인
    form_exists = await page.locator('input[name="attributeSet.0.value"]').count() > 0
    if not form_exists:
        # 폼 로딩 대기 (React 하이드레이션 등)
        try:
            await page.locator('input[name="attributeSet.0.value"]').wait_for(state="attached", timeout=10000)
            form_exists = True
        except Exception:
            pass

    if not form_exists:
        print(f"  ⚠ 고시정보 입력 폼을 찾을 수 없음 — 페이지 구조 확인 필요")
        await _save_debug_screenshot(page, f"gosi_{product_id}_no_form")
        # 폼이 없어도 계속 시도

    filled_count = 0

    # 고시 카테고리 (드롭다운)
    category = product.get("고시카테고리")
    if category:
        await select_dropdown(page, 'div[name="categoryName"] button', category)
        print(f"  ✓ 고시카테고리: {category}")
        filled_count += 1

    # attributeSet 필드들
    field_map = {
        "종류":           'input[name="attributeSet.0.value"]',
        "소재":           'input[name="attributeSet.1.value"]',
        "색상":           'input[name="attributeSet.2.value"]',
        "크기":           'input[name="attributeSet.3.value"]',
        "제조자_수입자":   'input[name="attributeSet.4.value"]',
        "제조국":         'input[name="attributeSet.5.value"]',
        "취급시_주의사항": 'input[name="attributeSet.6.value"]',
        "품질보증기준":    'input[name="attributeSet.7.value"]',
        "AS_전화번호":     'input[name="attributeSet.8.value"]',
    }

    for field_name, selector in field_map.items():
        value = product.get(field_name)
        if value and str(value).strip():
            try:
                el = page.locator(selector)
                if await el.count() > 0:
                    await react_clear_and_fill(page, selector, str(value))
                    print(f"  ✓ {field_name}: {str(value)[:40]}")
                    filled_count += 1
                else:
                    print(f"  ⚠ {field_name}: 입력 필드 '{selector}' 미발견")
            except Exception as e:
                print(f"  ⚠ {field_name}: 입력 실패 — {e}")

    # 원산지 (드롭다운)
    origin = product.get("원산지")
    if origin:
        await select_dropdown(page, 'div[name="countryOfOriginId"] button', origin)
        print(f"  ✓ 원산지: {origin}")
        filled_count += 1

    # HS코드 (드롭다운)
    hs_code = product.get("HS코드")
    if hs_code:
        await select_dropdown(page, 'div[name="hsCodeId"] button', str(hs_code))
        print(f"  ✓ HS코드: {hs_code}")
        filled_count += 1

    # 배송 정보
    shipping_map = {
        "상품무게_kg": 'input[name="productWeight"]',
        "박스가로_cm": 'input[name="boxWidth"]',
        "박스세로_cm": 'input[name="boxHeight"]',
        "박스높이_cm": 'input[name="boxDepth"]',
    }

    for field_name, selector in shipping_map.items():
        value = product.get(field_name)
        if value and str(value).strip():
            try:
                await react_clear_and_fill(page, selector, str(value))
                print(f"  ✓ {field_name}: {value}")
                filled_count += 1
            except Exception as e:
                print(f"  ⚠ {field_name}: 입력 실패 — {e}")

    print(f"  → 총 {filled_count}개 필드 입력 완료")

    # 저장하기 버튼 클릭 + 결과 확인
    save_btn = page.locator('button:has-text("저장하기")')
    if await save_btn.is_visible():
        # 저장 전 스크린샷
        await _save_debug_screenshot(page, f"gosi_{product_id}_before_save")

        await save_btn.click()
        print(f"  → '저장하기' 버튼 클릭")
        await page.wait_for_timeout(3000)

        # 저장 결과 확인
        saved = False

        # 성공 토스트/메시지 확인
        try:
            success_msg = page.locator('text=/저장.*완료|수정.*완료|성공/')
            if await success_msg.is_visible(timeout=2000):
                msg_text = await success_msg.first.inner_text()
                print(f"  ✓ 저장 성공 메시지: {msg_text}")
                saved = True
        except Exception:
            pass

        # 에러 메시지 확인
        if not saved:
            try:
                error_msg = page.locator('text=/실패|오류|에러|필수|입력해/, [class*="error"], [class*="alert-danger"]')
                if await error_msg.is_visible(timeout=1000):
                    err_text = await error_msg.first.inner_text()
                    print(f"  ❌ 저장 에러 메시지: {err_text}")
                    await _save_debug_screenshot(page, f"gosi_{product_id}_save_error")
                    raise Exception(f"고시정보 저장 실패: {err_text}")
            except Exception as e:
                if "고시정보 저장 실패" in str(e):
                    raise
                pass

        # 에러도 성공도 없으면 — URL 변화/화면 상태로 간접 확인
        if not saved:
            print(f"  ⚠ 명시적 성공/실패 메시지 없음 — 저장 상태 불확실")
            await _save_debug_screenshot(page, f"gosi_{product_id}_after_save")
            # URL이 바뀌었거나 토스트가 없었으면 일단 진행
            print(f"  💾 상품 #{product_id} 저장 시도 완료 (결과 불확실)")
        else:
            print(f"  💾 상품 #{product_id} 저장 완료!")
    else:
        print(f"  ❌ '저장하기' 버튼을 찾을 수 없음")
        await _save_debug_screenshot(page, f"gosi_{product_id}_no_save_btn")
        raise Exception("저장하기 버튼 미발견")

    await page.wait_for_timeout(delay * 1000)


# ═══════════════════════════════════════════
# 판매 입찰 자동 등록
# ═══════════════════════════════════════════

async def _save_debug_screenshot(page, name):
    """디버그용 스크린샷 저장"""
    try:
        debug_dir = Path("debug_screenshots")
        debug_dir.mkdir(exist_ok=True)
        from datetime import datetime as dt
        ts = dt.now().strftime("%Y%m%d_%H%M%S")
        path = debug_dir / f"{ts}_{name}.png"
        await page.screenshot(path=str(path))
        print(f"  📸 스크린샷: {path}")
    except Exception as e:
        print(f"  ⚠ 스크린샷 실패: {e}")


async def place_bid(page, bid, delay=3.0):
    """
    판매 입찰 플로우:
    1. 상품번호로 URL 직접 접근 (검색 대신 URL 파라미터 사용)
    2. "판매 입찰하기" 버튼 클릭
    3. 사이즈 선택 + 수량 설정
    4. "판매 입찰 계속" 클릭
    5. 판매 희망가 입력
    5-1. 입찰기한 설정 (30/60/90일)
    6. 체크박스 선택 → "판매 입찰하기" 최종 클릭
    7. 결과 확인 (성공/실패 판별)
    """
    product_id = str(bid["product_id"])
    size = str(bid.get("사이즈", "ONE SIZE"))
    price = str(int(bid.get("입찰가격", 0)))
    qty = int(bid.get("수량", 1))
    bid_days = int(bid.get("bid_days", 30))

    print(f"\n💰 입찰: 상품 #{product_id}, 사이즈={size}, {price}원, {qty}개, {bid_days}일")

    # ── 1단계: 상품번호로 검색 (전체 기간 + keyword) ──
    bid_url = (
        f"{PARTNER_URL}/business/products"
        f"?page=1&perPage=10"
        f"&startDate=&endDate="
        f"&keyword={product_id}"
        f"&categoryId=&brandId=&productId=&sort="
    )
    print(f"  → URL: {bid_url}")
    await page.goto(bid_url)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(3000)

    # 현재 URL 확인 (리다이렉트 감지)
    current_url = page.url
    print(f"  → 현재 URL: {current_url}")
    if "/sign-in" in current_url:
        print(f"  ❌ 세션 만료 — 로그인 페이지로 리다이렉트됨")
        await _save_debug_screenshot(page, f"bid_{product_id}_session_expired")
        return False

    # "전체" 기간 탭 클릭 (혹시 안 눌려있으면)
    try:
        all_tab = page.locator('button:has-text("전체")').first
        if await all_tab.is_visible(timeout=1000):
            await all_tab.click()
            await page.wait_for_timeout(1000)
    except Exception:
        pass

    # ── 2단계: "판매 입찰하기" 버튼 클릭 ──
    bid_buttons = page.locator('button:has-text("판매 입찰하기")')
    count = await bid_buttons.count()
    print(f"  → '판매 입찰하기' 버튼 {count}개 발견")
    if count == 0:
        print(f"  ❌ '판매 입찰하기' 버튼 없음 — 상품이 검색되지 않았거나 입찰 불가")
        await _save_debug_screenshot(page, f"bid_{product_id}_no_button")
        return False

    await bid_buttons.first.click()
    await page.wait_for_timeout(2000)

    # "상품 정보 입력 필요" 팝업 체크
    try:
        info_needed = page.locator('text="상품 정보 입력 필요"')
        if await info_needed.is_visible(timeout=1000):
            print(f"  ❌ 고시정보 미입력! 먼저 고시정보 등록 필요")
            await _save_debug_screenshot(page, f"bid_{product_id}_gosi_needed")
            await page.locator('button:has-text("취소")').first.click()
            return False
    except Exception:
        pass

    # ── 3단계: 옵션/수량 선택 모달 ──
    try:
        await page.locator('text="옵션/수량 선택"').wait_for(timeout=5000)
        print(f"  ✓ 옵션/수량 선택 모달 열림")
    except Exception:
        print(f"  ❌ 옵션/수량 선택 모달 안 열림")
        await _save_debug_screenshot(page, f"bid_{product_id}_no_modal")
        return False

    # ── 3-1단계: 사이즈 선택 ──
    if size and size != "ONE SIZE":
        print(f"  → 사이즈 '{size}' 선택 시도...")
        size_selected = False

        # 방법1: 사이즈 버튼/라벨 텍스트 매칭
        size_btns = page.locator(f'button:has-text("{size}"), label:has-text("{size}"), [data-size="{size}"]')
        sc = await size_btns.count()
        print(f"  → 사이즈 버튼 {sc}개 발견 (text 매칭)")
        if sc > 0:
            await size_btns.first.click()
            await page.wait_for_timeout(500)
            size_selected = True
            print(f"  ✓ 사이즈 '{size}' 선택 완료 (버튼)")

        # 방법2: 드롭다운/셀렉트 방식
        if not size_selected:
            try:
                size_select = page.locator('select').first
                if await size_select.count() > 0:
                    options = await size_select.locator('option').all_text_contents()
                    print(f"  → 셀렉트 옵션: {options}")
                    matching = [opt for opt in options if size in opt]
                    if matching:
                        await size_select.select_option(label=matching[0])
                        size_selected = True
                        print(f"  ✓ 사이즈 '{size}' 선택 완료 (select)")
            except Exception:
                pass

        # 방법3: 모달 내 사이즈 목록에서 정확 매칭
        if not size_selected:
            try:
                # 사이즈 칩/옵션 클릭 (partial text match)
                all_items = page.locator('[class*="size"], [class*="option"], [class*="chip"]')
                item_count = await all_items.count()
                for idx in range(item_count):
                    item = all_items.nth(idx)
                    txt = (await item.inner_text()).strip()
                    if size in txt or txt == size:
                        await item.click()
                        await page.wait_for_timeout(500)
                        size_selected = True
                        print(f"  ✓ 사이즈 '{size}' 선택 완료 (class 매칭: '{txt}')")
                        break
            except Exception:
                pass

        if not size_selected:
            print(f"  ⚠ 사이즈 '{size}' 선택 실패 — 기본 사이즈로 진행")
            await _save_debug_screenshot(page, f"bid_{product_id}_size_fail_{size}")
    else:
        print(f"  → ONE SIZE / 사이즈 선택 불필요")

    # ── 3-2단계: 수량 설정 (+ 버튼) ──
    plus_btn = page.locator('[class*="Counter_plus"]')
    if await plus_btn.count() == 0:
        plus_btn = page.locator('button:has-text("+")').first

    # 현재 수량값 확인해서 정확히 맞추기
    try:
        counter_input = page.locator('input[type="number"], [class*="Counter"] input, [class*="counter"] input')
        if await counter_input.count() > 0:
            current_qty = int(await counter_input.first.input_value() or "0")
            clicks_needed = max(0, qty - current_qty)
            print(f"  → 현재 수량: {current_qty}, 목표: {qty}, 클릭 필요: {clicks_needed}")
        else:
            clicks_needed = qty
            print(f"  → 수량 입력 필드 미발견, {qty}번 클릭")
    except Exception:
        clicks_needed = qty

    for i in range(clicks_needed):
        await plus_btn.first.click()
        await page.wait_for_timeout(200)
    print(f"  ✓ 수량 {qty}개 설정 완료")

    # "판매 입찰 계속" 클릭
    continue_btn = page.locator('button:has-text("판매 입찰 계속")')
    try:
        await continue_btn.click(timeout=3000)
        await page.wait_for_timeout(2000)
        print(f"  ✓ '판매 입찰 계속' 클릭")
    except Exception:
        print(f"  ❌ '판매 입찰 계속' 버튼 실패")
        await _save_debug_screenshot(page, f"bid_{product_id}_continue_fail")
        return False

    # ── 4단계: 판매 희망가 입력 (수량만큼 input이 여러 개일 수 있음) ──
    price_inputs = page.locator('input[placeholder*="판매 희망가"]')
    input_count = await price_inputs.count()
    if input_count == 0:
        price_inputs = page.locator('input[placeholder*="희망가"]')
        input_count = await price_inputs.count()
    if input_count == 0:
        # 추가 폴백: 가격 관련 input
        price_inputs = page.locator('input[type="text"]').filter(has=page.locator('..'))
        input_count = await price_inputs.count()
        print(f"  → 폴백: text input {input_count}개 발견")

    if input_count == 0:
        print(f"  ❌ 가격 입력 필드를 찾을 수 없음")
        await _save_debug_screenshot(page, f"bid_{product_id}_no_price_input")
        return False

    try:
        for i in range(input_count):
            inp = price_inputs.nth(i)
            await inp.click()
            await page.keyboard.press("Meta+a")
            await page.keyboard.press("Backspace")
            await inp.type(price, delay=50)
        print(f"  ✓ 판매 희망가: {price}원 ({input_count}개 입력)")
    except Exception as e:
        print(f"  ❌ 가격 입력 실패: {e}")
        await _save_debug_screenshot(page, f"bid_{product_id}_price_fail")
        return False

    # ── 4-1단계: 입찰기한 설정 (30/60/90일) ──
    try:
        deadline_label = f"{bid_days}일"
        # select 형태 드롭다운
        deadline_sel = page.locator('select').filter(has_text=re.compile(r'[369]0일'))
        if await deadline_sel.count() > 0:
            await deadline_sel.first.select_option(label=deadline_label)
            print(f"  ✓ 입찰기한 select: {bid_days}일")
        else:
            # 버튼/라디오 형태
            deadline_btn = page.locator(f'button:has-text("{deadline_label}"), label:has-text("{deadline_label}")')
            if await deadline_btn.count() > 0:
                await deadline_btn.first.click()
                print(f"  ✓ 입찰기한 버튼: {bid_days}일")
            else:
                print(f"  ℹ 입찰기한 선택 UI 없음 (기본값 사용)")
    except Exception as e:
        print(f"  ⚠ 입찰기한 설정 실패: {e}")

    # ── 5단계: 체크박스 선택 (동의 체크박스만 대상) ──
    try:
        checkboxes = page.locator('input[type="checkbox"]')
        cnt = await checkboxes.count()
        checked_count = 0
        print(f"  → 체크박스 {cnt}개 발견")

        for i in range(cnt):
            cb = checkboxes.nth(i)
            if not await cb.is_visible():
                continue
            if await cb.is_checked():
                checked_count += 1
                continue
            # React 체크박스는 부모 div를 클릭해야 state가 반영됨
            parent = page.locator(f'input[type="checkbox"] >> nth={i} >> xpath=..')
            try:
                await parent.click(timeout=2000)
                await page.wait_for_timeout(300)
                if await cb.is_checked():
                    checked_count += 1
            except Exception:
                # fallback: force click
                try:
                    await cb.click(force=True)
                    await page.wait_for_timeout(300)
                    if await cb.is_checked():
                        checked_count += 1
                except Exception:
                    pass

        await page.wait_for_timeout(500)
        print(f"  ✓ 체크박스 {checked_count}/{cnt}개 선택됨")
    except Exception as e:
        print(f"  ⚠ 체크박스 실패: {e}")

    # ── 6단계: 하단 "판매 입찰하기" 버튼 클릭 ──
    final_btn = page.locator('button:has-text("판매 입찰하기")').last
    try:
        is_disabled = await final_btn.is_disabled()
        print(f"  → 최종 버튼 상태: disabled={is_disabled}")
        if is_disabled:
            print(f"  ❌ '판매 입찰하기' 버튼이 비활성 — 필수 입력 누락 가능성")
            await _save_debug_screenshot(page, f"bid_{product_id}_btn_disabled")
            return False
        await final_btn.click(timeout=3000)
        await page.wait_for_timeout(2000)
        print(f"  ✓ '판매 입찰하기' 최종 클릭")
    except Exception as e:
        print(f"  ❌ 최종 버튼 실패: {e}")
        await _save_debug_screenshot(page, f"bid_{product_id}_final_btn_fail")
        return False

    # ── 7단계: 확인 팝업 ("총 N건의 판매 입찰하기" → 확인) ──
    try:
        confirm_btn = page.locator('button:has-text("확인")')
        await confirm_btn.wait_for(state="visible", timeout=3000)
        await confirm_btn.click()
        await page.wait_for_timeout(2000)
        print(f"  ✓ 확인 팝업 클릭")
    except Exception as e:
        print(f"  ⚠ 확인 팝업 없거나 실패: {e}")

    # ── 8단계: 입찰 신청 결과 확인 (성공 여부 판별) ──
    bid_success = False
    try:
        result = page.locator('text="입찰 신청 결과"')
        await result.wait_for(timeout=8000)
        print(f"  ✓ '입찰 신청 결과' 팝업 감지")

        # "성공 N건" 텍스트 확인
        try:
            result_text = await page.locator('text=/성공.*건/').first.inner_text(timeout=3000)
            print(f"  🎉 결과: {result_text}")
            if "성공" in result_text:
                bid_success = True
        except Exception:
            print(f"  ⚠ '성공 N건' 텍스트 못 찾음")

        # 결과 팝업 닫기
        close_btn = page.locator('button:has-text("확인")')
        if await close_btn.count() > 0:
            await close_btn.click()
            await page.wait_for_timeout(1000)
    except Exception:
        # 결과 팝업이 안 뜬 경우 — 에러 팝업이 있는지 확인
        print(f"  ⚠ '입찰 신청 결과' 팝업 미감지 — 에러 확인 중...")
        await _save_debug_screenshot(page, f"bid_{product_id}_no_result")
        try:
            # 에러 메시지 확인
            error_texts = await page.locator('[class*="error"], [class*="alert"], [class*="warning"]').all_text_contents()
            if error_texts:
                print(f"  ❌ 에러 메시지: {'; '.join(t.strip() for t in error_texts if t.strip())}")
        except Exception:
            pass

    if bid_success:
        print(f"  💾 입찰 등록 완료! #{product_id}, {size}, {price}원 × {qty}개")
    else:
        print(f"  ❌ 입찰 등록 실패 — 결과 확인 불가. #{product_id}")
        await _save_debug_screenshot(page, f"bid_{product_id}_failed")

    await page.wait_for_timeout(delay * 1000)
    return bid_success


# ═══════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════

async def main():
    parser = argparse.ArgumentParser(description="KREAM 판매자센터 자동화 v2")
    parser.add_argument("--mode", choices=["login", "login-kream", "product", "bid", "all"],
                        default="product",
                        help="login=판매자센터 로그인, login-kream=KREAM 로그인, product=고시정보, bid=입찰, all=전부")
    parser.add_argument("--excel", default=EXCEL_PATH, help="데이터 엑셀 파일 경로")
    args = parser.parse_args()

    print("=" * 50)
    print("🤖 KREAM 판매자센터 자동화 v2")
    print(f"   모드: {args.mode}")
    print(f"   엑셀: {args.excel}")
    print("=" * 50)

    async with async_playwright() as p:

        if args.mode == "login":
            await login_manual(p)
            return

        if args.mode == "login-kream":
            await login_kream(p)
            return

        settings = load_settings(args.excel)
        headless = str(settings.get("headless_mode", "false")).lower() == "true"
        delay = float(settings.get("delay_between_items", 3))

        browser = await create_browser(p, headless=headless)
        context = await create_context(browser, STATE_FILE)
        page = await context.new_page()
        await apply_stealth(page)

        if not await ensure_logged_in(page):
            await browser.close()
            return

        if args.mode in ("product", "all"):
            products = load_product_data(args.excel)
            print(f"\n📋 상품 고시정보 {len(products)}건 처리")
            for i, product in enumerate(products, 1):
                print(f"\n{'='*40} [{i}/{len(products)}] {'='*40}")
                try:
                    await fill_product_info(page, product, delay)
                except Exception as e:
                    print(f"  ❌ 오류: {e}")

        if args.mode in ("bid", "all"):
            bids = load_bid_data(args.excel)
            print(f"\n📋 판매 입찰 {len(bids)}건 처리")
            success, fail = 0, 0
            for i, bid in enumerate(bids, 1):
                print(f"\n{'='*40} [{i}/{len(bids)}] {'='*40}")
                try:
                    if await place_bid(page, bid, delay):
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    print(f"  ❌ 오류: {e}")
                    fail += 1
            print(f"\n📊 입찰 결과: 성공 {success}건, 실패 {fail}건")

        await context.storage_state(path=STATE_FILE)
        print("\n✅ 모든 작업 완료!")
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
